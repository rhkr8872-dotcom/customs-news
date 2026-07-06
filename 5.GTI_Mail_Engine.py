# -*- coding: utf-8 -*-
# GTI FINAL CORE v27 - Mail engine, 24h + original URL accuracy guard
"""
GTI STEP5 Mail Engine - LAW1/NEWSREST report quality form v27

Report form
-----------
1. 총평
2. Top3 Deep Analysis
3. Regulation
4. 주요뉴스

LAW1 rule: Regulation table uses only STEP4-1 output derived from 1-1.regulation_raw.xlsx; all other materials are News.

This step does not reselect STEP4 results. It keeps all selected regulation/news
items, then rewrites weak STEP4 text into an executive report style.
"""

from __future__ import annotations

import argparse
import html
import os
import re
import smtplib
import ssl
from datetime import datetime
from email.message import EmailMessage
from email.utils import formataddr
from pathlib import Path
from urllib.parse import unquote

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


REGULATION_INPUT_FILE = Path(os.getenv("GTI_REGULATION_INPUT_FILE", r"C:\Temp\4-1.regulation_ai_summary.xlsx"))
NEWS_INPUT_FILE = Path(os.getenv("GTI_NEWS_INPUT_FILE", r"C:\Temp\4-2.news_ai_summary.xlsx"))
OUTPUT_DIR = Path(os.getenv("GTI_OUTPUT_DIR", r"C:\Temp\12345\c_type_outputs"))
RUN_DATE = os.getenv("GTI_RUN_DATE", datetime.now().strftime("%Y-%m-%d"))

NEWS_MAX_ROWS = int(os.getenv("GTI_NEWS_MAX_ROWS", "0"))  # 0 = no cap
SEND_EMAIL = os.getenv("GTI_SEND_EMAIL", "Y").strip().upper() in {"Y", "YES", "TRUE", "1"}
SMTP_HOST = os.getenv("GTI_SMTP_HOST", "smtp.naver.com")
SMTP_PORT = int(os.getenv("GTI_SMTP_PORT", "465"))
SMTP_USER = os.getenv("GTI_SMTP_USER", "kch8872@naver.com").strip()
SMTP_PASS = (os.getenv("GTI_SMTP_PASS") or os.getenv("GTI_MAIL_PW") or "").strip()
MAIL_TO = os.getenv("GTI_MAIL_TO", "").strip()
MAIL_FROM_NAME = os.getenv("GTI_MAIL_FROM_NAME", "GTI Radar").strip()
RECIPIENT_FILE = Path(os.getenv("GTI_RECIPIENT_FILE", r"C:\Temp\00.xlsx"))


OUTPUT_COLUMNS = [
    "No", "Content Type", "Mail Group", "Samsung Impact", "Affected Subsidiary", "Impact Reason",
    "Date", "Headline", "Major Changes", "Summary", "AI Analysis", "Action Plan", "Country", "Agency",
    "Risk", "Importance Score", "Priority Group", "Issue", "Cluster", "URL", "Source", "Source File",
]

GROUP_REGULATION = "Regulation"
GROUP_NEWS = "주요뉴스"


def output_paths() -> dict[str, Path]:
    return {
        "analysis": OUTPUT_DIR / "4.news_ai_analysis.xlsx",
        "mail_xlsx": OUTPUT_DIR / f"[GTI Radar] Global Trade Intelligence({RUN_DATE}).xlsx",
        "mail_html": OUTPUT_DIR / f"[GTI Radar] Global Trade Intelligence({RUN_DATE}).html",
        "cumulative": OUTPUT_DIR / "gti_news_cumulative.xlsx",
    }


def clean(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return re.sub(r"\s+", " ", str(value)).strip()


def pick_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    lookup = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in lookup:
            return lookup[cand.lower()]
    return None


def safe_num(value) -> float:
    try:
        if value is None or pd.isna(value):
            return 0.0
    except Exception:
        pass
    text = re.sub(r"[^0-9.\-]", "", str(value))
    try:
        return float(text) if text else 0.0
    except Exception:
        return 0.0


def normalize_risk(value) -> str:
    raw = clean(value)
    low = raw.lower()
    if raw in {"상", "중", "하"}:
        return raw
    if low in {"high", "h", "red"}:
        return "상"
    if low in {"medium", "med", "m", "orange"}:
        return "중"
    if low in {"low", "l", "blue"}:
        return "하"
    return "중"


def risk_weight(value) -> int:
    return {"상": 300, "중": 150, "하": 0}.get(normalize_risk(value), 0)


def priority_weight(value) -> int:
    p = clean(value).upper()
    return {"CORE": 1000, "POLICY_WATCH": 850, "USABLE": 650, "REFERENCE": 300, "WATCH": 250}.get(p, 200)


def parse_date(value):
    dt = pd.to_datetime(value, errors="coerce")
    return pd.Timestamp.min if pd.isna(dt) else dt


def display_date(value) -> str:
    dt = pd.to_datetime(value, errors="coerce")
    if pd.isna(dt):
        return clean(value)[:16]
    if dt.hour == 0 and dt.minute == 0:
        return dt.strftime("%Y-%m-%d")
    return dt.strftime("%Y-%m-%d %H:%M")


def best_url_from_values(values) -> str:
    invalid = {
        "", "nan", "none", "null", "new", "https://new", "http://new",
        "https://news", "http://news", "https://news.google.com", "https://news.google.com/",
    }
    candidates: list[str] = []
    for value in values:
        text = clean(value)
        if not text:
            continue
        for item in [text] + re.findall(r"https?://[^'\"),\s]+", text):
            url = html.unescape(item).strip().strip("<>'\"").rstrip(".,);]}")
            if url.lower() in invalid:
                continue
            if re.match(r"^https?://", url, re.I) and url not in candidates:
                candidates.append(url)
    for url in candidates:
        low = url.lower()
        if "news.google.com/rss/articles/" not in low and "news.google.com/articles/" not in low:
            return url
    return candidates[0] if candidates else ""


def non_empty_hint(value: str) -> str:
    text = clean(value)
    if not text or text in {"본문에서 확인 불가", "nan", "None"}:
        return ""
    return text


def normalize_input(df: pd.DataFrame, content_type: str, source_file: Path) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    col_date = pick_col(df, ["Date", "date"])
    col_headline = pick_col(df, ["Headline", "Title", "headline"])
    col_country = pick_col(df, ["Country", "country"])
    col_agency = pick_col(df, ["Agency", "Publisher", "agency", "source"])
    col_risk = pick_col(df, ["Risk", "risk"])
    col_score = pick_col(df, ["Importance Score", "final_score", "samsung_score", "Score", "Importance"])
    col_priority = pick_col(df, ["Priority Group", "priority_group", "mail_section", "Tier"])
    col_issue = pick_col(df, ["Issue", "issue_type", "topic_keyword", "topic", "IssueKey"])
    col_cluster = pick_col(df, ["Cluster", "cluster_key", "ClusterHeadlines"])
    col_summary = pick_col(df, [
        "Summary", "summary", "ExecutiveMessage", "Executive Summary",
        "ArticleSummary", "article_summary", "PostSummary", "post_summary",
        "ContentSummary", "content_summary", "BodySummary", "body_summary",
        "본문요약", "요약", "뉴스요약", "원문요약",
        "Description", "description", "Snippet", "snippet",
        "News", "news", "본문", "content", "Content", "body", "Body",
        "raw_text", "RawText", "full_text", "FullText", "ArticleBody", "article_body"
    ])
    col_analysis = pick_col(df, ["AI Analysis", "analysis", "samsung_reason"])
    col_action = pick_col(df, ["Action Plan", "RequiredAction", "action"])
    col_source = pick_col(df, ["Source", "SourceFile", "source"])
    col_impact = pick_col(df, ["Samsung Impact", "samsung_impact"])
    col_subs = pick_col(df, ["Affected Subsidiary", "affected_subsidiary", "affected_subsidiaries"])
    col_reason = pick_col(df, ["Impact Reason", "subsidiary_reason", "samsung_reason", "SelectReason"])

    hint_cols = {
        "effective_date_hint": pick_col(df, ["effective_date_hint"]),
        "change_detail_hint": pick_col(df, ["change_detail_hint"]),
        "hs_hint": pick_col(df, ["hs_hint"]),
        "tariff_rate_hint": pick_col(df, ["tariff_rate_hint"]),
        "KeywordMatches": pick_col(df, ["KeywordMatches"]),
        "affected_products": pick_col(df, ["affected_products", "impact_products", "subsidiary_products"]),
        "fta_impact": pick_col(df, ["fta_impact"]),
        "export_control_impact": pick_col(df, ["export_control_impact"]),
        "hs_impact": pick_col(df, ["hs_impact"]),
        "tariff_impact": pick_col(df, ["tariff_impact"]),
    }

    url_cols = [
        pick_col(df, ["BestLinkURL"]),
        pick_col(df, ["OriginalURLCandidate"]),
        pick_col(df, ["original_url"]),
        pick_col(df, ["URL", "url", "Link"]),
        pick_col(df, ["GoogleURL"]),
        col_source,
    ]
    url_cols = [c for c in url_cols if c]

    out = pd.DataFrame()
    out["Date"] = df[col_date].apply(display_date) if col_date else ""
    out["_sort_date"] = df[col_date].apply(parse_date) if col_date else pd.Timestamp.min
    out["Headline"] = df[col_headline].apply(clean) if col_headline else ""
    out["Country"] = df[col_country].apply(clean) if col_country else ""
    out["Agency"] = df[col_agency].apply(clean) if col_agency else ""
    out["Risk"] = df[col_risk].apply(normalize_risk) if col_risk else "중"
    out["Importance Score"] = df[col_score].apply(safe_num) if col_score else 0
    out["Priority Group"] = df[col_priority].apply(lambda v: clean(v).upper()) if col_priority else ("CORE" if content_type == "Regulation" else "USABLE")
    out["Issue"] = df[col_issue].apply(clean) if col_issue else ""
    out["Issue"] = out["Issue"].replace({
        "TARIFF": "관세정책", "SECTION_301_232": "관세정책",
        "CUSTOMS": "통관", "CUSTOMS_CLEARANCE": "통관",
        "ORIGIN_FTA": "FTA/원산지", "CBAM_CARBON": "CBAM",
        "HS_CLASSIFICATION": "HS/품목분류", "AD_CVD": "AD/CVD",
        "EXPORT_CONTROL": "수출통제",
    })
    out["Cluster"] = df[col_cluster].apply(clean) if col_cluster else ""
    out["Summary"] = df[col_summary].apply(clean) if col_summary else ""
    out["AI Analysis"] = df[col_analysis].apply(clean) if col_analysis else ""
    out["Action Plan"] = df[col_action].apply(clean) if col_action else ""
    out["Samsung Impact"] = df[col_impact].apply(lambda v: clean(v).title() if clean(v).lower() in {"direct", "indirect", "watch"} else clean(v)) if col_impact else "Watch"
    out["Samsung Impact"] = out["Samsung Impact"].replace({"": "Watch", "직접": "Direct", "간접": "Indirect", "모니터링": "Watch"})
    out["Affected Subsidiary"] = df[col_subs].apply(clean) if col_subs else ""
    out["Impact Reason"] = df[col_reason].apply(clean) if col_reason else ""
    out["Source"] = df[col_source].apply(clean) if col_source else ""
    out["Source File"] = str(source_file)

    # Preserve possible original body/summary fields for STEP5 post summary.
    # These columns are not necessarily written to the final Excel, but are used
    # by major_changes() / Top3 Deep Analysis.
    extra_source_cols = {
        "Original Post Summary": pick_col(df, [
            "ArticleSummary", "article_summary", "PostSummary", "post_summary",
            "ContentSummary", "content_summary", "BodySummary", "body_summary",
            "본문요약", "뉴스요약", "원문요약", "Description", "description", "Snippet", "snippet"
        ]),
        "Original Body Text": pick_col(df, [
            "본문", "content", "Content", "body", "Body", "raw_text", "RawText",
            "full_text", "FullText", "ArticleBody", "article_body", "News", "news"
        ]),
    }
    for out_extra_col, src_extra_col in extra_source_cols.items():
        out[out_extra_col] = df[src_extra_col].apply(clean) if src_extra_col else ""

    out["Content Type"] = content_type
    out["Mail Group"] = GROUP_REGULATION if content_type == "Regulation" else GROUP_NEWS
    out["URL"] = df.apply(lambda r: best_url_from_values([r.get(c, "") for c in url_cols]), axis=1) if len(df) else ""

    for out_col, src_col in hint_cols.items():
        out[out_col] = df[src_col].apply(clean) if src_col else ""

    out = out[out["Headline"].astype(str).str.strip().ne("")]
    return out.reset_index(drop=True)


def read_step4_results() -> pd.DataFrame:
    frames = []
    if REGULATION_INPUT_FILE.exists():
        frames.append(normalize_input(pd.read_excel(REGULATION_INPUT_FILE), "Regulation", REGULATION_INPUT_FILE))
    if NEWS_INPUT_FILE.exists():
        news = normalize_input(pd.read_excel(NEWS_INPUT_FILE), "News", NEWS_INPUT_FILE)
        if NEWS_MAX_ROWS > 0:
            news = news.head(NEWS_MAX_ROWS)
        frames.append(news)
    if not frames:
        raise FileNotFoundError(f"STEP4 outputs not found: {REGULATION_INPUT_FILE}, {NEWS_INPUT_FILE}")

    rows = pd.concat(frames, ignore_index=True)
    rows["_dedup_key"] = rows.apply(
        lambda r: clean(r.get("URL")) or (clean(r.get("Headline"))[:160] + "|" + clean(r.get("Agency")) + "|" + clean(r.get("Date"))),
        axis=1,
    )
    rows = rows.drop_duplicates(subset=["_dedup_key"], keep="first").drop(columns=["_dedup_key"], errors="ignore")
    rows["_integrated_score"] = rows.apply(
        lambda r: priority_weight(r["Priority Group"]) + risk_weight(r["Risk"]) + (180 if r["Content Type"] == "Regulation" else 0) + safe_num(r["Importance Score"]),
        axis=1,
    )
    return rows.reset_index(drop=True)


def dedup_report_rows(rows: pd.DataFrame) -> pd.DataFrame:
    """Remove near-duplicate report items after STEP4 merge.

    This intentionally catches cases where the same policy appears through two
    official pages, such as CBAM certificate price "to be published" and
    "now available", or the same bonded warehouse notice from two boards.
    """
    if rows.empty:
        return rows
    rows = rows.copy()
    rows["_report_dedup_key"] = rows.apply(report_dedup_key, axis=1)
    rows["_dedup_rank"] = rows.apply(dedup_rank, axis=1)
    rows = rows.sort_values(["_dedup_rank", "_integrated_score", "_sort_date"], ascending=[False, False, False])
    rows = rows.drop_duplicates(subset=["_report_dedup_key"], keep="first")
    return rows.drop(columns=["_report_dedup_key", "_dedup_rank"], errors="ignore").reset_index(drop=True)


def report_dedup_key(row: pd.Series) -> str:
    issue = clean(row.get("Issue")) or issue_for(row)
    title = clean(row.get("Headline")).lower()
    source = clean(row.get("Agency")).lower()
    normalized = re.sub(r"\([^)]*\)|\[[^]]*\]", " ", title)
    normalized = re.sub(r"제출기한[:：]?\s*\d{4}[.\-/]\d{1,2}[.\-/]\d{1,2}.*", " ", normalized)
    normalized = re.sub(r"\b(to be published|now available|published|available|first)\b", " ", normalized)
    normalized = re.sub(r"[^0-9a-z가-힣]+", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()

    if "cbam" in normalized and "certificate price" in normalized:
        return "REG:CBAM_CERTIFICATE_PRICE"
    if "보세창고" in normalized and "특허" in normalized and "운영" in normalized:
        return "REG:BONDED_WAREHOUSE_LICENSE_OPERATION"
    if "환전영업자" in normalized and "관리" in normalized:
        return "REG:FX_BUSINESS_OPERATOR_MANAGEMENT"
    if "수입신고" in normalized and "가산세" in normalized:
        return "REG:IMPORT_DECLARATION_DELAY_SURCHARGE"

    tokens = [t for t in normalized.split() if len(t) >= 2]
    return f"{clean(row.get('Content Type'))}:{issue}:{' '.join(tokens[:9])}:{source[:24]}"


def dedup_rank(row: pd.Series) -> float:
    title = clean(row.get("Headline")).lower()
    rank = safe_num(row.get("Importance Score")) + priority_weight(row.get("Priority Group")) + risk_weight(row.get("Risk"))
    if "now available" in title or "successfully entered into force" in title:
        rank += 150
    if "to be published" in title or "reminder" in title:
        rank -= 80
    if "제출기한" in title:
        rank += 60
    if clean(row.get("Agency")).startswith("Korea Customs") or "관세청" in clean(row.get("Agency")):
        rank += 40
    return rank


def issue_for(row) -> str:
    issue = clean(row.get("Issue"))
    if issue and issue.lower() not in {"watch", "policy_watch", "usable", "core"}:
        return issue
    text = " ".join(clean(row.get(c)) for c in ["Headline", "Summary", "AI Analysis", "Action Plan", "KeywordMatches"]).lower()
    if any(k in text for k in ["section 301", "section 232", "tariff", "quota", "duty", "관세", "쿼터"]):
        return "관세정책"
    if any(k in text for k in ["anti-dumping", "antidumping", "countervailing", "ad/cvd", "반덤핑", "상계관세"]):
        return "AD/CVD"
    if any(k in text for k in ["cbam", "carbon border"]):
        return "CBAM"
    if any(k in text for k in ["fta", "cepa", "origin", "원산지"]):
        return "FTA/원산지"
    if any(k in text for k in ["export control", "entity list", "uflpa", "forced labor", "수출통제"]):
        return "수출통제"
    if any(k in text for k in ["hs code", "classification", "품목분류"]):
        return "HS/품목분류"
    if clean(row.get("Content Type")) == "Regulation":
        return "법규"
    return "Watch"


def prepare_rows(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    rows["Issue"] = rows.apply(issue_for, axis=1)
    rows = dedup_report_rows(rows)
    rows["Mail Group"] = rows["Content Type"].map({"Regulation": GROUP_REGULATION}).fillna(GROUP_NEWS)
    rows["Major Changes"] = rows.apply(major_changes, axis=1)
    rows["Summary"] = rows.apply(report_summary, axis=1)
    rows["AI Analysis"] = rows.apply(report_impact, axis=1)
    rows["Action Plan"] = rows.apply(report_action, axis=1)
    rows["_report_score"] = rows.apply(report_score, axis=1)
    rows = rows.sort_values(["_integrated_score", "_sort_date"], ascending=[False, False]).reset_index(drop=True)
    rows["No"] = range(1, len(rows) + 1)
    return rows


def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    pool = rows.copy()
    pool["_top3_score"] = pool.apply(top3_deep_score, axis=1)
    pool = pool.sort_values(["_top3_score", "_sort_date"], ascending=[False, False])
    selected = []
    used_issues = set()
    for _, row in pool.iterrows():
        issue = clean(row.get("Issue"))
        if issue in used_issues and len(selected) < 3:
            continue
        selected.append(row)
        used_issues.add(issue)
        if len(selected) == 3:
            break
    if len(selected) < 3:
        for _, row in pool.iterrows():
            if any(clean(row.get("Headline")) == clean(x.get("Headline")) for x in selected):
                continue
            selected.append(row)
            if len(selected) == 3:
                break
    out = pd.DataFrame(selected).reset_index(drop=True)
    if not out.empty:
        out["No"] = range(1, len(out) + 1)
    return out


def top3_deep_score(row: pd.Series) -> float:
    text = " ".join(clean(row.get(c)) for c in [
        "Headline", "Major Changes", "Summary", "AI Analysis", "Action Plan", "Issue"
    ]).lower()
    score = report_score(row)

    high_terms = [
        "anti-dumping", "antidumping", "countervailing", "ad/cvd", "반덤핑", "상계관세",
        "cbam", "carbon border", "탄소국경", "탄소세",
        "export control", "entity list", "forced labor", "uflpa", "수출통제", "강제노동",
        "section 301", "section 232", "tariff quota", "duty-free quota", "customs duty",
        "관세", "무관세", "쿼터", "원산지", "rules of origin", "hs code",
    ]
    medium_terms = [
        "fta", "cepa", "usmca", "통상협정", "fta", "통관", "보세", "신고", "classification",
    ]
    low_terms = [
        "수출 85.9", "수출입 현황", "잠정치", "재정적자", "refunds", "customs revenue",
        "브랜드", "주식", "전략회의", "칼럼", "market outlook",
    ]

    if any(t in text for t in high_terms):
        score += 1200
    if any(t in text for t in medium_terms):
        score += 450
    if clean(row.get("Content Type")) == "Regulation":
        score += 300
    if any(t in text for t in low_terms):
        score -= 900
    return score


def report_score(row: pd.Series) -> float:
    impact_weight = {"Direct": 2200, "Indirect": 900, "Watch": 0}.get(clean(row.get("Samsung Impact")), 0)
    type_weight = 350 if clean(row.get("Content Type")) == "Regulation" else 0
    issue_weight = {
        "관세정책": 500,
        "AD/CVD": 500,
        "반덤핑/상계관세": 500,
        "CBAM": 450,
        "수출통제": 450,
        "FTA/원산지": 350,
        "통관": 300,
        "통관/세관": 300,
        "HS/품목분류": 300,
    }.get(clean(row.get("Issue")), 150)
    return safe_num(row.get("Importance Score")) + priority_weight(row.get("Priority Group")) + risk_weight(row.get("Risk")) + impact_weight + type_weight + issue_weight


def hint_line(label: str, value: str) -> str:
    value = non_empty_hint(value)
    return f"{label}: {value}" if value else ""


def compact_parts(parts: list[str], fallback: str) -> str:
    parts = [p for p in parts if clean(p)]
    return "; ".join(parts) if parts else fallback


def major_changes(row: pd.Series) -> str:
    issue = clean(row.get("Issue"))
    headline = clean(row.get("Headline"))
    title_l = headline.lower()

    if "보세창고" in headline and "특허" in headline:
        return (
            "개정 사유: 자가용보세창고 특허요건 완화 및 불명확한 규정 보완 필요. "
            "주요 개정 내용: 자가용보세창고 반입 대상에 국제무역선·기 적재 자가화물 외 수리용 예비부분품 및 부속품 장치를 허용하고, "
            "관세법 제178조상 물품반입 정지기간을 오해 없이 적용할 수 있도록 규정을 명확화하는 내용입니다."
        )
    if "환전영업자" in headline and "관리" in headline:
        return (
            "주요 내용: 환전영업자의 등록·관리, 보고·자료제출, 영업장 운영 및 관세청 관리 기준과 관련된 고시입니다. "
            "해외출장·주재원·외환거래 지원 프로세스와 연결될 수 있어 실제 법인 업무 해당 여부 확인이 필요합니다."
        )
    if "cbam" in title_l and "certificate price" in title_l:
        return (
            "주요 내용: EU CBAM 인증서 가격이 공표되었거나 공표 일정이 확정된 사안입니다. "
            "EU 수입품의 내재배출량 신고, 인증서 구매 비용, 공급사 배출량 자료 확보 체계에 영향을 줄 수 있습니다."
        )
    if "customs enforcement" in title_l and "executive order" in title_l:
        return (
            "주요 내용: 미국 세관 집행 강화 행정명령 관련 사안입니다. "
            "수입신고 정확성, 저가신고·우회수입·전자상거래 물품 관리 및 CBP 심사 강화 가능성을 확인해야 합니다."
        )
    if "수입신고" in headline and "가산세" in headline:
        return (
            "주요 내용: 수입신고 지연 가산세 부과 대상이 되는 매점매석 금지 품목의 적용기간 연장 공고입니다. "
            "해당 품목 수입 시 신고 지연, 재고 운영, 통관 일정 관리 기준을 확인해야 합니다."
        )

    parts = [
        hint_line("시행/적용일", row.get("effective_date_hint")),
        hint_line("변경 내용", row.get("change_detail_hint")),
        hint_line("대상 HS", row.get("hs_hint")),
        hint_line("관세율/쿼터", row.get("tariff_rate_hint")),
        hint_line("키워드", row.get("KeywordMatches")),
    ]
    if any(parts):
        return compact_parts(parts, "")

    if issue == "관세정책":
        return "관세율, 쿼터, 면세/환급 또는 Section 301/232 등 관세 비용에 영향을 줄 수 있는 정책 변화입니다."
    if issue in {"AD/CVD", "반덤핑/상계관세"}:
        return "반덤핑 또는 상계관세 조사·판정·연장 가능성이 있는 사안입니다. 공급국, 대상 품목, 조사 기간과 관세율 확인이 필요합니다."
    if issue == "CBAM":
        return "CBAM 신고, 인증서 가격, 배출량 자료 또는 EU 수입통관 절차와 연결되는 탄소국경조정 변화입니다."
    if issue == "FTA/원산지":
        return "FTA/CEPA 협정, 원산지 기준, CO 발급 또는 특혜관세 적용 가능성에 영향을 주는 변화입니다."
    if issue == "수출통제":
        return "Entity List, ECCN, UFLPA, forced labor 또는 전략물자·제재 스크리닝 관련 변화입니다."
    if issue == "통관":
        return "보세, 통관, 신고, 세관 심사 또는 행정절차 기준에 영향을 줄 수 있는 공식 공지입니다."
    if issue == "HS/품목분류":
        return "HS 분류 기준 또는 품목 해석이 달라질 수 있어 품목 마스터와 신고 기준 점검이 필요한 사안입니다."
    return f"{headline} 관련 관세·통상 모니터링 사안입니다."


def report_summary(row: pd.Series) -> str:
    issue = clean(row.get("Issue"))
    country = clean(row.get("Country")) or "관련 국가"
    agency = clean(row.get("Agency")) or "관련 기관"
    change = major_changes(row)
    if clean(row.get("Content Type")) == "Regulation":
        return f"{agency}의 공식 법규/공지입니다. 핵심은 {change} 원문 기준으로 시행일, 적용 품목, HS, 세율 또는 신고 절차를 확인해야 합니다."
    return f"{country}에서 포착된 {issue} 뉴스입니다. 핵심은 {change} 삼성전자 관련 법인·품목에 직접 적용되는지 확인할 필요가 있습니다."


def report_impact(row: pd.Series) -> str:
    issue = clean(row.get("Issue"))
    impact = clean(row.get("Samsung Impact")) or "Watch"
    subs = clean(row.get("Affected Subsidiary")) or "SEC/HQ"
    products = non_empty_hint(row.get("affected_products"))
    product_txt = f" 대상 제품 후보는 {products}입니다." if products else ""
    if issue == "관세정책":
        return f"{subs} 기준 수입가격, 관세환급, 할당관세/쿼터, 공급국 선택에 영향을 줄 수 있습니다.{product_txt} Impact는 {impact}로 분류됩니다."
    if issue in {"AD/CVD", "반덤핑/상계관세"}:
        return f"{subs}의 철강·부품·원재료 조달에서 AD/CVD 추가관세 또는 조사 대응 자료 부담이 생길 수 있습니다.{product_txt} 공급국과 HS별 노출도를 확인해야 합니다."
    if issue == "CBAM":
        return f"{subs}의 EU향 판매·공급망에서 CBAM 신고자료, 배출량 증빙, 인증서 비용 관리가 필요할 수 있습니다.{product_txt}"
    if issue == "FTA/원산지":
        return f"{subs}의 FTA 활용, 원산지 판정, CO 발급, BOM 원산지 증빙 체계에 영향을 줄 수 있습니다.{product_txt}"
    if issue == "수출통제":
        return f"{subs}의 거래처 스크리닝, ECCN/전략물자 분류, 우회수출 통제와 연결될 수 있습니다.{product_txt}"
    if issue == "통관":
        return f"{subs}의 수입신고, 보세창고, 통관 심사, 세관 제출자료 운영 기준에 반영 여부를 확인해야 합니다.{product_txt}"
    if issue == "HS/품목분류":
        return f"{subs}의 HS 마스터, 품목 설명, 관세율 산정 및 신고 정확성에 영향을 줄 수 있습니다.{product_txt}"
    return f"{subs} 기준 관세·통상 리스크 모니터링 가치가 있습니다. Impact는 {impact}입니다."


def report_action(row: pd.Series) -> str:
    issue = clean(row.get("Issue"))
    subs = clean(row.get("Affected Subsidiary")) or "SEC/HQ"
    if issue == "관세정책":
        return f"{subs}: 대상 HS·공급국·거래금액을 매핑하고 세율/쿼터/환급 가능성을 산출해 관세비용 영향표에 반영하십시오."
    if issue in {"AD/CVD", "반덤핑/상계관세"}:
        return f"{subs}: 대상 품목·공급국·벤더를 확인하고 조사대응 자료, 원산지 증빙, 가격자료 보관 필요성을 점검하십시오."
    if issue == "CBAM":
        return f"{subs}: EU향 품목, 공급사 배출량 자료, CBAM 신고·인증서 비용 반영 여부를 ESG/구매/통관 담당과 확인하십시오."
    if issue == "FTA/원산지":
        return f"{subs}: BOM 원산지, CO 발급, 직접운송, 누적기준, 특혜세율 적용 가능성을 FTA 마스터와 대조하십시오."
    if issue == "수출통제":
        return f"{subs}: ECCN/전략물자 분류, 거래처·최종사용자 스크리닝, 제재국 우회거래 가능성을 재점검하십시오."
    if issue == "통관":
        return f"{subs}: 통관 SOP, 보세/신고 체크리스트, 관세사 안내문, 세관 제출자료 양식을 업데이트하십시오."
    if issue == "HS/품목분류":
        return f"{subs}: 관련 제품의 HS 설명, 판정 근거, 해외법인 신고코드와 한국 본사 마스터 간 차이를 점검하십시오."
    return f"{subs}: 원문 기준으로 대상 국가, 품목, 시행일, 담당 부서를 확인하고 후속 모니터링하십시오."


def html_link(title: str, url: str) -> str:
    title_e = html.escape(clean(title))
    url = best_url_from_values([url])
    if not url:
        return title_e
    return f'<a href="{html.escape(url)}" target="_blank">{title_e}</a>'


def risk_color(risk: str) -> str:
    return {"상": "#C00000", "중": "#C55A11", "하": "#4472C4"}.get(normalize_risk(risk), "#555")


def short_text(value, fallback: str, limit: int = 360) -> str:
    text = clean(value) or fallback
    return text[:limit] + ("..." if len(text) > limit else "")


def one_line(row: pd.Series) -> str:
    return f"{clean(row.get('Issue'))} / {clean(row.get('Country')) or '-'} / {clean(row.get('Samsung Impact'))}: {short_text(row.get('Major Changes'), '주요 변경내역 확인 필요', 130)}"


def top3_summary_sentence(row: pd.Series) -> str:
    issue = clean(row.get("Issue"))
    if issue in {"AD/CVD", "반덤핑/상계관세"}:
        return "반덤핑·상계관세 조사는 대상 품목의 HS, 원산지, 가격자료 방어체계를 우선 점검해야 합니다."
    if issue == "FTA/원산지":
        return "FTA·원산지 변경은 관련 법인의 CO 발급요건과 특혜세율 적용 가능성을 재검토해야 합니다."
    if issue in {"관세정책", "통관", "통관/세관", "HS/품목분류"}:
        return "관세·통상 정책 변화는 관련 품목의 HS, 원산지, 관세율 영향을 확인해야 합니다."
    if issue == "CBAM":
        return "CBAM 변화는 EU향 품목의 배출량 자료, 인증서 비용, 신고의무 반영 여부를 우선 점검해야 합니다."
    if issue == "수출통제":
        return "수출통제 변화는 ECCN·전략물자 분류와 거래처·최종사용자 스크리닝 체계를 우선 확인해야 합니다."
    return short_text(row.get("Major Changes"), "관세·통상 영향 여부를 원문 기준으로 확인해야 합니다.", 150)


def top3_summary_rows(rows: pd.DataFrame, top3: pd.DataFrame) -> pd.DataFrame:
    """Issue-level summary rows for the executive summary.

    The executive summary should be thematic and practical, not just repeat the
    three article titles. Prefer AD/CVD, FTA/origin, tariff/customs if present.
    """
    preferred_groups = [
        {"AD/CVD", "반덤핑/상계관세"},
        {"FTA/원산지", "ORIGIN", "원산지"},
        {"관세정책", "통관", "통관/세관", "HS/품목분류"},
        {"CBAM"},
        {"수출통제"},
    ]
    selected = []
    used = set()
    for group in preferred_groups:
        cand = rows[rows["Issue"].astype(str).isin(group)].copy()
        if cand.empty:
            continue
        cand = cand.sort_values(["_report_score", "_sort_date"], ascending=[False, False])
        row = cand.iloc[0]
        key = clean(row.get("Issue"))
        if key in used:
            continue
        selected.append(row)
        used.add(key)
        if len(selected) >= 3:
            break
    if len(selected) < 3:
        for _, row in top3.iterrows():
            key = clean(row.get("Issue"))
            if key in used:
                continue
            selected.append(row)
            used.add(key)
            if len(selected) >= 3:
                break
    return pd.DataFrame(selected)


def overall_html(rows: pd.DataFrame, top3: pd.DataFrame) -> str:
    reg = rows[rows["Content Type"].eq("Regulation")]
    news = rows[rows["Content Type"].eq("News")]
    direct = rows[rows["Samsung Impact"].eq("Direct")]
    indirect = rows[rows["Samsung Impact"].eq("Indirect")]
    watch = rows[rows["Samsung Impact"].eq("Watch")]
    issues = rows["Issue"].value_counts().head(6)
    issue_txt = ", ".join(f"{k} {v}건" for k, v in issues.items())
    summary_rows = top3_summary_rows(rows, top3)
    top_lines = "".join(f"<li>{html.escape(top3_summary_sentence(r))}</li>" for _, r in summary_rows.iterrows())
    return f"""
    <div style="padding:15px;background:#F4F6F8;border-left:6px solid #1F4E78;margin-bottom:18px;">
      <div style="font-size:14px;color:#555;margin-bottom:8px;">
        금일 선별 결과: 법규 {len(reg)}건, 주요뉴스 {len(news)}건 | Direct {len(direct)}건, Indirect {len(indirect)}건, Watch {len(watch)}건
      </div>
      <div style="font-size:15px;font-weight:bold;line-height:1.8;margin-bottom:8px;">
        금일 GTI Radar는 {html.escape(issue_txt)} 중심으로 관세·통상 변화가 포착되었습니다. 법규는 시행일·HS·세율·신고절차 반영 여부를, 뉴스는 실제 비용·원산지·수출통제 영향 가능성을 우선 확인해야 합니다.
      </div>
      <div style="margin-top:8px;"><b>Top3 요약</b><ol style="margin-top:6px;">{top_lines}</ol></div>
    </div>
    """


def top3_html(top3: pd.DataFrame) -> str:
    blocks = []
    for idx, row in top3.iterrows():
        blocks.append(f"""
        <div style="margin:14px 0 18px 0;padding:15px;border-left:5px solid #C00000;background:#FFF7F7;">
          <div style="font-size:15px;font-weight:bold;margin-bottom:6px;">Top {idx + 1}. {html_link(row.get('Headline'), row.get('URL'))}</div>
          <div style="font-size:12px;color:#555;margin-bottom:9px;">
            Type: {html.escape(clean(row.get('Content Type')))} | Topic: {html.escape(clean(row.get('Issue')))} |
            Samsung Impact: <b>{html.escape(clean(row.get('Samsung Impact')))}</b> |
            Subsidiary: {html.escape(clean(row.get('Affected Subsidiary')) or 'SEC/HQ')} |
            Agency: {html.escape(clean(row.get('Agency')))} | Publish Date: {html.escape(clean(row.get('Date')))} |
            Country: {html.escape(clean(row.get('Country')))} |
            Risk: <span style="color:{risk_color(row.get('Risk'))};font-weight:bold;">{html.escape(clean(row.get('Risk')))}</span> |
            Score: {safe_num(row.get('Importance Score')):.0f}
          </div>
          <div style="margin-top:8px;"><b>Executive Impact</b><br>{html.escape(one_line(row))}</div>
          <div style="margin-top:8px;"><b>주요 변경내역</b><br>{html.escape(short_text(row.get('Major Changes'), '주요 변경내역 확인 필요', 520))}</div>
          <div style="margin-top:8px;"><b>삼성 영향</b><br>{html.escape(short_text(row.get('AI Analysis'), '삼성 영향 검토 필요', 520))}</div>
          <div style="margin-top:8px;"><b>Action</b><br>{html.escape(short_text(row.get('Action Plan'), '담당 부서 확인 필요', 520))}</div>
        </div>
        """)
    return "".join(blocks)


def table_html(title: str, rows: pd.DataFrame, color: str) -> str:
    if rows.empty:
        return f"<h3 style='color:{color};'>{html.escape(title)} (0건)</h3>"
    trs = []
    for _, row in rows.iterrows():
        trs.append(f"""
        <tr>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;">{html.escape(str(row.get('No')))}</td>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;">{html.escape(clean(row.get('Issue')))}</td>
          <td style="padding:7px;border:1px solid #ddd;vertical-align:top;">{html_link(row.get('Headline'), row.get('URL'))}</td>
          <td style="padding:7px;border:1px solid #ddd;vertical-align:top;">{html.escape(short_text(row.get('Major Changes'), '주요 변경내역 확인 필요', 260))}</td>
          <td style="padding:7px;border:1px solid #ddd;vertical-align:top;">{html.escape(short_text(row.get('AI Analysis'), '영향 검토 필요', 260))}</td>
          <td style="padding:7px;border:1px solid #ddd;vertical-align:top;">{html.escape(short_text(row.get('Action Plan'), '담당 부서 확인 필요', 260))}</td>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;">{html.escape(clean(row.get('Country')))}</td>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;color:{risk_color(row.get('Risk'))};font-weight:bold;">{html.escape(clean(row.get('Risk')))}</td>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;">{html.escape(clean(row.get('Samsung Impact')))}</td>
        </tr>
        """)
    return f"""
    <h3 style="margin-top:24px;color:{color};">{html.escape(title)} ({len(rows)}건)</h3>
    <table style="border-collapse:collapse;width:100%;font-size:12px;table-layout:fixed;">
      <colgroup>
        <col style="width:3%;"><col style="width:8%;"><col style="width:22%;">
        <col style="width:22%;"><col style="width:18%;"><col style="width:18%;">
        <col style="width:5%;"><col style="width:4%;"><col style="width:5%;">
      </colgroup>
      <thead>
        <tr style="background:{color};color:white;">
          <th style="padding:7px;border:1px solid #ddd;">No</th>
          <th style="padding:7px;border:1px solid #ddd;">Issue</th>
          <th style="padding:7px;border:1px solid #ddd;">Headline</th>
          <th style="padding:7px;border:1px solid #ddd;">주요 변경내역</th>
          <th style="padding:7px;border:1px solid #ddd;">삼성 영향</th>
          <th style="padding:7px;border:1px solid #ddd;">Action</th>
          <th style="padding:7px;border:1px solid #ddd;">Country</th>
          <th style="padding:7px;border:1px solid #ddd;">Risk</th>
          <th style="padding:7px;border:1px solid #ddd;">Impact</th>
        </tr>
      </thead>
      <tbody>{''.join(trs)}</tbody>
    </table>
    """


def build_html(rows: pd.DataFrame, top3: pd.DataFrame) -> str:
    subject = f"[GTI Radar] Global Trade Intelligence | {RUN_DATE}"
    regulation = rows[rows["Content Type"].eq("Regulation")]
    news = rows[rows["Content Type"].eq("News")]
    return f"""<!DOCTYPE html>
<html lang="ko">
<head><meta charset="utf-8"><title>{html.escape(subject)}</title></head>
<body style="font-family:Arial,'Malgun Gothic',sans-serif;font-size:13px;color:#222;line-height:1.55;">
  <div style="max-width:1320px;margin:0 auto;">
    <h2 style="margin-bottom:3px;color:#1F4E78;">[GTI Radar] Global Trade Intelligence</h2>
    <div style="font-size:13px;color:#555;margin-bottom:16px;">{RUN_DATE} | Samsung Electronics Customs & Trade Intelligence</div>

    <h3 style="margin-top:18px;margin-bottom:6px;">1. 총평</h3>
    {overall_html(rows, top3)}

    <h3 style="margin-top:22px;color:#C00000;">2. Top3 Deep Analysis</h3>
    {top3_html(top3)}

    {table_html('3. Regulation', regulation, '#1F4E78')}
    {table_html('4. 주요뉴스', news, '#548235')}

    <p style="margin-top:18px;color:#666;font-size:12px;">첨부 Excel에는 전체 선별 결과와 원문 링크가 포함되어 있습니다.</p>
  </div>
</body>
</html>"""


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    widths = {
        "A": 5, "B": 13, "C": 13, "D": 13, "E": 18, "F": 30, "G": 15, "H": 48,
        "I": 46, "J": 46, "K": 46, "L": 46, "M": 14, "N": 18, "O": 8, "P": 12,
        "Q": 16, "R": 16, "S": 24, "T": 36, "U": 24, "V": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"


def append_output_row(ws, row: pd.Series) -> None:
    ws.append([row.get(c, "") for c in OUTPUT_COLUMNS])
    headline_col = OUTPUT_COLUMNS.index("Headline") + 1
    cell = ws.cell(row=ws.max_row, column=headline_col)
    url = best_url_from_values([row.get("URL")])
    if url:
        cell.hyperlink = url
        cell.font = Font(color="0563C1", underline="single", bold=True)


def save_excel(rows: pd.DataFrame, top3: pd.DataFrame, paths: dict[str, Path]) -> None:
    wb = Workbook()
    sheets = [
        ("GTI Radar", rows),
        ("Top3 Deep Analysis", top3),
        ("Regulation", rows[rows["Content Type"].eq("Regulation")]),
        ("주요뉴스", rows[rows["Content Type"].eq("News")]),
    ]
    first = True
    for name, frame in sheets:
        ws = wb.active if first else wb.create_sheet(name[:31])
        first = False
        ws.title = name[:31]
        ws.append(OUTPUT_COLUMNS)
        for _, row in frame.iterrows():
            append_output_row(ws, row)
        style_sheet(ws)

    runlog = wb.create_sheet("Run Log")
    runlog.append(["item", "value"])
    runlog.append(["regulation_input", str(REGULATION_INPUT_FILE)])
    runlog.append(["news_input", str(NEWS_INPUT_FILE)])
    runlog.append(["run_date", RUN_DATE])
    runlog.append(["total_rows", len(rows)])
    runlog.append(["regulation_rows", int(rows["Content Type"].eq("Regulation").sum())])
    runlog.append(["news_rows", int(rows["Content Type"].eq("News").sum())])
    runlog.append(["direct_rows", int(rows["Samsung Impact"].eq("Direct").sum())])
    runlog.append(["indirect_rows", int(rows["Samsung Impact"].eq("Indirect").sum())])
    runlog.append(["watch_rows", int(rows["Samsung Impact"].eq("Watch").sum())])
    style_sheet(runlog)

    wb.save(paths["mail_xlsx"])
    wb.save(paths["analysis"])
    rows[OUTPUT_COLUMNS].to_excel(paths["cumulative"], index=False)


def read_recipients() -> list[str]:
    recipients = []
    if MAIL_TO:
        recipients.extend([x.strip() for x in re.split(r"[;,]", MAIL_TO) if x.strip()])
    if RECIPIENT_FILE.exists():
        try:
            df = pd.read_excel(RECIPIENT_FILE)
            for col in df.columns:
                for value in df[col].dropna().astype(str):
                    text = clean(value)
                    if re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", text):
                        recipients.append(text)
        except Exception:
            pass
    seen, out = set(), []
    for email in recipients:
        low = email.lower()
        if low not in seen:
            seen.add(low)
            out.append(email)
    return out


def send_email(html_body: str, attachment: Path) -> None:
    if not SEND_EMAIL:
        print("[MAIL SKIP] GTI_SEND_EMAIL=N or --no-email")
        return
    recipients = read_recipients()
    if not recipients:
        print("[MAIL SKIP] recipients missing")
        return
    if not SMTP_USER or not SMTP_PASS:
        print("[MAIL SKIP] SMTP credential missing")
        return

    msg = EmailMessage()
    msg["Subject"] = f"[GTI Radar] Global Trade Intelligence({RUN_DATE})"
    msg["From"] = formataddr((MAIL_FROM_NAME, SMTP_USER))
    msg["To"] = ", ".join(recipients)
    msg.set_content("GTI Radar report is attached. HTML mail requires an HTML-capable client.")
    msg.add_alternative(html_body, subtype="html")
    data = attachment.read_bytes()
    msg.add_attachment(
        data,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=attachment.name,
    )

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, context=context, timeout=30) as server:
        server.login(SMTP_USER, SMTP_PASS)
        server.send_message(msg)
    print(f"[MAIL SENT] {len(recipients)} recipients")



# ======================================================================
# GTI STEP5 Executive Quality Patch v3
# ----------------------------------------------------------------------
# Purpose
# 1) Re-rank Top3 by Samsung relevance + customs actionability, not keywords only
# 2) Auto-demote low-relevance items such as wheat/agriculture/general economy to REFERENCE
# 3) Expand Top3 Deep Analysis into issue summary / Samsung impact / customs impact / risk / action / owner
# 4) Keep STEP4 selected rows, but improve mail report quality at STEP5
# ======================================================================

SAMSUNG_RELEVANCE_TERMS = {
    "high": [
        "semiconductor", "chip", "memory", "dram", "nand", "hbm", "foundry", "wafer",
        "반도체", "메모리", "파운드리", "웨이퍼",
        "display", "oled", "lcd", "디스플레이",
        "battery", "cell", "cathode", "anode", "lithium", "nickel", "cobalt", "graphite",
        "배터리", "양극재", "음극재", "리튬", "니켈", "코발트", "흑연",
        "electronics", "smartphone", "mobile", "tv", "appliance", "home appliance",
        "전자", "스마트폰", "모바일", "가전", "tv",
        "pcb", "substrate", "module", "camera module", "sensor", "mlcc",
        "기판", "모듈", "센서", "mlcc",
        "rare earth", "gallium", "germanium", "gan", "silicon carbide", "sic",
        "희토류", "갈륨", "게르마늄", "전략물자",
        "steel", "aluminum", "copper", "zinc", "cold-rolled", "galvanized",
        "철강", "알루미늄", "구리", "아연", "냉간압연", "도금강판",
    ],
    "medium": [
        "customs", "tariff", "duty", "origin", "fta", "cbam", "ad/cvd", "anti-dumping",
        "countervailing", "hs code", "classification", "importer", "export control",
        "forced labor", "uflpa", "section 301", "section 232", "usmca", "cepa",
        "관세", "통관", "원산지", "수출통제", "강제노동", "반덤핑", "상계관세",
        "품목분류", "수입자", "보세", "수입신고", "수출신고",
    ],
}

LOW_RELEVANCE_TERMS = [
    "wheat", "rice", "corn", "soybean", "sugar", "agriculture", "agricultural",
    "livestock", "pork", "beef", "fishery", "food", "grain", "flour",
    "밀", "쌀", "옥수수", "농산물", "농업", "축산", "식품", "곡물", "밀가루",
    "염소산업", "혈통관리", "소비자물가", "주식", "배당", "concert", "sports",
]

GENERAL_NEWS_TERMS = [
    "수출 85.9", "수출입 현황", "증시", "주가", "배당", "실적", "gdp", "환율",
    "market outlook", "stock", "shares", "dividend", "budget deficit",
]

CRITICAL_CUSTOMS_TERMS = [
    "anti-dumping", "antidumping", "countervailing", "ad/cvd", "반덤핑", "상계관세",
    "cbam", "carbon border", "탄소국경",
    "customs enforcement", "cbp", "ior", "importer of record", "bond",
    "관세청", "세관", "수입신고", "보세", "통관",
    "export control", "entity list", "strategic goods", "dual-use", "수출통제", "전략물자",
    "forced labor", "uflpa", "강제노동",
    "hs code", "classification", "품목분류", "hscode",
    "section 301", "section 232", "tariff", "관세율", "쿼터", "환급",
    "fta", "origin", "rules of origin", "certificate of origin", "원산지", "co 발급",
]

def _row_text(row: pd.Series, cols: list[str] | None = None) -> str:
    if cols is None:
        cols = [
            "Headline", "Summary", "AI Analysis", "Action Plan", "Major Changes", "Issue",
            "Country", "Agency", "KeywordMatches", "affected_products", "hs_hint",
            "tariff_rate_hint", "fta_impact", "export_control_impact", "hs_impact", "tariff_impact",
            "Impact Reason", "Affected Subsidiary",
        ]
    return " ".join(clean(row.get(c)) for c in cols).lower()

def _has_any(text: str, terms: list[str]) -> bool:
    return any(t.lower() in text for t in terms)

def samsung_relevance_score(row: pd.Series) -> int:
    """Samsung product / subsidiary relevance score.

    This prevents generic trade news from being promoted to Top3 only because it
    contains words like export, tariff, or FTA.
    """
    text = _row_text(row)
    score = 0

    if clean(row.get("Samsung Impact")) == "Direct":
        score += 2200
    elif clean(row.get("Samsung Impact")) == "Indirect":
        score += 1000
    elif clean(row.get("Samsung Impact")) == "Watch":
        score += 200

    if _has_any(text, SAMSUNG_RELEVANCE_TERMS["high"]):
        score += 1400
    if _has_any(text, SAMSUNG_RELEVANCE_TERMS["medium"]):
        score += 500

    subs = clean(row.get("Affected Subsidiary"))
    if subs and subs not in {"관련 법인 검토", "SEC/HQ", "HQ", "본사"}:
        score += 400
    products = clean(row.get("affected_products"))
    if products:
        score += 500

    if _has_any(text, LOW_RELEVANCE_TERMS):
        score -= 1600
    if _has_any(text, GENERAL_NEWS_TERMS):
        score -= 800

    return score

def customs_actionability_score(row: pd.Series) -> int:
    """Score whether the item requires customs/trade compliance action."""
    text = _row_text(row)
    issue = clean(row.get("Issue"))
    score = 0

    if issue in {"관세정책", "AD/CVD", "반덤핑/상계관세", "CBAM", "수출통제", "FTA/원산지", "통관", "통관/세관", "HS/품목분류"}:
        score += 700
    if clean(row.get("Content Type")) == "Regulation":
        score += 500
    if normalize_risk(row.get("Risk")) == "상":
        score += 500
    elif normalize_risk(row.get("Risk")) == "중":
        score += 200

    if _has_any(text, CRITICAL_CUSTOMS_TERMS):
        score += 900
    for col in ["hs_hint", "tariff_rate_hint", "effective_date_hint", "change_detail_hint"]:
        if non_empty_hint(row.get(col)):
            score += 250

    # Actionable only when there is a plausible internal task.
    if any(k in text for k in ["시행", "적용", "신고", "세율", "hs", "관세율", "원산지", "co ", "환급", "증빙", "허가", "license"]):
        score += 300

    if _has_any(text, LOW_RELEVANCE_TERMS):
        score -= 1200

    return score

def reference_reason(row: pd.Series) -> str:
    text = _row_text(row)
    if _has_any(text, LOW_RELEVANCE_TERMS):
        return "삼성전자 주요 제품·부품·원재료와 직접 관련성이 낮은 일반 품목/농산물성 규제입니다."
    if samsung_relevance_score(row) < 200 and customs_actionability_score(row) < 700:
        return "관세업무 실행 조치가 필요한 수준의 HS·세율·원산지·신고절차 변경이 확인되지 않았습니다."
    if _has_any(text, GENERAL_NEWS_TERMS):
        return "일반 경제/시장 동향 성격이 강해 임원 보고 Top3보다는 참고 모니터링에 적합합니다."
    return ""

def executive_priority(row: pd.Series) -> str:
    """CORE / WATCH / REFERENCE override for STEP5 mail quality."""
    ref = reference_reason(row)
    if ref:
        return "REFERENCE"
    srel = samsung_relevance_score(row)
    act = customs_actionability_score(row)
    if srel >= 1800 and act >= 1400:
        return "CORE"
    if act >= 1400:
        return "POLICY_WATCH"
    if srel >= 1000 and act >= 900:
        return "USABLE"
    return clean(row.get("Priority Group")) or "WATCH"

def report_score(row: pd.Series) -> float:
    """Override: report score based on relevance/actionability, not raw keyword score."""
    base = safe_num(row.get("Importance Score"))
    impact_weight = {"Direct": 2200, "Indirect": 900, "Watch": 100, "Reference": -800}.get(clean(row.get("Samsung Impact")), 0)
    risk = risk_weight(row.get("Risk"))
    type_weight = 350 if clean(row.get("Content Type")) == "Regulation" else 0
    priority = executive_priority(row)
    pri_weight = {"CORE": 1200, "POLICY_WATCH": 800, "USABLE": 450, "WATCH": 200, "REFERENCE": -1200}.get(priority, 0)
    return base + impact_weight + risk + type_weight + pri_weight + samsung_relevance_score(row) + customs_actionability_score(row)

def top3_deep_score(row: pd.Series) -> float:
    """Override: choose Top3 only when Samsung relevance + customs actionability are high."""
    score = report_score(row)
    text = _row_text(row)
    priority = executive_priority(row)

    if priority == "REFERENCE":
        score -= 10000
    if clean(row.get("Content Type")) == "Regulation":
        score += 300
    if normalize_risk(row.get("Risk")) == "상":
        score += 300

    # Strong boost for issues that usually require HQ action.
    if clean(row.get("Issue")) in {"AD/CVD", "반덤핑/상계관세", "수출통제", "CBAM", "관세정책", "HS/품목분류", "통관"}:
        score += 700

    # Explicitly avoid "export" only items without Samsung/customs action.
    if "export" in text and not _has_any(text, CRITICAL_CUSTOMS_TERMS) and samsung_relevance_score(row) < 800:
        score -= 2500

    return score

def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    """Override: exclude REFERENCE/low-actionability from Top3 and keep issue diversity."""
    pool = rows.copy()
    pool["Executive Priority"] = pool.apply(executive_priority, axis=1)
    pool["_top3_score"] = pool.apply(top3_deep_score, axis=1)

    # Top3 후보 필터: Reference 제외 + 업무 실행성 최소값
    candidate = pool[
        (pool["Executive Priority"].ne("REFERENCE")) &
        (pool.apply(customs_actionability_score, axis=1) >= 700)
    ].copy()

    if candidate.empty:
        candidate = pool[pool["Executive Priority"].ne("REFERENCE")].copy()
    if candidate.empty:
        candidate = pool.copy()

    candidate = candidate.sort_values(["_top3_score", "_sort_date"], ascending=[False, False])
    selected = []
    used_issues = set()
    for _, row in candidate.iterrows():
        issue = clean(row.get("Issue"))
        if issue in used_issues and len(selected) < 3:
            continue
        selected.append(row)
        used_issues.add(issue)
        if len(selected) == 3:
            break

    if len(selected) < 3:
        for _, row in candidate.iterrows():
            if any(clean(row.get("Headline")) == clean(x.get("Headline")) for x in selected):
                continue
            selected.append(row)
            if len(selected) == 3:
                break

    out = pd.DataFrame(selected).reset_index(drop=True)
    if not out.empty:
        out["No"] = range(1, len(out) + 1)
    return out

def prepare_rows(rows: pd.DataFrame) -> pd.DataFrame:
    """Override: rewrite report fields and demote low relevance rows at STEP5."""
    rows = rows.copy()
    rows["Issue"] = rows.apply(issue_for, axis=1)
    rows = dedup_report_rows(rows)
    rows["Mail Group"] = rows["Content Type"].map({"Regulation": GROUP_REGULATION}).fillna(GROUP_NEWS)
    rows["Executive Priority"] = rows.apply(executive_priority, axis=1)

    # Demote low relevance items so they remain in report table but do not become Top3.
    rows.loc[rows["Executive Priority"].eq("REFERENCE"), "Priority Group"] = "REFERENCE"
    rows.loc[rows["Executive Priority"].eq("REFERENCE"), "Samsung Impact"] = "Reference"

    rows["Major Changes"] = rows.apply(major_changes, axis=1)
    rows["Summary"] = rows.apply(report_summary, axis=1)
    rows["AI Analysis"] = rows.apply(report_impact, axis=1)
    rows["Action Plan"] = rows.apply(report_action, axis=1)
    rows["_report_score"] = rows.apply(report_score, axis=1)

    # Final mail ordering should reflect executive relevance.
    rows = rows.sort_values(["_report_score", "_sort_date"], ascending=[False, False]).reset_index(drop=True)
    rows["No"] = range(1, len(rows) + 1)
    return rows

def _issue_summary_detail(row: pd.Series) -> str:
    issue = clean(row.get("Issue"))
    country = clean(row.get("Country")) or "확인 필요"
    date = clean(row.get("Date")) or "확인 필요"
    hs = non_empty_hint(row.get("hs_hint")) or "원문/법인 품목 기준 확인 필요"
    rate = non_empty_hint(row.get("tariff_rate_hint")) or "해당 시 별도 산출 필요"
    change = short_text(row.get("Major Changes"), "주요 변경내역 확인 필요", 420)
    return (
        f"• 이슈구분: {issue}\n"
        f"• 대상국가: {country}\n"
        f"• 게시/시행일: {date}\n"
        f"• 대상 HS/품목: {hs}\n"
        f"• 세율/쿼터/허가 변화: {rate}\n"
        f"• 핵심내용: {change}"
    )

def _samsung_impact_detail(row: pd.Series) -> str:
    issue = clean(row.get("Issue"))
    subs = clean(row.get("Affected Subsidiary")) or "SEC/HQ 및 관련 해외법인"
    products = non_empty_hint(row.get("affected_products")) or "법인별 수입·수출 실적 기준 매핑 필요"
    impact = clean(row.get("Samsung Impact")) or "Watch"

    if impact == "Reference":
        ref = reference_reason(row)
        return (
            f"• 영향등급: Reference\n"
            f"• 판단사유: {ref or '삼성전자 직접 영향은 낮고 정책 방향성 모니터링 가치 중심입니다.'}\n"
            f"• 영향법인: 즉시 특정 불필요\n"
            f"• 영향품목: {products}"
        )

    base = [
        f"• 영향등급: {impact}",
        f"• 영향법인 후보: {subs}",
        f"• 영향제품 후보: {products}",
    ]

    if issue in {"AD/CVD", "반덤핑/상계관세"}:
        base.append("• 영향업무: 공급국 변경, 원산지 증빙, 가격자료 방어, AD/CVD 추가관세 비용 산출")
        base.append("• 리스크: 대상 HS를 사용하는 원재료·부품 수입 시 추가 관세 및 사후심사 대응 부담 발생 가능")
    elif issue == "CBAM":
        base.append("• 영향업무: EU향 품목의 배출량 자료 확보, CBAM 신고, 인증서 비용 반영")
        base.append("• 리스크: 공급사 배출량 자료 미확보 시 신고 오류 또는 비용 추정 누락 가능")
    elif issue == "수출통제":
        base.append("• 영향업무: ECCN/전략물자 분류, 최종사용자 확인, 우회수출 스크리닝")
        base.append("• 리스크: 허가 필요 품목을 무허가 수출하거나 제재 거래처와 거래할 가능성")
    elif issue == "FTA/원산지":
        base.append("• 영향업무: BOM 원산지 판정, CO 발급요건, 직접운송, 누적기준, FTA Master 정합성")
        base.append("• 리스크: 원산지 기준 미충족 상태에서 특혜세율 적용 또는 CO 발급 오류 가능")
    elif issue in {"통관", "통관/세관"}:
        base.append("• 영향업무: 수입신고, 보세운송/보세공장, 관세사 제출자료, 통관 SOP")
        base.append("• 리스크: 신고 지연, 자동수리 조건 오류, 세관 제출자료 누락 가능")
    elif issue == "HS/품목분류":
        base.append("• 영향업무: HS Master, 품목 설명, 관세율, FTA 판정 기준")
        base.append("• 리스크: 법인별 HS 불일치 및 관세율 오적용 가능")
    else:
        base.append("• 영향업무: 관련 국가·품목 기준 통상 리스크 모니터링")
        base.append("• 리스크: 현재 직접 영향은 제한적이나 정책 확산 여부 확인 필요")

    return "\n".join(base)

def _customs_impact_detail(row: pd.Series) -> str:
    issue = clean(row.get("Issue"))
    lines = []
    if issue in {"AD/CVD", "반덤핑/상계관세"}:
        lines = [
            "• 관세비용: 대상 HS 수입금액 × 추가관세율로 잠재 비용 산출 필요",
            "• 신고리스크: 공급국·원산지·제조자 기준이 불명확하면 AD/CVD 회피 의심 가능",
            "• 증빙자료: 구매계약서, 원산지증명, 가격결정자료, 공급자 진술서 보관 필요",
        ]
    elif issue == "CBAM":
        lines = [
            "• 관세/준조세 비용: CBAM 인증서 가격 및 내재배출량 기준 비용 반영 필요",
            "• 신고리스크: EU 수입자 신고자료와 공급사 배출량 자료 불일치 가능",
            "• 증빙자료: 배출량 산정서, 공급사 확인서, 품목별 CN/HS 매핑 필요",
        ]
    elif issue == "수출통제":
        lines = [
            "• 수출허가: ECCN/전략물자 해당 여부 및 최종사용자 확인 필요",
            "• 거래심사: 제재국·제재자·우회수출 경로 스크리닝 필요",
            "• 시스템: Item Master에 전략물자/ECCN/허가필요 여부 필드 반영 검토",
        ]
    elif issue == "FTA/원산지":
        lines = [
            "• FTA 비용: 특혜세율 적용 가능 여부 및 미적용 시 관세비용 차이 산출 필요",
            "• 원산지 리스크: BOM, Vendor 원산지확인서, HS 기준 불일치 가능",
            "• 시스템: FTA Master·HS Master·Item Master 정합성 점검 필요",
        ]
    elif issue in {"통관", "통관/세관"}:
        lines = [
            "• 신고절차: 관세사 신고 양식, 제출자료, 자동수리 조건 변경 여부 확인 필요",
            "• 운영리스크: 보세·수입신고 오류 또는 지연 시 비용/가산세 발생 가능",
            "• 시스템: 통관 체크리스트와 법인 SOP 업데이트 필요",
        ]
    elif issue == "HS/품목분류":
        lines = [
            "• HS 리스크: 동일 품목에 대한 법인·관세사별 HS 불일치 가능",
            "• 비용영향: HS 변경 시 기본세율, FTA 세율, AD/CVD 적용 여부 재산정 필요",
            "• 시스템: HS Master 변경 승인 Workflow 필요",
        ]
    else:
        lines = [
            "• 직접 관세비용 영향은 현재 낮음",
            "• 정책 방향성 모니터링 후 유사 규제가 전자부품·전략물자로 확대되는지 확인 필요",
        ]
    return "\n".join(lines)

def _action_detail(row: pd.Series) -> str:
    impact = clean(row.get("Samsung Impact"))
    issue = clean(row.get("Issue"))
    subs = clean(row.get("Affected Subsidiary")) or "SEC/HQ"

    if impact == "Reference":
        return (
            "• 즉시조치: 불필요\n"
            "• 모니터링: 동일 국가에서 전자부품·전략물자·관세율 관련 후속 공지가 나오는지 확인\n"
            "• GTI 처리: 본문 Top3 제외, Reference 뉴스로 보관\n"
            "• Owner: GTI 운영자"
        )

    if issue in {"AD/CVD", "반덤핑/상계관세"}:
        return (
            f"• 즉시조치: {subs} 대상 HS·공급국·벤더 매핑\n"
            "• 1주 내: 최근 12개월 수입실적 기준 잠재 AD/CVD 비용 산출\n"
            "• 1개월 내: 원산지/가격자료 방어 파일 구축 및 관세사 신고 기준 공유\n"
            "• Owner: HQ Customs + 구매 + 해당 법인 통관담당"
        )
    if issue == "CBAM":
        return (
            f"• 즉시조치: {subs} EU향 대상품목 및 공급사 배출량 자료 보유 여부 확인\n"
            "• 1주 내: CBAM 신고 대상 CN/HS와 공급사별 배출량 Gap List 작성\n"
            "• 1개월 내: 인증서 비용 반영 로직 및 ESG/통관 공동관리 체계 수립\n"
            "• Owner: HQ Customs + ESG + EU 판매법인"
        )
    if issue == "수출통제":
        return (
            f"• 즉시조치: {subs} 대상 제품의 ECCN/전략물자 분류 확인\n"
            "• 1주 내: 거래처·최종사용자·목적지 스크리닝 결과 재점검\n"
            "• 1개월 내: Item Master에 Export Control Flag 반영\n"
            "• Owner: HQ Export Control + 사업부 + 해외법인"
        )
    if issue == "FTA/원산지":
        return (
            f"• 즉시조치: {subs} 대상 품목의 FTA 적용 여부와 CO 발급/수취 현황 확인\n"
            "• 1주 내: BOM 원산지, Vendor 원산지확인서, HS 기준 일치 여부 점검\n"
            "• 1개월 내: FTA Master·HS Master·Item Master 업데이트\n"
            "• Owner: HQ Customs/FTA + 법인 구매/물류"
        )
    if issue in {"통관", "통관/세관"}:
        return (
            f"• 즉시조치: {subs} 관세사에 신고절차 변경 여부 확인\n"
            "• 1주 내: 통관 SOP, 보세운송/보세공장 체크리스트, 제출자료 양식 개정\n"
            "• 1개월 내: ERP/ONE-Origin 반영 필요 필드 정의\n"
            "• Owner: HQ Customs + 법인 통관담당 + 관세사"
        )
    if issue == "HS/품목분류":
        return (
            f"• 즉시조치: {subs} 품목별 HS Master와 신고 HS 비교\n"
            "• 1주 내: 불일치 품목 Root Cause 분석 및 변경 승인자료 확보\n"
            "• 1개월 내: HS 변경 Workflow 및 관세율 영향표 반영\n"
            "• Owner: HQ Customs + 법인 Master Data 담당"
        )
    return (
        f"• 즉시조치: {subs} 적용 가능성 확인\n"
        "• 1주 내: 대상 국가·품목·HS·법인 매핑\n"
        "• 1개월 내: 후속 공지 모니터링 및 필요 시 Master 반영\n"
        "• Owner: HQ Customs"
    )

def report_summary(row: pd.Series) -> str:
    return _issue_summary_detail(row).replace("\n", " ")

def report_impact(row: pd.Series) -> str:
    return _samsung_impact_detail(row).replace("\n", " ")

def report_action(row: pd.Series) -> str:
    return _action_detail(row).replace("\n", " ")

def top3_summary_sentence(row: pd.Series) -> str:
    title = clean(row.get("Headline"))
    issue = clean(row.get("Issue"))
    impact = clean(row.get("Samsung Impact"))
    if impact == "Reference":
        return f"{title}: 삼성전자 직접 영향 낮음. Reference로 관리하고 Top3에서는 제외하는 것이 적절합니다."
    if issue in {"AD/CVD", "반덤핑/상계관세"}:
        return f"{title}: 대상 HS·공급국·벤더 기준 AD/CVD 비용과 원산지 방어자료 점검이 필요합니다."
    if issue == "CBAM":
        return f"{title}: EU향 품목의 배출량 자료, CBAM 신고 및 인증서 비용 반영 여부를 확인해야 합니다."
    if issue == "수출통제":
        return f"{title}: ECCN·전략물자 분류와 최종사용자 스크리닝을 우선 점검해야 합니다."
    if issue == "FTA/원산지":
        return f"{title}: CO 발급요건, BOM 원산지, FTA Master 정합성 재검토가 필요합니다."
    if issue in {"관세정책", "통관", "통관/세관", "HS/품목분류"}:
        return f"{title}: 대상 HS, 신고절차, 관세율 및 법인 SOP 반영 여부를 확인해야 합니다."
    return f"{title}: 삼성 관련성과 관세업무 실행 필요성을 확인해야 합니다."

def top3_html(top3: pd.DataFrame) -> str:
    blocks = []
    for idx, row in top3.iterrows():
        blocks.append(f"""
        <div style="margin:14px 0 18px 0;padding:15px;border-left:5px solid #C00000;background:#FFF7F7;">
          <div style="font-size:15px;font-weight:bold;margin-bottom:6px;">Top {idx + 1}. {html_link(row.get('Headline'), row.get('URL'))}</div>
          <div style="font-size:12px;color:#555;margin-bottom:9px;">
            Type: {html.escape(clean(row.get('Content Type')))} | Topic: {html.escape(clean(row.get('Issue')))} |
            Samsung Impact: <b>{html.escape(clean(row.get('Samsung Impact')))}</b> |
            Executive Priority: <b>{html.escape(executive_priority(row))}</b> |
            Subsidiary: {html.escape(clean(row.get('Affected Subsidiary')) or 'SEC/HQ')} |
            Agency: {html.escape(clean(row.get('Agency')))} | Publish Date: {html.escape(clean(row.get('Date')))} |
            Country: {html.escape(clean(row.get('Country')))} |
            Risk: <span style="color:{risk_color(row.get('Risk'))};font-weight:bold;">{html.escape(clean(row.get('Risk')))}</span> |
            Score: {report_score(row):.0f}
          </div>
          <div style="margin-top:8px;"><b>1) 이슈 요약</b><br><pre style="white-space:pre-wrap;font-family:Arial,'Malgun Gothic',sans-serif;margin:4px 0 0 0;">{html.escape(_issue_summary_detail(row))}</pre></div>
          <div style="margin-top:8px;"><b>2) 삼성전자 영향</b><br><pre style="white-space:pre-wrap;font-family:Arial,'Malgun Gothic',sans-serif;margin:4px 0 0 0;">{html.escape(_samsung_impact_detail(row))}</pre></div>
          <div style="margin-top:8px;"><b>3) 관세업무 영향 / 리스크</b><br><pre style="white-space:pre-wrap;font-family:Arial,'Malgun Gothic',sans-serif;margin:4px 0 0 0;">{html.escape(_customs_impact_detail(row))}</pre></div>
          <div style="margin-top:8px;"><b>4) 대응방안</b><br><pre style="white-space:pre-wrap;font-family:Arial,'Malgun Gothic',sans-serif;margin:4px 0 0 0;">{html.escape(_action_detail(row))}</pre></div>
        </div>
        """)
    return "".join(blocks)

def table_html(title: str, rows: pd.DataFrame, color: str) -> str:
    """Override: add Executive Priority column to make REFERENCE demotion visible."""
    if rows.empty:
        return f"<h3 style='color:{color};'>{html.escape(title)} (0건)</h3>"
    trs = []
    for _, row in rows.iterrows():
        trs.append(f"""
        <tr>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;">{html.escape(str(row.get('No')))}</td>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;">{html.escape(clean(row.get('Issue')))}</td>
          <td style="padding:7px;border:1px solid #ddd;vertical-align:top;">{html_link(row.get('Headline'), row.get('URL'))}</td>
          <td style="padding:7px;border:1px solid #ddd;vertical-align:top;">{html.escape(short_text(row.get('Major Changes'), '주요 변경내역 확인 필요', 260))}</td>
          <td style="padding:7px;border:1px solid #ddd;vertical-align:top;">{html.escape(short_text(row.get('AI Analysis'), '영향 검토 필요', 300))}</td>
          <td style="padding:7px;border:1px solid #ddd;vertical-align:top;">{html.escape(short_text(row.get('Action Plan'), '담당 부서 확인 필요', 300))}</td>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;">{html.escape(clean(row.get('Country')))}</td>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;color:{risk_color(row.get('Risk'))};font-weight:bold;">{html.escape(clean(row.get('Risk')))}</td>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;">{html.escape(clean(row.get('Samsung Impact')))}</td>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;">{html.escape(executive_priority(row))}</td>
        </tr>
        """)
    return f"""
    <h3 style="margin-top:24px;color:{color};">{html.escape(title)} ({len(rows)}건)</h3>
    <table style="border-collapse:collapse;width:100%;font-size:12px;table-layout:fixed;">
      <colgroup>
        <col style="width:3%;"><col style="width:7%;"><col style="width:21%;">
        <col style="width:20%;"><col style="width:18%;"><col style="width:18%;">
        <col style="width:5%;"><col style="width:4%;"><col style="width:5%;"><col style="width:7%;">
      </colgroup>
      <thead>
        <tr style="background:{color};color:white;">
          <th style="padding:7px;border:1px solid #ddd;">No</th>
          <th style="padding:7px;border:1px solid #ddd;">Issue</th>
          <th style="padding:7px;border:1px solid #ddd;">Headline</th>
          <th style="padding:7px;border:1px solid #ddd;">주요 변경내역</th>
          <th style="padding:7px;border:1px solid #ddd;">삼성 영향</th>
          <th style="padding:7px;border:1px solid #ddd;">Action</th>
          <th style="padding:7px;border:1px solid #ddd;">Country</th>
          <th style="padding:7px;border:1px solid #ddd;">Risk</th>
          <th style="padding:7px;border:1px solid #ddd;">Impact</th>
          <th style="padding:7px;border:1px solid #ddd;">Priority</th>
        </tr>
      </thead>
      <tbody>{''.join(trs)}</tbody>
    </table>
    """

def overall_html(rows: pd.DataFrame, top3: pd.DataFrame) -> str:
    reg = rows[rows["Content Type"].eq("Regulation")]
    news = rows[rows["Content Type"].eq("News")]
    direct = rows[rows["Samsung Impact"].eq("Direct")]
    indirect = rows[rows["Samsung Impact"].eq("Indirect")]
    watch = rows[rows["Samsung Impact"].eq("Watch")]
    ref = rows[rows["Samsung Impact"].eq("Reference")]
    issues = rows["Issue"].value_counts().head(6)
    issue_txt = ", ".join(f"{k} {v}건" for k, v in issues.items())
    top_lines = "".join(f"<li>{html.escape(top3_summary_sentence(r))}</li>" for _, r in top3.iterrows())

    ref_note = ""
    if len(ref):
        ref_note = f"<div style='margin-top:8px;color:#666;'>Reference {len(ref)}건은 삼성 직접 영향 또는 관세업무 실행성이 낮아 Top3 후보에서 제외했습니다.</div>"

    return f"""
    <div style="padding:15px;background:#F4F6F8;border-left:6px solid #1F4E78;margin-bottom:18px;">
      <div style="font-size:14px;color:#555;margin-bottom:8px;">
        금일 선별 결과: 법규 {len(reg)}건, 주요뉴스 {len(news)}건 | Direct {len(direct)}건, Indirect {len(indirect)}건, Watch {len(watch)}건, Reference {len(ref)}건
      </div>
      <div style="font-size:15px;font-weight:bold;line-height:1.8;margin-bottom:8px;">
        금일 GTI Radar는 {html.escape(issue_txt)} 중심으로 관세·통상 변화가 포착되었습니다.
        Top3는 단순 키워드가 아니라 삼성 관련성, 관세업무 실행성, 시행 긴급성, 비용/리스크 규모를 기준으로 재선정했습니다.
      </div>
      <div style="margin-top:8px;"><b>Top3 요약</b><ol style="margin-top:6px;">{top_lines}</ol></div>
      {ref_note}
    </div>
    """

# ======================================================================
# End of GTI STEP5 Executive Quality Patch v3
# ======================================================================


# ======================================================================
# GTI STEP5 Executive Quality Patch v4
# ----------------------------------------------------------------------
# v3 보완사항
# 1) REFERENCE 판정이 과도하게 적용되어 AD/CVD, CBAM, 통관, 수출통제까지
#    Top3 후보에서 제외되는 문제 수정
# 2) Top3가 1건만 나오는 문제 수정: CORE/POLICY_WATCH/USABLE/WATCH 순서로
#    반드시 3건까지 보충
# 3) 공식 법규, 고위험 관세/통관 이슈는 삼성 직접법인이 특정되지 않아도
#    Top3 후보로 유지
# ======================================================================

ACTIONABLE_ISSUES = {
    "관세정책", "AD/CVD", "반덤핑/상계관세", "CBAM",
    "수출통제", "FTA/원산지", "통관", "통관/세관", "HS/품목분류"
}

REFERENCE_ONLY_ISSUES = {"무역일반", "일반경제", "시장동향", "기타"}

def reference_reason(row: pd.Series) -> str:
    """v4 override: only demote clearly non-actionable items.

    Do not demote AD/CVD, CBAM, customs, export control, HS, or official high-risk
    regulations simply because affected subsidiary is not specified yet.
    """
    text = _row_text(row)
    issue = clean(row.get("Issue"))
    title = clean(row.get("Headline")).lower()
    content_type = clean(row.get("Content Type"))
    risk = normalize_risk(row.get("Risk"))

    # 1) Explicit agriculture/food/general item with no Samsung/customs action.
    if _has_any(text, LOW_RELEVANCE_TERMS):
        if issue not in ACTIONABLE_ISSUES or issue in REFERENCE_ONLY_ISSUES or "wheat" in text or "염소산업" in text:
            return "삼성전자 주요 제품·부품·원재료와 직접 관련성이 낮은 일반 품목/농산물성 규제입니다."

    # 2) General stock/economy news, even if semiconductor appears as macro statistic.
    if _has_any(text, GENERAL_NEWS_TERMS):
        if not any(k in text for k in ["tariff rate", "관세율", "수출통제", "cbam", "ad/cvd", "반덤핑", "상계관세", "hs code"]):
            return "일반 경제/시장 동향 성격이 강해 임원 보고 Top3보다는 참고 모니터링에 적합합니다."

    # 3) Generic trade/export notice without HS/tariff/origin/customs/export-control action.
    if issue in REFERENCE_ONLY_ISSUES or issue == "무역일반":
        if not _has_any(text, CRITICAL_CUSTOMS_TERMS):
            return "관세업무 실행 조치가 필요한 수준의 HS·세율·원산지·신고절차 변경이 확인되지 않았습니다."

    # 4) Weak non-regulation item with no actionable issue.
    if content_type != "Regulation" and issue not in ACTIONABLE_ISSUES:
        if samsung_relevance_score(row) < 300 and customs_actionability_score(row) < 600:
            return "삼성 관련성과 관세업무 실행성이 낮아 Reference 관리가 적절합니다."

    # Keep official regulations and actionable issues as candidates.
    if content_type == "Regulation" and risk in {"상", "중"}:
        return ""
    if issue in ACTIONABLE_ISSUES:
        return ""

    return ""

def executive_priority(row: pd.Series) -> str:
    """v4 override: preserve actionable customs issues as candidates."""
    ref = reference_reason(row)
    if ref:
        return "REFERENCE"

    srel = samsung_relevance_score(row)
    act = customs_actionability_score(row)
    issue = clean(row.get("Issue"))
    content_type = clean(row.get("Content Type"))
    risk = normalize_risk(row.get("Risk"))

    if clean(row.get("Samsung Impact")) == "Direct":
        return "CORE"
    if clean(row.get("Samsung Impact")) == "Indirect" and issue in ACTIONABLE_ISSUES:
        return "CORE"
    if issue in {"AD/CVD", "반덤핑/상계관세", "수출통제", "CBAM", "HS/품목분류"}:
        return "POLICY_WATCH" if clean(row.get("Samsung Impact")) != "Indirect" else "CORE"
    if issue in {"관세정책", "통관", "통관/세관"}:
        return "POLICY_WATCH"
    if content_type == "Regulation" and risk == "상":
        return "POLICY_WATCH"
    if issue == "FTA/원산지" and (srel >= 900 or "battery" in _row_text(row) or risk == "상"):
        return "USABLE"
    if act >= 1100:
        return "USABLE"
    return clean(row.get("Priority Group")) or "WATCH"

def report_score(row: pd.Series) -> float:
    """v4 override: balanced scoring."""
    base = safe_num(row.get("Importance Score"))
    issue = clean(row.get("Issue"))
    priority = executive_priority(row)
    content_type = clean(row.get("Content Type"))

    issue_weight = {
        "AD/CVD": 1500,
        "반덤핑/상계관세": 1500,
        "수출통제": 1450,
        "CBAM": 1300,
        "관세정책": 1150,
        "통관": 1100,
        "통관/세관": 1100,
        "HS/품목분류": 1100,
        "FTA/원산지": 800,
        "무역일반": 100,
    }.get(issue, 250)

    pri_weight = {
        "CORE": 1800,
        "POLICY_WATCH": 1300,
        "USABLE": 800,
        "WATCH": 300,
        "REFERENCE": -2500,
    }.get(priority, 0)

    impact_weight = {
        "Direct": 2400,
        "Indirect": 1200,
        "Watch": 300,
        "Reference": -1000,
    }.get(clean(row.get("Samsung Impact")), 0)

    type_weight = 550 if content_type == "Regulation" else 0
    risk_w = risk_weight(row.get("Risk"))

    return (
        base + issue_weight + pri_weight + impact_weight + type_weight + risk_w
        + max(samsung_relevance_score(row), -1000)
        + max(customs_actionability_score(row), -500)
    )

def top3_deep_score(row: pd.Series) -> float:
    """v4 override: Top3 ranking with explicit non-reference and issue priority."""
    priority = executive_priority(row)
    if priority == "REFERENCE":
        return -999999

    issue = clean(row.get("Issue"))
    score = report_score(row)

    # Favor concrete cost/compliance topics.
    if issue in {"AD/CVD", "반덤핑/상계관세"}:
        score += 1800
    elif issue == "수출통제":
        score += 1600
    elif issue == "CBAM":
        score += 1400
    elif issue in {"관세정책", "통관", "통관/세관"}:
        score += 1200
    elif issue == "HS/품목분류":
        score += 1000
    elif issue == "FTA/원산지":
        score += 600

    if clean(row.get("Content Type")) == "Regulation":
        score += 500
    if normalize_risk(row.get("Risk")) == "상":
        score += 500

    return score

def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    """v4 override: always select up to 3 non-reference rows if available."""
    pool = rows.copy()
    pool["Executive Priority"] = pool.apply(executive_priority, axis=1)
    pool["_top3_score"] = pool.apply(top3_deep_score, axis=1)

    non_ref = pool[pool["Executive Priority"].ne("REFERENCE")].copy()
    if non_ref.empty:
        non_ref = pool.copy()

    non_ref = non_ref.sort_values(["_top3_score", "_sort_date"], ascending=[False, False])

    selected = []
    used_issues = set()

    # 1st pass: issue diversity
    for _, row in non_ref.iterrows():
        issue = clean(row.get("Issue"))
        if issue in used_issues:
            continue
        selected.append(row)
        used_issues.add(issue)
        if len(selected) == 3:
            break

    # 2nd pass: fill remaining even if same issue
    if len(selected) < 3:
        for _, row in non_ref.iterrows():
            title = clean(row.get("Headline"))
            if any(clean(x.get("Headline")) == title for x in selected):
                continue
            selected.append(row)
            if len(selected) == 3:
                break

    out = pd.DataFrame(selected).reset_index(drop=True)
    if not out.empty:
        out["No"] = range(1, len(out) + 1)
    return out

def prepare_rows(rows: pd.DataFrame) -> pd.DataFrame:
    """v4 override: demote only true references, keep actionable issues."""
    rows = rows.copy()
    rows["Issue"] = rows.apply(issue_for, axis=1)
    rows = dedup_report_rows(rows)
    rows["Mail Group"] = rows["Content Type"].map({"Regulation": GROUP_REGULATION}).fillna(GROUP_NEWS)
    rows["Executive Priority"] = rows.apply(executive_priority, axis=1)

    ref_mask = rows["Executive Priority"].eq("REFERENCE")
    rows.loc[ref_mask, "Priority Group"] = "REFERENCE"
    rows.loc[ref_mask, "Samsung Impact"] = "Reference"

    # Preserve Priority Group for actionable issues.
    rows.loc[~ref_mask, "Priority Group"] = rows.loc[~ref_mask, "Executive Priority"]

    rows["Major Changes"] = rows.apply(major_changes, axis=1)
    rows["Summary"] = rows.apply(report_summary, axis=1)
    rows["AI Analysis"] = rows.apply(report_impact, axis=1)
    rows["Action Plan"] = rows.apply(report_action, axis=1)
    rows["_report_score"] = rows.apply(report_score, axis=1)

    rows = rows.sort_values(["_report_score", "_sort_date"], ascending=[False, False]).reset_index(drop=True)
    rows["No"] = range(1, len(rows) + 1)
    return rows

# ======================================================================
# End of GTI STEP5 Executive Quality Patch v4
# ======================================================================


# ======================================================================
# GTI STEP5 Executive Quality Patch v5
# ----------------------------------------------------------------------
# v4 보완사항
# 1) 2.Top3 / 3.Regulation / 4.주요뉴스 모두에 게시물 요약 반영
# 2) "주요 변경내역" = 기존 변경내역 + 게시물 전체요약 2~3줄
# 3) STEP4 원본 Summary/AI Analysis를 보존하여 STEP5에서 덮어쓰기 전에 활용
# ======================================================================

def _source_summary_text(row: pd.Series, limit: int = 520) -> str:
    """Return original post/article summary from STEP4 before STEP5 rewrite."""
    candidates = [
        clean(row.get("Original Summary")),
        clean(row.get("Original AI Analysis")),
        clean(row.get("Impact Reason")),
        clean(row.get("Summary")),
    ]
    for text in candidates:
        if text and text not in {"본문에서 확인 불가", "nan", "None"}:
            # Remove repetitive generic STEP5-style sentences if already rewritten.
            bad = [
                "공식 법규/공지입니다. 핵심은",
                "에서 포착된",
                "삼성전자 관련 법인",
            ]
            if any(b in text for b in bad) and len(text) < 180:
                continue
            text = re.sub(r"\s+", " ", text).strip()
            return text[:limit] + ("..." if len(text) > limit else "")
    return "원문 요약 정보가 부족합니다. 원문 링크 기준으로 세부 내용 확인이 필요합니다."

def _two_three_line_summary(row: pd.Series) -> str:
    """Create 2~3 line bulletin summary for mail display."""
    src = _source_summary_text(row, 680)
    issue = clean(row.get("Issue"))
    country = clean(row.get("Country")) or "관련국"
    date = clean(row.get("Date")) or "게시일 확인 필요"
    headline = clean(row.get("Headline"))

    # If source summary is weak, create a structured fallback from available fields.
    if src.startswith("원문 요약 정보가 부족"):
        src = (
            f"{headline} 관련 {issue} 이슈입니다. "
            f"대상 국가는 {country}이며 게시/확인일은 {date}입니다. "
            "세부 대상 품목, HS, 세율, 시행일은 원문 및 법인 실적 기준으로 추가 확인이 필요합니다."
        )

    sentences = re.split(r"(?<=[.!?。？！])\s+|(?<=다\.)\s+|(?<=니다\.)\s+", src)
    sentences = [s.strip(" -•\n\t") for s in sentences if s.strip()]
    if len(sentences) >= 2:
        return "\n".join(f"• {s}" for s in sentences[:3])
    # Short text fallback: split by length.
    if len(src) > 180:
        return f"• {src[:180].strip()}\n• {src[180:360].strip()}"
    return f"• {src}"

def major_changes(row: pd.Series) -> str:
    """v5 override: current change detail + article/post full summary 2~3 lines."""
    issue = clean(row.get("Issue"))
    headline = clean(row.get("Headline"))
    title_l = headline.lower()

    if "보세창고" in headline and "특허" in headline:
        current = (
            "개정 사유: 자가용보세창고 특허요건 완화 및 불명확한 규정 보완 필요. "
            "주요 개정 내용: 자가용보세창고 반입 대상에 국제무역선·기 적재 자가화물 외 수리용 예비부분품 및 부속품 장치를 허용하고, "
            "관세법 제178조상 물품반입 정지기간을 오해 없이 적용할 수 있도록 규정을 명확화하는 내용입니다."
        )
    elif "환전영업자" in headline and "관리" in headline:
        current = (
            "주요 내용: 환전영업자의 등록·관리, 보고·자료제출, 영업장 운영 및 관세청 관리 기준과 관련된 고시입니다. "
            "해외출장·주재원·외환거래 지원 프로세스와 연결될 수 있어 실제 법인 업무 해당 여부 확인이 필요합니다."
        )
    elif "cbam" in title_l and "certificate price" in title_l:
        current = (
            "주요 내용: EU CBAM 인증서 가격이 공표되었거나 공표 일정이 확정된 사안입니다. "
            "EU 수입품의 내재배출량 신고, 인증서 구매 비용, 공급사 배출량 자료 확보 체계에 영향을 줄 수 있습니다."
        )
    elif "customs enforcement" in title_l and "executive order" in title_l:
        current = (
            "주요 내용: 미국 세관 집행 강화 행정명령 관련 사안입니다. "
            "수입신고 정확성, 저가신고·우회수입·전자상거래 물품 관리 및 CBP 심사 강화 가능성을 확인해야 합니다."
        )
    elif "수입신고" in headline and "가산세" in headline:
        current = (
            "주요 내용: 수입신고 지연 가산세 부과 대상이 되는 매점매석 금지 품목의 적용기간 연장 공고입니다. "
            "해당 품목 수입 시 신고 지연, 재고 운영, 통관 일정 관리 기준을 확인해야 합니다."
        )
    else:
        parts = [
            hint_line("시행/적용일", row.get("effective_date_hint")),
            hint_line("변경 내용", row.get("change_detail_hint")),
            hint_line("대상 HS", row.get("hs_hint")),
            hint_line("관세율/쿼터", row.get("tariff_rate_hint")),
            hint_line("키워드", row.get("KeywordMatches")),
        ]
        if any(parts):
            current = compact_parts(parts, "")
        elif issue == "관세정책":
            current = "관세율, 쿼터, 면세/환급 또는 Section 301/232 등 관세 비용에 영향을 줄 수 있는 정책 변화입니다."
        elif issue in {"AD/CVD", "반덤핑/상계관세"}:
            current = "반덤핑 또는 상계관세 조사·판정·연장 가능성이 있는 사안입니다. 공급국, 대상 품목, 조사 기간과 관세율 확인이 필요합니다."
        elif issue == "CBAM":
            current = "CBAM 신고, 인증서 가격, 배출량 자료 또는 EU 수입통관 절차와 연결되는 탄소국경조정 변화입니다."
        elif issue == "FTA/원산지":
            current = "FTA/CEPA 협정, 원산지 기준, CO 발급 또는 특혜관세 적용 가능성에 영향을 주는 변화입니다."
        elif issue == "수출통제":
            current = "Entity List, ECCN, UFLPA, forced labor 또는 전략물자·제재 스크리닝 관련 변화입니다."
        elif issue == "통관":
            current = "보세, 통관, 신고, 세관 심사 또는 행정절차 기준에 영향을 줄 수 있는 공식 공지입니다."
        elif issue == "HS/품목분류":
            current = "HS 분류 기준 또는 품목 해석이 달라질 수 있어 품목 마스터와 신고 기준 점검이 필요한 사안입니다."
        else:
            current = f"{headline} 관련 관세·통상 모니터링 사안입니다."

    post_summary = _two_three_line_summary(row)
    return f"{current}\n\n[게시물 요약]\n{post_summary}"

def _issue_summary_detail(row: pd.Series) -> str:
    """v5 override: Top3 issue summary includes post summary."""
    issue = clean(row.get("Issue"))
    country = clean(row.get("Country")) or "확인 필요"
    date = clean(row.get("Date")) or "확인 필요"
    hs = non_empty_hint(row.get("hs_hint")) or "원문/법인 품목 기준 확인 필요"
    rate = non_empty_hint(row.get("tariff_rate_hint")) or "해당 시 별도 산출 필요"
    change = clean(row.get("Major Changes")) or major_changes(row)
    return (
        f"• 이슈구분: {issue}\n"
        f"• 대상국가: {country}\n"
        f"• 게시/시행일: {date}\n"
        f"• 대상 HS/품목: {hs}\n"
        f"• 세율/쿼터/허가 변화: {rate}\n"
        f"• 주요 변경내역 및 게시물 요약:\n{change}"
    )

def prepare_rows(rows: pd.DataFrame) -> pd.DataFrame:
    """v5 override: preserve original summaries, then generate Major Changes with post summary."""
    rows = rows.copy()
    rows["Issue"] = rows.apply(issue_for, axis=1)
    rows = dedup_report_rows(rows)
    rows["Mail Group"] = rows["Content Type"].map({"Regulation": GROUP_REGULATION}).fillna(GROUP_NEWS)

    # Preserve STEP4 source texts before STEP5 overwrites them.
    if "Original Summary" not in rows.columns:
        rows["Original Summary"] = rows.get("Summary", "")
    if "Original AI Analysis" not in rows.columns:
        rows["Original AI Analysis"] = rows.get("AI Analysis", "")
    if "Original Action Plan" not in rows.columns:
        rows["Original Action Plan"] = rows.get("Action Plan", "")

    rows["Executive Priority"] = rows.apply(executive_priority, axis=1)

    ref_mask = rows["Executive Priority"].eq("REFERENCE")
    rows.loc[ref_mask, "Priority Group"] = "REFERENCE"
    rows.loc[ref_mask, "Samsung Impact"] = "Reference"
    rows.loc[~ref_mask, "Priority Group"] = rows.loc[~ref_mask, "Executive Priority"]

    rows["Major Changes"] = rows.apply(major_changes, axis=1)
    rows["Summary"] = rows.apply(report_summary, axis=1)
    rows["AI Analysis"] = rows.apply(report_impact, axis=1)
    rows["Action Plan"] = rows.apply(report_action, axis=1)
    rows["_report_score"] = rows.apply(report_score, axis=1)

    rows = rows.sort_values(["_report_score", "_sort_date"], ascending=[False, False]).reset_index(drop=True)
    rows["No"] = range(1, len(rows) + 1)
    return rows

def table_html(title: str, rows: pd.DataFrame, color: str) -> str:
    """v5 override: widen Major Changes and show line breaks for post summary."""
    if rows.empty:
        return f"<h3 style='color:{color};'>{html.escape(title)} (0건)</h3>"
    trs = []
    for _, row in rows.iterrows():
        major = html.escape(short_text(row.get('Major Changes'), '주요 변경내역 및 게시물 요약 확인 필요', 620)).replace("\n", "<br>")
        impact = html.escape(short_text(row.get('AI Analysis'), '영향 검토 필요', 360)).replace("\n", "<br>")
        action = html.escape(short_text(row.get('Action Plan'), '담당 부서 확인 필요', 360)).replace("\n", "<br>")
        trs.append(f"""
        <tr>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;">{html.escape(str(row.get('No')))}</td>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;">{html.escape(clean(row.get('Issue')))}</td>
          <td style="padding:7px;border:1px solid #ddd;vertical-align:top;">{html_link(row.get('Headline'), row.get('URL'))}</td>
          <td style="padding:7px;border:1px solid #ddd;vertical-align:top;">{major}</td>
          <td style="padding:7px;border:1px solid #ddd;vertical-align:top;">{impact}</td>
          <td style="padding:7px;border:1px solid #ddd;vertical-align:top;">{action}</td>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;">{html.escape(clean(row.get('Country')))}</td>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;color:{risk_color(row.get('Risk'))};font-weight:bold;">{html.escape(clean(row.get('Risk')))}</td>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;">{html.escape(clean(row.get('Samsung Impact')))}</td>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;">{html.escape(executive_priority(row))}</td>
        </tr>
        """)
    return f"""
    <h3 style="margin-top:24px;color:{color};">{html.escape(title)} ({len(rows)}건)</h3>
    <table style="border-collapse:collapse;width:100%;font-size:12px;table-layout:fixed;">
      <colgroup>
        <col style="width:3%;"><col style="width:7%;"><col style="width:19%;">
        <col style="width:25%;"><col style="width:16%;"><col style="width:16%;">
        <col style="width:5%;"><col style="width:4%;"><col style="width:5%;"><col style="width:7%;">
      </colgroup>
      <thead>
        <tr style="background:{color};color:white;">
          <th style="padding:7px;border:1px solid #ddd;">No</th>
          <th style="padding:7px;border:1px solid #ddd;">Issue</th>
          <th style="padding:7px;border:1px solid #ddd;">Headline</th>
          <th style="padding:7px;border:1px solid #ddd;">주요 변경내역 + 게시물 요약</th>
          <th style="padding:7px;border:1px solid #ddd;">삼성 영향</th>
          <th style="padding:7px;border:1px solid #ddd;">Action</th>
          <th style="padding:7px;border:1px solid #ddd;">Country</th>
          <th style="padding:7px;border:1px solid #ddd;">Risk</th>
          <th style="padding:7px;border:1px solid #ddd;">Impact</th>
          <th style="padding:7px;border:1px solid #ddd;">Priority</th>
        </tr>
      </thead>
      <tbody>{''.join(trs)}</tbody>
    </table>
    """

# ======================================================================
# End of GTI STEP5 Executive Quality Patch v5
# ======================================================================


# ======================================================================
# GTI STEP5 Executive Quality Patch v6
# ----------------------------------------------------------------------
# v5 보완사항
# 1) 게시물 요약이 제목만 반복되는 경우 자동 차단
# 2) STEP4 원문요약/본문 컬럼이 있으면 우선 사용
# 3) 원문요약이 없을 때는 "요약 부족"을 명확히 표시하고, 추정요약을 과장하지 않음
# 4) 주요 변경내역에 기존 내용 + 게시물 요약 2~3줄을 안정적으로 표시
# ======================================================================

def _looks_like_title_only(text: str, title: str) -> bool:
    t = clean(text)
    h = clean(title)
    if not t:
        return True
    if h and (t == h or t.replace(" ", "") == h.replace(" ", "")):
        return True
    # Very short snippets that just repeat title are not a real summary.
    if h and len(t) <= len(h) + 20 and h[:25] in t:
        return True
    generic_patterns = [
        "건은", "관련 공식 규제/공지 후보입니다", "원문 요약 정보가 부족합니다",
        "관련 관세·통상 모니터링 사안입니다"
    ]
    if any(p in t for p in generic_patterns) and len(t) < 180:
        return True
    return False

def _source_summary_text(row: pd.Series, limit: int = 700) -> str:
    """v6 override: get real post/article summary, not title repetition."""
    title = clean(row.get("Headline"))

    candidates = [
        clean(row.get("Original Post Summary")),
        clean(row.get("Original Body Text")),
        clean(row.get("Original Summary")),
        clean(row.get("Original AI Analysis")),
        clean(row.get("Impact Reason")),
        clean(row.get("Summary")),
    ]

    for text in candidates:
        if not text or text in {"본문에서 확인 불가", "nan", "None"}:
            continue
        text = re.sub(r"\s+", " ", text).strip()
        if _looks_like_title_only(text, title):
            continue
        # Avoid using STEP5 generated impact/action as article summary.
        generated_markers = ["• 영향등급:", "• 즉시조치:", "• 이슈구분:", "영향법인 후보"]
        if any(m in text for m in generated_markers):
            continue
        return text[:limit] + ("..." if len(text) > limit else "")

    return ""

def _two_three_line_summary(row: pd.Series) -> str:
    """v6 override: 2~3 line real source summary; avoid fake title summary."""
    src = _source_summary_text(row, 800)
    title = clean(row.get("Headline"))
    issue = clean(row.get("Issue"))
    country = clean(row.get("Country")) or "관련국"
    date = clean(row.get("Date")) or "게시일 확인 필요"

    if not src:
        return (
            "• 원문/STEP4 요약 본문이 없어 제목 수준 정보만 확인됩니다.\n"
            f"• 확인된 정보: {title} / 이슈구분 {issue} / 대상국가 {country} / 게시일 {date}\n"
            "• 정확한 게시물 전체요약을 위해 STEP4 결과에 ArticleSummary 또는 본문요약 컬럼을 저장해야 합니다."
        )

    sentences = re.split(r"(?<=[.!?。？！])\s+|(?<=다\.)\s+|(?<=니다\.)\s+", src)
    sentences = [s.strip(" -•\n\t") for s in sentences if s.strip()]
    # Remove title-only sentence if first sentence is same as headline.
    sentences = [s for s in sentences if not _looks_like_title_only(s, title)]

    if len(sentences) >= 2:
        return "\n".join(f"• {s}" for s in sentences[:3])
    if len(sentences) == 1:
        s = sentences[0]
        if len(s) > 220:
            return f"• {s[:220].strip()}\n• {s[220:440].strip()}"
        return f"• {s}"

    return (
        "• 원문/STEP4 요약 본문이 부족하여 상세 요약을 생성하지 않았습니다.\n"
        "• STEP4에서 기사 본문 또는 원문 요약을 별도 컬럼으로 저장하면 이 위치에 2~3줄 요약이 자동 반영됩니다."
    )

def major_changes(row: pd.Series) -> str:
    """v6 override: existing change detail + real post summary 2~3 lines."""
    issue = clean(row.get("Issue"))
    headline = clean(row.get("Headline"))
    title_l = headline.lower()

    if "보세창고" in headline and "특허" in headline:
        current = (
            "개정 사유: 자가용보세창고 특허요건 완화 및 불명확한 규정 보완 필요. "
            "주요 개정 내용: 자가용보세창고 반입 대상에 국제무역선·기 적재 자가화물 외 수리용 예비부분품 및 부속품 장치를 허용하고, "
            "관세법 제178조상 물품반입 정지기간을 오해 없이 적용할 수 있도록 규정을 명확화하는 내용입니다."
        )
    elif "환전영업자" in headline and "관리" in headline:
        current = (
            "주요 내용: 환전영업자의 등록·관리, 보고·자료제출, 영업장 운영 및 관세청 관리 기준과 관련된 고시입니다. "
            "해외출장·주재원·외환거래 지원 프로세스와 연결될 수 있어 실제 법인 업무 해당 여부 확인이 필요합니다."
        )
    elif "cbam" in title_l and "certificate price" in title_l:
        current = (
            "주요 내용: EU CBAM 인증서 가격이 공표되었거나 공표 일정이 확정된 사안입니다. "
            "EU 수입품의 내재배출량 신고, 인증서 구매 비용, 공급사 배출량 자료 확보 체계에 영향을 줄 수 있습니다."
        )
    elif "customs enforcement" in title_l and "executive order" in title_l:
        current = (
            "주요 내용: 미국 세관 집행 강화 행정명령 관련 사안입니다. "
            "수입신고 정확성, 저가신고·우회수입·전자상거래 물품 관리 및 CBP 심사 강화 가능성을 확인해야 합니다."
        )
    elif "수입신고" in headline and "가산세" in headline:
        current = (
            "주요 내용: 수입신고 지연 가산세 부과 대상이 되는 매점매석 금지 품목의 적용기간 연장 공고입니다. "
            "해당 품목 수입 시 신고 지연, 재고 운영, 통관 일정 관리 기준을 확인해야 합니다."
        )
    else:
        parts = [
            hint_line("시행/적용일", row.get("effective_date_hint")),
            hint_line("변경 내용", row.get("change_detail_hint")),
            hint_line("대상 HS", row.get("hs_hint")),
            hint_line("관세율/쿼터", row.get("tariff_rate_hint")),
            hint_line("키워드", row.get("KeywordMatches")),
        ]
        if any(parts):
            current = compact_parts(parts, "")
        elif issue == "관세정책":
            current = "관세율, 쿼터, 면세/환급 또는 Section 301/232 등 관세 비용에 영향을 줄 수 있는 정책 변화입니다."
        elif issue in {"AD/CVD", "반덤핑/상계관세"}:
            current = "반덤핑 또는 상계관세 조사·판정·연장 가능성이 있는 사안입니다. 공급국, 대상 품목, 조사 기간과 관세율 확인이 필요합니다."
        elif issue == "CBAM":
            current = "CBAM 신고, 인증서 가격, 배출량 자료 또는 EU 수입통관 절차와 연결되는 탄소국경조정 변화입니다."
        elif issue == "FTA/원산지":
            current = "FTA/CEPA 협정, 원산지 기준, CO 발급 또는 특혜관세 적용 가능성에 영향을 주는 변화입니다."
        elif issue == "수출통제":
            current = "Entity List, ECCN, UFLPA, forced labor 또는 전략물자·제재 스크리닝 관련 변화입니다."
        elif issue == "통관":
            current = "보세, 통관, 신고, 세관 심사 또는 행정절차 기준에 영향을 줄 수 있는 공식 공지입니다."
        elif issue == "HS/품목분류":
            current = "HS 분류 기준 또는 품목 해석이 달라질 수 있어 품목 마스터와 신고 기준 점검이 필요한 사안입니다."
        else:
            current = f"{headline} 관련 관세·통상 모니터링 사안입니다."

    return f"{current}\n\n[게시물 요약]\n{_two_three_line_summary(row)}"

def prepare_rows(rows: pd.DataFrame) -> pd.DataFrame:
    """v6 override: preserve raw summary/body columns before STEP5 rewrite."""
    rows = rows.copy()
    rows["Issue"] = rows.apply(issue_for, axis=1)
    rows = dedup_report_rows(rows)
    rows["Mail Group"] = rows["Content Type"].map({"Regulation": GROUP_REGULATION}).fillna(GROUP_NEWS)

    if "Original Summary" not in rows.columns:
        rows["Original Summary"] = rows.get("Summary", "")
    if "Original AI Analysis" not in rows.columns:
        rows["Original AI Analysis"] = rows.get("AI Analysis", "")
    if "Original Action Plan" not in rows.columns:
        rows["Original Action Plan"] = rows.get("Action Plan", "")
    if "Original Post Summary" not in rows.columns:
        rows["Original Post Summary"] = ""
    if "Original Body Text" not in rows.columns:
        rows["Original Body Text"] = ""

    rows["Executive Priority"] = rows.apply(executive_priority, axis=1)

    ref_mask = rows["Executive Priority"].eq("REFERENCE")
    rows.loc[ref_mask, "Priority Group"] = "REFERENCE"
    rows.loc[ref_mask, "Samsung Impact"] = "Reference"
    rows.loc[~ref_mask, "Priority Group"] = rows.loc[~ref_mask, "Executive Priority"]

    rows["Major Changes"] = rows.apply(major_changes, axis=1)
    rows["Summary"] = rows.apply(report_summary, axis=1)
    rows["AI Analysis"] = rows.apply(report_impact, axis=1)
    rows["Action Plan"] = rows.apply(report_action, axis=1)
    rows["_report_score"] = rows.apply(report_score, axis=1)

    rows = rows.sort_values(["_report_score", "_sort_date"], ascending=[False, False]).reset_index(drop=True)
    rows["No"] = range(1, len(rows) + 1)
    return rows

# ======================================================================
# End of GTI STEP5 Executive Quality Patch v6
# ======================================================================


# ======================================================================
# GTI STEP5 Executive Quality Patch v7
# ----------------------------------------------------------------------
# v6 보완사항
# 1) UNIPASS 공지 URL이 index.do?rowTitle=... 형태로 잘못 연결되는 문제 수정
# 2) openMYC0605014Q.do?ntarId=... 상세 URL을 우선 사용
# 3) URL 후보 중 rowTitle 목록 URL보다 ntarId/detail URL 우선 선택
# ======================================================================

def unipass_url_rank(url: str) -> int:
    low = clean(url).lower()
    if not low:
        return -999
    if "unipass.customs.go.kr" not in low:
        return 0
    # 상세 공지 URL 최우선
    if "openmyc0605014q.do" in low and "ntarid=" in low:
        return 1000
    if "ntarid=" in low:
        return 900
    # 상세 경로로 보이는 URL 우대
    if "/csp/myc/custsppt/" in low:
        return 800
    # index.do + rowTitle은 목록/검색용 URL이므로 낮은 점수
    if "index.do" in low and "rowtitle=" in low:
        return -500
    if "index.do" in low and "tgmenuid=" in low:
        return -300
    return 100

def best_url_from_values(values) -> str:
    """v7 override: prefer real article/detail URLs over search/list URLs.

    Especially for UNIPASS notice, prefer:
    https://unipass.customs.go.kr/csp/myc/custsppt/cmmn/NtarBrkdMtCtr/openMYC0605014Q.do?ntarId=...
    over:
    https://unipass.customs.go.kr/csp/index.do?tgMenuId=...&rowTitle=...
    """
    invalid = {
        "", "nan", "none", "null", "new", "https://new", "http://new",
        "https://news", "http://news", "https://news.google.com", "https://news.google.com/",
    }
    candidates: list[str] = []
    for value in values:
        text = clean(value)
        if not text:
            continue
        for item in [text] + re.findall(r"https?://[^'\"),\s]+", text):
            url = html.unescape(item).strip().strip("<>'\"").rstrip(".,);]}")
            if url.lower() in invalid:
                continue
            if re.match(r"^https?://", url, re.I) and url not in candidates:
                candidates.append(url)

    if not candidates:
        return ""

    def rank_url(url: str) -> tuple[int, int, int]:
        low = url.lower()
        rank = 0
        # Avoid Google RSS wrapper where possible
        if "news.google.com/rss/articles/" in low or "news.google.com/articles/" in low:
            rank -= 600
        # Prefer concrete detail pages
        if any(k in low for k in ["articleview", "article/", "/news/", "view", "open", "detail", "document"]):
            rank += 150
        # UNIPASS special handling
        rank += unipass_url_rank(url)
        # Prefer URL with query id over title-only search/list links
        if any(k in low for k in ["id=", "ntarid=", "idxno=", "articleid=", "no="]):
            rank += 80
        if "rowtitle=" in low:
            rank -= 250
        # Stable tie breaker: longer URLs often preserve detail ids
        return (rank, len(url), -candidates.index(url))

    return sorted(candidates, key=rank_url, reverse=True)[0]

# ======================================================================
# End of GTI STEP5 Executive Quality Patch v7
# ======================================================================


# ======================================================================
# GTI STEP5 Executive Selection & Summary Quality Patch v8
# ----------------------------------------------------------------------
# 보완사항
# 1) 일본 과세환율/환율 공지, 밀/농산물, 일반 기사 자동 Reference 강등
# 2) "글자크기 변경", "이전 기사보기" 등 언론사 UI 문구 제거
# 3) 53건 모두 Indirect로 분류되는 문제 보정: Direct / Indirect / Watch / Reference 재분류
# 4) Top3는 삼성 관련성 + 관세업무 실행성 + 비용/리스크 기준으로 재선정
# 5) Top3 요약은 임원 보고용 1줄로 압축
# ======================================================================

UI_NOISE_PHRASES = [
    "이전 기사보기", "다음 기사보기", "기사의 본문 내용은 이 글자크기로 변경됩니다",
    "본문 글씨 키우기", "본문 글씨 줄이기", "스크롤 이동 상태바", "바로가기 복사하기",
    "가 가", "댓글 0", "글자 크기", "글자크기", "본문영역", "기사원문",
    "공유 이메일에 공유하기", "카카오톡에 공유하기", "페이스북에 공유하기",
    "트위터에 공유하기", "링크 복사하기", "닫기", "번역 ENG JPN CHN",
    "편의기능", "AI기능", "추천질문", "관련종목", "AI해설", "에디터 픽", "추천기사",
]

LOW_VALUE_HEADLINE_TERMS = [
    "rate of exchange", "exchange rate", "과세환율", "환율정보", "환율 공지",
    "modalities for export of wheat", "export of wheat", "wheat reg", "밀 수출",
    "새우", "라이스페이퍼", "염소산업", "혈통관리", "세계 1위 패권", "일본은 어떻게 몰락",
    "인류의 살상을", "페라리", "로마 회동", "방위산업 공동", "파트너십", "협력 본격화",
]

HARD_CUSTOMS_ISSUES = {"AD/CVD", "반덤핑/상계관세", "CBAM", "수출통제", "FTA/원산지", "HS/품목분류", "관세정책"}
SOFT_CUSTOMS_ISSUES = {"통관", "통관/세관", "무역일반"}

COST_RISK_TERMS = [
    "33.67", "반덤핑", "덤핑방지", "상계관세", "ad/cvd", "anti-dumping", "countervailing",
    "cbam", "탄소국경", "section 301", "section 232", "관세율", "추가관세", "쿼터",
    "수출통제", "export control", "entity list", "forced labor", "uflpa", "원산지", "fta", "cepa",
    "hs code", "품목분류", "철강", "도금강판", "합판", "배터리", "반도체", "희토류",
]

def remove_ui_noise(text: str) -> str:
    t = clean(text)
    if not t:
        return ""
    t = html.unescape(t)
    t = re.sub(r"\s+", " ", t).strip()
    for p in UI_NOISE_PHRASES:
        t = t.replace(p, " ")
    # Remove isolated Korean font-size UI fragments like "가 가"
    t = re.sub(r"(?<![가-힣])가\s+가(?![가-힣])", " ", t)
    # Remove repeated share/navigation fragments
    t = re.sub(r"(공유|닫기|인쇄|즐겨찾기|댓글|추천기사)(\s+|$)", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t

def is_bad_summary_text(text: str) -> bool:
    t = clean(text)
    if not t:
        return True
    noise_hits = sum(1 for p in UI_NOISE_PHRASES if p in t)
    if noise_hits >= 2:
        return True
    if "기사의 본문 내용은 이 글자크기로 변경됩니다" in t:
        return True
    if len(remove_ui_noise(t)) < 40:
        return True
    return False

def is_exchange_rate_notice(row: pd.Series) -> bool:
    h = clean(row.get("Headline")).lower()
    u = clean(row.get("URL")).lower()
    ag = clean(row.get("Agency")).lower()
    return (
        "rate of exchange" in h
        or "exchange rate" in h
        or "kawase" in u
        or ("customs.go.jp" in u and ("exchange" in h or "kawase" in u))
        or "과세환율" in h
    )

def is_low_value_notice(row: pd.Series) -> bool:
    text = _row_text(row)
    headline = clean(row.get("Headline")).lower()
    if is_exchange_rate_notice(row):
        return True
    if any(k in headline for k in LOW_VALUE_HEADLINE_TERMS):
        return True
    if any(k in text for k in ["wheat", "밀 ", "새우", "라이스페이퍼", "염소산업", "혈통관리"]):
        # 단순 농수산물/일반 품목은 삼성전자 Top3 후보에서 제외
        return True
    return False

def has_real_customs_cost_signal(row: pd.Series) -> bool:
    text = _row_text(row)
    issue = clean(row.get("Issue"))
    if issue in HARD_CUSTOMS_ISSUES and any(k.lower() in text for k in COST_RISK_TERMS):
        return True
    if issue in {"AD/CVD", "반덤핑/상계관세", "CBAM", "수출통제"}:
        return True
    return False

def has_samsung_product_signal(row: pd.Series) -> bool:
    text = _row_text(row)
    return any(k in text for k in [
        "samsung", "삼성", "semiconductor", "반도체", "chip", "memory", "hbm",
        "battery", "배터리", "display", "oled", "mobile", "smartphone", "galaxy",
        "steel", "철강", "도금강판", "희토류", "rare earth", "pcb", "module",
    ])

def reference_reason(row: pd.Series) -> str:
    """v8 override: demote routine/low-value items and noisy articles."""
    if is_exchange_rate_notice(row):
        return "일본 세관의 주간 과세환율 공지로, 법인 실무 참고자료이나 HQ 임원 보고 Top3 대상은 아닙니다."
    if is_low_value_notice(row):
        return "삼성전자 주요 제품·부품·원재료와 직접 관련성이 낮거나 일반 품목/정기 공지 성격입니다."

    summary_blob = " ".join([
        clean(row.get("Summary")), clean(row.get("Original Summary")),
        clean(row.get("Major Changes")), clean(row.get("Original Body Text"))
    ])
    if is_bad_summary_text(summary_blob) and not has_real_customs_cost_signal(row):
        return "기사 본문 추출 품질이 낮아 UI 문구 중심으로 요약되었으며, 관세업무 실행성이 확인되지 않았습니다."

    issue = clean(row.get("Issue"))
    text = _row_text(row)

    # General cooperation, macro, diplomacy, market commentary without concrete customs policy
    general_noise = [
        "정상회담", "협력", "파트너십", "시장동향", "전망", "주가", "실적", "브랜드",
        "수출 85.9", "역대 최대", "경제안보", "로마 회동", "방위산업 공동",
    ]
    if any(x in text for x in general_noise) and not has_real_customs_cost_signal(row):
        return "일반 산업·외교·시장 동향 성격으로 관세/통상 실행 조치가 불명확합니다."

    if issue in SOFT_CUSTOMS_ISSUES and not has_real_customs_cost_signal(row) and not has_samsung_product_signal(row):
        return "통관/무역 키워드는 있으나 대상 HS·세율·원산지·수출통제 등 실행 조치가 확인되지 않았습니다."

    return ""

def infer_samsung_impact(row: pd.Series) -> str:
    """Direct/Indirect/Watch/Reference 재분류."""
    if reference_reason(row):
        return "Reference"
    current = clean(row.get("Samsung Impact"))
    text = _row_text(row)
    issue = clean(row.get("Issue"))

    if "samsung electronics" in text or "삼성전자" in text:
        if issue in HARD_CUSTOMS_ISSUES:
            return "Indirect"
        return "Watch"

    if issue in {"AD/CVD", "반덤핑/상계관세", "CBAM", "수출통제"}:
        return "Indirect" if has_samsung_product_signal(row) or has_real_customs_cost_signal(row) else "Watch"

    if issue in {"FTA/원산지", "HS/품목분류", "관세정책"}:
        return "Indirect" if has_samsung_product_signal(row) or has_real_customs_cost_signal(row) else "Watch"

    if issue in SOFT_CUSTOMS_ISSUES:
        return "Watch"

    return current if current in {"Direct", "Indirect", "Watch", "Reference"} else "Watch"

def executive_priority(row: pd.Series) -> str:
    """v8 override: CORE 남발 방지."""
    if reference_reason(row):
        return "REFERENCE"

    impact = infer_samsung_impact(row)
    issue = clean(row.get("Issue"))
    risk = normalize_risk(row.get("Risk"))

    if impact == "Direct":
        return "CORE"
    if issue in {"AD/CVD", "반덤핑/상계관세"} and has_real_customs_cost_signal(row):
        return "CORE"
    if issue in {"CBAM", "수출통제"} and (risk in {"상", "중"} or has_samsung_product_signal(row)):
        return "CORE"
    if issue in {"FTA/원산지", "관세정책", "HS/품목분류"} and has_real_customs_cost_signal(row):
        return "POLICY_WATCH"
    if impact == "Indirect":
        return "USABLE"
    return "WATCH"

def report_score(row: pd.Series) -> float:
    """v8 override: routine notices/noisy pages penalty."""
    priority = executive_priority(row)
    if priority == "REFERENCE":
        return -10000 + safe_num(row.get("Importance Score"))

    base = safe_num(row.get("Importance Score"))
    issue = clean(row.get("Issue"))
    impact = infer_samsung_impact(row)

    issue_weight = {
        "AD/CVD": 1800,
        "반덤핑/상계관세": 1800,
        "수출통제": 1700,
        "CBAM": 1600,
        "관세정책": 1100,
        "FTA/원산지": 1000,
        "HS/품목분류": 1000,
        "통관": 300,
        "통관/세관": 300,
        "무역일반": 0,
    }.get(issue, 150)

    impact_weight = {"Direct": 2500, "Indirect": 1200, "Watch": 300, "Reference": -3000}.get(impact, 0)
    priority_weight = {"CORE": 1800, "POLICY_WATCH": 1200, "USABLE": 600, "WATCH": 100}.get(priority, 0)

    score = base + issue_weight + impact_weight + priority_weight + risk_weight(row.get("Risk"))
    if has_samsung_product_signal(row):
        score += 600
    if has_real_customs_cost_signal(row):
        score += 900
    if is_bad_summary_text(" ".join([clean(row.get("Summary")), clean(row.get("Original Summary"))])):
        score -= 700
    if is_low_value_notice(row):
        score -= 5000
    return score

def top3_deep_score(row: pd.Series) -> float:
    if executive_priority(row) == "REFERENCE":
        return -999999
    score = report_score(row)
    issue = clean(row.get("Issue"))

    # Top3 should favor actionable strategic customs issues.
    if issue in {"AD/CVD", "반덤핑/상계관세"}:
        score += 1200
    elif issue == "수출통제":
        score += 1000
    elif issue == "CBAM":
        score += 900
    elif issue in {"FTA/원산지", "관세정책", "HS/품목분류"}:
        score += 500
    elif issue in {"통관", "통관/세관"}:
        score -= 600

    return score

def _source_summary_text(row: pd.Series, limit: int = 700) -> str:
    """v8 override: STEP4 Summary 보존하되 UI 노이즈 제거."""
    title = clean(row.get("Headline"))
    candidates = [
        clean(row.get("Original Post Summary")),
        clean(row.get("Original Summary")),
        clean(row.get("Summary")),
        clean(row.get("Original Body Text")),
        clean(row.get("Original AI Analysis")),
        clean(row.get("Impact Reason")),
    ]
    for text in candidates:
        if not text or text in {"본문에서 확인 불가", "nan", "None"}:
            continue
        text = remove_ui_noise(text)
        if _looks_like_title_only(text, title):
            continue
        if len(text) < 40:
            continue
        generated_markers = ["• 영향등급:", "• 즉시조치:", "• 이슈구분:", "영향법인 후보"]
        if any(m in text for m in generated_markers):
            continue
        return text[:limit] + ("..." if len(text) > limit else "")
    return ""

def _two_three_line_summary(row: pd.Series) -> str:
    src = _source_summary_text(row, 850)
    title = clean(row.get("Headline"))
    issue = clean(row.get("Issue"))
    country = clean(row.get("Country")) or "관련국"
    date = clean(row.get("Date")) or "게시일 확인 필요"

    if not src:
        # routine notice fallback
        if is_exchange_rate_notice(row):
            return "• 일본 세관의 주간 과세환율 공지입니다.\n• 통관신고 적용 환율 참고자료로, HQ 임원 보고 대상 중요 이슈는 아닙니다."
        return (
            f"• 원문 본문 요약이 부족합니다. 확인된 정보: {title} / {issue} / {country} / {date}\n"
            "• 대상 HS·세율·시행일·삼성 영향은 원문 또는 법인 실적 기준 추가 확인이 필요합니다."
        )

    sentences = re.split(r"(?<=[.!?。？！])\s+|(?<=다\.)\s+|(?<=니다\.)\s+", src)
    sentences = [remove_ui_noise(s.strip(" -•\n\t")) for s in sentences if remove_ui_noise(s.strip())]
    sentences = [s for s in sentences if not _looks_like_title_only(s, title) and len(s) >= 20]
    if len(sentences) >= 2:
        return "\n".join(f"• {s}" for s in sentences[:3])
    if len(sentences) == 1:
        s = sentences[0]
        return f"• {s[:260].strip()}" + (f"\n• {s[260:520].strip()}" if len(s) > 260 else "")
    return "• 원문 본문 요약이 부족하여 제목 수준 정보만 확인됩니다."

def major_changes(row: pd.Series) -> str:
    """v8 override: Summary 표시 품질 개선."""
    headline = clean(row.get("Headline"))
    issue = clean(row.get("Issue"))
    if is_exchange_rate_notice(row):
        current = "일본 세관의 주간 과세환율 공지로, 수입신고 과세가격 환산 시 참고하는 정기 고시입니다."
    elif "wheat" in headline.lower() or "밀" in headline:
        current = "인도 DGFT의 밀 수출 관련 절차 공지로, 삼성전자 주요 품목과 직접 관련성은 낮습니다."
    elif issue in {"AD/CVD", "반덤핑/상계관세"}:
        current = "반덤핑/상계관세 관련 조치로, 대상 HS·공급국·관세율·수입실적 기준 비용 영향 확인이 필요합니다."
    elif issue == "CBAM":
        current = "CBAM 관련 조치로, EU향 품목의 배출량 자료·인증서 비용·공급사 데이터 확보 여부 확인이 필요합니다."
    elif issue == "수출통제":
        current = "수출통제 관련 조치로, 대상 품목/ECCN·최종사용자·목적지 스크리닝 필요성이 있습니다."
    elif issue == "FTA/원산지":
        current = "FTA/원산지 관련 조치로, 협정 적용 가능성·CO 발급요건·BOM 원산지 정합성 확인이 필요합니다."
    else:
        parts = [
            hint_line("시행/적용일", row.get("effective_date_hint")),
            hint_line("변경 내용", row.get("change_detail_hint")),
            hint_line("대상 HS", row.get("hs_hint")),
            hint_line("관세율/쿼터", row.get("tariff_rate_hint")),
            hint_line("키워드", row.get("KeywordMatches")),
        ]
        current = compact_parts(parts, "") if any(parts) else f"{headline} 관련 관세·통상 모니터링 사안입니다."

    return f"{current}\n\n[게시물 요약]\n{_two_three_line_summary(row)}"

def report_impact(row: pd.Series) -> str:
    """v8 override: Reference/Watch 반영."""
    row = row.copy()
    row["Samsung Impact"] = infer_samsung_impact(row)
    if row["Samsung Impact"] == "Reference":
        ref = reference_reason(row)
        return f"• 영향등급: Reference • 판단사유: {ref} • 즉시 조치 불필요, 참고 모니터링 대상"
    return _samsung_impact_detail(row).replace("\n", " ")

def report_action(row: pd.Series) -> str:
    impact = infer_samsung_impact(row)
    if impact == "Reference":
        return "• 즉시조치: 불필요 • 처리: Reference 뉴스로 보관 • 후속: 동일 국가에서 전자부품·전략물자·관세율 관련 후속 공지 발생 시 재검토"
    if impact == "Watch":
        return "• 즉시조치: 본사 모니터링 • 1주 내: 대상 국가·품목·HS·시행일 확인 • 필요 시: 관련 법인/관세사에 영향 여부 확인"
    return _action_detail(row).replace("\n", " ")

def prepare_rows(rows: pd.DataFrame) -> pd.DataFrame:
    """v8 override: impact 재분류 및 Reference 강등."""
    rows = rows.copy()
    rows["Issue"] = rows.apply(issue_for, axis=1)
    rows = dedup_report_rows(rows)
    rows["Mail Group"] = rows["Content Type"].map({"Regulation": GROUP_REGULATION}).fillna(GROUP_NEWS)

    if "Original Summary" not in rows.columns:
        rows["Original Summary"] = rows.get("Summary", "")
    if "Original AI Analysis" not in rows.columns:
        rows["Original AI Analysis"] = rows.get("AI Analysis", "")
    if "Original Action Plan" not in rows.columns:
        rows["Original Action Plan"] = rows.get("Action Plan", "")
    if "Original Post Summary" not in rows.columns:
        rows["Original Post Summary"] = ""
    if "Original Body Text" not in rows.columns:
        rows["Original Body Text"] = ""

    rows["Samsung Impact"] = rows.apply(infer_samsung_impact, axis=1)
    rows["Executive Priority"] = rows.apply(executive_priority, axis=1)
    rows["Priority Group"] = rows["Executive Priority"]

    rows["Major Changes"] = rows.apply(major_changes, axis=1)
    rows["Summary"] = rows.apply(report_summary, axis=1)
    rows["AI Analysis"] = rows.apply(report_impact, axis=1)
    rows["Action Plan"] = rows.apply(report_action, axis=1)
    rows["_report_score"] = rows.apply(report_score, axis=1)

    rows = rows.sort_values(["_report_score", "_sort_date"], ascending=[False, False]).reset_index(drop=True)
    rows["No"] = range(1, len(rows) + 1)
    return rows

def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    """v8 override: exclude routine/reference and select real strategic issues."""
    pool = rows.copy()
    pool["Executive Priority"] = pool.apply(executive_priority, axis=1)
    pool["_top3_score"] = pool.apply(top3_deep_score, axis=1)

    candidate = pool[
        (pool["Executive Priority"].isin(["CORE", "POLICY_WATCH", "USABLE"])) &
        (~pool.apply(lambda r: bool(reference_reason(r)), axis=1)) &
        (pool["_top3_score"] > 0)
    ].copy()

    if candidate.empty:
        candidate = pool[pool["Executive Priority"].ne("REFERENCE")].copy()
    if candidate.empty:
        candidate = pool.copy()

    candidate = candidate.sort_values(["_top3_score", "_sort_date"], ascending=[False, False])

    selected = []
    used_norm_titles = set()
    used_issues = set()
    # issue diversity first
    for _, row in candidate.iterrows():
        title = normalize_title_key(clean(row.get("Headline")))
        issue = clean(row.get("Issue"))
        if title in used_norm_titles:
            continue
        if issue in used_issues and len(selected) < 3:
            continue
        selected.append(row)
        used_norm_titles.add(title)
        used_issues.add(issue)
        if len(selected) == 3:
            break
    # fill if needed
    if len(selected) < 3:
        for _, row in candidate.iterrows():
            title = normalize_title_key(clean(row.get("Headline")))
            if title in used_norm_titles:
                continue
            selected.append(row)
            used_norm_titles.add(title)
            if len(selected) == 3:
                break

    out = pd.DataFrame(selected).reset_index(drop=True)
    if not out.empty:
        out["No"] = range(1, len(out) + 1)
    return out

def normalize_title_key(title: str) -> str:
    t = clean(title).lower()
    t = re.sub(r"[-–—|].*$", "", t)
    t = re.sub(r"[^a-z0-9가-힣]+", " ", t)
    return re.sub(r"\s+", " ", t).strip()[:80]

def top3_summary_sentence(row: pd.Series) -> str:
    title = clean(row.get("Headline"))
    issue = clean(row.get("Issue"))
    if is_exchange_rate_notice(row):
        return f"{title}: 주간 과세환율 참고자료로 Top3 제외 대상입니다."
    if issue in {"AD/CVD", "반덤핑/상계관세"}:
        return f"{title}: 추가관세에 따른 조달비용 및 원산지 방어 리스크 점검이 필요합니다."
    if issue == "수출통제":
        return f"{title}: AI·반도체 등 전략기술 수출통제 확산 가능성에 대한 스크리닝 강화가 필요합니다."
    if issue == "CBAM":
        return f"{title}: EU향 품목의 배출량 데이터와 CBAM 비용 대응이 필요합니다."
    if issue == "FTA/원산지":
        return f"{title}: 공급망 다변화와 FTA 특혜관세 활용 가능성을 검토해야 합니다."
    if issue == "관세정책":
        return f"{title}: 관세율·쿼터·시행일 변화에 따른 비용 영향 확인이 필요합니다."
    return f"{title}: 삼성 관련성과 관세업무 실행 필요성을 추가 확인해야 합니다."

# ======================================================================
# End of GTI STEP5 Executive Selection & Summary Quality Patch v8
# ======================================================================


# ======================================================================
# GTI STEP5 UNIPASS URL & Top3 Detail Patch v9 - 2026-06-14
# ----------------------------------------------------------------------
# 1) UNIPASS rowTitle 목록 URL을 ntarId 직접열람 URL로 보정
# 2) Top3 상세분석을 제도/기간/영향/Action 중심으로 상세화
# 3) PN51 Advance Authorization/EPCG 수출의무 연장 공지 특화 분석 추가
# ======================================================================

UNIPASS_NOTICE_URL_PREFIX = (
    "https://unipass.customs.go.kr/csp/myc/custsppt/cmmn/"
    "NtarBrkdMtCtr/openMYC0605014Q.do?ntarId="
)

UNIPASS_NOTICE_ID_BY_TITLE = {
    "다수 사업장 운영 사업자의 전자상거래업자 등록 신청 방법": "202606122928",
}


def normalize_korean_title(text: str) -> str:
    text = clean(text)
    text = re.sub(r"\([^)]*\)|\[[^]]*\]", " ", text)
    text = re.sub(r"[^0-9A-Za-z가-힣]+", " ", text)
    return re.sub(r"\s+", " ", text).strip().lower()


def unipass_direct_url_from_title(title: str) -> str:
    norm_title = normalize_korean_title(title)
    for key, ntar_id in UNIPASS_NOTICE_ID_BY_TITLE.items():
        norm_key = normalize_korean_title(key)
        if norm_key and (norm_key in norm_title or norm_title in norm_key):
            return UNIPASS_NOTICE_URL_PREFIX + ntar_id
    return ""


def fix_unipass_url(row: pd.Series) -> str:
    url = clean(row.get("URL"))
    title = clean(row.get("Headline"))
    agency = clean(row.get("Agency"))
    direct = unipass_direct_url_from_title(title)
    if direct:
        return direct

    low = url.lower()
    # 이미 직접열람 URL이면 그대로 사용
    if "unipass.customs.go.kr" in low and "openmyc0605014q.do" in low and "ntarid=" in low:
        return url
    # rowTitle 목록 URL은 클릭 성공률이 낮으므로, 알려진 ntarId가 없으면 기존 URL 유지
    # 단, 이후 매핑을 추가하기 쉽도록 이 함수 한 곳에서 관리한다.
    if "unipass" in agency.lower() or "유니패스" in agency or "unipass.customs.go.kr" in low:
        return direct or url
    return url


def is_pn51_export_obligation(row: pd.Series) -> bool:
    text = _text_blob(row).lower() if "_text_blob" in globals() else " ".join(clean(row.get(c)) for c in ["Headline", "Summary", "Major Changes"]).lower()
    return (
        "pn 51" in text
        or "export obligation period" in text
        or ("advance authorization" in text and "epcg" in text)
        or ("수출의무" in text and "epcg" in text.lower())
    )


def detailed_pn51_summary() -> str:
    return (
        "인도 정부가 수출촉진을 위해 Advance Authorization 및 EPCG 수출의무 이행기간을 "
        "2026년 8월까지 자동 연장함에 따라 인도 생산법인의 수출의무 이행 부담이 완화될 전망입니다.\n\n"
        "주요 내용은 다음과 같습니다.\n"
        "- Advance Authorization EO 만료기간이 2026.03.01~2026.05.31인 경우 → 2026.08.31까지 자동 연장\n"
        "- EPCG(Block-wise EO) 만료기간이 2026.03.01~2026.05.31인 경우 → 2026.08.31까지 자동 연장\n"
        "- 별도 신청서 제출 불필요\n"
        "- 연장 수수료(Composition Fee) 면제\n"
        "- 세관은 연장된 EO 기준으로 수출을 인정\n"
        "- EO 충족 여부는 EODC 발급 시 최종 검증"
    )


def detailed_pn51_impact() -> str:
    return (
        "간접 영향 (Watch Level)\n"
        "본 공고는 인도 정부의 수출지원 조치로 삼성전자에 직접적인 관세 인상 또는 수입규제 영향을 주는 정책은 아닙니다. "
        "다만 삼성전자 인도 생산법인 또는 협력업체가 Advance Authorization 또는 EPCG(Export Promotion Capital Goods)를 "
        "활용하는 경우 실무 영향이 발생할 수 있습니다.\n\n"
        "관세업무 관점 영향\n"
        "1. EO 미충족 리스크 완화\n"
        "- 인도 생산법인이 원자재를 무관세 수입 후 수출의무를 부여받은 경우 EO 기간이 자동 연장됩니다.\n"
        "- EO 미충족에 따른 추징관세, 이자부담, 허가 취소 리스크가 감소합니다.\n\n"
        "2. 공급망 운영 유연성 증가\n"
        "- 홍해 사태, 중동 물류 리스크, 글로벌 공급망 재편으로 수출계획 달성이 어려운 기업의 부담을 완화합니다.\n\n"
        "3. 삼성전자 인도 생산법인 영향\n"
        "- 휴대폰, TV, 가전제품 생산 시 수입부품에 Advance Authorization 또는 EPCG를 활용하고 있다면 "
        "EO 달성 일정에 추가 여유가 확보됩니다."
    )


def detailed_pn51_action() -> str:
    return (
        "삼성전자 인도 법인의 Advance Authorization 및 EPCG 활용 현황을 점검하고 "
        "EO 만료 예정 건에 대한 연장 적용 여부를 확인할 것을 권고합니다.\n\n"
        "즉시 조치\n"
        "- 인도 법인 확인: Advance Authorization 사용 여부, EPCG 사용 여부, EO 만료 예정 허가 현황\n\n"
        "1주 내\n"
        "- Authorization No, 제도구분(AA/EPCG), 기존 EO 만료일, 연장적용 여부(Y/N), 예상 EO 달성률 리스트 작성\n\n"
        "1개월 내\n"
        "- ONE-Origin 시스템에 Authorization 번호, EO 만료일, EO 달성률, EODC 발급 여부 관리항목 추가 검토\n\n"
        "Owner: HQ Customs / India subsidiary trade compliance"
    )


def _v9_generic_top3_summary(row: pd.Series) -> str:
    issue = clean(row.get("Issue"))
    title = clean(row.get("Headline"))
    text = _text_blob(row).lower() if "_text_blob" in globals() else title.lower()
    base = clean(row.get("Major Changes")) or clean(row.get("Summary"))

    if issue in {"AD/CVD", "반덤핑/상계관세"} or any(k in text for k in ["anti-dumping", "antidumping", "countervailing", "덤핑", "상계관세"]):
        return (
            f"{base}\n\n"
            "추가 확인 포인트\n"
            "- 대상 HS 및 제품 사양이 삼성전자 또는 협력사 조달품목과 겹치는지 확인\n"
            "- 중국산/제3국산 우회수출 여부와 원산지 증빙자료 방어 가능성 점검\n"
            "- 잠정관세율 또는 최종판정 관세율을 기준으로 원가 영향 시뮬레이션 필요"
        )
    if issue == "CBAM" or "cbam" in text:
        return (
            f"{base}\n\n"
            "추가 확인 포인트\n"
            "- EU 수출품의 CBAM 대상 여부, 내재배출량 산정자료, 인증서 구매비용 영향 확인\n"
            "- 철강·알루미늄·부품 공급망 내 벤더별 탄소자료 확보 가능성 점검"
        )
    if issue in {"FTA/원산지"} or any(k in text for k in ["fta", "cepa", "origin", "원산지"]):
        return (
            f"{base}\n\n"
            "추가 확인 포인트\n"
            "- 대상 법인의 BOM 원산지, CO 발급요건, 직접운송 요건 충족 여부 확인\n"
            "- FTA Master 및 협정세율 적용 가능성을 구매/물류/관세 데이터와 대사"
        )
    if issue in {"통관", "통관/세관"} or any(k in text for k in ["customs", "통관", "보세", "과세가격"]):
        return (
            f"{base}\n\n"
            "추가 확인 포인트\n"
            "- 통관신고, 보세운송, 반출입신고, 과세가격 자료 제출 프로세스 변경 여부 확인\n"
            "- 국내외 법인의 신고 자동화/마스터 데이터/증빙 보관 기준 반영 필요"
        )
    return base or f"{title} 관련 관세·통상 영향과 삼성전자 적용 가능성을 추가 확인해야 합니다."


def _v9_generic_top3_impact(row: pd.Series) -> str:
    issue = clean(row.get("Issue"))
    impact = clean(row.get("Samsung Impact")) or "Watch"
    country = clean(row.get("Country")) or "관련국"
    return (
        f"영향등급: {impact}\n"
        f"대상 국가/지역: {country}\n"
        "삼성전자 본사 관세담당자 관점에서는 대상 HS, 공급국, 벤더, 법인별 수입·수출 실적을 먼저 매핑해야 합니다. "
        "직접 관세 인상 여부가 명확하지 않더라도, 원산지 증빙, 통관신고, 협정세율 적용, 공급망 비용에 영향을 줄 수 있는 이슈로 관리가 필요합니다."
    )


def _v9_generic_top3_action(row: pd.Series) -> str:
    issue = clean(row.get("Issue"))
    if issue in {"AD/CVD", "반덤핑/상계관세", "AD_CVD"}:
        return (
            "즉시 조치: 대상 HS·공급국·벤더 리스트 확인 및 중국산/우회수출 가능성 점검\n"
            "1주 내: 최근 수입실적 기준 잠정/최종 AD-CVD 비용 영향 산출\n"
            "1개월 내: 원산지 증빙, 가격자료, 공급계약 Incoterms 및 관세부담 주체 정비\n"
            "Owner: HQ Customs / Procurement / Regional trade compliance"
        )
    if issue == "CBAM":
        return (
            "즉시 조치: EU 수출품 중 CBAM 대상 품목 여부 확인\n"
            "1주 내: 벤더별 내재배출량 자료 확보 가능성 및 인증서 비용 추정\n"
            "1개월 내: CBAM 신고자료 수집 체계와 구매계약상 탄소자료 제출 의무 반영 검토\n"
            "Owner: HQ Customs / ESG / EU subsidiary"
        )
    if issue in {"통관", "통관/세관"}:
        return (
            "즉시 조치: 대상 법인 및 신고 프로세스 적용 여부 확인\n"
            "1주 내: 보세·통관·과세가격 관련 마스터 데이터와 증빙자료 점검\n"
            "1개월 내: 시스템 변경사항과 업무 SOP 반영 여부 확인\n"
            "Owner: HQ Customs / Customs broker / Relevant subsidiary"
        )
    return (
        "즉시 조치: 관련 법인 적용 가능성 확인\n"
        "1주 내: 대상 국가·품목·HS·법인 매핑\n"
        "1개월 내: 후속 공지 모니터링 및 필요 시 Master 반영\n"
        "Owner: HQ Customs"
    )


def major_changes(row: pd.Series) -> str:
    if is_pn51_export_obligation(row):
        return detailed_pn51_summary()
    return _v9_generic_top3_summary(row)


def report_impact(row: pd.Series) -> str:
    if is_pn51_export_obligation(row):
        return detailed_pn51_impact()
    return _v9_generic_top3_impact(row)


def report_action(row: pd.Series) -> str:
    if is_pn51_export_obligation(row):
        return detailed_pn51_action()
    return _v9_generic_top3_action(row)


def prepare_rows(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    rows["Issue"] = rows.apply(issue_for, axis=1)
    rows = dedup_report_rows(rows)
    rows["Mail Group"] = rows["Content Type"].map({"Regulation": GROUP_REGULATION}).fillna(GROUP_NEWS)

    if "Original Summary" not in rows.columns:
        rows["Original Summary"] = rows.get("Summary", "")
    if "Original AI Analysis" not in rows.columns:
        rows["Original AI Analysis"] = rows.get("AI Analysis", "")
    if "Original Action Plan" not in rows.columns:
        rows["Original Action Plan"] = rows.get("Action Plan", "")
    if "Original Post Summary" not in rows.columns:
        rows["Original Post Summary"] = ""
    if "Original Body Text" not in rows.columns:
        rows["Original Body Text"] = ""

    rows["URL"] = rows.apply(fix_unipass_url, axis=1)
    rows["Samsung Impact"] = rows.apply(infer_samsung_impact, axis=1)
    rows["Executive Priority"] = rows.apply(executive_priority, axis=1)
    rows["Priority Group"] = rows["Executive Priority"]
    rows["Major Changes"] = rows.apply(major_changes, axis=1)
    rows["Summary"] = rows.apply(report_summary, axis=1)
    rows["AI Analysis"] = rows.apply(report_impact, axis=1)
    rows["Action Plan"] = rows.apply(report_action, axis=1)
    rows["_report_score"] = rows.apply(report_score, axis=1)

    rows = rows.sort_values(["_report_score", "_sort_date"], ascending=[False, False]).reset_index(drop=True)
    rows["No"] = range(1, len(rows) + 1)
    return rows


# ======================================================================
# End of GTI STEP5 UNIPASS URL & Top3 Detail Patch v9 - 2026-06-14
# ======================================================================


# ======================================================================
# GTI STEP5 Recovery Patch v10 - 2026-06-14
# ----------------------------------------------------------------------
# v9가 상세분석을 강화하면서 이전의 뉴스 보강/법규 전체 유지 로직을
# 덮어쓴 문제를 복구한다.
# 1) 법규는 신규/변경 법규 전체 유지
# 2) 뉴스는 4-2 audit candidates에서 보강하여 30~50건 유지
# 3) Top3는 Reference/일반뉴스 제외, 실행형 관세·통상 이슈 중심
# ======================================================================

NEWS_MIN_REPORT_ROWS = int(os.getenv("GTI_NEWS_MIN_REPORT_ROWS", "30"))
NEWS_MAX_REPORT_ROWS = int(os.getenv("GTI_NEWS_MAX_REPORT_ROWS", "50"))
NEWS_AUDIT_INPUT_FILE = Path(os.getenv("GTI_NEWS_AUDIT_INPUT_FILE", r"C:\Temp\4-2.news_ai_audit_candidates.xlsx"))


def _blob_v10(row: pd.Series) -> str:
    return " ".join(
        clean(row.get(c))
        for c in [
            "Headline", "Issue", "Summary", "Major Changes", "AI Analysis", "Action Plan",
            "Original Summary", "Original Post Summary", "Original Body Text", "Impact Reason",
            "KeywordMatches", "Agency", "Country"
        ]
    )


def bad_url_v10(url: str) -> bool:
    low = clean(url).lower()
    if not low:
        return True
    return "news.google.com" in low or low in {
        "https://news.google.com", "https://news.google.com/",
        "https://google.com", "https://www.google.com", "https://www.google.com/",
    }


def hard_reference_v10(row: pd.Series) -> bool:
    text = _blob_v10(row).lower()
    if clean(row.get("Content Type")) == "Regulation":
        return is_exchange_rate_notice(row) or "modalities for export of wheat" in text or "export of wheat" in text
    weak = [
        "신간", "서평", "bookreview", "/culture/", "문화", "혈통관리",
        "세관인", "주무관 선정", "공무원 선정", "표창", "수상",
        "주가", "증시", "코스피", "코스닥", "환율", "금리", "부동산",
        "미토스", "주술", "길 잃은 삼성", "ax",
    ]
    trade = [
        "tariff", "관세", "customs", "통관", "보세", "fta", "cepa", "origin", "원산지",
        "cbam", "anti-dumping", "antidumping", "countervailing", "덤핑", "상계관세",
        "export control", "수출통제", "entity list", "forced labor", "uflpa",
    ]
    return any(k in text for k in weak) and not any(k in text for k in trade)


def normalized_issue_v10(row: pd.Series) -> str:
    issue = clean(row.get("Issue")) or issue_for(row)
    text = _blob_v10(row).lower()
    if issue in {"AD/CVD", "반덤핑/상계관세"} or any(k in text for k in ["anti-dumping", "antidumping", "countervailing", "ad/cvd", "덤핑방지", "상계관세"]):
        return "AD_CVD"
    if issue in {"통관", "통관/세관"} or any(k in text for k in ["customs", "통관", "보세", "과세가격", "반출입신고"]):
        return "CUSTOMS"
    if issue == "CBAM" or "cbam" in text:
        return "CBAM"
    if issue == "FTA/원산지" or any(k in text for k in ["fta", "cepa", "origin", "원산지"]):
        return "FTA_ORIGIN"
    if issue == "수출통제" or any(k in text for k in ["export control", "entity list", "uflpa", "forced labor", "수출통제", "강제노동"]):
        return "EXPORT_CONTROL"
    if issue == "관세정책" or any(k in text for k in ["tariff", "관세", "quota", "section 301", "section 232"]):
        return "TARIFF_POLICY"
    return issue


def display_issue_v10(row: pd.Series) -> str:
    norm = normalized_issue_v10(row)
    return {
        "AD_CVD": "AD/CVD",
        "CUSTOMS": "통관/세관",
        "CBAM": "CBAM",
        "FTA_ORIGIN": "FTA/원산지",
        "EXPORT_CONTROL": "수출통제",
        "TARIFF_POLICY": "관세정책",
    }.get(norm, clean(row.get("Issue")) or issue_for(row))


def issue_key_v10(row: pd.Series) -> str:
    text = _blob_v10(row).lower()
    norm = normalized_issue_v10(row)
    if norm == "AD_CVD" and any(k in text for k in ["zinc", "아연", "galvanized", "도금", "cold-rolled", "냉간압연", "steel", "철강"]):
        return "AD_CVD_STEEL_ZINC"
    if norm == "CBAM" and ("certificate" in text or "인증서" in text):
        return "CBAM_CERTIFICATE"
    if norm == "FTA_ORIGIN" and ("morocco" in text or "모로코" in text):
        return "FTA_MOROCCO_CEPA"
    if norm == "CUSTOMS" and "보세공장" in text:
        return "CUSTOMS_BONDED_FACTORY"
    if norm == "CUSTOMS" and "보세창고" in text:
        return "CUSTOMS_BONDED_WAREHOUSE"
    title = normalize_korean_title(clean(row.get("Headline")))
    return f"{norm}:{title[:90]}"


def news_score_v10(row: pd.Series) -> float:
    text = _blob_v10(row).lower()
    score = safe_num(row.get("Importance Score")) + risk_weight(row.get("Risk")) + priority_weight(row.get("Priority Group"))
    for term in [
        "anti-dumping", "antidumping", "countervailing", "덤핑", "상계관세",
        "cbam", "section 301", "section 232", "tariff", "관세", "quota",
        "export control", "entity list", "forced labor", "uflpa", "수출통제",
        "fta", "cepa", "origin", "원산지", "customs", "통관", "보세",
        "battery", "배터리", "semiconductor", "반도체", "steel", "철강", "rare earth", "희토류",
    ]:
        if term in text:
            score += 220
    if bad_url_v10(row.get("URL")):
        score -= 2000
    if hard_reference_v10(row):
        score -= 3000
    return score


def read_step4_results() -> pd.DataFrame:
    frames = []
    if REGULATION_INPUT_FILE.exists():
        frames.append(normalize_input(pd.read_excel(REGULATION_INPUT_FILE), "Regulation", REGULATION_INPUT_FILE))

    news_frames = []
    if NEWS_INPUT_FILE.exists():
        news_frames.append(normalize_input(pd.read_excel(NEWS_INPUT_FILE), "News", NEWS_INPUT_FILE))
    if NEWS_AUDIT_INPUT_FILE.exists():
        try:
            news_frames.append(normalize_input(pd.read_excel(NEWS_AUDIT_INPUT_FILE), "News", NEWS_AUDIT_INPUT_FILE))
        except Exception as exc:
            print(f"[WARN] news audit top-up skipped: {NEWS_AUDIT_INPUT_FILE} / {exc}")

    if news_frames:
        news = pd.concat(news_frames, ignore_index=True)
        news["URL"] = news.apply(fix_unipass_url, axis=1)
        news = news[~news["URL"].apply(bad_url_v10)].copy()
        news["_issue_key"] = news.apply(issue_key_v10, axis=1)
        news["_news_score"] = news.apply(news_score_v10, axis=1)
        news = news.sort_values(["_news_score", "_sort_date"], ascending=[False, False])
        news = news.drop_duplicates(subset=["URL"], keep="first")
        news = news.drop_duplicates(subset=["_issue_key"], keep="first")
        max_rows = NEWS_MAX_ROWS if NEWS_MAX_ROWS > 0 else NEWS_MAX_REPORT_ROWS
        max_rows = max(NEWS_MIN_REPORT_ROWS, min(NEWS_MAX_REPORT_ROWS, max_rows))
        frames.append(news.head(max_rows).drop(columns=["_issue_key", "_news_score"], errors="ignore"))

    if not frames:
        raise FileNotFoundError(f"STEP4 outputs not found: {REGULATION_INPUT_FILE}, {NEWS_INPUT_FILE}")

    rows = pd.concat(frames, ignore_index=True)
    rows["URL"] = rows.apply(fix_unipass_url, axis=1)
    rows["_dedup_key"] = rows.apply(lambda r: clean(r.get("URL")).lower() or clean(r.get("Headline"))[:160], axis=1)
    rows = rows.drop_duplicates(subset=["_dedup_key"], keep="first").drop(columns=["_dedup_key"], errors="ignore")
    rows["_integrated_score"] = rows.apply(
        lambda r: priority_weight(r["Priority Group"]) + risk_weight(r["Risk"]) + (180 if r["Content Type"] == "Regulation" else 0) + safe_num(r["Importance Score"]),
        axis=1,
    )
    return rows.reset_index(drop=True)


def infer_samsung_impact(row: pd.Series) -> str:
    text = _blob_v10(row).lower()
    norm = normalized_issue_v10(row)
    if hard_reference_v10(row):
        return "Reference"
    if norm in {"AD_CVD", "CBAM", "FTA_ORIGIN", "EXPORT_CONTROL", "TARIFF_POLICY", "CUSTOMS"}:
        return "Indirect"
    if any(k in text for k in ["tariff", "관세", "customs", "통관", "origin", "원산지", "fta", "cepa"]):
        return "Watch"
    return clean(row.get("Samsung Impact")) or "Watch"


def executive_priority(row: pd.Series) -> str:
    impact = infer_samsung_impact(row)
    norm = normalized_issue_v10(row)
    if impact == "Reference":
        return "REFERENCE"
    if norm in {"AD_CVD", "CBAM", "EXPORT_CONTROL", "TARIFF_POLICY", "CUSTOMS"}:
        return "CORE"
    if norm in {"FTA_ORIGIN"}:
        return "POLICY_WATCH"
    return "WATCH"


def report_score(row: pd.Series) -> float:
    norm = normalized_issue_v10(row)
    impact = infer_samsung_impact(row)
    impact_w = {"Direct": 2500, "Indirect": 1200, "Watch": 200, "Reference": -2000}.get(impact, 0)
    norm_w = {"AD_CVD": 1800, "EXPORT_CONTROL": 1600, "CBAM": 1500, "TARIFF_POLICY": 1300, "CUSTOMS": 1200, "FTA_ORIGIN": 1000}.get(norm, 200)
    type_w = 800 if clean(row.get("Content Type")) == "Regulation" else 0
    return safe_num(row.get("Importance Score")) + risk_weight(row.get("Risk")) + impact_w + norm_w + type_w + news_score_v10(row) / 5


def prepare_rows(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    rows["Issue"] = rows.apply(issue_for, axis=1)
    rows["Issue"] = rows.apply(display_issue_v10, axis=1)

    if "Original Summary" not in rows.columns:
        rows["Original Summary"] = rows.get("Summary", "")
    if "Original AI Analysis" not in rows.columns:
        rows["Original AI Analysis"] = rows.get("AI Analysis", "")
    if "Original Action Plan" not in rows.columns:
        rows["Original Action Plan"] = rows.get("Action Plan", "")
    if "Original Post Summary" not in rows.columns:
        rows["Original Post Summary"] = ""
    if "Original Body Text" not in rows.columns:
        rows["Original Body Text"] = ""

    reg = rows[rows["Content Type"].eq("Regulation")].copy()
    news = rows[rows["Content Type"].eq("News")].copy()
    news = news[~news["URL"].apply(bad_url_v10)].copy()
    news["_issue_key"] = news.apply(issue_key_v10, axis=1)
    news["_report_score_pre"] = news.apply(report_score, axis=1)
    news = news.sort_values(["_report_score_pre", "_sort_date"], ascending=[False, False])
    news = news.drop_duplicates(subset=["URL"], keep="first")
    news = news.drop_duplicates(subset=["_issue_key"], keep="first").drop(columns=["_issue_key", "_report_score_pre"], errors="ignore")
    news = news.head(NEWS_MAX_REPORT_ROWS)

    rows = pd.concat([reg, news], ignore_index=True)
    rows["URL"] = rows.apply(fix_unipass_url, axis=1)
    rows["Mail Group"] = rows["Content Type"].map({"Regulation": GROUP_REGULATION}).fillna(GROUP_NEWS)
    rows["Samsung Impact"] = rows.apply(infer_samsung_impact, axis=1)
    rows["Executive Priority"] = rows.apply(executive_priority, axis=1)
    rows["Priority Group"] = rows["Executive Priority"]
    rows["Major Changes"] = rows.apply(major_changes, axis=1)
    rows["Summary"] = rows.apply(report_summary, axis=1)
    rows["AI Analysis"] = rows.apply(report_impact, axis=1)
    rows["Action Plan"] = rows.apply(report_action, axis=1)
    rows["_report_score"] = rows.apply(report_score, axis=1)
    rows = rows.sort_values(["_report_score", "_sort_date"], ascending=[False, False]).reset_index(drop=True)
    rows["No"] = range(1, len(rows) + 1)
    return rows


def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    pool = rows.copy()
    pool["Samsung Impact"] = pool.apply(infer_samsung_impact, axis=1)
    pool["Executive Priority"] = pool.apply(executive_priority, axis=1)
    pool["_issue_key"] = pool.apply(issue_key_v10, axis=1)
    pool["_issue_type"] = pool.apply(normalized_issue_v10, axis=1)
    pool["_top3_score"] = pool.apply(report_score, axis=1)
    pool = pool[(pool["Executive Priority"].ne("REFERENCE")) & (~pool.apply(hard_reference_v10, axis=1))].copy()
    if pool.empty:
        pool = rows.copy()
        pool["_issue_key"] = pool.apply(issue_key_v10, axis=1)
        pool["_issue_type"] = pool.apply(normalized_issue_v10, axis=1)
        pool["_top3_score"] = pool.apply(report_score, axis=1)
    pool = pool.sort_values(["_top3_score", "_sort_date"], ascending=[False, False])

    selected, used_types, used_keys = [], set(), set()
    for _, row in pool.iterrows():
        typ = clean(row.get("_issue_type"))
        key = clean(row.get("_issue_key"))
        if typ in used_types or key in used_keys:
            continue
        selected.append(row)
        used_types.add(typ)
        used_keys.add(key)
        if len(selected) == 3:
            break
    if len(selected) < 3:
        for _, row in pool.iterrows():
            key = clean(row.get("_issue_key"))
            if key in used_keys:
                continue
            selected.append(row)
            used_keys.add(key)
            if len(selected) == 3:
                break
    out = pd.DataFrame(selected).drop(columns=["_issue_key", "_issue_type", "_top3_score"], errors="ignore").reset_index(drop=True)
    if not out.empty:
        out["No"] = range(1, len(out) + 1)
    return out


# ======================================================================
# End of GTI STEP5 Recovery Patch v10 - 2026-06-14
# ======================================================================


# ======================================================================
# GTI STEP5 Report Sensing Patch v11 - 2026-06-14
# ----------------------------------------------------------------------
# 목적: 업무 변화 센싱 보고서 기준 고정
# - 법규는 Step4-1 daily에 있는 항목 모두 기재
# - 뉴스는 30건 이하
# - Top3는 관세/통상 법규성 30 + 정책 20 + 직접영향 40 + 간접영향 10 기준
# - 게시날짜를 Publish Date로 표시
# - 원문 URL 필수, Google Alert/Google News wrapper 제외
# - Summary 출력 직전 UI 문구 제거
# ======================================================================

NEWS_MAX_REPORT_ROWS = min(int(os.getenv("GTI_NEWS_MAX_REPORT_ROWS", "30")), 30)
NEWS_MIN_REPORT_ROWS = min(int(os.getenv("GTI_NEWS_MIN_REPORT_ROWS", "30")), NEWS_MAX_REPORT_ROWS)

if "Publish Date" not in OUTPUT_COLUMNS:
    try:
        OUTPUT_COLUMNS.insert(OUTPUT_COLUMNS.index("Date") + 1, "Publish Date")
    except Exception:
        OUTPUT_COLUMNS.append("Publish Date")


_REPORT_UI_NOISE_V11 = [
    "이전 기사보기", "다음 기사보기", "기사의 본문 내용은 이 글자크기로 변경됩니다",
    "본문 글씨 키우기", "본문 글씨 줄이기", "스크롤 이동 상태바", "가 가",
    "바로가기 복사하기", "공유하기", "본문영역", "기사원문", "추천기사",
]


def clean_report_text_v11(value: str) -> str:
    text = clean(value)
    for phrase in _REPORT_UI_NOISE_V11:
        text = text.replace(phrase, " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def bad_url_v11(url: str) -> bool:
    low = clean(url).lower()
    if not low:
        return True
    return any(k in low for k in [
        "news.google.com",
        "google.co.kr/alerts/feeds",
        "google.com/alerts/feeds",
        "google.co.kr/rss",
        "google.com/rss",
    ])


def blob_v11(row: pd.Series) -> str:
    return " ".join(clean(row.get(c)) for c in [
        "Headline", "Issue", "Summary", "Major Changes", "AI Analysis", "Action Plan",
        "Original Summary", "Original Post Summary", "Original Body Text", "Impact Reason",
        "Agency", "Country", "KeywordMatches", "URL"
    ]).lower()


def hard_reference_v11(row: pd.Series) -> bool:
    text = blob_v11(row)
    if bad_url_v11(row.get("URL")):
        return True
    if clean(row.get("Content Type")) == "Regulation":
        if any(k in text for k in ["rate of exchange", "exchange rate", "과세환율", "export of wheat", "wheat reg"]):
            return True
        title = clean(row.get("Headline")).lower()
        source_text = " ".join([
            clean(row.get("Headline")),
            clean(row.get("Major Changes")),
            clean(row.get("Summary")),
            clean(row.get("OriginalArticle")),
            clean(row.get("article_body")),
        ]).lower()
        concrete_reg_change = any(k in source_text for k in [
            "anti-dumping", "antidumping", "countervailing", "ad/cvd", "덤핑", "상계관세",
            "export obligation", "advance authorization", "epcg",
            "export control", "entity list", "forced labor", "uflpa", "수출통제", "전략물자", "강제노동",
            "fta", "cepa", "tepa", "rules of origin", "certificate of origin", "origin", "원산지", "협정세율",
            "customs duty", "import duty", "tariff rate", "tariff quota", "hs code", "classification",
            "customs clearance", "customs declaration", "bonded", "bonded warehouse",
            "관세율", "할당관세", "품목분류", "통관절차", "수입신고", "수출신고", "보세", "보세창고", "과세가격",
            "e-commerce exporter", "electronic commerce exporter", "전자상거래업자",
            "cbam", "carbon border", "탄소국경",
        ])
        generic_notice_title = (
            re.search(r"\b(public|publick|trade|trrade)\s+notice\b", title)
            or title in {"public notice eng", "trade notice"}
            or "credit assistance" in title
            or "emerging export opportunities" in title
            or "interest subvention" in title
            or "collateral support" in title
            or "bank validation" in title
            or "testing inspections" in title
            or "labsetu" in title
        )
        if any(k in title for k in [
            "credit assistance", "emerging export opportunities", "interest subvention",
            "collateral support", "bank validation", "alternative trade instruments",
            "testing inspections", "labsetu",
        ]):
            return True
        if generic_notice_title:
            # Top3 should not be promoted by generic fallback phrases such as
            # "check HS/tariff/customs impact".  Generic DGFT notices need a
            # concrete operational keyword in the title or source summary.
            if not concrete_reg_change:
                return True
        if not concrete_reg_change:
            return True
        return False
    # 제목 자체가 산업/외교/행사/시장 일반뉴스이면 Issue 컬럼의 오분류보다 우선해 제외한다.
    hard_weak_first = [
        "보안시장", "인터롭", "전방위 협력", "유럽순방", "순방", "정상회의",
        "교황", "피렌체", "면담", "방문",
        "미토스", "주술", "길 잃은 삼성", "신간", "서평", "bookreview",
        "주가", "증시", "record margins", "memory costs",
        "industrial ecosystems", "finance must be a partner",
        "몰카", "범죄", "청년인턴", "채용", "합격자",
        "laboratory system in libya", "developing laboratory system",
    ]
    if any(k in text for k in hard_weak_first):
        title_only = clean(row.get("Headline")).lower()
        concrete_in_text = any(k in title_only for k in [
            "tariff", "관세", "customs duty", "anti-dumping", "antidumping", "countervailing",
            "cbam", "forced labor", "uflpa", "section 301", "section 232", "fta", "cepa",
        ])
        if not concrete_in_text:
            return True
    if not has_strong_trade_action_v11(row):
        return True
    weak = [
        "신간", "서평", "bookreview", "/culture/", "문화", "혈통관리",
        "세관인", "주무관 선정", "표창", "수상",
        "주가", "증시", "코스피", "환율", "금리", "부동산",
        "미토스", "주술", "길 잃은 삼성", "apple", "record margins", "memory costs",
        "보안시장", "인터롭", "전방위 협력", "순방", "정상회의",
        "몰카", "범죄", "청년인턴", "채용", "합격자", "industrial ecosystems",
        "finance must be a partner",
    ]
    policy = ["관세", "통관", "fta", "원산지", "cbam", "수출통제", "anti-dumping", "덤핑", "상계관세", "tariff", "customs", "quota"]
    return any(k in text for k in weak) and not any(k in text for k in policy)


def has_strong_trade_action_v11(row: pd.Series) -> bool:
    text = blob_v11(row)
    return any(k in text for k in [
        "tariff", "tariffs", "customs duty", "import duty", "관세", "관세율", "쿼터", "quota",
        "customs", "clearance", "declaration", "통관", "보세", "수입신고", "수출신고",
        "anti-dumping", "antidumping", "countervailing", "ad/cvd", "반덤핑", "상계관세", "덤핑방지",
        "cbam", "carbon border", "탄소국경",
        "fta", "cepa", "rules of origin", "origin", "원산지", "협정세율",
        "export control", "entity list", "forced labor", "uflpa", "수출통제", "전략물자", "강제노동",
        "section 301", "section 232", "301조", "232조",
        "hs code", "classification", "품목분류",
    ])


def norm_issue_v11(row: pd.Series) -> str:
    issue = clean(row.get("Issue"))
    text = blob_v11(row)
    if issue in {"AD/CVD", "반덤핑/상계관세"} or any(k in text for k in ["anti-dumping", "antidumping", "countervailing", "ad/cvd", "반덤핑", "상계관세"]):
        return "AD_CVD"
    if issue in {"통관", "통관/세관"} or any(k in text for k in ["customs", "통관", "보세", "과세가격", "반출입신고"]):
        return "CUSTOMS"
    if issue == "CBAM" or "cbam" in text:
        return "CBAM"
    if issue == "FTA/원산지" or any(k in text for k in ["fta", "cepa", "origin", "원산지"]):
        return "FTA_ORIGIN"
    if issue == "수출통제" or any(k in text for k in ["export control", "entity list", "uflpa", "forced labor", "수출통제", "강제노동"]):
        return "EXPORT_CONTROL"
    if issue == "관세정책" or any(k in text for k in ["tariff", "관세", "quota", "section 301", "section 232"]):
        return "TARIFF_POLICY"
    return issue or "기타"


def issue_key_v11(row: pd.Series) -> str:
    title = clean(row.get("Headline")).lower()
    title = re.sub(r"\([^)]*\)|\[[^]]*\]", " ", title)
    title = re.sub(r"[-|].*$", " ", title)
    title = re.sub(r"[^0-9a-z가-힣]+", " ", title)
    title = re.sub(r"\s+", " ", title).strip()
    return f"{norm_issue_v11(row)}:{title[:80]}"


def report_score(row: pd.Series) -> float:
    text = blob_v11(row)
    law_news = 100 if any(k in text for k in ["law", "regulation", "notice", "고시", "공고", "법령", "anti-dumping", "상계관세"]) else 0
    policy = 100 if norm_issue_v11(row) in {"AD_CVD", "CUSTOMS", "CBAM", "FTA_ORIGIN", "EXPORT_CONTROL", "TARIFF_POLICY"} else 0
    direct = 100 if ("samsung" in text or "삼성" in text) and policy else 70 if any(k in text for k in ["semiconductor", "반도체", "battery", "배터리", "display", "steel", "철강", "rare earth", "희토류"]) else 20
    indirect = 70 if any(k in text for k in ["china", "중국", "vietnam", "베트남", "india", "인도", "eu", "usa", "미국", "korea", "한국", "supply chain", "공급망"]) else 20
    score = law_news * 0.30 + policy * 0.20 + direct * 0.40 + indirect * 0.10
    score += safe_num(row.get("Importance Score")) / 10
    if hard_reference_v11(row):
        score -= 100
    if clean(row.get("Content Type")) == "Regulation" and not hard_reference_v11(row):
        score += 15
    return score


def infer_samsung_impact(row: pd.Series) -> str:
    if hard_reference_v11(row):
        return "Reference"
    score = report_score(row)
    if score >= 75:
        return "Indirect"
    if score >= 45:
        return "Watch"
    return "Reference"


def executive_priority(row: pd.Series) -> str:
    if hard_reference_v11(row):
        return "REFERENCE"
    score = report_score(row)
    if score >= 80:
        return "CORE"
    if score >= 60:
        return "POLICY_WATCH"
    if score >= 45:
        return "WATCH"
    return "REFERENCE"


def prepare_rows(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    if "Publish Date" not in rows.columns:
        rows["Publish Date"] = rows.get("Date", "")
    else:
        rows["Publish Date"] = rows["Publish Date"].where(rows["Publish Date"].astype(str).str.strip().ne(""), rows.get("Date", ""))
    rows["Publish Date"] = rows["Publish Date"].apply(lambda v: clean(v) if clean(v) and clean(v).lower() != "nan" else "확인 필요")

    if "Major Changes" not in rows.columns:
        rows["Major Changes"] = rows.get("Summary", "")
    else:
        rows["Major Changes"] = rows["Major Changes"].where(
            rows["Major Changes"].astype(str).str.strip().ne(""),
            rows.get("Summary", "")
        )

    for col in ["Major Changes", "Summary", "AI Analysis", "Action Plan"]:
        if col in rows.columns:
            rows[col] = rows[col].apply(clean_report_text_v11)

    rows["Issue"] = rows.apply(lambda r: {
        "AD_CVD": "AD/CVD",
        "CUSTOMS": "통관/세관",
        "CBAM": "CBAM",
        "FTA_ORIGIN": "FTA/원산지",
        "EXPORT_CONTROL": "수출통제",
        "TARIFF_POLICY": "관세정책",
    }.get(norm_issue_v11(r), clean(r.get("Issue"))), axis=1)

    reg = rows[rows["Content Type"].eq("Regulation")].copy()
    news = rows[rows["Content Type"].eq("News")].copy()
    news = news[~news["URL"].apply(bad_url_v11)].copy()
    news = news[~news.apply(hard_reference_v11, axis=1)].copy()
    news["_issue_key"] = news.apply(issue_key_v11, axis=1)
    news["_report_score"] = news.apply(report_score, axis=1)
    news = news.sort_values(["_report_score", "_sort_date"], ascending=[False, False])
    news = news.drop_duplicates(subset=["URL"], keep="first")
    news = news.drop_duplicates(subset=["_issue_key"], keep="first")
    news = news.head(NEWS_MAX_REPORT_ROWS).drop(columns=["_issue_key"], errors="ignore")

    rows = pd.concat([reg, news], ignore_index=True)
    rows["Samsung Impact"] = rows.apply(infer_samsung_impact, axis=1)
    rows["Executive Priority"] = rows.apply(executive_priority, axis=1)
    rows["Priority Group"] = rows["Executive Priority"]
    rows["Mail Group"] = rows["Content Type"].map({"Regulation": GROUP_REGULATION}).fillna(GROUP_NEWS)
    rows["_report_score"] = rows.apply(report_score, axis=1)
    rows = rows.sort_values(["_report_score", "_sort_date"], ascending=[False, False]).reset_index(drop=True)
    rows["No"] = range(1, len(rows) + 1)
    return rows


def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    pool = rows.copy()
    pool["_score"] = pool.apply(report_score, axis=1)
    pool["_issue_type"] = pool.apply(norm_issue_v11, axis=1)
    pool["_issue_key"] = pool.apply(issue_key_v11, axis=1)
    pool = pool[(pool["Content Type"].isin(["Regulation", "News"])) & (~pool.apply(hard_reference_v11, axis=1))].copy()
    pool = pool.sort_values(["_score", "_sort_date"], ascending=[False, False])
    selected, used_types, used_keys = [], set(), set()
    for _, row in pool.iterrows():
        typ = clean(row.get("_issue_type"))
        key = clean(row.get("_issue_key"))
        if typ in used_types or key in used_keys:
            continue
        selected.append(row)
        used_types.add(typ)
        used_keys.add(key)
        if len(selected) == 3:
            break
    if len(selected) < 3:
        for _, row in pool.iterrows():
            key = clean(row.get("_issue_key"))
            title = clean(row.get("Headline"))
            if key in used_keys or any(clean(x.get("Headline")) == title for x in selected):
                continue
            selected.append(row)
            used_keys.add(key)
            if len(selected) == 3:
                break
    out = pd.DataFrame(selected).drop(columns=["_score", "_issue_type", "_issue_key"], errors="ignore").reset_index(drop=True)
    if not out.empty:
        out["No"] = range(1, len(out) + 1)
    return out


def table_html(title: str, rows: pd.DataFrame, color: str) -> str:
    if rows.empty:
        return f"<h3 style='color:{color};'>{html.escape(title)} (0건)</h3>"
    trs = []
    for _, row in rows.iterrows():
        summary = html.escape(short_text(clean_report_text_v11(row.get("Major Changes")), "Summary 확인 필요", 650)).replace("\n", "<br>")
        impact = html.escape(short_text(clean_report_text_v11(row.get("AI Analysis")), "영향 검토 필요", 360)).replace("\n", "<br>")
        action = html.escape(short_text(clean_report_text_v11(row.get("Action Plan")), "담당 부서 확인 필요", 360)).replace("\n", "<br>")
        trs.append(f"""
        <tr>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;">{html.escape(str(row.get('No')))}</td>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;">{html.escape(clean(row.get('Issue')))}</td>
          <td style="padding:7px;border:1px solid #ddd;vertical-align:top;">{html_link(row.get('Headline'), row.get('URL'))}</td>
          <td style="padding:7px;border:1px solid #ddd;vertical-align:top;">{summary}</td>
          <td style="padding:7px;border:1px solid #ddd;vertical-align:top;">{impact}</td>
          <td style="padding:7px;border:1px solid #ddd;vertical-align:top;">{action}</td>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;">{html.escape(clean(row.get('Country')))}</td>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;color:{risk_color(row.get('Risk'))};font-weight:bold;">{html.escape(clean(row.get('Risk')))}</td>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;">{html.escape(clean(row.get('Samsung Impact')))}</td>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;">{html.escape(clean(row.get('Publish Date') or row.get('Date')))}</td>
        </tr>
        """)
    return f"""
    <h3 style="margin-top:24px;color:{color};">{html.escape(title)} ({len(rows)}건)</h3>
    <table style="border-collapse:collapse;width:100%;font-size:12px;table-layout:fixed;">
      <thead>
        <tr style="background:{color};color:white;">
          <th style="padding:7px;border:1px solid #ddd;">No</th>
          <th style="padding:7px;border:1px solid #ddd;">Issue</th>
          <th style="padding:7px;border:1px solid #ddd;">Headline</th>
          <th style="padding:7px;border:1px solid #ddd;">Summary</th>
          <th style="padding:7px;border:1px solid #ddd;">삼성 영향</th>
          <th style="padding:7px;border:1px solid #ddd;">Action</th>
          <th style="padding:7px;border:1px solid #ddd;">Country</th>
          <th style="padding:7px;border:1px solid #ddd;">Risk</th>
          <th style="padding:7px;border:1px solid #ddd;">Impact</th>
          <th style="padding:7px;border:1px solid #ddd;">Publish Date</th>
        </tr>
      </thead>
      <tbody>{''.join(trs)}</tbody>
    </table>
    """


# ======================================================================
# GTI STEP5 UNIPASS URL & TOP3 DETAIL Patch v13 - 2026-06-14
# ----------------------------------------------------------------------
# - UNIPASS rowTitle URL을 ntarId 직접열람 URL로 최종 보정
# - PN51 / Advance Authorization / EPCG 수출의무 연장 건 Top3 상세분석 강화
# ======================================================================

UNIPASS_NOTICE_URL_PREFIX_V13 = (
    "https://unipass.customs.go.kr/csp/myc/custsppt/cmmn/"
    "NtarBrkdMtCtr/openMYC0605014Q.do?ntarId="
)

UNIPASS_NOTICE_ID_BY_TITLE_V13 = {
    "다수 사업장 운영 사업자의 전자상거래업자 등록 신청 방법": "202606122928",
}


def _v13_norm_title(text: str) -> str:
    try:
        text = unquote(clean(text))
    except Exception:
        text = clean(text)
    text = re.sub(r"\([^)]*\)|\[[^]]*\]", " ", text)
    text = re.sub(r"[^0-9A-Za-z가-힣]+", " ", text)
    return re.sub(r"\s+", " ", text).strip().lower()


def _v13_row_title_from_url(url: str) -> str:
    text = clean(url)
    m = re.search(r"[?&]rowTitle=([^&]+)", text, flags=re.I)
    if not m:
        return ""
    try:
        return unquote(m.group(1))
    except Exception:
        return m.group(1)


def _v13_unipass_direct_url(row: pd.Series) -> str:
    values = [
        row.get("Headline", ""),
        row.get("Title", ""),
        row.get("Summary", ""),
        _v13_row_title_from_url(row.get("URL", "")),
        _v13_row_title_from_url(row.get("Source", "")),
    ]
    normalized = [_v13_norm_title(v) for v in values if clean(v)]
    for key, ntar_id in UNIPASS_NOTICE_ID_BY_TITLE_V13.items():
        norm_key = _v13_norm_title(key)
        if any(norm_key and (norm_key in v or v in norm_key) for v in normalized):
            return UNIPASS_NOTICE_URL_PREFIX_V13 + ntar_id
    return ""


_PREV_FIX_UNIPASS_URL_V13 = fix_unipass_url


def fix_unipass_url(row: pd.Series) -> str:
    direct = _v13_unipass_direct_url(row)
    if direct:
        return direct
    url = clean(row.get("URL"))
    low = url.lower()
    if "unipass.customs.go.kr" in low and "openmyc0605014q.do" in low and "ntarid=" in low:
        return url
    try:
        return _PREV_FIX_UNIPASS_URL_V13(row)
    except Exception:
        return url


def detailed_pn51_summary() -> str:
    return (
        "인도 정부가 수출촉진을 위해 Advance Authorization 및 EPCG 수출의무 이행기간을 "
        "2026년 8월까지 자동 연장함에 따라, 인도 생산법인 또는 협력업체의 수출의무(EO) "
        "이행 부담이 완화될 전망입니다.\n\n"
        "주요 내용은 다음과 같습니다.\n"
        "- Advance Authorization EO 만료기간이 2026.03.01~2026.05.31인 경우 → 2026.08.31까지 자동 연장\n"
        "- EPCG(Block-wise EO) 만료기간이 2026.03.01~2026.05.31인 경우 → 2026.08.31까지 자동 연장\n"
        "- 별도 신청서 제출 불필요\n"
        "- 연장 수수료(Composition Fee) 면제\n"
        "- 세관은 연장된 EO 기준으로 수출을 인정\n"
        "- EO 충족 여부는 EODC 발급 시 최종 검증\n\n"
        "따라서 본 건은 단순 정책 뉴스가 아니라, 인도 내 수입부품 무관세 활용 제도와 "
        "수출의무 관리 일정에 영향을 줄 수 있는 공식 공지로 보아야 합니다."
    )


def detailed_pn51_impact() -> str:
    return (
        "간접 영향 (Watch Level)\n"
        "본 공고는 인도 정부의 수출지원 조치로 삼성전자에 직접적인 관세 인상 또는 수입규제 영향을 주는 정책은 아닙니다. "
        "다만 삼성전자 인도 생산법인 또는 협력업체가 Advance Authorization 또는 "
        "EPCG(Export Promotion Capital Goods)를 활용하는 경우 관세업무상 영향이 발생할 수 있습니다.\n\n"
        "관세업무 관점 영향\n"
        "1. EO 미충족 리스크 완화\n"
        "- 인도 생산법인이 원자재·부품을 무관세 또는 감면 조건으로 수입하고 수출의무를 부여받은 경우, "
        "EO 기간이 자동 연장됩니다.\n"
        "- 이에 따라 EO 미충족에 따른 추징관세, 이자부담, 허가 취소 리스크가 감소합니다.\n\n"
        "2. 공급망 운영 유연성 증가\n"
        "- 홍해 사태, 중동 물류 리스크, 글로벌 공급망 재편 등으로 수출계획 달성이 지연되는 기업의 "
        "수출실적 관리 부담을 완화하는 효과가 있습니다.\n\n"
        "3. 삼성전자 인도 생산법인 영향\n"
        "- 휴대폰, TV, 가전제품 등 인도 생산 제품에 투입되는 수입부품이 Advance Authorization 또는 EPCG와 "
        "연계되어 있다면 EO 달성 일정에 추가 여유가 확보됩니다.\n"
        "- 직접 관세비용 증가 이슈는 아니지만, 허가번호별 EO 만료일·달성률·EODC 발급 상태 관리가 필요합니다."
    )


def detailed_pn51_action() -> str:
    return (
        "삼성전자 인도 법인의 Advance Authorization 및 EPCG 활용 현황을 점검하고 "
        "EO 만료 예정 건에 대한 자동 연장 적용 여부를 확인할 것을 권고합니다.\n\n"
        "즉시 조치\n"
        "- 인도 법인에 Advance Authorization 사용 여부 확인\n"
        "- EPCG 사용 여부 확인\n"
        "- 2026.03.01~2026.05.31 사이 EO 만료 예정 허가 현황 확인\n\n"
        "1주 내 확인 리스트\n"
        "- Authorization No: 허가번호\n"
        "- 제도구분: AA / EPCG\n"
        "- EO 만료일: 기존 만료일\n"
        "- 연장적용 여부: Y/N\n"
        "- 예상 EO 달성률: %\n\n"
        "1개월 내 관리항목 검토\n"
        "- ONE-Origin 또는 내부 관세관리 Master에 Authorization 번호, EO 만료일, EO 달성률, "
        "EODC 발급 여부를 관리항목으로 추가 검토\n"
        "- Owner: HQ Customs / India Customs 담당"
    )


# ======================================================================
# GTI STEP5 Emergency Report Guard v14 - 2026-06-14
# ----------------------------------------------------------------------
# The executive mail must not show Reference/old/body-missing rows as if they
# were actionable sensing results.  Keep them in source/audit files, but remove
# them from the mail report and from Top3.
# ======================================================================

def _v14_text(v) -> str:
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    return clean(v)


def _v14_blob(row: pd.Series) -> str:
    return " ".join(_v14_text(row.get(c, "")) for c in [
        "Headline", "Issue", "Major Changes", "Summary", "AI Analysis", "Action Plan",
        "Samsung Impact", "Priority Group", "URL", "Source",
    ]).lower()


def _v14_parse_date(row: pd.Series):
    for c in ["Publish Date", "Date"]:
        v = _v14_text(row.get(c, ""))
        if not v or v in {"확인 필요", "nan", "NaT"}:
            continue
        try:
            return pd.to_datetime(v, errors="coerce")
        except Exception:
            pass
    return pd.NaT


def _v14_is_pn51(row: pd.Series) -> bool:
    blob = _v14_blob(row)
    return (
        "pn 51" in blob
        or "export obligation period" in blob
        or ("advance authorization" in blob and "epcg" in blob)
        or ("수출의무" in blob and "epcg" in blob)
    )


def _v14_is_reportable(row: pd.Series) -> bool:
    blob = _v14_blob(row)
    impact = _v14_text(row.get("Samsung Impact"))
    issue = _v14_text(row.get("Issue"))
    ctype = _v14_text(row.get("Content Type"))
    title = _v14_text(row.get("Headline")).lower()
    focus = " ".join(_v14_text(row.get(c, "")) for c in ["Headline", "Major Changes", "Summary"]).lower()

    if impact == "Reference" or issue == "Reference":
        return False
    if any(k in title for k in [
        "public notice no.26", "public notice no 26", "public notice no. 26",
        "rate of exchange", "과세환율",
    ]):
        return False
    if any(k in blob for k in [
        "본문 내용 확인 불가", "본문 확인 불가", "원문 내용이 파싱되지 않아",
        "상세 분석이 어렵습니다", "구체적인 영향을 분석할 수 없습니다",
    ]) and not _v14_is_pn51(row):
        return False
    if ctype == "News" and any(k in title for k in [
        "청년인턴", "채용", "몰카", "범죄", "모닝뉴스", "보안시장",
        "주가", "증시", "미토스", "주술", "페라리", "유럽순방",
        "aeo strategy", "ai search", "crm purchase", "unemployment rate",
        "retailer stocks", "american eagle", "hubspot", "hockney",
        "low tariff coverage", "corporate rules", "노사 갈등", "임협",
    ]):
        return False
    if ctype == "News":
        if any(k in blob for k in [
            "직접적인 영향은 확인되지", "직접적인 관련성은 낮", "직접적인 연관성은 확인되지",
            "관세/통상 업무와 직접적인 연관성은 확인되지", "업무 관련성이 있는지 확인",
            "해당 없음",
        ]):
            return False
        strong_news = any(k in focus for k in [
            "관세 인하", "관세 부과", "관세율", "상호관세", "tariff cut", "tariff hike", "tariff increase",
            "customs clearance", "customs declaration", "통관", "수입신고", "수출신고",
            "free trade agreement", "fta", "cepa", "tepa", "원산지", "rules of origin",
            "export control", "수출 통제", "수출통제", "entity list", "forced labor", "uflpa",
            "cbam", "carbon border", "탄소국경", "k스틸법", "steel act",
            "anti-dumping", "antidumping", "countervailing", "반덤핑", "상계관세",
            "one-gate export", "one gate export", "원스톱 수출",
        ])
        if not strong_news:
            return False
    return True


def _v14_quality_score(row: pd.Series) -> float:
    blob = _v14_blob(row)
    score = safe_num(row.get("Importance Score"))
    if _v14_is_pn51(row):
        score += 5000
    if any(k in blob for k in ["anti-dumping", "antidumping", "countervailing", "반덤핑", "상계관세"]):
        score += 900
    if any(k in blob for k in ["fta", "cepa", "tepa", "origin", "원산지"]):
        score += 700
    if any(k in blob for k in ["cbam", "탄소국경"]):
        score += 650
    if any(k in blob for k in ["export control", "수출통제", "entity list", "uflpa", "forced labor"]):
        score += 650
    if any(k in blob for k in ["tariff", "관세", "customs", "통관", "보세"]):
        score += 300
    if _v14_text(row.get("Content Type")) == "Regulation":
        score += 250
    return score


_PREV_PREPARE_ROWS_V14 = prepare_rows


def prepare_rows(rows: pd.DataFrame) -> pd.DataFrame:
    rows = _PREV_PREPARE_ROWS_V14(rows).copy()
    if rows.empty:
        return rows

    rows["URL"] = rows.apply(fix_unipass_url, axis=1)
    rows["Publish Date"] = rows["Publish Date"].apply(lambda v: _v14_text(v) or "확인 필요")
    rows["Date"] = rows["Date"].apply(lambda v: _v14_text(v) or "확인 필요")

    # Force known high-value issue details.
    pn_mask = rows.apply(_v14_is_pn51, axis=1)
    if pn_mask.any():
        rows.loc[pn_mask, "Major Changes"] = detailed_pn51_summary()
        rows.loc[pn_mask, "Summary"] = detailed_pn51_summary()
        rows.loc[pn_mask, "AI Analysis"] = detailed_pn51_impact()
        rows.loc[pn_mask, "Action Plan"] = detailed_pn51_action()
        rows.loc[pn_mask, "Samsung Impact"] = "Indirect"
        rows.loc[pn_mask, "Issue"] = "통관/세관"

    # Report only actionable sensing rows.  Reference rows remain in STEP4 files.
    rows = rows[rows.apply(_v14_is_reportable, axis=1)].copy()

    # News should never exceed 30. Regulations are all reportable regulations.
    reg = rows[rows["Content Type"].eq("Regulation")].copy()
    news = rows[rows["Content Type"].eq("News")].copy()
    if not news.empty:
        news["_v14_score"] = news.apply(_v14_quality_score, axis=1)
        news = news.sort_values(["_v14_score", "_sort_date"], ascending=[False, False]).head(30)
        news = news.drop(columns=["_v14_score"], errors="ignore")
    rows = pd.concat([reg, news], ignore_index=True)
    rows["_v14_score"] = rows.apply(_v14_quality_score, axis=1)
    rows = rows.sort_values(["_v14_score", "_sort_date"], ascending=[False, False]).drop(columns=["_v14_score"], errors="ignore")
    rows = rows.reset_index(drop=True)
    rows["No"] = range(1, len(rows) + 1)
    return rows


def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    if rows.empty:
        return rows.copy()
    pool = rows[rows.apply(_v14_is_reportable, axis=1)].copy()
    if pool.empty:
        return pool
    pool["_v14_score"] = pool.apply(_v14_quality_score, axis=1)
    pool = pool.sort_values(["_v14_score", "_sort_date"], ascending=[False, False])
    selected = []
    used_issue = set()
    used_title = set()
    for _, row in pool.iterrows():
        issue = _v14_text(row.get("Issue"))
        title_key = re.sub(r"\s+", " ", _v14_text(row.get("Headline")).lower())[:80]
        if title_key in used_title:
            continue
        if issue in used_issue and not _v14_is_pn51(row):
            continue
        selected.append(row)
        used_issue.add(issue)
        used_title.add(title_key)
        if len(selected) >= 3:
            break
    out = pd.DataFrame(selected).drop(columns=["_v14_score"], errors="ignore").reset_index(drop=True)
    if not out.empty:
        out["No"] = range(1, len(out) + 1)
    return out



# ======================================================================
# GTI STEP5 Weighted Score Patch v18
# ----------------------------------------------------------------------
# - Use WeightedScore from STEP4 as the primary report sort/top3 basis.
# - Preserve score breakdown columns in Excel output when present.
# - Reclassify noisy/low-action articles as lower priority for mail layout.
# ======================================================================

def weighted_score_value(row: pd.Series) -> float:
    return safe_num(row.get("WeightedScore")) or safe_num(row.get("Importance Score")) or safe_num(row.get("final_score"))

def weighted_report_score(row: pd.Series) -> float:
    score = weighted_score_value(row)
    # Regulation still gets small boost because it is official source.
    if clean(row.get("Content Type")) == "Regulation":
        score += 5
    # Penalize obvious non-reportable phrases that may still pass.
    blob = " ".join(clean(row.get(c)) for c in ["Headline", "Summary", "Major Changes", "AI Analysis"]).lower()
    noise = ["마케팅", "신약", "비비드", "축제", "브랜드", "주가", "전략회의", "칼럼", "스포츠"]
    if any(x in blob for x in noise):
        score -= 50
    return score

def report_score(row: pd.Series) -> float:
    """v18 override: STEP5 follows STEP4 WeightedScore."""
    return weighted_report_score(row)

def top3_deep_score(row: pd.Series) -> float:
    """v18 override: TOP3 by weighted score + strategic issue boost."""
    score = weighted_report_score(row)
    issue = clean(row.get("Issue"))
    if issue in {"AD/CVD", "반덤핑/상계관세"}:
        score += 12
    elif issue == "수출통제":
        score += 10
    elif issue == "CBAM":
        score += 9
    elif issue in {"FTA/원산지", "관세정책", "HS/품목분류"}:
        score += 6
    return score

def prepare_rows(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    rows["Issue"] = rows.apply(issue_for, axis=1)
    rows = dedup_report_rows(rows)
    rows["Mail Group"] = rows["Content Type"].map({"Regulation": GROUP_REGULATION}).fillna(GROUP_NEWS)
    rows["Major Changes"] = rows.apply(major_changes, axis=1)
    rows["Summary"] = rows.apply(report_summary, axis=1)
    rows["AI Analysis"] = rows.apply(report_impact, axis=1)
    rows["Action Plan"] = rows.apply(report_action, axis=1)
    rows["_report_score"] = rows.apply(report_score, axis=1)
    rows = rows.sort_values(["_report_score", "_sort_date"], ascending=[False, False]).reset_index(drop=True)
    rows["No"] = range(1, len(rows) + 1)
    return rows

# Extend Excel output columns if weighted columns exist.
for _col in ["WeightedScore", "ScoreBreakdown", "CustomsTradeLawScore", "CustomsTradePolicyScore", "DirectImpactScore", "IndirectImpactScore"]:
    if _col not in OUTPUT_COLUMNS:
        OUTPUT_COLUMNS.append(_col)

# ======================================================================
# End of GTI STEP5 Weighted Score Patch v18
# ======================================================================


# ======================================================================
# GTI STEP5 Weighted Score Patch v19
# ----------------------------------------------------------------------
# Fix:
# - save_excel() KeyError when OUTPUT_COLUMNS contains weighted columns
#   but rows does not include them.
# - Ensure missing output columns are created before Excel/HTML generation.
# - Recalculate weighted columns in STEP5 when STEP4 file does not contain them.
# ======================================================================

WEIGHTED_COLS_V19 = [
    "Publish Date",
    "CustomsTradeLawScore",
    "CustomsTradePolicyScore",
    "DirectImpactScore",
    "IndirectImpactScore",
    "WeightedScore",
    "ScoreBreakdown",
]

def ensure_output_columns_v19(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in WEIGHTED_COLS_V19:
        if col not in df.columns:
            df[col] = ""
    for col in OUTPUT_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df

def _v19_blob(row: pd.Series) -> str:
    return " ".join(clean(row.get(c)) for c in [
        "Headline", "Issue", "Major Changes", "Summary", "AI Analysis",
        "Action Plan", "KeywordMatches", "Country", "Agency"
    ]).lower()

def _v19_contains(blob: str, terms: list[str]) -> bool:
    return any(t.lower() in blob for t in terms)

def recalc_weighted_score_v19(row: pd.Series) -> pd.Series:
    """Fallback weighted score calculation inside STEP5."""
    blob = _v19_blob(row)
    issue = clean(row.get("Issue"))
    impact = clean(row.get("Samsung Impact"))

    law_terms = [
        "법령", "고시", "공고", "규칙", "관세법", "federal register", "regulation",
        "anti-dumping", "antidumping", "countervailing", "ad/cvd", "반덤핑", "덤핑방지",
        "상계관세", "cbam", "carbon border", "fta", "cepa", "rules of origin",
        "원산지", "hs code", "품목분류", "customs", "통관", "세관", "관세청",
    ]
    policy_terms = [
        "관세", "tariff", "section 301", "section 232", "수출통제", "export control",
        "entity list", "cbam", "탄소국경", "fta", "cepa", "반덤핑", "상계관세",
        "quota", "쿼터", "제재", "sanction",
    ]
    direct_terms = [
        "samsung electronics", "samsung sdi", "samsung display", "삼성전자",
        "삼성sdi", "삼성디스플레이", "semiconductor", "반도체", "ai chip",
        "배터리", "battery", "display", "oled", "스마트폰", "galaxy",
    ]
    indirect_terms = [
        "steel", "철강", "알루미늄", "aluminum", "희토류", "rare earth",
        "리튬", "lithium", "공급망", "supply chain", "조달", "원가",
        "중국", "미국", "eu", "베트남", "인도", "멕시코", "폴란드",
    ]

    law = safe_num(row.get("CustomsTradeLawScore"))
    policy = safe_num(row.get("CustomsTradePolicyScore"))
    direct = safe_num(row.get("DirectImpactScore"))
    indirect = safe_num(row.get("IndirectImpactScore"))

    if not law:
        if issue in {"AD/CVD", "반덤핑/상계관세", "CBAM", "FTA/원산지", "HS/품목분류"}:
            law = 30
        elif _v19_contains(blob, law_terms):
            law = 20
        else:
            law = 0

    if not policy:
        if issue in {"수출통제", "관세정책", "AD/CVD", "반덤핑/상계관세", "CBAM"}:
            policy = 20
        elif _v19_contains(blob, policy_terms):
            policy = 12
        else:
            policy = 0

    if not direct:
        if impact == "Direct":
            direct = 40
        elif _v19_contains(blob, direct_terms) and (law + policy) > 0:
            direct = 30
        else:
            direct = 0

    if not indirect:
        if impact == "Indirect":
            indirect = 10
        elif _v19_contains(blob, indirect_terms) and (law + policy) > 0:
            indirect = 6
        elif impact == "Watch" and (law + policy) > 0:
            indirect = 3
        else:
            indirect = 0

    weighted = law + policy + direct + indirect
    row["CustomsTradeLawScore"] = int(law)
    row["CustomsTradePolicyScore"] = int(policy)
    row["DirectImpactScore"] = int(direct)
    row["IndirectImpactScore"] = int(indirect)
    row["WeightedScore"] = int(weighted)
    row["ScoreBreakdown"] = f"법규30={int(law)}; 정책20={int(policy)}; 직접40={int(direct)}; 간접10={int(indirect)}"
    if not clean(row.get("Publish Date")):
        row["Publish Date"] = clean(row.get("Date"))
    return row

def prepare_rows(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    rows["Issue"] = rows.apply(issue_for, axis=1)
    rows = dedup_report_rows(rows)
    rows["Mail Group"] = rows["Content Type"].map({"Regulation": GROUP_REGULATION}).fillna(GROUP_NEWS)
    rows["Major Changes"] = rows.apply(major_changes, axis=1)
    rows["Summary"] = rows.apply(report_summary, axis=1)
    rows["AI Analysis"] = rows.apply(report_impact, axis=1)
    rows["Action Plan"] = rows.apply(report_action, axis=1)

    for col in WEIGHTED_COLS_V19:
        if col not in rows.columns:
            rows[col] = ""

    rows = rows.apply(recalc_weighted_score_v19, axis=1)
    rows["_report_score"] = rows.apply(report_score, axis=1)
    rows = ensure_output_columns_v19(rows)
    rows = rows.sort_values(["_report_score", "_sort_date"], ascending=[False, False]).reset_index(drop=True)
    rows["No"] = range(1, len(rows) + 1)
    return rows

def save_excel(rows: pd.DataFrame, top3: pd.DataFrame, paths: dict[str, Path]) -> None:
    """v19 override: ensure all configured OUTPUT_COLUMNS exist before saving."""
    rows = ensure_output_columns_v19(rows)
    top3 = ensure_output_columns_v19(top3)
    paths["mail_xlsx"].parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(paths["mail_xlsx"], engine="openpyxl") as writer:
        rows[OUTPUT_COLUMNS].to_excel(writer, index=False, sheet_name="GTI Radar")
        top3[OUTPUT_COLUMNS].to_excel(writer, index=False, sheet_name="Top3")
        ws = writer.book["GTI Radar"]
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_cells in ws.columns:
            width = min(max(len(str(c.value or "")) for c in col_cells) + 2, 60)
            ws.column_dimensions[col_cells[0].column_letter].width = width

    rows[OUTPUT_COLUMNS].to_excel(paths["cumulative"], index=False)

# Ensure OUTPUT_COLUMNS includes weighted columns safely.
for _col in WEIGHTED_COLS_V19:
    if _col not in OUTPUT_COLUMNS:
        OUTPUT_COLUMNS.append(_col)

# ======================================================================
# End of GTI STEP5 Weighted Score Patch v19
# ======================================================================


# ======================================================================
# GTI STEP5 News Recovery Patch v20
# ----------------------------------------------------------------------
# Fix:
# - STEP5 output shows news=0 although STEP4 selected 30 news.
# - Root cause is usually NEWS_INPUT_FILE path/read/filter mismatch or
#   later STEP5 candidate filtering dropping all News rows.
#
# What v20 does:
# 1) Reads news from multiple safe fallback sources:
#    - C:\Temp\4-2.news_ai_summary.xlsx
#    - C:\Temp\4.news_ai_analysis.xlsx
#    - C:\Temp\4-2.news_ai_audit_candidates.xlsx
#    - output_dir\4.news_ai_analysis.xlsx
# 2) Preserves at least GTI_NEWS_MIN_REPORT_ROWS news rows when available.
# 3) Uses WeightedScore / final_score to sort news.
# 4) Logs the source and row counts so you can verify what STEP5 actually read.
# ======================================================================

NEWS_FALLBACK_FILES_V20 = [
    NEWS_INPUT_FILE,
    Path(os.getenv("GTI_NEWS_LEGACY_FILE", r"C:\Temp\4.news_ai_analysis.xlsx")),
    NEWS_AUDIT_INPUT_FILE if "NEWS_AUDIT_INPUT_FILE" in globals() else Path(r"C:\Temp\4-2.news_ai_audit_candidates.xlsx"),
    OUTPUT_DIR / "4.news_ai_analysis.xlsx",
]

def _v20_existing_files(paths: list[Path]) -> list[Path]:
    out = []
    seen = set()
    for p in paths:
        try:
            p = Path(p)
            key = str(p).lower()
            if key in seen:
                continue
            seen.add(key)
            if p.exists():
                out.append(p)
        except Exception:
            pass
    return out

def _v20_sort_cols(df: pd.DataFrame) -> tuple[list[str], list[bool]]:
    cols, asc = [], []
    for col, ascending in [
        ("WeightedScore", False),
        ("final_score", False),
        ("Importance Score", False),
        ("_sort_date", False),
        ("Date", False),
    ]:
        if col in df.columns:
            cols.append(col)
            asc.append(ascending)
    return cols, asc

def _v20_normalize_news_file(path: Path) -> pd.DataFrame:
    try:
        raw = pd.read_excel(path)
    except Exception as exc:
        print(f"[WARN] v20 news read failed: {path} / {type(exc).__name__}: {exc}")
        return pd.DataFrame()

    if raw is None or raw.empty:
        print(f"[WARN] v20 news file empty: {path}")
        return pd.DataFrame()

    # If this is already a STEP5 output, keep News rows only before re-normalizing.
    if "Content Type" in raw.columns:
        raw_news = raw[raw["Content Type"].astype(str).str.lower().eq("news")].copy()
        if not raw_news.empty:
            raw = raw_news

    try:
        news = normalize_input(raw, "News", path)
    except Exception as exc:
        print(f"[WARN] v20 news normalize failed: {path} / {type(exc).__name__}: {exc}")
        return pd.DataFrame()

    if news.empty:
        print(f"[WARN] v20 normalized news empty: {path}")
        return pd.DataFrame()

    # Preserve weighted columns from raw if normalize_input did not carry them.
    for col in [
        "WeightedScore", "ScoreBreakdown", "CustomsTradeLawScore", "CustomsTradePolicyScore",
        "DirectImpactScore", "IndirectImpactScore", "final_score", "priority_group", "mail_section",
        "selected",
    ]:
        if col in raw.columns and col not in news.columns and len(raw) == len(news):
            news[col] = raw[col].values

    # If raw was already final STEP5 output, WeightedScore columns may exist but Date order differs.
    for col in ["WeightedScore", "CustomsTradeLawScore", "CustomsTradePolicyScore", "DirectImpactScore", "IndirectImpactScore"]:
        if col not in news.columns:
            news[col] = ""

    news["Content Type"] = "News"
    news["Mail Group"] = GROUP_NEWS
    if "Priority Group" not in news.columns or news["Priority Group"].astype(str).str.strip().eq("").all():
        news["Priority Group"] = "CORE"

    return news

def _v20_collect_news_rows() -> pd.DataFrame:
    frames = []
    for path in _v20_existing_files(NEWS_FALLBACK_FILES_V20):
        df = _v20_normalize_news_file(path)
        if not df.empty:
            print(f"[INFO] v20 news source loaded: {path} rows={len(df)}")
            frames.append(df)

    if not frames:
        print("[WARN] v20 no news source loaded")
        return pd.DataFrame()

    news = pd.concat(frames, ignore_index=True, sort=False)

    # Remove rows with no headline.
    news = news[news["Headline"].astype(str).str.strip().ne("")].copy()

    # Dedup by URL first, then headline.
    if "URL" in news.columns:
        news["_url_key"] = news["URL"].astype(str).str.lower().str.strip()
        news = news.sort_values(_v20_sort_cols(news)[0], ascending=_v20_sort_cols(news)[1]) if _v20_sort_cols(news)[0] else news
        news = news.drop_duplicates(subset=["_url_key"], keep="first").drop(columns=["_url_key"], errors="ignore")
    news["_headline_key"] = news["Headline"].astype(str).str.lower().str.replace(r"[^0-9a-z가-힣]+", " ", regex=True).str.strip().str[:140]
    news = news.drop_duplicates(subset=["_headline_key"], keep="first").drop(columns=["_headline_key"], errors="ignore")

    cols, asc = _v20_sort_cols(news)
    if cols:
        news = news.sort_values(cols, ascending=asc)

    target_min = int(os.getenv("GTI_NEWS_MIN_REPORT_ROWS", "30"))
    target_max = int(os.getenv("GTI_NEWS_MAX_REPORT_ROWS", "50"))
    if NEWS_MAX_ROWS > 0:
        target_max = min(target_max, NEWS_MAX_ROWS)
    target = max(target_min, min(target_max, len(news)))
    news = news.head(target).reset_index(drop=True)

    print(f"[INFO] v20 news recovered={len(news)} target={target} from_sources={len(frames)}")
    return news

def read_step4_results() -> pd.DataFrame:
    frames = []

    if REGULATION_INPUT_FILE.exists():
        try:
            reg = normalize_input(pd.read_excel(REGULATION_INPUT_FILE), "Regulation", REGULATION_INPUT_FILE)
            frames.append(reg)
            print(f"[INFO] v20 regulation loaded: {REGULATION_INPUT_FILE} rows={len(reg)}")
        except Exception as exc:
            print(f"[WARN] v20 regulation read failed: {REGULATION_INPUT_FILE} / {type(exc).__name__}: {exc}")

    news = _v20_collect_news_rows()
    if not news.empty:
        frames.append(news)

    if not frames:
        raise FileNotFoundError(f"STEP4 outputs not found: {REGULATION_INPUT_FILE}, {NEWS_INPUT_FILE}")

    rows = pd.concat(frames, ignore_index=True, sort=False)
    rows["URL"] = rows.apply(lambda r: best_url_from_values([r.get("URL"), r.get("Source")]), axis=1)

    rows["_dedup_key"] = rows.apply(
        lambda r: clean(r.get("URL")).lower() or (
            clean(r.get("Headline"))[:160] + "|" + clean(r.get("Agency")) + "|" + clean(r.get("Date"))
        ),
        axis=1,
    )
    rows = rows.drop_duplicates(subset=["_dedup_key"], keep="first").drop(columns=["_dedup_key"], errors="ignore")

    rows["_integrated_score"] = rows.apply(
        lambda r: priority_weight(r.get("Priority Group")) + risk_weight(r.get("Risk")) +
                  (180 if clean(r.get("Content Type")) == "Regulation" else 0) +
                  safe_num(r.get("Importance Score")) + safe_num(r.get("WeightedScore")),
        axis=1,
    )

    print(
        f"[INFO] v20 total input rows={len(rows)} / "
        f"regulation={int(rows['Content Type'].eq('Regulation').sum())} / "
        f"news={int(rows['Content Type'].eq('News').sum())}"
    )
    return rows.reset_index(drop=True)

# ======================================================================
# End of GTI STEP5 News Recovery Patch v20
# ======================================================================


# ======================================================================
# GTI STEP5 Freshness & Executive Summary Patch v21
# ----------------------------------------------------------------------
# 목적:
# 1) 오래된 Regulation/News가 Top3에 선정되는 문제 방지
#    - 2025년 기사, 수개월 전 DGFT 공지 등 자동 Reference/제외
# 2) 총평 문구 수정:
#    - "금일 GTI Radar는 [Top3 핵심내용을 1문장으로] ..."
#    - 선별결과는 회색 작은 글씨로 하단 표시
# 3) News 30건이 STEP4에 있으면 STEP5에서 최대한 유지
#
# 주요 환경변수:
# - GTI_MAX_REPORT_AGE_DAYS=45         전체 메일 본문 허용 최대 일수
# - GTI_MAX_TOP3_AGE_DAYS=14           Top3 허용 최대 일수
# - GTI_ALLOW_STALE_ITEMS=N            Y면 오래된 건도 유지
# ======================================================================

GTI_MAX_REPORT_AGE_DAYS = int(os.getenv("GTI_MAX_REPORT_AGE_DAYS", "45"))
GTI_MAX_TOP3_AGE_DAYS = int(os.getenv("GTI_MAX_TOP3_AGE_DAYS", "14"))
GTI_ALLOW_STALE_ITEMS = os.getenv("GTI_ALLOW_STALE_ITEMS", "N").strip().upper() in {"Y", "YES", "TRUE", "1"}

_MONTH_MAP_V21 = {
    "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
    "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6,
    "jul": 7, "july": 7, "aug": 8, "august": 8, "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10, "nov": 11, "november": 11, "dec": 12, "december": 12,
}

def _run_date_ts_v21():
    dt = pd.to_datetime(RUN_DATE, errors="coerce")
    if pd.isna(dt):
        return pd.Timestamp(datetime.now().date())
    return pd.Timestamp(dt.date())

def extract_date_from_text_v21(text: str):
    t = clean(text)
    if not t:
        return pd.NaT

    # 2026-06-14 / 2026.06.14 / 2026/06/14
    m = re.search(r"(20\d{2})[.\-/](\d{1,2})[.\-/](\d{1,2})", t)
    if m:
        return pd.to_datetime(f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}", errors="coerce")

    # August 20, 2025
    m = re.search(r"\b([A-Za-z]{3,9})\s+(\d{1,2}),?\s+(20\d{2})\b", t)
    if m:
        mon = _MONTH_MAP_V21.get(m.group(1).lower())
        if mon:
            return pd.to_datetime(f"{m.group(3)}-{mon:02d}-{int(m.group(2)):02d}", errors="coerce")

    # 6March2026 / 6 March 2026
    m = re.search(r"\b(\d{1,2})\s*([A-Za-z]{3,9})\s*(20\d{2})\b", t)
    if m:
        mon = _MONTH_MAP_V21.get(m.group(2).lower())
        if mon:
            return pd.to_datetime(f"{m.group(3)}-{mon:02d}-{int(m.group(1)):02d}", errors="coerce")

    return pd.NaT

def effective_publish_date_v21(row: pd.Series):
    for col in ["Publish Date", "Date", "published", "PublishedAt", "CollectedAt"]:
        dt = pd.to_datetime(row.get(col), errors="coerce")
        if not pd.isna(dt):
            return pd.Timestamp(dt.date())
    dt = extract_date_from_text_v21(" ".join(clean(row.get(c)) for c in ["Headline", "Summary", "Major Changes", "URL"]))
    if not pd.isna(dt):
        return pd.Timestamp(dt.date())
    return pd.NaT

def item_age_days_v21(row: pd.Series) -> float:
    dt = effective_publish_date_v21(row)
    if pd.isna(dt):
        return 9999.0
    return float((_run_date_ts_v21() - dt).days)

def is_stale_item_v21(row: pd.Series, max_days: int = None) -> bool:
    if GTI_ALLOW_STALE_ITEMS:
        return False
    max_days = GTI_MAX_REPORT_AGE_DAYS if max_days is None else max_days
    age = item_age_days_v21(row)
    # future dates are not stale here
    if age < 0:
        return False
    return age > max_days

def stale_reason_v21(row: pd.Series) -> str:
    age = item_age_days_v21(row)
    dt = effective_publish_date_v21(row)
    dt_txt = "확인 불가" if pd.isna(dt) else dt.strftime("%Y-%m-%d")
    return f"게시일 {dt_txt}, 경과 {int(age) if age < 9999 else '확인불가'}일로 금일 보고/Top3 기준에서 제외"

def _issue_sentence_v21(row: pd.Series) -> str:
    title = clean(row.get("Headline"))
    issue = clean(row.get("Issue"))
    if issue in {"AD/CVD", "반덤핑/상계관세"}:
        return "반덤핑·상계관세 확대에 따른 추가관세 비용과 원산지 방어 리스크 점검이 필요합니다"
    if issue == "수출통제":
        return "수출통제 강화에 따른 전략물자·AI/반도체 거래 스크리닝 강화가 필요합니다"
    if issue == "CBAM":
        return "EU CBAM 대응을 위한 배출량 자료와 인증서 비용 관리가 필요합니다"
    if issue == "FTA/원산지":
        return "FTA 특혜관세 활용 가능성과 CO/BOM 원산지 정합성 점검이 필요합니다"
    if issue in {"관세정책", "통관", "통관/세관", "HS/품목분류"}:
        return "관세율·HS·통관 절차 변경에 따른 법인별 비용 및 신고 영향 확인이 필요합니다"
    # fallback from title
    return f"{title[:70]} 관련 관세·통상 영향 확인이 필요합니다"

def one_sentence_overall_v21(top3: pd.DataFrame) -> str:
    if top3 is None or top3.empty:
        return "금일 GTI Radar는 임원 보고 대상 핵심 관세·통상 뉴스가 제한적이며, 후속 모니터링 중심으로 관리가 필요합니다."
    phrases = []
    for _, r in top3.head(3).iterrows():
        s = _issue_sentence_v21(r)
        if s not in phrases:
            phrases.append(s)
    if len(phrases) == 1:
        body = phrases[0]
    elif len(phrases) == 2:
        body = f"{phrases[0]} 또한 {phrases[1]}"
    else:
        body = f"{phrases[0]} 또한 {phrases[1]} 동시에 {phrases[2]}"
    return f"금일 GTI Radar는 {body}"

def prepare_rows(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    rows["Issue"] = rows.apply(issue_for, axis=1)
    rows = dedup_report_rows(rows)
    rows["Mail Group"] = rows["Content Type"].map({"Regulation": GROUP_REGULATION}).fillna(GROUP_NEWS)
    rows["Major Changes"] = rows.apply(major_changes, axis=1)
    rows["Summary"] = rows.apply(report_summary, axis=1)
    rows["AI Analysis"] = rows.apply(report_impact, axis=1)
    rows["Action Plan"] = rows.apply(report_action, axis=1)

    # Freshness filter: do not report stale items unless explicitly allowed.
    if not GTI_ALLOW_STALE_ITEMS:
        before = len(rows)
        rows["_stale"] = rows.apply(lambda r: is_stale_item_v21(r, GTI_MAX_REPORT_AGE_DAYS), axis=1)
        stale = rows[rows["_stale"]].copy()
        if not stale.empty:
            print("[INFO] v21 stale items removed from report:")
            for _, r in stale.head(10).iterrows():
                print(f"  - {clean(r.get('Headline'))[:120]} / {stale_reason_v21(r)}")
        rows = rows[~rows["_stale"]].drop(columns=["_stale"], errors="ignore").copy()
        print(f"[INFO] v21 freshness filter: before={before}, after={len(rows)}, max_age_days={GTI_MAX_REPORT_AGE_DAYS}")

    # Recalculate weighted fallback columns if v19 exists.
    try:
        for col in WEIGHTED_COLS_V19:
            if col not in rows.columns:
                rows[col] = ""
        rows = rows.apply(recalc_weighted_score_v19, axis=1)
    except Exception:
        pass

    rows["_report_score"] = rows.apply(report_score, axis=1)
    try:
        rows = ensure_output_columns_v19(rows)
    except Exception:
        for col in OUTPUT_COLUMNS:
            if col not in rows.columns:
                rows[col] = ""

    rows = rows.sort_values(["_report_score", "_sort_date"], ascending=[False, False]).reset_index(drop=True)
    rows["No"] = range(1, len(rows) + 1)
    return rows

def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    pool = rows.copy()
    if pool.empty:
        return pool

    # Top3 freshness is stricter than full report.
    if not GTI_ALLOW_STALE_ITEMS:
        pool["_stale_top3"] = pool.apply(lambda r: is_stale_item_v21(r, GTI_MAX_TOP3_AGE_DAYS), axis=1)
        stale_top = pool[pool["_stale_top3"]].copy()
        if not stale_top.empty:
            print("[INFO] v21 stale items excluded from Top3:")
            for _, r in stale_top.head(10).iterrows():
                print(f"  - {clean(r.get('Headline'))[:120]} / {stale_reason_v21(r)}")
        pool = pool[~pool["_stale_top3"]].drop(columns=["_stale_top3"], errors="ignore").copy()

    if pool.empty:
        return pool

    pool["_top3_score"] = pool.apply(top3_deep_score, axis=1)
    pool = pool.sort_values(["_top3_score", "_sort_date"], ascending=[False, False])
    selected = []
    used_issues = set()
    for _, row in pool.iterrows():
        issue = clean(row.get("Issue"))
        if issue in used_issues and len(selected) < 3:
            continue
        selected.append(row)
        used_issues.add(issue)
        if len(selected) == 3:
            break
    if len(selected) < 3:
        for _, row in pool.iterrows():
            if any(clean(row.get("Headline")) == clean(x.get("Headline")) for x in selected):
                continue
            selected.append(row)
            if len(selected) == 3:
                break
    out = pd.DataFrame(selected).reset_index(drop=True)
    if not out.empty:
        out["No"] = range(1, len(out) + 1)
    return out

def overall_html(rows: pd.DataFrame, top3: pd.DataFrame) -> str:
    reg = rows[rows["Content Type"].eq("Regulation")]
    news = rows[rows["Content Type"].eq("News")]
    direct = rows[rows["Samsung Impact"].eq("Direct")]
    indirect = rows[rows["Samsung Impact"].eq("Indirect")]
    watch = rows[rows["Samsung Impact"].eq("Watch")]
    reference = rows[rows["Samsung Impact"].eq("Reference")] if "Reference" in rows["Samsung Impact"].unique() else rows.iloc[0:0]

    overall = one_sentence_overall_v21(top3)

    if top3 is None or top3.empty:
        top_lines = "<li>Top3 후보 없음: 최신성 기준을 충족하는 핵심 뉴스가 없습니다.</li>"
    else:
        top_lines = "".join(
            f"<li>{html.escape(clean(r.get('Headline'))[:120])}: {html.escape(_issue_sentence_v21(r))}</li>"
            for _, r in top3.head(3).iterrows()
        )

    result_line = (
        f"금일 선별 결과: 법규 {len(reg)}건, 주요뉴스 {len(news)}건 | "
        f"Direct {len(direct)}건, Indirect {len(indirect)}건, Watch {len(watch)}건, Reference {len(reference)}건"
    )

    return f"""
    <div style="padding:15px;background:#F4F6F8;border-left:6px solid #1F4E78;margin-bottom:18px;">
      <div style="font-size:15px;font-weight:bold;line-height:1.8;margin-bottom:10px;">
        {html.escape(overall)}
      </div>
      <div style="margin-top:8px;"><b>Top3 요약</b><ol style="margin-top:6px;">{top_lines}</ol></div>
      <div style="font-size:12px;color:#888;margin-top:12px;">
        *{html.escape(result_line)}
      </div>
    </div>
    """

# ======================================================================
# End of GTI STEP5 Freshness & Executive Summary Patch v21
# ======================================================================


# ======================================================================
# GTI STEP5 STRICT FINAL GUARD v22 - integrated into Mail Engine
# ----------------------------------------------------------------------
# Purpose
# - Do NOT read legacy/audit news fallback files at Step5.
# - Use only 4-1.regulation_ai_summary.xlsx and 4-2.news_ai_summary.xlsx.
# - Apply final mail-grade hard reject after Step4.
# - Recalibrate Samsung Impact before Top3 and mail table generation.
# ======================================================================

MAIL_REG_TARGET_MAX_V22 = int(os.getenv("GTI_MAIL_REG_TARGET_MAX", "25"))
MAIL_NEWS_TARGET_MAX_V22 = int(os.getenv("GTI_MAIL_NEWS_TARGET_MAX", "30"))
MAIL_MAX_PER_NEWS_ISSUE_V22 = int(os.getenv("GTI_MAIL_MAX_PER_NEWS_ISSUE", "2"))

V22_TRADE_REG_TERMS = [
    "관세", "통관", "수입", "수출", "수출입", "fta", "원산지", "환급", "보세", "외환",
    "덤핑", "상계관세", "세이프가드", "수출통제", "전략물자", "hs", "품목분류", "관세율",
    "무역구제", "세관", "할당관세", "쿼터", "cbam", "탄소국경", "철강", "알루미늄",
    "customs", "tariff", "duty", "import", "export", "origin", "rules of origin", "drawback",
    "anti-dumping", "antidumping", "countervailing", "safeguard", "export control", "entity list",
    "hs code", "classification", "carbon border", "melt and pour", "section 301", "section 232",
]

V22_REG_NOISE_TERMS = [
    "제대군인", "특수의료장비", "소방시설", "고용보험", "청년고용", "장애인복지", "농촌융복합",
    "농어촌정비", "인민군", "측량 및 지도", "상표법", "디자인보호법", "직제", "복지", "군 사관",
    "청년 실업", "병원선", "소방", "의료장비", "농로", "장애", "고용", "국가보훈", "과학기술정보통신부와 그 소속기관",
    "기초연금", "농어업", "평생 직업능력", "전통시장", "생태", "우체국보험", "보험료", "환급금",
    "국립수목원", "반부패", "청렴", "식품", "농심", "식약", "중기부", "문체", "총리령", "조세특례제한법",
]

V22_STRONG_POLICY_TERMS = [
    "section 301", "301조", "section 232", "232조", "anti-dumping", "anti dumping", "antidumping",
    "countervailing", "ad/cvd", "safeguard", "cbam", "carbon border", "tariff-rate quota",
    "tariff quota", "duty-free quota", "export control", "entity list", "uflpa", "forced labor",
    "rules of origin", "hs code", "classification", "customs duty", "import duty", "melt and pour",
    "반덤핑", "상계관세", "무역구제", "세이프가드", "탄소국경", "수출통제", "강제노동",
    "원산지", "품목분류", "할당관세", "관세율", "무관세", "쿼터", "통관", "보세", "환급", "관세",
]

V22_SAMSUNG_EXACT_TERMS = ["samsung", "samsung electronics", "samsung sdi", "samsung display", "삼성", "삼성전자", "삼성sdi", "삼성디스플레이"]
V22_PRODUCT_TERMS = [
    "semiconductor", "chip", "chips", "memory", "hbm", "battery", "display", "oled", "smartphone", "mobile", "appliance",
    "steel", "aluminum", "abs", "resin", "pcb", "wafer", "copper", "zinc", "rare earth",
    "반도체", "칩", "메모리", "배터리", "디스플레이", "스마트폰", "모바일", "가전", "철강", "알루미늄",
    "합성수지", "수지", "웨이퍼", "구리", "아연", "희토류",
]

V22_NEWS_NOISE_TERMS = [
    "손바닥뉴스", "시장동향", "경제 아카데미", "포항상의", "염전 노예", "교황", "대통령", "순방",
    "호르무즈", "사설", "칼럼", "opinion", "editorial", "youtube", "뉴스) - youtube", "운임 인상",
    "기자회견", "정상회담", "외교", "business trip", "g7 정상", "기업 해결사", "한반도 구상",
    "로펌 경쟁", "주가", "증시", "시장 전망", "market outlook", "자동차 시장", "수소배관 국산화",
    "돼지고기", "고등어", "오징어", "농산물", "쇠고기", "쌀", "설탕", "cheese", "홍콩",
    "타이어 생산거점", "車산업", "자동차 산업", "포항시", "아카데미", "지역경제", "대장간",
]

V22_HARD_REJECT_REASONS = [
    "event_training_tender_noise", "financial_industry_noise_without_trade_policy", "samsung_general_business_noise",
    "general_economy_without_samsung_policy", "low_value_general_news", "bilateral_industry_news_without_trade_policy",
    "ai_chip_industry_without_control_signal", "export_control_industry_without_control_signal", "google_news_original_url_unresolved",
    "future_date_abnormal", "no_valid_url", "v12_hard_reference_or_noise", "v12_no_customs_trade_action_signal",
    "strict_bad_or_unresolved_url", "strict_digest_politics_market_noise", "strict_no_concrete_customs_trade_signal",
    "mail_guard_bad_or_unresolved_url", "mail_guard_existing_hard_reject", "mail_guard_digest_politics_market_noise",
    "mail_guard_no_concrete_customs_trade_signal", "mail_guard_reference_not_reportable",
]

V22_BAD_URL_STATUSES = ["SEARCH_NO_GOOD_RESULT", "NO_ORIGINAL_URL", "GOOGLE_UNRESOLVED", "GOOGLE_HOME", "EMPTY_OR_BAD_LINK"]


def _v22_has_any(text: str, terms: list[str]) -> bool:
    low = str(text or "").lower()
    return any(str(t).lower() in low for t in terms if str(t).strip())


def _v22_text(row: pd.Series, cols: list[str]) -> str:
    return " ".join(clean(row.get(c, "")) for c in cols).lower()


def _v22_good_url(row: pd.Series) -> bool:
    u = best_url_from_values([row.get("URL", ""), row.get("Source", "")])
    q = clean(row.get("URL_Quality", "")).upper()
    ul = u.lower()
    if not ul.startswith(("http://", "https://")):
        return False
    if "youtube.com" in ul or "youtu.be" in ul:
        return False
    if "news.google.com" in ul and not ("/rss/articles/" in ul or "/articles/" in ul):
        return False
    return not any(x in q for x in V22_BAD_URL_STATUSES)


def _v22_recalibrate_impact(row: pd.Series) -> str:
    original = _v22_text(row, [
        "Headline", "Major Changes", "Original Post Summary", "Original Body Text", "Cluster", "Agency", "Source", "URL",
        "Issue", "KeywordMatches",
    ])
    strong = _v22_has_any(original, V22_STRONG_POLICY_TERMS)
    samsung = _v22_has_any(original, V22_SAMSUNG_EXACT_TERMS)
    product = _v22_has_any(original, V22_PRODUCT_TERMS)
    if samsung and strong:
        return "Direct"
    if product and strong:
        return "Indirect"
    if strong:
        return "Watch"
    return "Reference"


def _v22_reg_keep(row: pd.Series) -> bool:
    primary = _v22_text(row, ["Headline", "URL", "Source", "Agency"])
    title = clean(row.get("Headline", "")).lower()
    has_trade = _v22_has_any(primary, V22_TRADE_REG_TERMS)
    is_noise = _v22_has_any(title, V22_REG_NOISE_TERMS)
    if "환급" in title and not _v22_has_any(title, ["관세", "수출용", "원재료", "drawback", "customs"]):
        is_noise = True
    return has_trade and not is_noise


def _v22_news_keep(row: pd.Series) -> bool:
    # Trust STEP4-selected news unless a real hard blocker remains. Do not use RejectReason
    # itself as policy text, because labels such as "no_customs" can distort the signal.
    text = _v22_text(row, ["Headline", "Major Changes", "Summary", "AI Analysis", "Action Plan", "Issue", "Impact Reason", "KeywordMatches"])
    rr = clean(row.get("RejectReason", ""))
    rr_set = {x.strip() for x in rr.split(";") if x.strip()}
    broad_reasons = {
        "weighted_v18_not_topN_or_noise",
        "v20_major_title_or_weighted_below_topN_or_noise",
        "weak_samsung_relevance",
        "report_issue_duplicate_compressed",
        "expanded_policy_watch",
        "v12_hard_reference_or_noise",
        "v12_no_customs_trade_action_signal",
        "strict_existing_hard_reject",
        "strict_backfill_reportable_policy",
    }
    if not _v22_good_url(row):
        return False
    hard_hits = [x for x in V22_HARD_REJECT_REASONS if x in rr and x not in broad_reasons]
    if hard_hits:
        return False
    if rr_set and not rr_set <= broad_reasons and any(x in rr for x in V22_HARD_REJECT_REASONS):
        return False
    if _v22_has_any(text, V22_NEWS_NOISE_TERMS):
        return False
    if not _v22_has_any(text, V22_STRONG_POLICY_TERMS):
        return False
    return _v22_recalibrate_impact(row) != "Reference"


def _v22_issue_key(row: pd.Series) -> str:
    t = _v22_text(row, ["Headline", "Major Changes", "Summary", "AI Analysis", "Issue", "Cluster"])
    if "india" in t and ("uk" in t or "britain" in t) and ("fta" in t or "ceta" in t or "trade agreement" in t):
        return "india_uk_fta"
    if "eu" in t and ("steel" in t or "철강" in t) and ("safeguard" in t or "세이프가드" in t or "quota" in t or "쿼터" in t):
        return "eu_steel_safeguard"
    if "section 301" in t or "301조" in t:
        return "section_301"
    if "section 232" in t or "232조" in t:
        return "section_232"
    if "cbam" in t or "탄소국경" in t:
        return "cbam"
    if "korea" in t and ("mongol" in t or "몽골" in t) and "cepa" in t:
        return "korea_mongolia_cepa"
    return re.sub(r"[^0-9a-z가-힣]+", " ", clean(row.get("Headline", "")).lower()).strip()[:90]


def _v22_filter_mail_rows(rows: pd.DataFrame) -> pd.DataFrame:
    if rows.empty:
        return rows
    rows = rows.copy()
    reg = rows[rows["Content Type"].eq("Regulation")].copy()
    news = rows[rows["Content Type"].eq("News")].copy()

    before_reg, before_news = len(reg), len(news)
    if not reg.empty:
        reg = reg[reg.apply(_v22_reg_keep, axis=1)].copy()
        if "_integrated_score" in reg.columns:
            reg = reg.sort_values(["_integrated_score", "_sort_date"], ascending=[False, False])
        reg = reg.head(MAIL_REG_TARGET_MAX_V22)

    if not news.empty:
        news["Samsung Impact"] = news.apply(_v22_recalibrate_impact, axis=1)
        news = news[news.apply(_v22_news_keep, axis=1)].copy()
        if not news.empty:
            news["_v22_issue_key"] = news.apply(_v22_issue_key, axis=1)
            sort_cols = [c for c in ["WeightedScore", "Importance Score", "_integrated_score", "_sort_date"] if c in news.columns]
            if sort_cols:
                news = news.sort_values(sort_cols, ascending=[False] * len(sort_cols))
            dup_mask = news.groupby("_v22_issue_key").cumcount() >= MAIL_MAX_PER_NEWS_ISSUE_V22
            news = news[~dup_mask].copy()
            news = news.head(MAIL_NEWS_TARGET_MAX_V22).drop(columns=["_v22_issue_key"], errors="ignore")

    out = pd.concat([reg, news], ignore_index=True, sort=False)
    print(
        f"[INFO] v22 strict mail guard: regulation {before_reg}->{len(reg)}, "
        f"news {before_news}->{len(news)}, total={len(out)}"
    )
    return out.reset_index(drop=True)


def read_step4_results() -> pd.DataFrame:
    """v22 override: read only official Step4 selected outputs, not legacy/audit fallback files."""
    frames = []
    if REGULATION_INPUT_FILE.exists():
        reg = normalize_input(pd.read_excel(REGULATION_INPUT_FILE), "Regulation", REGULATION_INPUT_FILE)
        frames.append(reg)
        print(f"[INFO] v22 regulation loaded: {REGULATION_INPUT_FILE} rows={len(reg)}")
    else:
        print(f"[WARN] v22 regulation missing: {REGULATION_INPUT_FILE}")

    if NEWS_INPUT_FILE.exists():
        news = normalize_input(pd.read_excel(NEWS_INPUT_FILE), "News", NEWS_INPUT_FILE)
        if NEWS_MAX_ROWS > 0:
            news = news.head(NEWS_MAX_ROWS)
        frames.append(news)
        print(f"[INFO] v22 news loaded: {NEWS_INPUT_FILE} rows={len(news)}")
    else:
        print(f"[WARN] v22 news missing: {NEWS_INPUT_FILE}")

    if not frames:
        raise FileNotFoundError(f"STEP4 outputs not found: {REGULATION_INPUT_FILE}, {NEWS_INPUT_FILE}")

    rows = pd.concat(frames, ignore_index=True, sort=False)
    rows["URL"] = rows.apply(lambda r: best_url_from_values([r.get("URL"), r.get("Source")]), axis=1)
    rows["_dedup_key"] = rows.apply(
        lambda r: clean(r.get("URL")).lower() or (
            clean(r.get("Headline"))[:160] + "|" + clean(r.get("Agency")) + "|" + clean(r.get("Date"))
        ),
        axis=1,
    )
    rows = rows.drop_duplicates(subset=["_dedup_key"], keep="first").drop(columns=["_dedup_key"], errors="ignore")
    rows["_integrated_score"] = rows.apply(
        lambda r: priority_weight(r.get("Priority Group")) + risk_weight(r.get("Risk")) +
                  (180 if clean(r.get("Content Type")) == "Regulation" else 0) +
                  safe_num(r.get("Importance Score")) + safe_num(r.get("WeightedScore")),
        axis=1,
    )
    rows = _v22_filter_mail_rows(rows)
    print(
        f"[INFO] v22 total input rows={len(rows)} / "
        f"regulation={int(rows['Content Type'].eq('Regulation').sum())} / "
        f"news={int(rows['Content Type'].eq('News').sum())}"
    )
    return rows.reset_index(drop=True)


def prepare_rows(rows: pd.DataFrame) -> pd.DataFrame:
    """v22 override: prepare after final mail-grade filtering and impact recalibration."""
    rows = rows.copy()
    if rows.empty:
        return rows
    news_mask = rows["Content Type"].eq("News")
    rows.loc[news_mask, "Samsung Impact"] = rows[news_mask].apply(_v22_recalibrate_impact, axis=1)
    rows["Issue"] = rows.apply(issue_for, axis=1)
    rows = dedup_report_rows(rows)
    rows["Mail Group"] = rows["Content Type"].map({"Regulation": GROUP_REGULATION}).fillna(GROUP_NEWS)
    rows["Major Changes"] = rows.apply(major_changes, axis=1)
    rows["Summary"] = rows.apply(report_summary, axis=1)
    rows["AI Analysis"] = rows.apply(report_impact, axis=1)
    rows["Action Plan"] = rows.apply(report_action, axis=1)
    for col in WEIGHTED_COLS_V19:
        if col not in rows.columns:
            rows[col] = ""
    rows = rows.apply(recalc_weighted_score_v19, axis=1)
    rows["_report_score"] = rows.apply(report_score, axis=1)
    rows = ensure_output_columns_v19(rows)
    rows = rows.sort_values(["_report_score", "_sort_date"], ascending=[False, False]).reset_index(drop=True)
    rows["No"] = range(1, len(rows) + 1)
    return rows


def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    """v22 override: choose Top3 only from non-Reference, actionable customs/trade rows."""
    if rows.empty:
        return rows
    pool = rows.copy()
    pool = pool[pool["Samsung Impact"].ne("Reference")].copy()
    if pool.empty:
        pool = rows.copy()
    pool["_v22_top3_actionable"] = pool.apply(
        lambda r: _v22_has_any(_v22_text(r, ["Headline", "Major Changes", "Summary", "AI Analysis", "Action Plan", "Issue"]), V22_STRONG_POLICY_TERMS),
        axis=1,
    )
    cand = pool[pool["_v22_top3_actionable"]].copy()
    if cand.empty:
        cand = pool.copy()
    cand["_top3_score"] = cand.apply(top3_deep_score, axis=1)
    cand = cand.sort_values(["_top3_score", "_sort_date"], ascending=[False, False])
    selected, used = [], set()
    for _, row in cand.iterrows():
        key = clean(row.get("Issue")) or _v22_issue_key(row)
        if key in used:
            continue
        selected.append(row)
        used.add(key)
        if len(selected) == 3:
            break
    if len(selected) < 3:
        for _, row in cand.iterrows():
            if any(clean(row.get("Headline")) == clean(x.get("Headline")) for x in selected):
                continue
            selected.append(row)
            if len(selected) == 3:
                break
    out = pd.DataFrame(selected).drop(columns=["_v22_top3_actionable"], errors="ignore").reset_index(drop=True)
    if not out.empty:
        out["No"] = range(1, len(out) + 1)
    return out

# ======================================================================
# End of GTI STEP5 STRICT FINAL GUARD v22
# ======================================================================


# ======================================================================
# GTI STEP5 STRICT FINAL GUARD v23
# ----------------------------------------------------------------------
# Added on top of v22:
# 1) Final freshness filter by Date / URL / title-embedded date
# 2) Remove event/seminar/webinar/forum/training type news from mail report
# 3) Re-number after final filtering, before Top3 selection
# ======================================================================

V23_MAIL_MAX_AGE_DAYS = int(os.getenv("GTI_MAIL_MAX_AGE_DAYS", "45"))
V23_MAIL_MAX_AGE_DAYS_REG = int(os.getenv("GTI_MAIL_MAX_AGE_DAYS_REG", str(V23_MAIL_MAX_AGE_DAYS)))
V23_MAIL_MAX_AGE_DAYS_NEWS = int(os.getenv("GTI_MAIL_MAX_AGE_DAYS_NEWS", str(V23_MAIL_MAX_AGE_DAYS)))

V23_EVENT_SEMINAR_NOISE_TERMS = [
    "seminar", "webinar", "conference", "forum", "symposium", "workshop", "training",
    "세미나", "웨비나", "컨퍼런스", "콘퍼런스", "포럼", "심포지엄", "워크숍", "교육", "설명회",
    "esg 온", "on 세미나", "제39회 esg", "발제", "패널토론", "기조연설",
]

V23_MONTHS = {
    "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
    "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6,
    "jul": 7, "july": 7, "aug": 8, "august": 8, "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10, "nov": 11, "november": 11, "dec": 12, "december": 12,
}

def _v23_text(row: pd.Series, cols: list[str] | None = None) -> str:
    if cols is None:
        cols = [
            "Headline", "Summary", "AI Analysis", "Action Plan", "Major Changes", "Issue",
            "Country", "Agency", "URL", "Original Body Text", "Original Post Summary",
        ]
    return " ".join(clean(row.get(c)) for c in cols).lower()

def _v23_has_any(text: str, terms: list[str]) -> bool:
    return any(str(t).lower() in text for t in terms if str(t).strip())

def _v23_run_date_ts() -> pd.Timestamp:
    dt = pd.to_datetime(RUN_DATE, errors="coerce")
    return pd.Timestamp(datetime.now().date()) if pd.isna(dt) else pd.Timestamp(dt).normalize()

def _v23_parse_embedded_date(text_value: str):
    text_value = clean(text_value)
    if not text_value:
        return pd.NaT

    # 2026-06-17, 2026/06/17, 2026.06.17, or URL /2026/06/17/
    m = re.search(r"(20\d{2})[.\-/](\d{1,2})[.\-/](\d{1,2})", text_value)
    if m:
        return pd.to_datetime(f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}", errors="coerce")

    # August 20, 2025
    m = re.search(r"\b([A-Za-z]{3,9})\s+(\d{1,2}),?\s+(20\d{2})\b", text_value)
    if m:
        mon = V23_MONTHS.get(m.group(1).lower())
        if mon:
            return pd.to_datetime(f"{m.group(3)}-{mon:02d}-{int(m.group(2)):02d}", errors="coerce")

    # 6March2026 or 6 March 2026
    m = re.search(r"\b(\d{1,2})\s*([A-Za-z]{3,9})\s*(20\d{2})\b", text_value)
    if m:
        mon = V23_MONTHS.get(m.group(2).lower())
        if mon:
            return pd.to_datetime(f"{m.group(3)}-{mon:02d}-{int(m.group(1)):02d}", errors="coerce")

    return pd.NaT

def _v23_best_row_date(row: pd.Series):
    # Prefer normalized Date, then Publish Date if present, then URL/title embedded dates.
    for c in ["Date", "Publish Date", "date", "published_at"]:
        if c in row.index:
            dt = pd.to_datetime(clean(row.get(c)), errors="coerce")
            if not pd.isna(dt):
                return pd.Timestamp(dt).normalize()
    embedded = _v23_parse_embedded_date(" ".join(clean(row.get(c)) for c in ["URL", "Headline", "Summary", "Major Changes"]))
    if not pd.isna(embedded):
        return pd.Timestamp(embedded).normalize()
    return pd.NaT

def _v23_is_stale(row: pd.Series) -> bool:
    dt = _v23_best_row_date(row)
    if pd.isna(dt):
        return False
    max_days = V23_MAIL_MAX_AGE_DAYS_REG if clean(row.get("Content Type")) == "Regulation" else V23_MAIL_MAX_AGE_DAYS_NEWS
    age_days = (_v23_run_date_ts() - pd.Timestamp(dt).normalize()).days
    return age_days > max_days

def _v23_is_future_abnormal(row: pd.Series) -> bool:
    dt = _v23_best_row_date(row)
    if pd.isna(dt):
        return False
    return (_v23_run_date_ts() - pd.Timestamp(dt).normalize()).days < -2

def _v23_is_event_seminar_noise(row: pd.Series) -> bool:
    if clean(row.get("Content Type")) != "News":
        return False
    # Event noise must be title/metadata based. Do not use AI-generated Summary/Action Plan,
    # because trade-law articles can mention training/confirmation text in generated prose.
    title_text = _v23_text(row, ["Headline", "Agency", "Source", "URL"])
    if not _v23_has_any(title_text, V23_EVENT_SEMINAR_NOISE_TERMS):
        return False
    # Strong customs/trade-law titles such as preliminary anti-subsidy duty, tariff,
    # export control, origin, or CBAM should not be removed as event/seminar noise.
    if _v23_has_any(title_text, V22_STRONG_POLICY_TERMS):
        return False
    # Keep official customs/trade notices even if they include "comment request";
    # remove secondary event/seminar articles.
    official_source = _v23_has_any(clean(row.get("Agency")).lower() + " " + clean(row.get("URL")).lower(), [
        "federalregister.gov", "customs", "gov", "europa.eu", "wto.org", "ustr.gov", "cbp.gov"
    ])
    return not official_source

def _v23_final_report_filter(rows: pd.DataFrame) -> pd.DataFrame:
    if rows.empty:
        return rows
    before = len(rows)
    rows = rows.copy()
    reasons = []
    keep_mask = []
    for _, row in rows.iterrows():
        r = []
        if _v23_is_stale(row):
            r.append("v23_stale_over_max_age")
        if _v23_is_future_abnormal(row):
            r.append("v23_future_date_abnormal")
        if _v23_is_event_seminar_noise(row):
            r.append("v23_event_seminar_noise")
        reasons.append("; ".join(r))
        keep_mask.append(not r)

    dropped = rows.loc[[not x for x in keep_mask]].copy()
    kept = rows.loc[keep_mask].copy()

    if len(dropped) > 0:
        try:
            print("[INFO] v23 final mail guard removed:")
            for _, r in dropped.head(20).iterrows():
                print(f"  - {clean(r.get('Headline'))[:110]} / date={clean(r.get('Date')) or clean(_v23_best_row_date(r))} / reason={reasons[int(r.name)] if isinstance(r.name, int) and r.name < len(reasons) else ''}")
        except Exception:
            pass

    kept = kept.reset_index(drop=True)
    if "No" in kept.columns:
        kept["No"] = range(1, len(kept) + 1)
    print(f"[INFO] v23 final mail guard: before={before}, after={len(kept)}, removed={before-len(kept)}, max_age_days={V23_MAIL_MAX_AGE_DAYS}")
    return kept

_v23_prepare_rows_base = prepare_rows

def prepare_rows(rows: pd.DataFrame) -> pd.DataFrame:
    rows = _v23_prepare_rows_base(rows)
    rows = _v23_final_report_filter(rows)
    # Re-sort after final filtering to avoid gaps and stale Top3 candidates.
    if "_report_score" in rows.columns and "_sort_date" in rows.columns:
        rows = rows.sort_values(["_report_score", "_sort_date"], ascending=[False, False]).reset_index(drop=True)
    elif "_integrated_score" in rows.columns and "_sort_date" in rows.columns:
        rows = rows.sort_values(["_integrated_score", "_sort_date"], ascending=[False, False]).reset_index(drop=True)
    if "No" in rows.columns:
        rows["No"] = range(1, len(rows) + 1)
    return rows

_v23_choose_top3_base = choose_top3

def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    pool = _v23_final_report_filter(rows)
    if pool.empty:
        pool = rows.copy()
    out = _v23_choose_top3_base(pool)
    if not out.empty:
        out["No"] = range(1, len(out) + 1)
    return out

# ======================================================================
# End of GTI STEP5 STRICT FINAL GUARD v23
# ======================================================================


# ======================================================================
# GTI STEP5 STRICT FINAL GUARD v25
# ----------------------------------------------------------------------
# Added on top of v23:
# 1) Remove news rows where article body was not actually available.
# 2) Remove Google Alert/feed URLs when original URL recovery failed.
# 3) Compress remaining same-issue duplicates: steel-law regulations and India-UK FTA.
# 4) Sanitize hs_hint values that are actually years/dates, e.g. 2026.0.
# ======================================================================

V24_MAX_INDIA_UK_FTA = int(os.getenv("GTI_MAIL_MAX_INDIA_UK_FTA", "1"))
V24_MAX_STEEL_LAW_REG = int(os.getenv("GTI_MAIL_MAX_STEEL_LAW_REG", "1"))

V24_BODY_UNAVAILABLE_TERMS = [
    "본문 확인 불가", "본문 내용이 없어", "본문을 가져오지 못했습니다", "원문 확인이 불가능",
    "제목만으로 요약하지 않았습니다", "body unavailable", "no article body", "could not fetch",
]

V24_GOOGLE_ALERT_URL_TERMS = [
    "google.co.kr/alerts/feeds", "google.com/alerts/feeds", "google.co.kr/alerts", "google.com/alerts",
    "alerts/feeds/",
]


def _v24_low(text: str) -> str:
    return clean(text).lower()


def _v24_is_bad_hs_hint(value) -> bool:
    txt = clean(value)
    if not txt:
        return False
    low = txt.lower()
    if low in {"nan", "none", "null", "본문에서 확인 불가"}:
        return True
    # Only a year or year-like float: 2026, 2026.0
    if re.fullmatch(r"20\d{2}(\.0+)?", txt):
        return True
    # Date-like values accidentally mapped into HS field.
    if re.fullmatch(r"20\d{2}[.\-/]\d{1,2}([.\-/]\d{1,2})?", txt):
        return True
    # Excel serial/date-shaped junk is not an HS unless it contains a plausible HS delimiter/length.
    digits = re.sub(r"\D", "", txt)
    if len(digits) == 4 and digits.startswith("20"):
        return True
    return False


def _v24_sanitize_rows(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    for col in ["hs_hint", "tariff_rate_hint", "effective_date_hint", "change_detail_hint"]:
        if col not in rows.columns:
            rows[col] = ""
    rows["hs_hint"] = rows["hs_hint"].apply(lambda v: "" if _v24_is_bad_hs_hint(v) else v)
    return rows


def _v24_has_unavailable_body(row: pd.Series) -> bool:
    if clean(row.get("Content Type")) != "News":
        return False
    text = " ".join(clean(row.get(c)) for c in [
        "Headline", "Summary", "Major Changes", "AI Analysis", "Action Plan", "Original Body Text", "Original Post Summary"
    ])
    return _v22_has_any(text, V24_BODY_UNAVAILABLE_TERMS)


def _v24_google_alert_unresolved(row: pd.Series) -> bool:
    if clean(row.get("Content Type")) != "News":
        return False
    url = clean(row.get("URL")) or clean(row.get("Source"))
    low = url.lower()
    if not _v22_has_any(low, V24_GOOGLE_ALERT_URL_TERMS):
        return False
    # If the URL is still a Google Alert/feed link at Step5, original URL recovery failed.
    return True


def _v24_special_issue_key(row: pd.Series) -> str:
    t = _v22_text(row, ["Headline", "Summary", "Major Changes", "AI Analysis", "Action Plan", "Issue", "URL", "Agency"])
    content_type = clean(row.get("Content Type"))
    if content_type == "Regulation" and "철강산업" in t and ("특별법" in t or "시행규칙" in t or "탄소중립" in t):
        return "reg_korea_steel_special_act"
    if "india" in t and ("uk" in t or "britain" in t or "united kingdom" in t) and ("fta" in t or "ceta" in t or "trade pact" in t or "trade agreement" in t):
        return "news_india_uk_fta"
    return ""


def _v24_dedup_rank(row: pd.Series) -> float:
    score = safe_num(row.get("_report_score")) + safe_num(row.get("Importance Score")) + risk_weight(row.get("Risk"))
    title = _v24_low(row.get("Headline"))
    url = _v24_low(row.get("URL"))
    # Prefer official law/current detailed row over truncated duplicate.
    if "law.go.kr" in url or "federalregister.gov" in url or "customs" in url:
        score += 80
    if "시행규칙" in title and "제16호" in title:
        score += 40
    if title.endswith("...") or title.endswith("…"):
        score -= 80
    # For India-UK cluster, prefer articles with article body and clearer trade-policy signal.
    if _v24_has_unavailable_body(row):
        score -= 500
    if "steel" in title and ("safeguard" in title or "british" in title):
        score += 50
    return score


def _v24_final_report_filter(rows: pd.DataFrame) -> pd.DataFrame:
    if rows.empty:
        return rows
    before = len(rows)
    rows = _v24_sanitize_rows(rows).copy()

    drop_reasons = []
    keep = []
    for _, row in rows.iterrows():
        r = []
        if _v24_has_unavailable_body(row):
            r.append("v24_body_unavailable_news")
        if _v24_google_alert_unresolved(row):
            r.append("v24_google_alert_url_unresolved")
        drop_reasons.append("; ".join(r))
        keep.append(not r)

    dropped = rows.loc[[not x for x in keep]].copy()
    kept = rows.loc[keep].copy()
    if len(dropped) > 0:
        print("[INFO] v24 final mail guard removed:")
        for pos, (_, r) in enumerate(dropped.head(20).iterrows()):
            idx = r.name if isinstance(r.name, int) and r.name < len(drop_reasons) else pos
            reason = drop_reasons[idx] if isinstance(idx, int) and idx < len(drop_reasons) else ""
            print(f"  - {clean(r.get('Headline'))[:110]} / reason={reason}")

    if not kept.empty:
        kept["_v24_special_issue_key"] = kept.apply(_v24_special_issue_key, axis=1)
        kept["_v24_dedup_rank"] = kept.apply(_v24_dedup_rank, axis=1)
        kept = kept.sort_values(["_v24_dedup_rank", "_sort_date"], ascending=[False, False])
        selected_parts = []
        regular = kept[kept["_v24_special_issue_key"].eq("")].copy()
        selected_parts.append(regular)
        for key, frame in kept[kept["_v24_special_issue_key"].ne("")].groupby("_v24_special_issue_key", sort=False):
            limit = V24_MAX_INDIA_UK_FTA if key == "news_india_uk_fta" else V24_MAX_STEEL_LAW_REG
            selected_parts.append(frame.head(max(1, limit)))
        kept2 = pd.concat(selected_parts, ignore_index=True, sort=False) if selected_parts else kept
        duplicate_removed = len(kept) - len(kept2)
        kept = kept2.drop(columns=["_v24_special_issue_key", "_v24_dedup_rank"], errors="ignore")
    else:
        duplicate_removed = 0

    # Restore executive order after filtering/compression.
    if "_report_score" in kept.columns and "_sort_date" in kept.columns:
        kept = kept.sort_values(["_report_score", "_sort_date"], ascending=[False, False]).reset_index(drop=True)
    else:
        kept = kept.reset_index(drop=True)
    if "No" in kept.columns:
        kept["No"] = range(1, len(kept) + 1)
    print(f"[INFO] v24 final mail guard: before={before}, after={len(kept)}, removed={before-len(kept)}, duplicates_removed={duplicate_removed}")
    return kept


_v24_prepare_rows_base = prepare_rows


def prepare_rows(rows: pd.DataFrame) -> pd.DataFrame:
    rows = _v24_prepare_rows_base(rows)
    rows = _v24_final_report_filter(rows)
    if "No" in rows.columns:
        rows["No"] = range(1, len(rows) + 1)
    return rows


_v24_choose_top3_base = choose_top3


def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    pool = _v24_final_report_filter(rows)
    if pool.empty:
        pool = rows.copy()
    out = _v24_choose_top3_base(pool)
    if not out.empty:
        out["No"] = range(1, len(out) + 1)
    return out



# ======================================================================
# GTI STEP5 TITLE KEYWORD STRICT GUARD v26
# - Top3: title must contain customs/trade keyword.
# - News table: title must contain customs/trade keyword; otherwise excluded.
# - This guard is intentionally title-based to avoid Gemini/summary-driven over-selection.
# ======================================================================

GTI_TITLE_KEYWORD_STRICT = os.getenv("GTI_TITLE_KEYWORD_STRICT", "Y").strip().upper() not in {"N", "NO", "0", "FALSE"}
GTI_NEWS_TITLE_KEYWORD_REQUIRED = os.getenv("GTI_NEWS_TITLE_KEYWORD_REQUIRED", "Y").strip().upper() not in {"N", "NO", "0", "FALSE"}
GTI_TOP3_TITLE_KEYWORD_REQUIRED = os.getenv("GTI_TOP3_TITLE_KEYWORD_REQUIRED", "Y").strip().upper() not in {"N", "NO", "0", "FALSE"}

_V26_BASE_TITLE_KEYWORDS = [
    # Korean customs / trade / law
    "관세", "통관", "세관", "수입", "수출", "수출입", "무역", "통상", "관세율", "추가관세",
    "반덤핑", "덤핑", "상계관세", "무역구제", "세이프가드", "쿼터", "무관세",
    "원산지", "품목분류", "hs", "hs코드", "fta", "cepa", "협정", "특혜관세", "관세환급", "환급",
    "보세", "수입신고", "수출신고", "전략물자", "수출통제", "제재", "강제노동", "cbam", "탄소국경",
    "고시", "공고", "입법예고", "행정예고", "관보", "시행규칙", "시행령",
    # English customs / trade / law
    "customs", "tariff", "tariffs", "duty", "duties", "import", "export", "trade", "section 301", "section 232",
    "anti-dumping", "anti dumping", "antidumping", "countervailing", "ad/cvd", "safeguard", "quota", "duty-free",
    "rules of origin", "origin", "hs code", "classification", "fta", "cepa", "usmca", "cbam", "carbon border",
    "export control", "entity list", "forced labor", "uflpa", "federal register", "notice", "regulation",
]


def _v26_title_keywords() -> list[str]:
    extra = os.getenv("GTI_TITLE_KEYWORDS", "").strip()
    terms = list(_V26_BASE_TITLE_KEYWORDS)
    if extra:
        terms.extend([x.strip() for x in re.split(r"[;,|]", extra) if x.strip()])
    # unique, longer terms first to improve diagnostics if needed
    out = []
    seen = set()
    for t in terms:
        k = clean(t).lower()
        if k and k not in seen:
            out.append(k)
            seen.add(k)
    return sorted(out, key=len, reverse=True)


def _v26_title_keyword_matches(row: pd.Series) -> list[str]:
    title = clean(row.get("Headline")).lower()
    if not title:
        return []
    return [kw for kw in _v26_title_keywords() if kw in title]


def _v26_title_has_keyword(row: pd.Series) -> bool:
    return bool(_v26_title_keyword_matches(row))


def _v26_filter_news_by_title_keyword(rows: pd.DataFrame) -> pd.DataFrame:
    if rows.empty or not GTI_TITLE_KEYWORD_STRICT or not GTI_NEWS_TITLE_KEYWORD_REQUIRED:
        return rows
    rows = rows.copy()
    is_news = rows.get("Content Type", "").astype(str).str.lower().eq("news") if "Content Type" in rows.columns else pd.Series(False, index=rows.index)
    no_kw = rows.apply(lambda r: not _v26_title_has_keyword(r), axis=1)
    drop_mask = is_news & no_kw
    dropped = rows.loc[drop_mask].copy()
    if not dropped.empty:
        print("[INFO] v26 title keyword guard removed from news:")
        for _, r in dropped.head(30).iterrows():
            print(f"  - {clean(r.get('Headline'))[:120]} / reason=v26_no_title_keyword")
    kept = rows.loc[~drop_mask].copy().reset_index(drop=True)
    if "No" in kept.columns:
        kept["No"] = range(1, len(kept) + 1)
    print(f"[INFO] v26 title keyword news guard: before={len(rows)}, after={len(kept)}, removed={len(dropped)}")
    return kept


_v26_prepare_rows_base = prepare_rows


def prepare_rows(rows: pd.DataFrame) -> pd.DataFrame:
    rows = _v26_prepare_rows_base(rows)
    rows = _v26_filter_news_by_title_keyword(rows)
    if "No" in rows.columns:
        rows["No"] = range(1, len(rows) + 1)
    return rows


_v26_choose_top3_base = choose_top3


def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    if rows.empty or not GTI_TITLE_KEYWORD_STRICT or not GTI_TOP3_TITLE_KEYWORD_REQUIRED:
        return _v26_choose_top3_base(rows)
    pool = rows[rows.apply(_v26_title_has_keyword, axis=1)].copy()
    removed = len(rows) - len(pool)
    if removed:
        print(f"[INFO] v26 Top3 title keyword guard: candidates={len(rows)}, keyword_candidates={len(pool)}, removed={removed}")
    if pool.empty:
        return pool
    out = _v26_choose_top3_base(pool)
    # Base chooser can still fill with non-keyword rows only if pool was not used; enforce again.
    out = out[out.apply(_v26_title_has_keyword, axis=1)].copy().reset_index(drop=True)
    if not out.empty:
        out["No"] = range(1, len(out) + 1)
    return out

# ======================================================================
# End of GTI STEP5 TITLE KEYWORD STRICT GUARD v26
# ======================================================================

# ======================================================================
# End of GTI STEP5 STRICT FINAL GUARD v25
# ======================================================================


# ======================================================================
# GTI STEP5 STRICT FINAL GUARD v27
# ----------------------------------------------------------------------
# 24-hour mail exposure rule + original URL accuracy + official LAW1 source gate
# - 72h collection is allowed upstream; final mail is limited to current daily window.
# - EXCLUDED/REJECT rows are never mailed.
# - Regulation rows must be from official source domains only.
# - News rows must have an original/non-Google URL and known publish date.
# - Top3 has stricter gates: title keyword + 24h + URL OK + not EXCLUDED.
# - Add audit columns to Excel for traceability.
# ======================================================================

GTI_MAIL_MAX_AGE_HOURS_V27 = float(os.getenv("GTI_MAIL_MAX_AGE_HOURS", "24"))
GTI_MAIL_ALLOW_PREVIOUS_DATE_ONLY_V27 = os.getenv("GTI_MAIL_ALLOW_PREVIOUS_DATE_ONLY", "Y").strip().upper() not in {"N", "NO", "0", "FALSE"}
GTI_MAIL_DROP_UNKNOWN_DATE_V27 = os.getenv("GTI_MAIL_DROP_UNKNOWN_DATE", "Y").strip().upper() not in {"N", "NO", "0", "FALSE"}
GTI_MAIL_STRICT_ORIGINAL_URL_V27 = os.getenv("GTI_MAIL_STRICT_ORIGINAL_URL", "Y").strip().upper() not in {"N", "NO", "0", "FALSE"}
GTI_MAIL_OFFICIAL_REG_ONLY_V27 = os.getenv("GTI_MAIL_OFFICIAL_REG_ONLY", "Y").strip().upper() not in {"N", "NO", "0", "FALSE"}

# Extend Excel output with audit columns. Keep original order and append trace fields.
_V27_AUDIT_COLUMNS = [
    "GoogleURL", "OriginalURLCandidate", "BestLinkURL", "URLRestoreStatus", "URLDecodeStatus",
    "FinalURLDomain", "FreshnessAgeHours", "FreshnessStatus", "OfficialSourceFlag", "RejectReason", "MailGuardReason",
]
for _v27_col in _V27_AUDIT_COLUMNS:
    if _v27_col not in OUTPUT_COLUMNS:
        OUTPUT_COLUMNS.append(_v27_col)

V27_OFFICIAL_REG_DOMAINS = [
    # Korea official law/customs sources
    "customs.go.kr", "unipass.customs.go.kr", "law.go.kr", "moleg.go.kr", "gwanbo.go.kr",
    "moef.go.kr", "motie.go.kr", "korea.kr",
    # US official sources
    "federalregister.gov", "cbp.gov", "ustr.gov", "bis.gov", "trade.gov", "usitc.gov", "commerce.gov",
    # EU / international / major production-country official sources
    "europa.eu", "taxation-customs.ec.europa.eu", "policy.trade.ec.europa.eu",
    "wto.org", "wcoomd.org", "gov.uk", "customs.go.jp", "mof.go.jp", "meti.go.jp",
    "cbic.gov.in", "dgft.gov.in", "mofcom.gov.cn", "customs.gov.cn", "gacc.gov.cn",
    "moit.gov.vn", "customs.gov.vn", "gov.br", "sat.gob.mx", "gob.mx",
]

V27_BAD_URL_TERMS = [
    "google.co.kr/alerts/feeds", "google.com/alerts/feeds", "alerts/feeds/",
    "news.google.com/rss/articles", "news.google.com/articles", "news.google.com/",
    "https://news.google.com", "http://news.google.com",
]

V27_ALLOWED_URL_STATUSES = {
    "", "ORIGINAL_INPUT", "RESOLVED_DIRECT", "RESOLVED_REDIRECT", "RESOLVED_SELENIUM",
    "RESTORED_ORIGINAL_CANDIDATE", "RESTORED_CANONICAL_CANDIDATE", "RESTORED_GOOGLE_QUERY",
}

V27_EXCLUDED_PRIORITY_TERMS = ["EXCLUDED", "REJECT", "REJECTED", "DROP", "NOISE"]
V27_HARD_REJECT_TERMS = [
    "event_training_tender_noise", "financial_industry_noise_without_trade_policy",
    "samsung_general_business_noise", "bilateral_industry_news_without_trade_policy",
    "body_unavailable", "google_alert_url_unresolved", "search_no_good_result",
    "bad_url", "unresolved", "strict_bad_or_unresolved_url", "strict_digest_politics_market_noise",
    "strict_reference_not_reportable", "strict_no_concrete_customs_trade_signal",
]


def _v27_domain(url: object) -> str:
    try:
        from urllib.parse import urlparse
        raw = best_url_from_values([url]) or clean(url)
        host = urlparse(raw).netloc.lower().removeprefix("www.")
        return host
    except Exception:
        return ""


def _v27_is_bad_google_or_generic_url(url: object) -> bool:
    raw = clean(url).lower()
    if not raw or raw in {"nan", "none", "null", "https://news", "http://news", "https://new", "http://new"}:
        return True
    if not re.match(r"^https?://", raw):
        return True
    return any(term in raw for term in V27_BAD_URL_TERMS)


def _v27_url_quality_ok(row: pd.Series) -> bool:
    url = clean(row.get("URL"))
    status = clean(row.get("URLDecodeStatus")) or clean(row.get("URLRestoreStatus")) or clean(row.get("URL_Quality"))
    status_upper = status.upper()
    if _v27_is_bad_google_or_generic_url(url):
        return False
    if any(term in status_upper for term in ["GOOGLE_REMAINED", "GOOGLE_UNRESOLVED", "FAILED", "SEARCH_NO_GOOD_RESULT", "BAD_URL"]):
        return False
    # Do not reject blank legacy status when URL is clearly original/non-Google.
    if status and status not in V27_ALLOWED_URL_STATUSES and status_upper not in {s.upper() for s in V27_ALLOWED_URL_STATUSES}:
        # Keep if it is a non-Google URL and the status is informational, but discount unknown bad statuses.
        if any(x in status_upper for x in ["ERROR", "TIMEOUT", "EMPTY"]):
            return False
    return True


def _v27_is_official_reg_source(row: pd.Series) -> bool:
    content_type = clean(row.get("Content Type"))
    if content_type != "Regulation":
        return True
    if not GTI_MAIL_OFFICIAL_REG_ONLY_V27:
        return True
    text = " ".join(clean(row.get(c)) for c in ["URL", "BestLinkURL", "OriginalURLCandidate", "Source", "Agency", "Source File"]).lower()
    return any(d in text for d in V27_OFFICIAL_REG_DOMAINS)


def _v27_reference_datetime() -> pd.Timestamp:
    ref = os.getenv("GTI_MAIL_REFERENCE_DATETIME", "").strip()
    if ref:
        dt = pd.to_datetime(ref, errors="coerce")
        if not pd.isna(dt):
            return pd.Timestamp(dt).tz_localize(None) if getattr(dt, "tzinfo", None) else pd.Timestamp(dt)
    # Default to actual execution time. This makes the mail truly 24-hour based.
    return pd.Timestamp(datetime.now())


def _v27_date_candidates(row: pd.Series) -> str:
    return " ".join(clean(row.get(c)) for c in [
        "Date", "Publish Date", "Published", "CollectedAt", "Headline", "Summary", "Major Changes", "URL"
    ])


def _v27_parse_publish_dt(row: pd.Series):
    # Prefer _sort_date if it is a real timestamp and not Timestamp.min.
    try:
        dt = pd.to_datetime(row.get("_sort_date"), errors="coerce")
        if not pd.isna(dt) and dt.year > 1970:
            return pd.Timestamp(dt), False
    except Exception:
        pass

    for col in ["Date", "Publish Date", "Published", "CollectedAt"]:
        raw = clean(row.get(col))
        if raw and raw not in {"확인 필요", "미확인", "nan", "None"}:
            dt = pd.to_datetime(raw, errors="coerce")
            if not pd.isna(dt) and dt.year > 1970:
                # If raw is date-only, treat as date-only for calendar-day allowance.
                date_only = bool(re.fullmatch(r"\d{4}[-./]\d{1,2}[-./]\d{1,2}", raw[:10])) and len(raw.strip()) <= 10
                return pd.Timestamp(dt), date_only

    # Last resort: find YYYY-MM-DD-like string in text.
    text = _v27_date_candidates(row)
    m = re.search(r"(20\d{2})[-./](\d{1,2})[-./](\d{1,2})", text)
    if m:
        dt = pd.to_datetime(f"{m.group(1)}-{m.group(2)}-{m.group(3)}", errors="coerce")
        if not pd.isna(dt):
            return pd.Timestamp(dt), True
    return pd.NaT, False


def _v27_freshness(row: pd.Series) -> tuple[bool, str, str]:
    dt, date_only = _v27_parse_publish_dt(row)
    if pd.isna(dt):
        return (not GTI_MAIL_DROP_UNKNOWN_DATE_V27), "UNKNOWN_DATE", ""
    ref = _v27_reference_datetime()
    age_hours = (ref - dt).total_seconds() / 3600.0
    # Future-dated within one day can happen because of timezone/date-only parsing; allow but mark.
    if age_hours < -24:
        return False, "FUTURE_DATE_OVER_24H", f"{age_hours:.1f}"
    if date_only and GTI_MAIL_ALLOW_PREVIOUS_DATE_ONLY_V27:
        min_date = (ref.normalize() - pd.Timedelta(days=1)).date()
        if dt.date() >= min_date:
            return True, "OK_DATE_WINDOW_24H", f"{age_hours:.1f}"
    if age_hours <= GTI_MAIL_MAX_AGE_HOURS_V27:
        return True, "OK_24H", f"{age_hours:.1f}"
    return False, "STALE_OVER_24H", f"{age_hours:.1f}"


def _v27_priority_excluded(row: pd.Series) -> bool:
    p = clean(row.get("Priority Group")).upper()
    tier = clean(row.get("Tier")).upper()
    return any(t in p for t in V27_EXCLUDED_PRIORITY_TERMS) or any(t in tier for t in V27_EXCLUDED_PRIORITY_TERMS)


def _v27_reject_reason_hard(row: pd.Series) -> bool:
    rr = clean(row.get("RejectReason"))
    if not rr:
        return False
    rr_set = {x.strip().lower() for x in rr.split(";") if x.strip()}
    broad_prior_reasons = {
        "v12_hard_reference_or_noise",
        "v12_no_customs_trade_action_signal",
        "weighted_v18_not_topn_or_noise",
        "v20_major_title_or_weighted_below_topn_or_noise",
        "weak_samsung_relevance",
        "report_issue_duplicate_compressed",
        "expanded_policy_watch",
        "strict_existing_hard_reject",
    }
    if rr_set and rr_set <= broad_prior_reasons:
        return False
    low = rr.lower()
    return any(t in low for t in V27_HARD_REJECT_TERMS)


def _v27_add_audit_fields(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    for col in _V27_AUDIT_COLUMNS:
        if col not in rows.columns:
            rows[col] = ""
    rows["FinalURLDomain"] = rows["URL"].apply(_v27_domain)
    rows["OfficialSourceFlag"] = rows.apply(lambda r: "Y" if _v27_is_official_reg_source(r) else "N", axis=1)
    statuses = []
    ages = []
    for _, r in rows.iterrows():
        ok, status, age = _v27_freshness(r)
        statuses.append(status)
        ages.append(age)
    rows["FreshnessStatus"] = statuses
    rows["FreshnessAgeHours"] = ages
    return rows


def _v27_final_mail_filter(rows: pd.DataFrame, top3_mode: bool = False) -> pd.DataFrame:
    if rows.empty:
        return rows
    before = len(rows)
    rows = _v27_add_audit_fields(rows)
    keep_flags = []
    reasons_all = []
    for _, row in rows.iterrows():
        reasons = []
        fresh_ok, fresh_status, _age = _v27_freshness(row)
        if not fresh_ok:
            reasons.append(f"v27_{fresh_status.lower()}")
        if _v27_priority_excluded(row):
            reasons.append("v27_priority_excluded")
        if _v27_reject_reason_hard(row):
            reasons.append("v27_reject_reason_hard")
        if not _v27_url_quality_ok(row):
            reasons.append("v27_original_url_not_verified")
        if clean(row.get("Content Type")) == "Regulation" and not _v27_is_official_reg_source(row):
            reasons.append("v27_non_official_regulation_source")
        if top3_mode and not _v26_title_has_keyword(row):
            reasons.append("v27_top3_no_title_keyword")
        reasons_all.append("; ".join(reasons))
        keep_flags.append(not reasons)

    rows["MailGuardReason"] = reasons_all
    dropped = rows.loc[[not x for x in keep_flags]].copy()
    kept = rows.loc[keep_flags].copy().reset_index(drop=True)
    label = "Top3" if top3_mode else "final mail"
    if not dropped.empty:
        print(f"[INFO] v27 {label} guard removed:")
        for _, r in dropped.head(40).iterrows():
            print(
                f"  - {clean(r.get('Headline'))[:120]} / date={clean(r.get('Date'))} / "
                f"domain={clean(r.get('FinalURLDomain'))} / reason={clean(r.get('MailGuardReason'))}"
            )
    print(f"[INFO] v27 {label} guard: before={before}, after={len(kept)}, removed={len(dropped)}, max_age_hours={GTI_MAIL_MAX_AGE_HOURS_V27:g}")
    if "No" in kept.columns:
        kept["No"] = range(1, len(kept) + 1)
    return kept



# ======================================================================
# GTI STEP5 v28: date-only freshness fix
# ----------------------------------------------------------------------
# If STEP4 gives Date/Publish Date as YYYY-MM-DD, treat it as a date-only
# signal before using internal _sort_date. Otherwise previous-day official
# regulation rows are incorrectly dropped as STALE_OVER_24H at morning runs.
# ======================================================================

_v28_parse_publish_dt_base = _v27_parse_publish_dt

def _v27_parse_publish_dt(row: pd.Series):
    for col in ["Date", "Publish Date", "Published", "CollectedAt"]:
        raw = clean(row.get(col))
        if raw and raw not in {"?", "nan", "None", "Nat", "NaT"}:
            head = raw.strip()[:10]
            if re.fullmatch(r"\d{4}[-./]\d{1,2}[-./]\d{1,2}", head) and len(raw.strip()) <= 10:
                dt = pd.to_datetime(head, errors="coerce")
                if not pd.isna(dt) and dt.year > 1970:
                    return pd.Timestamp(dt), True
    return _v28_parse_publish_dt_base(row)

# ======================================================================
# End of GTI STEP5 v28
# ======================================================================

# Preserve audit columns from STEP4 inputs. The base normalizer intentionally kept the report schema small;
# v27 extends it for source/freshness traceability.
_v27_normalize_input_base = normalize_input

def normalize_input(df: pd.DataFrame, content_type: str, source_file: Path) -> pd.DataFrame:
    out = _v27_normalize_input_base(df, content_type, source_file)
    df2 = df.copy()
    df2.columns = [str(c).strip() for c in df2.columns]
    audit_map = {
        "GoogleURL": ["GoogleURL", "google_url"],
        "OriginalURLCandidate": ["OriginalURLCandidate", "original_url", "original_url_candidate"],
        "BestLinkURL": ["BestLinkURL", "best_link_url"],
        "URLRestoreStatus": ["URLRestoreStatus", "url_restore_status", "URL_Quality", "url_quality"],
        "URLDecodeStatus": ["URLDecodeStatus", "url_decode_status"],
        "RejectReason": ["RejectReason", "reject_reason"],
        "Tier": ["Tier", "tier"],
        "Publish Date": ["Publish Date", "Published", "published", "publish_date"],
    }
    for out_col, candidates in audit_map.items():
        src = pick_col(df2, candidates)
        if src and len(df2[src]) >= len(out):
            out[out_col] = df2[src].iloc[:len(out)].apply(clean).values
        elif out_col not in out.columns:
            out[out_col] = ""
    return out


_v27_prepare_rows_base = prepare_rows

def prepare_rows(rows: pd.DataFrame) -> pd.DataFrame:
    rows = _v27_prepare_rows_base(rows)
    rows = _v27_final_mail_filter(rows, top3_mode=False)
    if "_report_score" in rows.columns and "_sort_date" in rows.columns:
        rows = rows.sort_values(["_report_score", "_sort_date"], ascending=[False, False]).reset_index(drop=True)
    if "No" in rows.columns:
        rows["No"] = range(1, len(rows) + 1)
    return rows


_v27_choose_top3_base = choose_top3

def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    pool = _v27_final_mail_filter(rows, top3_mode=True)
    if pool.empty:
        return pool
    out = _v27_choose_top3_base(pool)
    out = _v27_final_mail_filter(out, top3_mode=True)
    if not out.empty:
        out["No"] = range(1, len(out) + 1)
    return out

# ======================================================================
# End of GTI STEP5 STRICT FINAL GUARD v27
# ======================================================================

def main() -> None:
    paths = output_paths()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    rows = prepare_rows(read_step4_results())
    top3 = choose_top3(rows)
    html_body = build_html(rows, top3)
    save_excel(rows, top3, paths)
    paths["mail_html"].write_text(html_body, encoding="utf-8")
    send_email(html_body, paths["mail_xlsx"])

    reg_n = int(rows["Content Type"].eq("Regulation").sum())
    news_n = int(rows["Content Type"].eq("News").sum())
    print(f"[DONE] HTML: {paths['mail_html']}")
    print(f"[DONE] XLSX: {paths['mail_xlsx']}")
    print(
        f"[ROWS] total={len(rows)}, regulation={reg_n}, news={news_n}, "
        f"direct={(rows['Samsung Impact'] == 'Direct').sum()}, "
        f"indirect={(rows['Samsung Impact'] == 'Indirect').sum()}, "
        f"watch={(rows['Samsung Impact'] == 'Watch').sum()}"
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", default=None)
    parser.add_argument("--regulation-input", default=None)
    parser.add_argument("--news-input", default=None)
    parser.add_argument("--output-dir", default=None)
    parser.add_argument("--no-email", action="store_true")
    args = parser.parse_args()
    if args.date:
        RUN_DATE = args.date
    if args.regulation_input:
        REGULATION_INPUT_FILE = Path(args.regulation_input)
    if args.news_input:
        NEWS_INPUT_FILE = Path(args.news_input)
    if args.output_dir:
        OUTPUT_DIR = Path(args.output_dir)
    if args.no_email:
        SEND_EMAIL = False
    main()
