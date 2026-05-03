# -*- coding: utf-8 -*-
r"""
5.GTI Mail Engine.py
GTI Radar STEP5 Mail Engine

Input
- C:\temp\news_raw.xlsx
- C:\temp\news_cumulative.xlsx (optional)
- C:\temp\mail.xlsx (recipient list, optional)

Output
- C:\temp\GTI_Radar_YYYY-MM-DD_Top25.xlsx
- C:\temp\GTI_Radar_YYYY-MM-DD_Top25_Email.html
- C:\temp\mail_cumulative.xlsx

Email Subject
- [GTI Radar] Global Trade Intelligence | yyyy-mm-dd

SMTP
- Naver SMTP 기본값 사용
- 보안을 위해 비밀번호는 환경변수 사용 권장
  set GTI_SMTP_USER=kch8872@naver.com
  set GTI_SMTP_PASS=네이버_SMTP_앱비밀번호
  set GTI_SEND_EMAIL=Y

GitHub Actions에서도 동일하게 Secrets 사용:
- GTI_SMTP_USER
- GTI_SMTP_PASS
- GTI_MAIL_TO (00.xlsx가 없을 때 대체 수신자, 콤마 구분)
"""

from __future__ import annotations

import os
import re
import ssl
import html
import smtplib
import traceback
from datetime import datetime
from email.message import EmailMessage
from email.utils import formataddr
from urllib.parse import urlparse

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# =========================
# 0. CONFIG
# =========================
BASE_DIR = r"C:\temp"
INPUT_RAW = os.path.join(BASE_DIR, "news_raw.xlsx")
INPUT_CUMULATIVE = os.path.join(BASE_DIR, "news_cumulative.xlsx")
RECIPIENT_FILE = os.path.join(BASE_DIR, "00.xlsx")
OUTPUT_DIR = BASE_DIR
MAIL_CUMULATIVE = os.path.join(BASE_DIR, "mail_cumulative.xlsx")

TODAY = datetime.now().strftime("%Y-%m-%d")
SUBJECT = f"[GTI Radar] Global Trade Intelligence | {TODAY}"

OUTPUT_XLSX = os.path.join(OUTPUT_DIR, f"GTI_Radar_{TODAY}_Top25.xlsx")
OUTPUT_HTML = os.path.join(OUTPUT_DIR, f"GTI_Radar_{TODAY}_Top25_Email.html")

SMTP_HOST = os.getenv("GTI_SMTP_HOST", "smtp.naver.com")
SMTP_PORT = int(os.getenv("GTI_SMTP_PORT", "465"))
SMTP_USER = os.getenv("GTI_SMTP_USER", "")  # 예: kch8872@naver.com
SMTP_PASS = os.getenv("GTI_SMTP_PASS", "")  # 네이버 SMTP 앱 비밀번호
MAIL_FROM_NAME = os.getenv("GTI_MAIL_FROM_NAME", "GTI Radar")
SEND_EMAIL = os.getenv("GTI_SEND_EMAIL", "N").strip().upper() == "Y"

# 00.xlsx가 없을 때 대체 수신자. 예: lifepal.kwak@samsung.com,kch8872@naver.com
FALLBACK_TO = os.getenv("GTI_MAIL_TO", "").strip()

FOCUS_COUNTRIES = ["KR", "CN", "VN", "IN", "US", "MX", "BR"]

RISK_ORDER = {"상": 1, "중": 2, "하": 3, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "": 4}
SECTION_ORDER = {
    "1.당사 영향": 1,
    "2.통상 정책": 2,
    "3.규제 변화": 3,
    "4.경쟁사 동향": 4,
}

# =========================
# 1. CLEAN / NORMALIZE
# =========================
def clean_text(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    s = str(v)
    s = html.unescape(s)
    s = re.sub(r"\*\*", "", s)
    s = re.sub(r"^[\s\-•·]+", "", s, flags=re.MULTILINE)
    s = re.sub(r"\n{3,}", "\n\n", s)
    s = re.sub(r"[ \t]+", " ", s)
    return s.strip()


def compact_text(v) -> str:
    return re.sub(r"\s+", " ", clean_text(v)).strip()


def safe_date(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    try:
        dt = pd.to_datetime(v, errors="coerce")
        if pd.isna(dt):
            return ""
        return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return compact_text(v)


def normalize_risk(v) -> str:
    s = compact_text(v)
    if s in ["상", "HIGH", "High", "high"]:
        return "상"
    if s in ["중", "MEDIUM", "Medium", "medium"]:
        return "중"
    if s in ["하", "LOW", "Low", "low"]:
        return "하"
    if "상" in s or "직접" in s or "HIGH" in s.upper():
        return "상"
    if "중" in s or "간접" in s or "MED" in s.upper():
        return "중"
    if "하" in s or "기타" in s or "LOW" in s.upper():
        return "하"
    return "중"


def infer_country(text: str, current: str = "") -> str:
    raw = f"{current} {text}"
    low = raw.lower()
    countries = []
    def add(code):
        if code not in countries:
            countries.append(code)
    if any(x in raw for x in ["한국", "관세청", "산업통상자원부", "기획재정부"]) or "korea" in low:
        add("KR")
    if any(x in raw for x in ["미국", "유에스", "USTR", "CBP", "상무부"] ) or "united states" in low or " u.s." in low or " us " in f" {low} ":
        add("US")
    if any(x in raw for x in ["중국", "중화인민공화국"]) or "china" in low or "mofcom" in low or "gacc" in low:
        add("CN")
    if "베트남" in raw or "vietnam" in low:
        add("VN")
    if "인도" in raw or "india" in low:
        add("IN")
    if "멕시코" in raw or "mexico" in low:
        add("MX")
    if "브라질" in raw or "brazil" in low:
        add("BR")
    if "eu" in low or "유럽연합" in raw or "european commission" in low:
        add("EU")
    return " / ".join(countries) if countries else compact_text(current) or "MULTI"


def infer_agency(text: str, current: str = "") -> str:
    if compact_text(current) and compact_text(current) not in ["nan", "None", "Unknown", "Google News"]:
        cur = compact_text(current)
    else:
        cur = ""
    low = text.lower()
    if "ustr" in low or "미 무역대표부" in text:
        return "USTR"
    if "cbp" in low or "u.s. customs" in low or "미 세관" in text:
        return "U.S. Customs and Border Protection (CBP)"
    if "commerce" in low or "상무부" in text:
        return "U.S. Department of Commerce"
    if "federal register" in low:
        return "U.S. Federal Register"
    if "관세청" in text:
        return "관세청"
    if "산업통상자원부" in text or "산업부" in text:
        return "산업통상자원부"
    if "기획재정부" in text or "재정경제부" in text:
        return "기획재정부"
    if "무역위원회" in text:
        return "무역위원회"
    if "european commission" in low or "eu commission" in low or "유럽연합" in text:
        return "European Commission Trade"
    if "mofcom" in low or "중국 상무부" in text:
        return "MOFCOM"
    if "gacc" in low or "중국 해관" in text or "해관총서" in text:
        return "GACC"
    if "vietnam customs" in low or "베트남 세관" in text:
        return "Vietnam Customs"
    if "india" in low or "인도" in text:
        return "Ministry of Commerce & Industry, India"
    if "mexico" in low or "멕시코" in text:
        return "Mexico Ministry of Economy / SAT"
    if "brazil" in low or "브라질" in text:
        return "Brazil MDIC / Receita Federal"
    if "wco" in low or "세계관세기구" in text:
        return "WCO"
    if "wto" in low or "세계무역기구" in text:
        return "WTO"
    return cur or "관련 정부/국제기관"


def infer_section(row: dict) -> str:
    text = f"{row.get('Headline','')} {row.get('Summary','')} {row.get('AI Analysis','')} {row.get('Action Plan','')}".lower()
    kor = f"{row.get('Headline','')} {row.get('Summary','')} {row.get('AI Analysis','')} {row.get('Action Plan','')}"
    if any(k in text for k in ["samsung", "semiconductor", "smartphone", "display", "appliance", "hbm", "chip"]) or any(k in kor for k in ["삼성", "반도체", "스마트폰", "디스플레이", "가전", "생산기지", "생산 거점"]):
        return "1.당사 영향"
    if any(k in text for k in ["tariff", "fta", "trade", "ustr", "cbp", "cbam", "section 232", "section 301"]) or any(k in kor for k in ["관세", "통상", "FTA", "무역협정", "수입규제", "환급"]):
        return "2.통상 정책"
    if any(k in text for k in ["customs", "export control", "origin", "hs", "valuation", "anti-dumping", "countervailing"]) or any(k in kor for k in ["통관", "세관", "원산지", "품목분류", "HS", "과세가격", "수출통제", "전략물자", "반덤핑", "상계관세", "규제"]):
        return "3.규제 변화"
    return "4.경쟁사 동향"


def domain_of(url: str) -> str:
    try:
        return urlparse(str(url)).netloc.lower()
    except Exception:
        return ""


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    # 다양한 원천 컬럼명을 STEP5 표준 컬럼으로 통일
    colmap = {}
    for c in df.columns:
        lc = str(c).strip().lower()
        if lc in ["date", "publish date", "publish_date", "published", "news date", "뉴스 원본 게시일시", "원문등록일"]:
            colmap[c] = "date"
        elif lc in ["headline", "title", "news title", "주요 뉴스", "뉴스 제목"]:
            colmap[c] = "Headline"
        elif lc in ["summary", "news", "뉴스 본문", "뉴스 본문요약", "주요내용"]:
            colmap[c] = "Summary"
        elif lc in ["ai analysis", "analysis", "impact", "ai_analysis", "전문관세사 분석"]:
            colmap[c] = "AI Analysis"
        elif lc in ["action plan", "action", "action_plan", "대응방안"]:
            colmap[c] = "Action Plan"
        elif lc in ["country", "국가", "대상 국가"]:
            colmap[c] = "Country"
        elif lc in ["agency", "관련 기관", "정책기관", "관련기관"]:
            colmap[c] = "agency"
        elif lc in ["risk", "importance", "중요도"]:
            colmap[c] = "importance"
        elif lc in ["url", "link", "source url", "출처url", "링크"]:
            colmap[c] = "URL"
        elif lc in ["source", "출처"]:
            colmap[c] = "source"
        elif lc in ["score", "risk_score", "importance_score"]:
            colmap[c] = "score"
    df = df.rename(columns=colmap).copy()
    for c in ["date", "Headline", "Summary", "AI Analysis", "Action Plan", "Country", "agency", "importance", "URL", "source", "score"]:
        if c not in df.columns:
            df[c] = ""
    return df


def headline_fallback(row) -> str:
    h = compact_text(row.get("Headline", ""))
    if h and h not in ["0", "0.0", "nan", "None"]:
        return h
    # 수집 단계에서 Headline이 0으로 들어온 경우 최소한 요약 첫 문장 사용
    s = compact_text(row.get("Summary", ""))
    if s:
        first = re.split(r"(?<=[.!?。])\s+|\n", s)[0]
        return first[:80].strip() or "뉴스 제목 확인 필요"
    url = compact_text(row.get("URL", ""))
    return domain_of(url) or "뉴스 제목 확인 필요"


def make_summary(summary: str, headline: str = "") -> str:
    s = clean_text(summary)
    h = compact_text(headline)
    # 이미 요약이 충분하면 문장 2~3개만 정리
    if not s:
        return "본문 정보가 제한적입니다. 원문 확인 후 관세·통상 영향 여부를 재검토해야 합니다."
    s = re.sub(r"^Summary\s*[:：]", "", s, flags=re.I).strip()
    # 제목만 반복된 경우
    if h and compact_text(s).replace(" ", "") == h.replace(" ", ""):
        return "수집 요약문이 제목 수준으로 제한되어 있습니다. 원문 확인 후 관세·통상 영향 여부를 재검토해야 합니다."
    # bullet/markdown 제거 후 2~3문장으로 제한
    sentences = re.split(r"(?<=[.!?。])\s+|\n+", s)
    sentences = [compact_text(x) for x in sentences if compact_text(x)]
    out = "\n".join(sentences[:3])
    if len(out) > 280:
        out = out[:280].rstrip() + "..."
    return out


def make_analysis(row: dict) -> str:
    v = clean_text(row.get("AI Analysis", ""))
    if v and len(compact_text(v)) > 30:
        return v
    text = f"{row.get('Headline','')} {row.get('Summary','')}"
    country = row.get("Country", "")
    if row.get("Risk") == "하":
        return "영향없음(직접). 삼성전자 주요 생산거점·제품·관세 이슈와의 직접 관련성은 낮아 모니터링 수준으로 관리 가능합니다."
    return f"삼성전자 생산거점({country}) 및 반도체·스마트폰·디스플레이·가전 공급망 관점에서 HS, 원산지, 과세가격, 수입규제 변동 가능성을 점검해야 합니다."


def make_action(row: dict) -> str:
    v = clean_text(row.get("Action Plan", ""))
    if v and len(compact_text(v)) > 20:
        return v
    if row.get("Risk") == "하":
        return "정기 모니터링 대상으로 유지하고, 동일 이슈가 관세·통상 조치로 확대될 경우 재분류합니다."
    return "① 대상 국가·HS·제품군 매핑 ② 원산지·거래가격 증빙 점검 ③ 관련 법령/공고 원문 확인 후 법인별 대응계획 수립"


# =========================
# 2. SCORE / SELECT
# =========================
def score_row(row: dict) -> int:
    text = f"{row.get('Headline','')} {row.get('Summary','')} {row.get('AI Analysis','')} {row.get('Action Plan','')} {row.get('Country','')} {row.get('agency','')}".lower()
    score = 0
    # 생산거점
    for k in ["vietnam", "베트남", "india", "인도", "china", "중국", "korea", "한국", "united states", "미국", "mexico", "멕시코", "brazil", "브라질"]:
        if k in text:
            score += 10
    # 제품
    for k in ["semiconductor", "반도체", "hbm", "chip", "smartphone", "스마트폰", "mobile", "display", "디스플레이", "oled", "appliance", "가전", "tv", "network", "네트워크"]:
        if k in text:
            score += 8
    # 관세 영향
    for k in ["tariff", "관세", "customs", "통관", "세관", "fta", "origin", "원산지", "hs", "품목분류", "valuation", "과세가격", "export control", "수출통제", "전략물자", "anti-dumping", "반덤핑", "countervailing", "상계관세", "cbam", "section 232", "section 301", "ieepa", "refund", "환급"]:
        if k in text:
            score += 7
    # 기관/정책성
    for k in ["ustr", "cbp", "wco", "wto", "mofcom", "gacc", "관세청", "산업통상자원부", "기획재정부", "european commission"]:
        if k in text:
            score += 5
    # 기존 score 반영
    try:
        score += int(float(row.get("score", 0)))
    except Exception:
        pass
    # 기존 중요도 반영
    risk = normalize_risk(row.get("importance", ""))
    if risk == "상":
        score += 25
    elif risk == "중":
        score += 12
    else:
        score += 3
    # 노이즈 감점
    noise = ["american eagle", "aeo", "osteoporosis", "concert", "festival", "youtube", "stock", "주가", "연예", "부동산"]
    if any(n in text for n in noise):
        score -= 30
    return score


def prepare_top25(raw: pd.DataFrame) -> pd.DataFrame:
    df = normalize_columns(raw)
    rows = []
    for _, r in df.iterrows():
        row = {c: r.get(c, "") for c in df.columns}
        headline = headline_fallback(row)
        url = compact_text(row.get("URL", "")) or compact_text(row.get("source", ""))
        summary = make_summary(row.get("Summary", ""), headline)
        full_text = f"{headline} {summary} {row.get('AI Analysis','')} {row.get('Action Plan','')} {row.get('Country','')} {row.get('agency','')}"
        country = infer_country(full_text, row.get("Country", ""))
        agency = infer_agency(full_text, row.get("agency", ""))
        risk = normalize_risk(row.get("importance", ""))
        section = infer_section({**row, "Headline": headline, "Summary": summary})
        out = {
            "date": safe_date(row.get("date", "")),
            "Headline": headline,
            "Summary": summary,
            "AI Analysis": "",
            "Action Plan": "",
            "Country": country,
            "agency": agency,
            "Risk": risk,
            "URL": url,
            "Section": section,
            "source": compact_text(row.get("source", "")),
            "score": 0,
        }
        out["AI Analysis"] = make_analysis({**row, **out})
        out["Action Plan"] = make_action({**row, **out})
        out["score"] = score_row({**row, **out})
        rows.append(out)

    outdf = pd.DataFrame(rows)
    # URL 또는 Headline 기준 중복 제거
    outdf["_url_key"] = outdf["URL"].fillna("").astype(str).str.lower().str.strip()
    outdf["_title_key"] = outdf["Headline"].fillna("").astype(str).str.lower().str.strip()
    outdf = outdf.sort_values(["score"], ascending=False)
    outdf = outdf.drop_duplicates(subset=["_url_key"], keep="first")
    outdf = outdf.drop_duplicates(subset=["_title_key"], keep="first")
    top25 = outdf.head(25).copy()
    top25["_section_order"] = top25["Section"].map(SECTION_ORDER).fillna(9)
    top25["_risk_order"] = top25["Risk"].map(RISK_ORDER).fillna(4)
    top25 = top25.sort_values(["_section_order", "_risk_order", "Country", "score"], ascending=[True, True, True, False]).reset_index(drop=True)
    top25["No"] = range(1, len(top25) + 1)
    # Top3는 최종 표시 순서와 별도로 score 기준 Top3
    score_top = top25.sort_values(["score"], ascending=False).head(3).copy()
    top25["Top3"] = top25["Headline"].isin(score_top["Headline"]).map(lambda x: "Y" if x else "")
    return top25[["No", "Section", "date", "Headline", "Summary", "AI Analysis", "Action Plan", "Country", "agency", "Risk", "URL", "score", "Top3"]]


# =========================
# 3. EXCEL GENERATION
# =========================
def save_excel(top25: pd.DataFrame, path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "GTI Radar Top25"

    headers = ["No", "Section", "date", "Headline", "Summary", "AI Analysis", "Action Plan", "Country", "agency", "Risk"]
    ws.append(headers)

    for _, r in top25.iterrows():
        row_idx = ws.max_row + 1
        ws.append([
            r["No"], r["Section"], r["date"], r["Headline"], r["Summary"], r["AI Analysis"], r["Action Plan"], r["Country"], r["agency"], r["Risk"]
        ])
        url = compact_text(r.get("URL", ""))
        if url:
            cell = ws.cell(row=row_idx, column=4)
            cell.hyperlink = url
            cell.font = Font(color="0563C1", underline="single")

    # Style
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    widths = {
        "A": 6, "B": 16, "C": 18, "D": 42, "E": 48, "F": 52, "G": 46, "H": 15, "I": 28, "J": 10
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        ws.row_dimensions[cell.row].height = 90
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Risk color
    for row in range(2, ws.max_row + 1):
        risk = ws.cell(row=row, column=10).value
        if risk == "상":
            ws.cell(row=row, column=10).fill = PatternFill("solid", fgColor="F8CBAD")
        elif risk == "중":
            ws.cell(row=row, column=10).fill = PatternFill("solid", fgColor="FFF2CC")
        else:
            ws.cell(row=row, column=10).fill = PatternFill("solid", fgColor="E2F0D9")

    ws2 = wb.create_sheet("Mail Summary")
    ws2["A1"] = "Subject"
    ws2["B1"] = SUBJECT
    ws2["A2"] = "HTML File"
    ws2["B2"] = OUTPUT_HTML
    ws2["A3"] = "Generated At"
    ws2["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for c in ["A", "B"]:
        ws2.column_dimensions[c].width = 45
    for row in ws2.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font

    wb.save(path)


# =========================
# 4. HTML GENERATION
# =========================
def top3_short(summary: str) -> str:
    s = compact_text(summary)
    return s[:50]


def build_total_review(top3: pd.DataFrame) -> str:
    countries = ", ".join([x for x in top3["Country"].dropna().astype(str).tolist() if x])
    return f"미국·EU·아시아 주요 통상 리스크가 삼성전자 생산거점과 제품별 관세·원산지 대응 필요성을 높이고 있습니다."


def html_link(headline: str, url: str) -> str:
    h = html.escape(compact_text(headline))
    u = html.escape(compact_text(url))
    if u:
        return f"<a href='{u}' style='color:#0563c1;text-decoration:underline;'>{h}</a>"
    return h


def build_html(top25: pd.DataFrame) -> str:
    # Top3는 score 기준
    top3 = top25.sort_values(["score"], ascending=False).head(3).copy()
    rest = top25[~top25["Headline"].isin(top3["Headline"])].copy()

    total_review = build_total_review(top3)
    bullet_lines = "\n".join([f"<li>{html.escape(top3_short(r['Summary']))}</li>" for _, r in top3.iterrows()])

    top3_blocks = []
    for i, (_, r) in enumerate(top3.iterrows(), start=1):
        meta = f"Publish Date: {html.escape(compact_text(r['date']))} | Country: {html.escape(compact_text(r['Country']))} | Agency: {html.escape(compact_text(r['agency']))} | Risk: {html.escape(compact_text(r['Risk']))}"
        top3_blocks.append(f"""
        <div style="margin:18px 0 22px 0;padding:14px;border-left:5px solid #C00000;background:#fff7f7;">
          <div style="font-size:15px;font-weight:bold;margin-bottom:6px;">{i}️⃣ {html_link(r['Headline'], r['URL'])}</div>
          <div style="font-size:12px;color:#555;margin-bottom:10px;">{meta}</div>
          <div style="margin-top:8px;"><b>Summary</b><br>{html.escape(clean_text(r['Summary'])).replace(chr(10), '<br>')}</div>
          <div style="margin-top:8px;"><b>AI Analysis</b><br>{html.escape(clean_text(r['AI Analysis'])).replace(chr(10), '<br>')}</div>
          <div style="margin-top:8px;"><b>Action</b><br>{html.escape(clean_text(r['Action Plan'])).replace(chr(10), '<br>')}</div>
        </div>
        """)

    rows = []
    for idx, (_, r) in enumerate(rest.iterrows(), start=4):
        rows.append(f"""
        <tr>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;">{idx}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;">{html_link(r['Headline'], r['URL'])}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;">{html.escape(clean_text(r['Summary'])).replace(chr(10), '<br>')}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;">{html.escape(clean_text(r['AI Analysis'])).replace(chr(10), '<br>')}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;">{html.escape(clean_text(r['Action Plan'])).replace(chr(10), '<br>')}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;">{html.escape(compact_text(r['Country']))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;">{html.escape(compact_text(r['agency']))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;">{html.escape(compact_text(r['Risk']))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;">{html.escape(compact_text(r['date']))}</td>
        </tr>
        """)

    html_body = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8">
<title>{html.escape(SUBJECT)}</title>
</head>
<body style="font-family:Arial,'Malgun Gothic',sans-serif;font-size:13px;color:#222;line-height:1.5;">
  <div style="max-width:1200px;margin:0 auto;">
    <h2 style="margin-bottom:4px;">[GTI Radar] Global Trade Intelligence</h2>
    <div style="font-size:14px;margin-bottom:4px;"><b>Date:</b> {TODAY}</div>
    <div style="font-size:12px;color:#555;margin-bottom:16px;">Coverage: Last 24 Hours | Focus: Samsung Electronics (KR / CN / VN / IN / US / MX / BR)</div>

    <h3 style="margin-top:18px;margin-bottom:6px;">총평</h3>
    <p style="margin-top:0;">{html.escape(total_review)}</p>
    <ul style="margin-top:6px;margin-bottom:18px;">
      {bullet_lines}
    </ul>

    <h3 style="color:#C00000;margin-top:22px;">🔴 TOP POLICY EVENTS (Top 3)</h3>
    {''.join(top3_blocks)}

    <h3 style="color:#1F4E78;margin-top:24px;">🟦 EVENT LIST (22)</h3>
    <table style="border-collapse:collapse;width:100%;font-size:12px;">
      <tr style="background:#1F4E78;color:white;">
        <th style="padding:7px;border:1px solid #d9d9d9;">No</th>
        <th style="padding:7px;border:1px solid #d9d9d9;">Headline</th>
        <th style="padding:7px;border:1px solid #d9d9d9;">Summary</th>
        <th style="padding:7px;border:1px solid #d9d9d9;">Impact</th>
        <th style="padding:7px;border:1px solid #d9d9d9;">Action</th>
        <th style="padding:7px;border:1px solid #d9d9d9;">Country</th>
        <th style="padding:7px;border:1px solid #d9d9d9;">Agency</th>
        <th style="padding:7px;border:1px solid #d9d9d9;">Risk</th>
        <th style="padding:7px;border:1px solid #d9d9d9;">Publish Date</th>
      </tr>
      {''.join(rows)}
    </table>

    <p style="margin-top:18px;color:#666;font-size:12px;">※ 첨부 Excel 파일에 전체 Top25 분석표가 포함되어 있습니다.</p>
  </div>
</body>
</html>"""
    return html_body


# =========================
# 5. RECIPIENT / SEND
# =========================
def load_recipients() -> list[str]:
    recipients = []
    if os.path.exists(RECIPIENT_FILE):
        try:
            rdf = pd.read_excel(RECIPIENT_FILE)
            # 모든 셀에서 이메일 형식 추출
            text = "\n".join(rdf.astype(str).fillna("").values.ravel().tolist())
            recipients.extend(re.findall(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}", text))
        except Exception as e:
            print(f"[WARN] recipient file read failed: {e}")
    if not recipients and FALLBACK_TO:
        recipients.extend([x.strip() for x in FALLBACK_TO.split(",") if x.strip()])
    # 중복 제거
    return list(dict.fromkeys(recipients))


def send_email(html_body: str, attachments: list[str]) -> None:
    recipients = load_recipients()
    if not recipients:
        print("[MAIL SKIP] 수신자 없음: C:\\temp\\00.xlsx 또는 GTI_MAIL_TO 환경변수를 확인하세요.")
        return
    if not SMTP_USER or not SMTP_PASS:
        print("[MAIL SKIP] SMTP 계정/비밀번호 없음: GTI_SMTP_USER, GTI_SMTP_PASS 환경변수를 설정하세요.")
        return

    msg = EmailMessage()
    msg["Subject"] = SUBJECT
    msg["From"] = formataddr((MAIL_FROM_NAME, SMTP_USER))
    msg["To"] = ", ".join(recipients)
    msg.set_content("GTI Radar 메일입니다. HTML 메일을 지원하는 클라이언트에서 확인해 주세요.")
    msg.add_alternative(html_body, subtype="html")

    for fp in attachments:
        if not os.path.exists(fp):
            continue
        with open(fp, "rb") as f:
            data = f.read()
        filename = os.path.basename(fp)
        maintype = "application"
        subtype = "vnd.openxmlformats-officedocument.spreadsheetml.sheet" if filename.lower().endswith(".xlsx") else "octet-stream"
        msg.add_attachment(data, maintype=maintype, subtype=subtype, filename=filename)

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, context=context) as server:
        server.login(SMTP_USER, SMTP_PASS)
        server.send_message(msg)
    print(f"[MAIL SENT] {len(recipients)} recipients")


# =========================
# 6. CUMULATIVE
# =========================
def update_mail_cumulative(top25: pd.DataFrame) -> None:
    data = top25.copy()
    data.insert(0, "mail_date", TODAY)
    data.insert(1, "subject", SUBJECT)
    if os.path.exists(MAIL_CUMULATIVE):
        try:
            old = pd.read_excel(MAIL_CUMULATIVE)
            data = pd.concat([old, data], ignore_index=True)
            data = data.drop_duplicates(subset=["mail_date", "Headline", "URL"], keep="last")
        except Exception:
            pass
    data.to_excel(MAIL_CUMULATIVE, index=False)


# =========================
# 7. MAIN
# =========================
def main() -> None:
    print("[START] GTI Mail Engine")
    if not os.path.exists(INPUT_RAW):
        raise FileNotFoundError(f"입력 파일 없음: {INPUT_RAW}")

    raw = pd.read_excel(INPUT_RAW)
    print(f"[LOAD] {INPUT_RAW} rows={len(raw)}")

    top25 = prepare_top25(raw)
    print(f"[SELECT] Top rows={len(top25)}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    save_excel(top25, OUTPUT_XLSX)
    print(f"[SAVE] Excel: {OUTPUT_XLSX}")

    html_body = build_html(top25)
    with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
        f.write(html_body)
    print(f"[SAVE] HTML: {OUTPUT_HTML}")

    update_mail_cumulative(top25)
    print(f"[SAVE] Cumulative: {MAIL_CUMULATIVE}")

    if SEND_EMAIL:
        send_email(html_body, [OUTPUT_XLSX])
    else:
        print("[MAIL SKIP] GTI_SEND_EMAIL=Y 설정 시 실제 발송")

    print("[DONE] GTI Mail Engine")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("[ERROR]", e)
        traceback.print_exc()
        raise
