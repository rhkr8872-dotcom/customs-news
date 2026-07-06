# -*- coding: utf-8 -*-
# GTI FINAL CORE v9 - News merge, title-keyword priority + URL quality guard
"""
STEP3-2 : news_merge.py

Input:
- C:\\Temp\\1-2.site_news_raw.xlsx   # STEP1 non-LAW1 official/news rows
- C:\\Temp\\2-1.naver_news_raw.xlsx
- C:\\Temp\\2-2.google_news_raw.xlsx
- C:\\Temp\\2-3.rss_news_raw.xlsx

Output:
- C:\\Temp\\3-2.news_summary.xlsx
- C:\\Temp\\3-2.news_cumulative.xlsx

Role:
- 언론/포털/RSS/사이트뉴스 후보 정리
- 법규 원문 전용 로직 제거
- 48시간 기준은 collected_at 우선 적용
- keyword.xlsx 기준 무역/통상/관세 관련 뉴스 선별
- 삼성 영향도와 중복 제거 scoring 강화
- 공공요금 tariff / 여행·스포츠·범죄 등 오탐 필터 강화
- Google Alert/RSS Agency를 실제 기사 URL 도메인 기준으로 보정
- Direct 영향도 과대 산정 완화(생산국가 단독 Direct 방지)
- cumulative는 URL 기준으로만 비교하며 기존 행을 절대 줄이지 않고 신규 행만 추가
"""

from __future__ import annotations

import argparse
import os
import re
import shutil
from difflib import SequenceMatcher
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import parse_qs, parse_qsl, unquote, urlencode, urlparse

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(os.getenv("STEP3_BASE_DIR", r"C:\Temp"))

INPUT_FILES = [
    # STEP1에서 LAW1(공인 법규)로 확정되지 않은 공식기관 공지/보도/정책자료도 뉴스 후보로 병합
    BASE_DIR / "1-2.site_news_raw.xlsx",
    BASE_DIR / "2-1.naver_news_raw.xlsx",
    BASE_DIR / "2-2.google_news_raw.xlsx",
    BASE_DIR / "2-3.rss_news_raw.xlsx",
]

OUTPUT_FILE = BASE_DIR / "3-2.news_summary.xlsx"
CUMULATIVE_FILE = BASE_DIR / "3-2.news_cumulative.xlsx"

KEYWORD_FILES = [
    BASE_DIR / "keyword.xlsx",
    BASE_DIR / "KEYWORD.xlsx",
    BASE_DIR / "keyword_master_trade_policy_150.xlsx",
]

RECENT_HOURS = int(os.getenv("GTI_STEP3_RECENT_HOURS", "72"))
MAX_OUTPUT = 300
MIN_SCORE = 20
MAX_PER_ISSUE_CLUSTER = 1
MAX_TRADE_GENERAL_OUTPUT = 60

# STEP3 최종 Summary Tier 배분 목표
# - STEP3는 임원보고 최종본이 아니라 STEP4 AI 분석 후보군입니다.
# - 모든 행이 CORE가 되는 것을 방지하고, STEP4가 Top30을 안정적으로 선별하도록
#   CORE/USABLE/REFERENCE를 강제 배분합니다.
TIER_CORE_LIMIT = int(os.getenv("GTI_STEP3_CORE_LIMIT", "50"))
TIER_USABLE_LIMIT = int(os.getenv("GTI_STEP3_USABLE_LIMIT", "100"))
TIER_REFERENCE_LIMIT = int(os.getenv("GTI_STEP3_REFERENCE_LIMIT", "150"))

URL_ID_QUERY_KEYS = {
    "article",
    "articleid",
    "articleno",
    "articleseq",
    "bbsid",
    "boardno",
    "boardseq",
    "idx",
    "idxno",
    "mi",
    "mode",
    "ncd",
    "news_id",
    "no",
    "noticeno",
    "nttsn",
    "nttsnurl",
    "pageid",
    "seq",
    "serial",
    "sn",
    "uid",
}

URL_DROP_QUERY_PREFIXES = (
    "utm_",
)

URL_DROP_QUERY_KEYS = {
    "fbclid",
    "gclid",
    "igshid",
    "mc_cid",
    "mc_eid",
    "ocid",
    "ref",
    "spm",
    "utm",
}

FINAL_COLS = [
    "Date",
    "CollectedAt",
    "Headline",
    "URL",
    "GoogleURL",
    "OriginalURLCandidate",
    "BestLinkURL",
    "Country",
    "Agency",
    "Risk",
    "TopicScore",
    "SamsungImpactScore",
    "RiskScore",
    "FinalScore",
    "Score",
    "Priority",
    "Tier",
    "NewsType",
    "IssueKey",
    "SamsungSignal",
    "SamsungImpact",
    "RegulationRelated",
    "RegulationTransferType",
    "KeywordMatches",
    "TitleKeywordFlag",
    "TitleKeywordMatches",
    "OriginalURLVerified",
    "URLQuality",
    "IssueClusterKey",
    "ClusterSize",
    "DuplicateCount",
    "ClusterSources",
    "ClusterHeadlines",
    "RepresentativeReason",
    "SelectReason",
    "RejectReason",
    "Source",
    "SourceFile",
    "Publisher",
    "Importance",
    "Category",
    "URLRestoreStatus",
    "URLDecodeStatus",
    "Step4Hint",
    "SourceScoreReason",
]

SOURCE_PRIORITY_DOMAINS = [
    "reuters.com",
    "bloomberg.com",
    "apnews.com",
    "ft.com",
    "wsj.com",
    "cnbc.com",
    "gov",
    "europa.eu",
    "wto.org",
    "wcoomd.org",
]

LOW_PRIORITY_REPUBLISHERS = [
    "yahoo.com",
    "msn.com",
    "aol.com",
    "nate.com",
    "news.google",
]

SAMSUNG_DIRECT_COUNTRIES = {
    "Korea",
    "United States",
    "China",
    "Vietnam",
    "India",
    "Mexico",
    "Brazil",
    "Poland",
    "Hungary",
    "Slovakia",
    "Indonesia",
    "Thailand",
}

TOPIC_SCORE_MAP = {
    "EXPORT_CONTROL": 100,
    "AD_CVD": 96,
    "TARIFF": 94,
    "CBAM_CARBON": 90,
    "ORIGIN_FTA": 86,
    "HS_CLASSIFICATION": 84,
    "CUSTOMS": 72,
    "TRADE_GENERAL": 45,
}

RISK_SCORE_MAP = {"상": 100, "중": 70, "하": 40}

ISSUE_CLUSTER_SIMILARITY = float(os.getenv("GTI_ISSUE_CLUSTER_SIMILARITY", "0.76"))

FALLBACK_TRADE_KEYWORDS = [
    "관세", "통관", "세관", "수입", "수출", "수출입", "무역",
    "무역규제", "수입규제", "수출통제", "전략물자", "제재",
    "원산지", "품목분류", "hs code", "hs코드", "관세율",
    "반덤핑", "상계관세", "세이프가드", "외국환거래",
    "customs", "tariff", "tariffs", "duty", "duties",
    "import", "export", "origin", "rules of origin", "trade",
    "fta", "cepa", "usmca", "anti-dumping", "antidumping",
    "countervailing", "safeguard", "export control", "sanction",
    "restriction", "classification", "hs", "hs code", "cbam",
    "carbon border", "trade agreement",
]

# 제목 keyword는 GTI 관세/통상 핵심어만 사용한다.
# 삼성/반도체/시총/주가 같은 일반 기업·시장 키워드는 제목 keyword로 인정하지 않는다.
TITLE_KEYWORD_TERMS = sorted(set(FALLBACK_TRADE_KEYWORDS + [
    "section 301", "section 232", "301조", "232조", "ieepa",
    "reciprocal tariff", "additional tariff", "customs duty", "import duty",
    "anti dumping", "anti-dumping", "antidumping", "countervailing duty", "ad/cvd", "cvd",
    "trade remedy", "export controls", "entity list", "uflpa", "forced labor",
    "free trade agreement", "rules of origin", "origin rule", "preference",
    "classification", "hs classification", "quota", "safeguard",
    "관세법", "관세청", "상호관세", "추가관세", "덤핑방지관세", "무역구제",
    "자유무역협정", "협정세율", "특혜관세", "원산지증명", "원산지 기준",
    "수출관리", "전략물자", "강제노동", "품목번호", "품목분류", "탄소국경",
]), key=lambda x: x.lower())

POLICY_RULES = [
    ("AD_CVD", [
        "anti-dumping", "anti dumping", "antidumping", "countervailing",
        "countervailing duty", "countervailing duties", "dumping", "ad/cvd",
        "cvd", "반덤핑", "덤핑방지관세", "상계관세", "무역구제",
    ]),
    ("TARIFF", ["tariff", "tariffs", "duty", "duties", "관세", "관세율", "additional tariff", "tariff cap", "tariff ceiling", "관세상한"]),
    ("ORIGIN_FTA", ["fta", "cepa", "usmca", "rules of origin", "origin", "원산지", "자유무역협정"]),
    ("EXPORT_CONTROL", ["export control", "export controls", "sanction", "restriction", "수출통제", "제재"]),
    ("CBAM_CARBON", ["cbam", "carbon border", "carbon tariff", "탄소국경"]),
    ("CUSTOMS", ["customs", "clearance", "declaration", "통관", "세관"]),
    ("HS_CLASSIFICATION", ["hs code", "classification", "품목분류", "hs코드"]),
]

AD_CVD_FORCE_TERMS = [
    "anti-dumping", "anti dumping", "antidumping", "countervailing",
    "countervailing duty", "countervailing duties", "ad/cvd", "cvd",
    "dumping duties", "반덤핑", "덤핑방지관세", "상계관세", "무역구제",
]

STEP4_REVIEW_TERMS = [
    "forced labor", "uflpa", "section 301", "301조", "section 232",
    "cbam", "carbon border", "fta", "cepa", "rules of origin",
    "tariff cap", "tariff ceiling", "관세상한", "강제노동", "원산지",
]

STANDALONE_NOISE_KEYWORDS = ["bis", "aeo", "수출", "관세", "customs", "export", "tariff"]

BIS_VALID_CONTEXT = [
    "export control", "entity list", "denied persons", "bureau of industry and security",
    "department of commerce", "commerce department", "수출통제", "산업안보국", "산업보안국",
    "미 상무부", "미국 상무부", "전략물자", "반도체", "semiconductor", "chip",
]

AEO_VALID_CONTEXT = [
    "authorized economic operator", "mutual recognition", "mra", "customs",
    "관세청", "통관", "수출입안전관리", "종합인증우수업체",
]

GENERIC_EXPORT_TARIFF_NOISE = [
    "수출 호조", "수출 증가", "수출 감소", "수출 실적", "수출액", "수출입 동향",
    "import price", "export growth", "export data", "수입물가", "관세청 통계",
    "customs seizure", "airport seizure", "drug seizure",
]

SAMSUNG_RULES = [
    ("SEMICONDUCTOR", ["semiconductor", "chip", "chips", "memory", "hbm", "반도체", "칩", "메모리"]),
    ("MOBILE", ["smartphone", "mobile phone", "handset", "galaxy", "스마트폰", "휴대폰", "갤럭시"]),
    ("BATTERY", ["battery", "batteries", "ev battery", "배터리", "이차전지"]),
    ("DISPLAY", ["display", "oled", "디스플레이"]),
    ("SAMSUNG_MENTION", ["samsung", "삼성", "삼성전자"]),
    ("PRODUCTION_COUNTRY", [
        "vietnam", "india", "china", "mexico", "brazil", "poland",
        "hungary", "slovakia", "korea", "indonesia", "thailand",
        "베트남", "인도", "중국", "멕시코", "브라질", "폴란드",
        "헝가리", "슬로바키아", "한국", "인도네시아", "태국",
    ]),
]

COUNTRY_HINTS = [
    ("Korea", ["korea", "south korea", "korean", "한국", "대한민국"]),
    ("United States", ["united states", "u.s.", "usa", "america", "미국", "트럼프", "biden"]),
    ("China", ["china", "chinese", "중국", "시진핑"]),
    ("EU", ["european union", "european commission", "eu ", "유럽연합"]),
    ("Japan", ["japan", "japanese", "일본"]),
    ("Vietnam", ["vietnam", "viet nam", "베트남"]),
    ("India", ["india", "indian", "인도"]),
    ("Mexico", ["mexico", "mexican", "멕시코"]),
    ("Brazil", ["brazil", "brazilian", "브라질"]),
    ("Turkey", ["turkey", "turkiye", "türkiye", "튀르키예"]),
    ("United Kingdom", ["united kingdom", "uk", "britain", "영국"]),
    ("Canada", ["canada", "canadian", "캐나다"]),
    ("Indonesia", ["indonesia", "indonesian", "인도네시아"]),
    ("Thailand", ["thailand", "thai", "태국"]),
    ("Pakistan", ["pakistan", "파키스탄"]),
    ("Nigeria", ["nigeria", "나이지리아"]),
]

NOISE_KEYWORDS = [
    "youtube", "facebook", "instagram", "tiktok", "reddit",
    "threads", "shorts", "reels", "x.com", "twitter",
    "celebrity", "movie", "music", "entertainment", "sports",
    "baseball", "football", "basketball", "tournament", "playoffs",
    "stock price", "price target", "dividend", "earnings call",
    "real estate", "housing", "apartment", "coupon", "shopping",
    "restaurant", "recipe", "fashion", "opinion", "editorial",
    "cocaine", "firearm", "handgun", "ammunition", "drug seizure",
    "scam", "fraud", "smuggling", "fleeced", "bribe", "bribery",
    "연예", "배우", "드라마", "영화", "음악", "야구", "축구", "농구",
    "주가", "증권", "부동산", "아파트", "맛집", "여행", "마약",
]

# 관세가 아닌 tariff(공공요금/요금제/에너지 가격) 오탐 차단
FALSE_TARIFF_CONTEXT = [
    "electricity tariff", "power tariff", "energy tariff", "gas tariff",
    "water tariff", "utility tariff", "telecom tariff", "mobile tariff",
    "internet tariff", "rail tariff", "bus tariff", "taxi tariff",
    "freight tariff", "shipping tariff", "cng", "png", "natural gas tariff",
    "tariff plan", "retail tariff", "domestic tariff", "solar tariff petition",
    "전기요금", "전력요금", "가스요금", "수도요금", "통신요금", "요금제",
    "택시요금", "버스요금", "철도요금", "도시가스", "천연가스", "공공요금",
]

# 무역/관세 단어가 일부 포함되어도 GTI Top30에 부적합한 생활/범죄/지역행사성 기사
GTI_LOW_VALUE_CONTEXT = [
    "drug bust", "drug seizure", "cocaine", "methamphetamine", "narcotics",
    "smuggling gang", "airport seizure", "counterfeit cigarettes", "gold smuggling",
    "restaurant", "recipe", "tourism", "festival", "sports day", "baseball", "football",
    "crime", "fraud", "bribery", "arrested", "sentenced", "murder", "shooting",
    "마약", "코카인", "필로폰", "밀수범", "금괴", "담배 밀수", "위조상품",
    "맛집", "축제", "관광", "도민체전", "체전", "야구", "축구", "범죄", "사기", "구속", "검거",
]

# 오탐 문맥이어도 살릴 수 있는 강한 정책/무역 신호
STRONG_POLICY_CONTEXT = [
    "section 301", "section 232", "ieepa", "reciprocal tariff", "additional tariff",
    "anti-dumping", "antidumping", "countervailing", "safeguard", "export control",
    "entity list", "bis", "ustr", "cbp", "usitc", "wto", "wco", "eu commission",
    "fta", "cepa", "usmca", "rules of origin", "hs code", "classification",
    "trade remedy", "customs duty", "import duty", "관세법", "관세청", "무역법 301조",
    "232조", "상호관세", "추가관세", "반덤핑", "상계관세", "세이프가드", "수출통제",
    "전략물자", "제재", "품목분류", "원산지", "자유무역협정",
]

REGULATION_SOURCE_HINTS = [
    "law.go.kr",
    "moleg.go.kr",
    "gwanbo.go.kr",
    "법령",
    "행정규칙",
    "국가법령정보센터",
]


NOTICE_NEWS_HINTS = [
    "notice",
    "notification",
    "announcement",
    "news",
    "press release",
    "press",
    "bulletin",
    "update",
    "공지",
    "공지사항",
    "공고",
    "알림",
    "뉴스",
    "보도",
    "보도자료",
    "동향",
]

MAJOR_TARIFF_REGULATION_TERMS = [
    "tariff",
    "tariffs",
    "customs",
    "duty",
    "duties",
    "import duty",
    "export control",
    "sanction",
    "restriction",
    "anti-dumping",
    "antidumping",
    "countervailing",
    "safeguard",
    "fta",
    "cepa",
    "usmca",
    "rules of origin",
    "origin",
    "hs code",
    "classification",
    "cbam",
    "regulation",
    "law",
    "act",
    "decree",
    "amendment",
    "관세",
    "관세율",
    "세율",
    "통관",
    "세관",
    "관세청",
    "수입",
    "수출",
    "수출입",
    "수출통제",
    "제재",
    "반덤핑",
    "상계관세",
    "세이프가드",
    "원산지",
    "품목분류",
    "자유무역협정",
    "법률",
    "법령",
    "시행령",
    "시행규칙",
    "고시",
    "개정",
    "입법예고",
    "행정예고",
]


OFFICIAL_REGULATION_SOURCE_HINTS = [
    "customs.go.kr",
    "cbp.gov",
    "ustr.gov",
    "bis.gov",
    "trade.gov",
    "usitc.gov",
    "ec.europa.eu",
    "taxation-customs.ec.europa.eu",
    "wto.org",
    "wcoomd.org",
    "gov.uk",
    "customs.go.jp",
    "law.go.kr",
    "moleg.go.kr",
    "gwanbo.go.kr",
    "unipass.customs.go.kr",
    "korea customs",
    "customs and border protection",
    "u.s. customs",
    "us customs",
    "japan customs",
    "eu commission",
    "european commission",
    "world customs organization",
    "관세청",
    "국가법령정보센터",
    "법제처",
    "관보",
]

MEDIA_NEWS_SOURCE_HINTS = [
    "naver_news",
    "google_news",
    "rss_news",
    "reuters",
    "bloomberg",
    "yonhap",
    "연합뉴스",
    "news",
    "신문",
    "일보",
    "경제",
]


def log(message: str) -> None:
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {message}")


def clean(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip().lower() for c in df.columns]
    return df.loc[:, ~df.columns.duplicated()]


def pick_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    for candidate in candidates:
        if candidate in df.columns:
            return candidate
    return None


def contains_any(text: str, keywords: list[str] | set[str]) -> bool:
    lowered = str(text).lower()
    return any(str(k).lower() in lowered for k in keywords if clean(k))


def is_google_alert_or_news_source(value: object) -> bool:
    text = clean(value).lower()
    return "google.co" in text or "google.com" in text or "news.google" in text or "google alert" in text or "google news rss" in text


def domain_from_url(value: object) -> str:
    raw = unwrap_google_url(value)
    parsed = urlparse(raw)
    if parsed.netloc:
        return parsed.netloc.lower().removeprefix("www.")
    return ""


def headline_publisher_hint(headline: object) -> str:
    # Google RSS 제목은 보통 '기사 제목 - 매체명' 형태임
    text = clean(headline)
    if " - " in text:
        hint = text.rsplit(" - ", 1)[-1].strip()
        if 2 <= len(hint) <= 60:
            return hint
    return ""


def is_false_positive_policy_news(row: pd.Series) -> bool:
    text = " ".join([
        analysis_text(row),
        clean(row.get("summary", "")),
        clean(row.get("Summary", "")),
        clean(row.get("Category", "")),
        clean(row.get("InputKeyword", "")),
    ]).lower()

    # Section 301/232, 반덤핑, FTA 등 강한 정책 신호가 있으면 보존
    if contains_any(text, STRONG_POLICY_CONTEXT):
        return False

    # tariff가 관세가 아닌 요금/가격 문맥이면 제외
    if contains_any(text, FALSE_TARIFF_CONTEXT):
        return True

    # 단순 범죄/지역행사/생활 기사: customs/tariff 키워드가 있어도 제외
    if contains_any(text, GTI_LOW_VALUE_CONTEXT):
        return True

    return False


def parse_datetime(value: object) -> pd.Timestamp | pd.NaT:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return pd.NaT
    if getattr(parsed, "tzinfo", None) is not None:
        parsed = parsed.tz_convert(None)
    return parsed


def normalize_title(title: object) -> str:
    text = clean(title).lower()
    text = re.sub(r"&quot;|&#39;|&amp;|nbsp;", " ", text)
    text = re.sub(r"\[[^\]]*\]", " ", text)
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"[^a-z0-9가-힣]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()

    stop_words = {
        "the", "a", "an", "new", "latest", "update", "updates",
        "news", "press", "release", "says", "said",
    }
    tokens = [token for token in text.split() if token not in stop_words]
    return " ".join(tokens).strip()


def normalize_issue_title(title: object) -> str:
    text = normalize_title(title)
    text = re.sub(r"\b\d{1,4}([.,]\d+)?%?\b", " ", text)
    text = re.sub(r"\b(core|explained|analysis|exclusive)\b", " ", text)
    text = re.sub(r"\b(news|daily|site|times|businessline|mint|nate|머니투데이|네이트)\b", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    tokens = [token for token in text.split() if len(token) >= 2]
    return " ".join(tokens[:12])


def unwrap_google_url(url: object) -> str:
    raw = clean(url)
    parsed = urlparse(raw)
    host = parsed.netloc.lower()
    if "google.com" not in host and "google.co" not in host and "news.google" not in host:
        return raw

    params = parse_qs(parsed.query)
    for key in ["url", "q", "u"]:
        if key in params and params[key]:
            return unquote(params[key][0])
    return raw


def is_google_unresolved_url(url: object) -> bool:
    raw = clean(url)
    if not raw:
        return False
    parsed = urlparse(raw)
    host = parsed.netloc.lower()
    return "news.google" in host or "google.com" in host or "google.co" in host


def is_google_article_redirect_url(url: object) -> bool:
    """Return True for Google News article redirect URLs that open the original article in a browser.

    Important:
    - https://news.google.com/ is NOT useful.
    - https://news.google.com/rss/articles/... or /articles/... is useful as a report link.
    """
    raw = clean(url).lower()
    if not raw.startswith(("http://", "https://")):
        return False
    parsed = urlparse(raw)
    host = parsed.netloc.lower()
    path = parsed.path.lower()
    if "news.google" not in host:
        return False
    return "/rss/articles/" in path or "/articles/" in path


def choose_article_url(
    url: object,
    original_candidate: object = "",
    canonical_candidate: object = "",
    google_candidate: object = "",
) -> tuple[str, str]:
    """Choose the best report link and preserve Google News redirect when needed.

    Priority:
    1) verified original_url_candidate
    2) verified canonical_url
    3) Google query parameter unwrap, if it becomes a non-Google URL
    4) Google News article redirect URL, e.g. /rss/articles/CBMi...
    5) non-Google raw URL

    Do not return generic https://news.google.com because it does not open the article.
    """
    raw = clean(url)
    original = clean(original_candidate)
    canonical = clean(canonical_candidate)
    google = clean(google_candidate)

    for candidate, status in [
        (original, "RESTORED_ORIGINAL_CANDIDATE"),
        (canonical, "RESTORED_CANONICAL_CANDIDATE"),
        (unwrap_google_url(raw), "RESTORED_GOOGLE_QUERY"),
    ]:
        if not candidate:
            continue
        if not is_google_unresolved_url(candidate):
            return candidate, status

    for candidate, status in [
        (google, "NEEDS_SELENIUM_RESOLVE"),
        (raw, "NEEDS_SELENIUM_RESOLVE"),
    ]:
        if is_google_article_redirect_url(candidate):
            return clean(candidate), status

    if raw and not is_google_unresolved_url(raw):
        return raw, "ORIGINAL_INPUT"

    if raw:
        return raw, "GOOGLE_UNRESOLVED"
    return "", "EMPTY_URL"


def normalize_query(query: str) -> str:
    pairs = []
    for key, value in parse_qsl(query, keep_blank_values=False):
        lower_key = key.lower().strip()
        if not lower_key:
            continue
        if lower_key in URL_DROP_QUERY_KEYS:
            continue
        if any(lower_key.startswith(prefix) for prefix in URL_DROP_QUERY_PREFIXES):
            continue
        if lower_key in URL_ID_QUERY_KEYS:
            pairs.append((lower_key, value.strip()))

    if not pairs:
        return ""

    pairs = sorted(set(pairs), key=lambda item: (item[0], item[1]))
    return urlencode(pairs, doseq=True)


def normalize_url(url: object) -> str:
    raw = unwrap_google_url(url)
    if not raw:
        return ""

    parsed = urlparse(raw)
    if not parsed.netloc:
        return raw.lower().rstrip("/")

    host = parsed.netloc.lower().removeprefix("www.")
    path = parsed.path.rstrip("/")

    if "news.google.com" in host:
        return f"{host}{path}?{parsed.query}".rstrip("?")

    query = normalize_query(parsed.query)
    normalized = f"{host}{path}".rstrip("/")
    if query:
        return f"{normalized}?{query}"
    return normalized


def analysis_url_text(url: object) -> str:
    raw = unwrap_google_url(url)
    parsed = urlparse(raw)
    host = parsed.netloc.lower()
    if "news.google.com" in host:
        return ""
    return raw


def analysis_text(row: pd.Series) -> str:
    return " ".join(
        [
            clean(row.get("Headline", "")),
            clean(row.get("Summary", "")),
            analysis_url_text(row.get("Source", "")),
            clean(row.get("Publisher", "")),
            clean(row.get("InputKeyword", "")),
            analysis_url_text(row.get("URL", "")),
        ]
    ).lower()


def load_keywords() -> list[str]:
    keywords: list[str] = []

    for path in KEYWORD_FILES:
        if not path.exists():
            continue

        try:
            kdf = normalize_columns(pd.read_excel(path))
            active_col = pick_column(kdf, ["active", "use", "enabled"])
            if active_col:
                active = kdf[active_col].fillna("Y").astype(str).str.upper().str.strip()
                kdf = kdf[active.isin(["Y", "YES", "TRUE", "1"])]

            keyword_cols = [
                col for col in kdf.columns
                if "keyword" in col or col in ["kr", "en", "cn", "vi", "hi", "tr", "es", "pt"]
            ]
            for col in keyword_cols:
                keywords.extend(kdf[col].dropna().astype(str).str.strip().tolist())

            log(f"keyword loaded: {path.name} / {len(keywords)}")
            break

        except Exception as exc:
            log(f"keyword load skip: {path} / {exc}")

    keywords.extend(FALLBACK_TRADE_KEYWORDS)
    keywords = [k.strip() for k in keywords if len(k.strip()) >= 2]
    return sorted(set(keywords), key=lambda x: x.lower())


def keyword_matches(text: str, keywords: list[str]) -> list[str]:
    lowered = text.lower()
    matches = [keyword for keyword in keywords if keyword.lower() in lowered]
    return sorted(set(matches), key=lambda x: x.lower())


def title_keyword_matches(headline: object) -> list[str]:
    """Return GTI trade/customs keywords that appear in the original headline.

    This is intentionally stricter than the general keyword match.  A title with
    only Samsung/semiconductor/market terms should not become CORE/Top3 material.
    """
    title = clean(headline).lower()
    if not title:
        return []
    return sorted({term for term in TITLE_KEYWORD_TERMS if clean(term) and clean(term).lower() in title}, key=lambda x: x.lower())


def has_title_keyword(row: pd.Series) -> bool:
    if clean(row.get("TitleKeywordFlag", "")).upper() == "Y":
        return True
    if clean(row.get("TitleKeywordMatches", "")):
        return True
    return bool(title_keyword_matches(row.get("Headline", "")))


def url_quality_from_status(status: object, url: object) -> tuple[str, str]:
    """Return (OriginalURLVerified, URLQuality).

    Google article redirect links are useful for follow-up, but they are not
    treated as confirmed original URLs.  STEP3 keeps them as candidates and lets
    STEP3-1/STEP4 perform Selenium recovery.
    """
    status_text = clean(status).upper()
    raw = clean(url)
    if not raw:
        return "N", "EMPTY_URL"
    if status_text in {"RESTORED_ORIGINAL_CANDIDATE", "RESTORED_CANONICAL_CANDIDATE", "RESTORED_GOOGLE_QUERY", "ORIGINAL_INPUT"} and not is_google_unresolved_url(raw):
        return "Y", "ORIGINAL_VERIFIED"
    if status_text == "NEEDS_SELENIUM_RESOLVE" or is_google_article_redirect_url(raw):
        return "N", "NEEDS_SELENIUM_RESOLVE"
    if is_google_unresolved_url(raw):
        return "N", "GOOGLE_UNRESOLVED"
    return "Y", "ORIGINAL_VERIFIED"


def load_one_file(path: Path) -> pd.DataFrame:
    if not path.exists():
        log(f"input missing skip: {path.name}")
        return pd.DataFrame()

    raw = normalize_columns(pd.read_excel(path))
    if raw.empty:
        log(f"input empty: {path.name}")
        return pd.DataFrame()

    title_col = pick_column(raw, ["title", "headline", "subject", "name"])
    if title_col is None:
        log(f"title column missing skip: {path.name}")
        return pd.DataFrame()

    date_col = pick_column(raw, ["date", "posted_date", "publish_date", "published", "pubdate"])
    collected_col = pick_column(raw, ["collected_at", "checked_at", "created_at"])
    url_col = pick_column(raw, ["url", "link", "href"])
    source_col = pick_column(raw, ["source", "site", "publisher"])
    agency_col = pick_column(raw, ["agency"])
    publisher_col = pick_column(raw, ["publisher"])
    keyword_col = pick_column(raw, ["keyword"])
    category_col = pick_column(raw, ["category"])
    importance_col = pick_column(raw, ["importance"])
    summary_col = pick_column(raw, ["summary", "description", "body"])
    google_url_col = pick_column(raw, ["google_url"])
    original_url_col = pick_column(raw, ["original_url_candidate", "original_url", "article_url"])
    canonical_url_col = pick_column(raw, ["canonical_url"])
    score_reason_col = pick_column(raw, ["score_reason"])
    site_type_col = pick_column(raw, ["site_type"])
    date_status_col = pick_column(raw, ["date_status"])

    df = pd.DataFrame()
    df["Date"] = raw[date_col].apply(parse_datetime) if date_col else pd.NaT
    df["CollectedAt"] = raw[collected_col].apply(parse_datetime) if collected_col else pd.NaT
    df["Headline"] = raw[title_col].astype(str).str.strip()
    raw_url = raw[url_col].astype(str).str.strip() if url_col else pd.Series([""] * len(raw), index=raw.index)
    original_candidate = raw[original_url_col].astype(str).str.strip() if original_url_col else pd.Series([""] * len(raw), index=raw.index)
    canonical_candidate = raw[canonical_url_col].astype(str).str.strip() if canonical_url_col else pd.Series([""] * len(raw), index=raw.index)
    google_candidate = raw[google_url_col].astype(str).str.strip() if google_url_col else raw_url
    restored = [
        choose_article_url(url, original, canonical, google)
        for url, original, canonical, google in zip(raw_url, original_candidate, canonical_candidate, google_candidate)
    ]
    df["URL"] = [item[0] for item in restored]
    df["BestLinkURL"] = df["URL"]
    df["URLRestoreStatus"] = [item[1] for item in restored]
    df["GoogleURL"] = google_candidate
    df["OriginalURLCandidate"] = original_candidate
    df["Source"] = raw[source_col].astype(str).str.strip() if source_col else ""
    df["AgencyRaw"] = raw[agency_col].astype(str).str.strip() if agency_col else ""
    df["Publisher"] = raw[publisher_col].astype(str).str.strip() if publisher_col else ""
    df["InputKeyword"] = raw[keyword_col].astype(str).str.strip() if keyword_col else ""
    df["Category"] = raw[category_col].astype(str).str.strip() if category_col else ""
    df["Importance"] = raw[importance_col].astype(str).str.strip() if importance_col else ""
    df["Summary"] = raw[summary_col].astype(str).str.strip() if summary_col else ""
    df["SourceScoreReason"] = raw[score_reason_col].astype(str).str.strip() if score_reason_col else ""
    df["site_type"] = raw[site_type_col].astype(str).str.strip() if site_type_col else ""
    df["date_status"] = raw[date_status_col].astype(str).str.strip() if date_status_col else ""
    df["SourceFile"] = path.name

    df = df[df["Headline"].astype(str).str.len() > 5].reset_index(drop=True)
    log(f"LOAD {path.name}: {len(df)} rows")
    return df


def load_data(input_files: list[Path]) -> pd.DataFrame:
    frames = [load_one_file(path) for path in input_files]
    frames = [frame for frame in frames if not frame.empty]
    if not frames:
        return pd.DataFrame(columns=FINAL_COLS)
    result = pd.concat(frames, ignore_index=True, sort=False)
    log(f"LOAD TOTAL: {len(result)} rows")
    return result


def is_notice_or_news_row(row: pd.Series) -> bool:
    source_file = clean(row.get("SourceFile", "")).lower()
    metadata = " ".join(
        [
            source_file,
            clean(row.get("Source", "")),
            clean(row.get("AgencyRaw", "")),
            clean(row.get("Publisher", "")),
            clean(row.get("Category", "")),
            clean(row.get("site_type", "")),
        ]
    ).lower()

    if any(name in source_file for name in ["site_news", "naver_news", "google_news", "rss_news"]):
        return True
    return contains_any(metadata, NOTICE_NEWS_HINTS)


def is_major_tariff_regulation_news(row: pd.Series) -> bool:
    text = " ".join(
        [
            analysis_text(row),
            clean(row.get("Category", "")),
            clean(row.get("Importance", "")),
            clean(row.get("site_type", "")),
        ]
    ).lower()
    return is_notice_or_news_row(row) and contains_any(text, MAJOR_TARIFF_REGULATION_TERMS)


def regulation_related_flag(row: pd.Series) -> str:
    return "Y" if is_major_tariff_regulation_news(row) else "N"


def classify_regulation_transfer_type(row: pd.Series) -> str:
    if not is_major_tariff_regulation_news(row):
        return "None"

    source_file = clean(row.get("SourceFile", "")).lower()
    metadata = " ".join(
        [
            analysis_url_text(row.get("URL", "")),
            analysis_url_text(row.get("Source", "")),
            clean(row.get("AgencyRaw", "")),
            clean(row.get("Agency", "")),
            clean(row.get("Publisher", "")),
            clean(row.get("Category", "")),
            clean(row.get("site_type", "")),
        ]
    ).lower()

    if contains_any(metadata, OFFICIAL_REGULATION_SOURCE_HINTS):
        return "OfficialNotice"
    if contains_any(source_file, MEDIA_NEWS_SOURCE_HINTS):
        return "MediaNews"
    if contains_any(metadata, MEDIA_NEWS_SOURCE_HINTS):
        return "MediaNews"
    return "MediaNews"


def remove_regulation_rows(df: pd.DataFrame) -> pd.DataFrame:
    before = len(df)
    keep = []
    for idx, row in df.iterrows():
        text = f"{row.get('Headline', '')} {analysis_url_text(row.get('URL', ''))} {row.get('Source', '')} {row.get('AgencyRaw', '')} {row.get('site_type', '')}".lower()
        if contains_any(text, REGULATION_SOURCE_HINTS):
            if is_major_tariff_regulation_news(row):
                keep.append(idx)
                continue
            continue
        keep.append(idx)
    result = df.loc[keep].reset_index(drop=True)
    log(f"법규 원문 행 제거: {before - len(result)}")
    return result


def remove_old_data(df: pd.DataFrame, recent_hours: int) -> pd.DataFrame:
    before = len(df)
    cutoff = pd.Timestamp(datetime.now() - timedelta(hours=recent_hours))

    result = df.copy()
    result["FilterDate"] = result["CollectedAt"]
    result.loc[result["FilterDate"].isna(), "FilterDate"] = result.loc[result["FilterDate"].isna(), "Date"]

    known = result["FilterDate"].notna()
    recent = result["FilterDate"] >= cutoff
    status_recent = result["date_status"].astype(str).str.lower().eq("recent") if "date_status" in result.columns else False

    result = result[(~known) | recent | status_recent].reset_index(drop=True)
    log(f"최근 {recent_hours}시간 초과 제거: {before - len(result)}")
    return result


def remove_noise(df: pd.DataFrame) -> pd.DataFrame:
    before = len(df)
    keep = []
    reject_reasons = {}

    for idx, row in df.iterrows():
        text = analysis_text(row)
        if is_false_positive_policy_news(row):
            reject_reasons[idx] = "false_positive_policy_context"
            continue
        if contains_any(text, NOISE_KEYWORDS):
            if is_major_tariff_regulation_news(row) and not is_false_positive_policy_news(row):
                keep.append(idx)
                continue
            reject_reasons[idx] = "noise_keyword"
            continue
        keep.append(idx)

    result = df.loc[keep].reset_index(drop=True)
    log(f"Noise 제거: {before - len(result)}")
    return result


def add_signals(df: pd.DataFrame, keywords: list[str]) -> pd.DataFrame:
    if df.empty:
        return df.copy()

    rows = []
    for _, row in df.iterrows():
        text = analysis_text(row)
        matches = keyword_matches(text, keywords)
        title_matches = title_keyword_matches(row.get("Headline", ""))
        policy_signals = [label for label, terms in POLICY_RULES if contains_any(text, terms)]
        samsung_signals = [label for label, terms in SAMSUNG_RULES if contains_any(text, terms)]

        if contains_any(text, AD_CVD_FORCE_TERMS):
            policy_signals = ["AD_CVD"] + [x for x in policy_signals if x != "AD_CVD"]

        item = row.to_dict()
        item["KeywordMatches"] = "; ".join(matches)
        item["TitleKeywordMatches"] = "; ".join(title_matches)
        item["TitleKeywordFlag"] = "Y" if title_matches else "N"
        item["PolicySignals"] = "; ".join(policy_signals)
        item["SamsungSignal"] = "; ".join(samsung_signals) if samsung_signals else "None"
        item["IssueKey"] = policy_signals[0] if policy_signals else "TRADE_GENERAL"
        rows.append(item)

    return pd.DataFrame(rows)


def filter_relevant_news(df: pd.DataFrame) -> pd.DataFrame:
    before = len(df)
    rows = []

    for _, row in df.iterrows():
        keyword_ok = bool(clean(row.get("KeywordMatches", "")))
        policy_ok = bool(clean(row.get("PolicySignals", "")))
        input_keyword_ok = bool(clean(row.get("InputKeyword", "")))

        if keyword_ok or policy_ok or input_keyword_ok:
            rows.append(row)

    result = pd.DataFrame(rows)
    log(f"뉴스 키워드/정책 신호 선별: {len(result)} / 제거: {before - len(result)}")
    return result.reset_index(drop=True)


def infer_country(row: pd.Series) -> str:
    existing = clean(row.get("Country", ""))
    if existing:
        return existing

    text = analysis_text(row)
    found = [country for country, hints in COUNTRY_HINTS if contains_any(text, hints)]
    return ", ".join(found[:3]) if found else "Global"


def infer_agency(row: pd.Series) -> str:
    agency = clean(row.get("AgencyRaw", ""))
    if agency and agency.lower() != "nan" and not is_google_alert_or_news_source(agency):
        return agency

    publisher = clean(row.get("Publisher", ""))
    if publisher and publisher.lower() != "nan" and not is_google_alert_or_news_source(publisher):
        return publisher

    url_domain = domain_from_url(row.get("URL", ""))
    if url_domain and not is_google_alert_or_news_source(url_domain):
        return url_domain

    source = clean(row.get("Source", ""))
    source_domain = domain_from_url(source)
    if source_domain and not is_google_alert_or_news_source(source_domain):
        return source_domain

    hint = headline_publisher_hint(row.get("Headline", ""))
    if hint:
        return hint

    return "Unknown"


def infer_risk(row: pd.Series) -> str:
    issue = clean(row.get("IssueKey", ""))
    samsung = clean(row.get("SamsungSignal", ""))
    importance = clean(row.get("Importance", "")).upper()

    if issue in ["TARIFF", "AD_CVD", "EXPORT_CONTROL", "CBAM_CARBON"]:
        return "상" if samsung != "None" or importance == "HIGH" else "중"
    if issue in ["ORIGIN_FTA", "HS_CLASSIFICATION"]:
        return "중"
    return "하"


def samsung_signal_parts(samsung: object) -> list[str]:
    text = clean(samsung)
    if not text or text == "None":
        return []
    return [part.strip() for part in text.split(";") if part.strip()]


def has_product_samsung_signal(samsung: object) -> bool:
    product_signals = {"SEMICONDUCTOR", "MOBILE", "BATTERY", "DISPLAY", "SAMSUNG_MENTION"}
    return any(part in product_signals for part in samsung_signal_parts(samsung))



def infer_samsung_impact(row: pd.Series) -> str:
    """Direct/Indirect impact flag for Samsung-facing Top30 selection."""
    country_text = clean(row.get("Country", ""))
    issue = clean(row.get("IssueKey", ""))
    samsung = clean(row.get("SamsungSignal", ""))
    signals = set(samsung_signal_parts(samsung))

    country_list = [part.strip() for part in country_text.split(",") if part.strip()]
    direct_country = any(country in SAMSUNG_DIRECT_COUNTRIES for country in country_list)
    product_signal = has_product_samsung_signal(samsung)
    high_policy = issue in {"TARIFF", "AD_CVD", "EXPORT_CONTROL", "CBAM_CARBON", "ORIGIN_FTA", "HS_CLASSIFICATION"}

    # Direct는 삼성 명시 또는 '생산/판매국가 + 삼성 관련 제품군 + 고위험 정책'일 때만 부여
    # 생산국가 단독 + 일반 tariff 기사까지 Direct로 잡히던 과대 산정을 방지한다.
    if "SAMSUNG_MENTION" in signals:
        return "Direct"
    if direct_country and product_signal and high_policy:
        return "Direct"
    if product_signal or high_policy or "PRODUCTION_COUNTRY" in signals:
        return "Indirect"
    return "Reference"


def calculate_topic_score(row: pd.Series) -> int:
    issue = clean(row.get("IssueKey", "")) or "TRADE_GENERAL"
    keyword_count = len([x for x in clean(row.get("KeywordMatches", "")).split(";") if x.strip()])
    base = TOPIC_SCORE_MAP.get(issue, 45)

    if is_major_tariff_regulation_news(row):
        base += 6
    base += min(keyword_count * 2, 10)
    return max(0, min(int(base), 100))


def calculate_samsung_impact_score(row: pd.Series) -> int:
    impact = clean(row.get("SamsungImpact", ""))
    samsung = clean(row.get("SamsungSignal", ""))
    signals = set(samsung_signal_parts(samsung))

    if impact == "Direct":
        score = 100
    elif impact == "Indirect":
        score = 72
    else:
        score = 35

    product_signals = signals.intersection({"SEMICONDUCTOR", "MOBILE", "BATTERY", "DISPLAY", "SAMSUNG_MENTION"})
    if product_signals:
        score += min(len(product_signals) * 4, 12)
    if "PRODUCTION_COUNTRY" in signals:
        score += 4

    return max(0, min(int(score), 100))


def calculate_risk_score(row: pd.Series) -> int:
    risk = clean(row.get("Risk", ""))
    issue = clean(row.get("IssueKey", ""))
    score = RISK_SCORE_MAP.get(risk, 40)
    if issue in {"EXPORT_CONTROL", "AD_CVD", "TARIFF"} and risk == "상":
        score = 100
    return max(0, min(int(score), 100))


def calculate_final_score(row: pd.Series) -> int:
    topic = int(row.get("TopicScore", 0) or 0)
    samsung = int(row.get("SamsungImpactScore", 0) or 0)
    risk = int(row.get("RiskScore", 0) or 0)
    score = int(round(topic * 0.50 + samsung * 0.30 + risk * 0.20))
    adjusted = score + step3_score_adjustment(row)
    if has_title_keyword(row):
        adjusted += 8
    else:
        adjusted -= 12
    return max(0, min(adjusted, 100))


def has_step4_review_signal(row: pd.Series) -> bool:
    return contains_any(analysis_text(row), STEP4_REVIEW_TERMS)


def has_standalone_noise_signal(row: pd.Series) -> bool:
    text = analysis_text(row)
    input_keyword = clean(row.get("InputKeyword", "")).lower()
    issue = clean(row.get("IssueKey", ""))

    if input_keyword == "bis" and not contains_any(text, BIS_VALID_CONTEXT):
        return True
    if input_keyword == "aeo" and not contains_any(text, AEO_VALID_CONTEXT):
        return True
    if input_keyword in ["수출", "관세", "customs", "export", "tariff"] and contains_any(text, GENERIC_EXPORT_TARIFF_NOISE):
        return True
    if issue in ["CUSTOMS", "TRADE_GENERAL"] and contains_any(text, GENERIC_EXPORT_TARIFF_NOISE) and not has_step4_review_signal(row):
        return True
    return False


def google_url_penalty(row: pd.Series) -> int:
    status = clean(row.get("URLRestoreStatus", ""))
    if status in {"GOOGLE_UNRESOLVED", "NEEDS_SELENIUM_RESOLVE", "GOOGLE_ARTICLE_REDIRECT"}:
        return -8
    return 0


def step3_score_adjustment(row: pd.Series) -> int:
    adjust = 0
    text = analysis_text(row)

    if contains_any(text, AD_CVD_FORCE_TERMS):
        adjust += 8
    if has_step4_review_signal(row):
        adjust += 6
    if has_standalone_noise_signal(row):
        adjust -= 18
    adjust += google_url_penalty(row)

    return adjust


def step4_hint(row: pd.Series) -> str:
    hints = []
    if clean(row.get("URLRestoreStatus", "")) in {"GOOGLE_UNRESOLVED", "NEEDS_SELENIUM_RESOLVE", "GOOGLE_ARTICLE_REDIRECT"}:
        hints.append("google_url_needs_selenium_resolve")
    if not has_title_keyword(row):
        hints.append("no_title_keyword_core_blocked")
    if contains_any(analysis_text(row), AD_CVD_FORCE_TERMS):
        hints.append("force_ad_cvd_review")
    if has_step4_review_signal(row):
        hints.append("review_upgrade_signal")
    if has_standalone_noise_signal(row):
        hints.append("standalone_keyword_noise_penalty")
    return "; ".join(hints)


def source_priority(row: pd.Series) -> int:
    text = " ".join([
        clean(row.get("URL", "")),
        clean(row.get("Source", "")),
        clean(row.get("Agency", "")),
        clean(row.get("Publisher", "")),
        clean(row.get("SourceFile", "")),
    ]).lower()

    priority = 50
    for idx, domain in enumerate(SOURCE_PRIORITY_DOMAINS):
        if domain in text:
            priority = max(priority, 100 - idx * 3)
    for domain in LOW_PRIORITY_REPUBLISHERS:
        if domain in text:
            priority -= 18
    if clean(row.get("RegulationTransferType", "")) == "OfficialNotice":
        priority += 15
    return max(0, min(priority, 100))

def calculate_score(row: pd.Series) -> int:
    text = analysis_text(row)
    score = 0
    issue = clean(row.get("IssueKey", ""))
    keyword_count = len([x for x in clean(row.get("KeywordMatches", "")).split(";") if x.strip()])
    samsung = clean(row.get("SamsungSignal", ""))
    source_file = clean(row.get("SourceFile", "")).lower()
    importance = clean(row.get("Importance", "")).upper()

    if issue in ["TARIFF", "AD_CVD", "EXPORT_CONTROL", "CBAM_CARBON"]:
        score += 35
    elif issue in ["ORIGIN_FTA", "HS_CLASSIFICATION"]:
        score += 28
    elif issue == "CUSTOMS":
        score += 18
    else:
        score += 10

    score += min(keyword_count * 4, 24)

    # 제목에 관세/통상 핵심 keyword가 있으면 우선권을 주고,
    # 제목 keyword가 없으면 삼성/반도체/시장 기사 과대 선정을 방지한다.
    title_kw = has_title_keyword(row)
    if title_kw:
        score += 18
    else:
        score -= 25

    if samsung != "None":
        signals = samsung_signal_parts(samsung)
        product_signals = [x for x in signals if x != "PRODUCTION_COUNTRY"]
        if product_signals:
            score += min(len(product_signals) * 10, 30)
            score += 12
        elif "PRODUCTION_COUNTRY" in signals:
            score += 3

    if importance == "HIGH":
        score += 10
    elif importance in ["MID", "MEDIUM"]:
        score += 5

    if "google" in source_file:
        score += 2
    if "rss" in source_file:
        score += 3
    if "site_news" in source_file:
        score += 5
    if is_major_tariff_regulation_news(row):
        score += 6

    if issue == "TRADE_GENERAL":
        score = min(score, 44)
    elif issue == "CUSTOMS" and not has_product_samsung_signal(samsung):
        score = min(score, 58)

    if not title_kw:
        # 제목에 관세/통상 keyword가 없으면 STEP3 후보에는 남길 수 있으나 CORE급으로 올리지 않는다.
        score = min(score, 68)

    return min(max(score, 0), 100)


def add_analysis_fields(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        log("분석필드 생성 skip: 0 rows")
        return pd.DataFrame(columns=FINAL_COLS)

    result = df.copy()
    result["Country"] = result.apply(infer_country, axis=1)
    result["Agency"] = result.apply(infer_agency, axis=1)
    result["Risk"] = result.apply(infer_risk, axis=1)
    result["SamsungImpact"] = result.apply(infer_samsung_impact, axis=1)
    result["TopicScore"] = result.apply(calculate_topic_score, axis=1)
    result["SamsungImpactScore"] = result.apply(calculate_samsung_impact_score, axis=1)
    result["RiskScore"] = result.apply(calculate_risk_score, axis=1)
    result["FinalScore"] = result.apply(calculate_final_score, axis=1)
    result["Score"] = result["FinalScore"]
    if "TitleKeywordMatches" not in result.columns:
        result["TitleKeywordMatches"] = result["Headline"].apply(lambda v: "; ".join(title_keyword_matches(v)))
    result["TitleKeywordFlag"] = result["TitleKeywordMatches"].apply(lambda v: "Y" if clean(v) else "N")
    url_quality_pairs = [url_quality_from_status(status, url) for status, url in zip(result.get("URLRestoreStatus", ""), result.get("URL", ""))]
    result["OriginalURLVerified"] = [p[0] for p in url_quality_pairs]
    result["URLQuality"] = [p[1] for p in url_quality_pairs]
    result["Step4Hint"] = result.apply(step4_hint, axis=1)
    result["NewsType"] = "NEWS"
    result["RegulationRelated"] = result.apply(regulation_related_flag, axis=1)
    result["RegulationTransferType"] = result.apply(classify_regulation_transfer_type, axis=1)
    result["Priority"] = result["Score"].rank(method="first", ascending=False).astype(int)
    if "BestLinkURL" not in result.columns:
        result["BestLinkURL"] = result["URL"]
    if "GoogleURL" not in result.columns:
        result["GoogleURL"] = ""
    if "OriginalURLCandidate" not in result.columns:
        result["OriginalURLCandidate"] = ""
    result["title_norm"] = result["Headline"].apply(normalize_title)
    result["url_norm"] = result["URL"].apply(normalize_url)
    return result


def best_link_for_dedup(row: pd.Series) -> str:
    """Return a stable URL key source for STEP3 duplicate removal.

    The report click link may be a Google News /rss/articles/ redirect. That is
    useful for mail, but it is not enough for de-duplication because the same
    article can appear through multiple Google/portal/RSS URLs. Therefore STEP3
    de-duplicates by:
      1) real OriginalURLCandidate when available,
      2) non-Google BestLinkURL/URL,
      3) Google article redirect URL only as a fallback,
      4) title_norm as the second pass for cross-source duplicates.
    """
    for col in ["OriginalURLCandidate", "BestLinkURL", "URL", "GoogleURL"]:
        v = clean(row.get(col, ""))
        if not v:
            continue
        # Prefer non-Google original URLs for dedup keys.
        if not is_google_unresolved_url(v):
            return v
    for col in ["BestLinkURL", "GoogleURL", "URL"]:
        v = clean(row.get(col, ""))
        if is_google_article_redirect_url(v):
            return v
    return clean(row.get("URL", ""))


def dedup_news(df: pd.DataFrame) -> pd.DataFrame:
    """Two-pass de-duplication.

    v3.4 fix:
    - Keep GoogleURL/BestLinkURL for mail links.
    - Do not let unique Google RSS article IDs prevent duplicate removal.
    - After URL de-dup, always run title_norm de-dup across all rows.
    This restores the expected 250~350 duplicate reduction range.
    """
    if df.empty:
        log("뉴스 중복 제거 skip: 0 rows")
        return df

    before = len(df)
    result = df.copy()
    result["dedup_url_source"] = result.apply(best_link_for_dedup, axis=1)
    result["url_norm"] = result["dedup_url_source"].apply(normalize_url)
    result["title_norm"] = result["Headline"].apply(normalize_title)

    result = result.sort_values(["Score", "FilterDate"], ascending=[False, False])

    # 1차: URL 기준 중복 제거. URL이 없는 건은 title 기준으로 처리.
    has_url = result[result["url_norm"] != ""].drop_duplicates(subset=["url_norm"], keep="first")
    no_url = result[result["url_norm"] == ""].drop_duplicates(subset=["title_norm"], keep="first")
    result = pd.concat([has_url, no_url], ignore_index=True, sort=False)

    # 2차: 동일 기사 제목이 Naver/Google/RSS/portal별로 남는 것을 제거.
    result = result.sort_values(["Score", "FilterDate"], ascending=[False, False])
    result = result.drop_duplicates(subset=["title_norm"], keep="first")
    result = result.drop(columns=["dedup_url_source"], errors="ignore")
    result = result.sort_values(["Score", "FilterDate"], ascending=[False, False]).reset_index(drop=True)

    log(f"뉴스 중복 제거: {before - len(result)}")
    return result

def make_issue_cluster_key(row: pd.Series) -> str:
    issue = clean(row.get("IssueKey", "")) or "TRADE_GENERAL"
    country = clean(row.get("Country", "")) or "Global"
    title = normalize_issue_title(row.get("Headline", ""))
    tokens = title.split()

    anchor_terms = []
    for term in [
        "cepa", "usmca", "fta", "tariff", "tariffs", "duty", "customs",
        "anti", "dumping", "countervailing", "export", "control", "cbam",
        "관세", "수출", "수입", "무역", "원산지", "반덤핑", "품목분류", "제재",
    ]:
        if term in title:
            anchor_terms.append(term)

    if anchor_terms:
        base = " ".join(anchor_terms[:6])
    else:
        base = " ".join(tokens[:8])

    return f"{issue}|{country}|{base}".lower()


def title_similarity(a: object, b: object) -> float:
    left = normalize_issue_title(a)
    right = normalize_issue_title(b)
    if not left or not right:
        return 0.0

    left_tokens = set(left.split())
    right_tokens = set(right.split())
    if not left_tokens or not right_tokens:
        return 0.0

    seq = SequenceMatcher(None, left, right).ratio()
    jaccard = len(left_tokens & right_tokens) / max(1, len(left_tokens | right_tokens))
    return max(seq, jaccard)


def same_policy_issue(row: pd.Series, rep: pd.Series) -> bool:
    if clean(row.get("IssueKey", "")) != clean(rep.get("IssueKey", "")):
        return False

    row_country = set(x.strip() for x in clean(row.get("Country", "")).split(",") if x.strip())
    rep_country = set(x.strip() for x in clean(rep.get("Country", "")).split(",") if x.strip())
    country_overlap = bool(row_country & rep_country) or "Global" in row_country or "Global" in rep_country
    if not country_overlap:
        return False

    if make_issue_cluster_key(row) == make_issue_cluster_key(rep):
        return True
    return title_similarity(row.get("Headline", ""), rep.get("Headline", "")) >= ISSUE_CLUSTER_SIMILARITY


def choose_representative(group: pd.DataFrame) -> pd.Series:
    ranked = group.copy()
    ranked["_source_priority"] = ranked.apply(source_priority, axis=1)
    ranked = ranked.sort_values(
        ["FinalScore", "_source_priority", "FilterDate"],
        ascending=[False, False, False],
    )
    return ranked.iloc[0]


def compress_cluster(group: pd.DataFrame) -> dict:
    rep = choose_representative(group).to_dict()
    sources = []
    headlines = []
    for _, row in group.iterrows():
        agency = clean(row.get("Agency", "")) or clean(row.get("Publisher", "")) or clean(row.get("SourceFile", ""))
        if agency and agency not in sources:
            sources.append(agency)
        headline = clean(row.get("Headline", ""))
        if headline and headline not in headlines:
            headlines.append(headline)

    rep["ClusterSize"] = len(group)
    rep["DuplicateCount"] = max(len(group) - 1, 0)
    rep["ClusterSources"] = "; ".join(sources[:12])
    rep["ClusterHeadlines"] = " | ".join(headlines[:8])
    rep["IssueClusterKey"] = make_issue_cluster_key(pd.Series(rep))
    rep["RepresentativeReason"] = (
        f"대표기사 선정: FinalScore={int(rep.get('FinalScore', 0) or 0)}, "
        f"SourcePriority={source_priority(pd.Series(rep))}, "
        f"ClusterSize={len(group)}"
    )
    return rep


def cluster_policy_events(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        log("정책 이슈 clustering skip: 0 rows")
        return df

    before = len(df)
    result = df.copy()
    result["IssueClusterKey"] = result.apply(make_issue_cluster_key, axis=1)
    result = result.sort_values(["FinalScore", "FilterDate"], ascending=[False, False]).reset_index(drop=True)

    # Previous near-similarity clustering compared each row with many representatives.
    # With large daily candidate pools this can run for hours.  Use the deterministic
    # issue key as the primary grouping key; title-level dedup already ran earlier.
    compressed_rows = []
    for _, group in result.groupby("IssueClusterKey", sort=False, dropna=False):
        compressed_rows.append(compress_cluster(group.copy()))

    compressed = pd.DataFrame(compressed_rows)
    if compressed.empty:
        log(f"정책 이슈 clustering 제거: {before} / 대표기사=0")
        return compressed

    compressed = compressed.sort_values(["FinalScore", "FilterDate"], ascending=[False, False]).reset_index(drop=True)
    log(f"정책 이슈 clustering 제거: {before - len(compressed)} / 대표기사={len(compressed)} / mode=fast_key")
    return compressed

def limit_trade_general(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    trade_general = df[df["IssueKey"].eq("TRADE_GENERAL")].copy()
    focused = df[~df["IssueKey"].eq("TRADE_GENERAL")].copy()

    if len(trade_general) > MAX_TRADE_GENERAL_OUTPUT:
        trade_general = trade_general.sort_values(["Score", "FilterDate"], ascending=[False, False]).head(MAX_TRADE_GENERAL_OUTPUT)

    result = pd.concat([focused, trade_general], ignore_index=True, sort=False)
    return result.sort_values(["Score", "FilterDate"], ascending=[False, False]).reset_index(drop=True)


def assign_tier_and_reasons(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    result = df.copy()
    tiers = []
    select_reasons = []
    reject_reasons = []

    for _, row in result.iterrows():
        score = int(row.get("Score", 0))
        issue = clean(row.get("IssueKey", ""))
        samsung = clean(row.get("SamsungSignal", ""))
        product_samsung = has_product_samsung_signal(samsung)
        major_notice_policy = is_major_tariff_regulation_news(row)
        title_keyword = has_title_keyword(row)
        unresolved_google = clean(row.get("URLRestoreStatus", "")) in {"GOOGLE_UNRESOLVED", "NEEDS_SELENIUM_RESOLVE", "GOOGLE_ARTICLE_REDIRECT"}
        standalone_noise = has_standalone_noise_signal(row)
        review_signal = has_step4_review_signal(row)
        ad_cvd_forced = contains_any(analysis_text(row), AD_CVD_FORCE_TERMS)

        reasons = []
        rejects = []

        if issue:
            reasons.append(f"issue={issue}")
        if samsung != "None":
            reasons.append(f"samsung={samsung}")
        if clean(row.get("KeywordMatches", "")):
            reasons.append("keyword_match")
        if title_keyword:
            reasons.append("title_keyword_match")
        else:
            reasons.append("no_title_keyword_core_blocked")
        if major_notice_policy:
            reasons.append("notice_news_tariff_regulation")
        if unresolved_google:
            reasons.append("google_url_unresolved")
        if standalone_noise:
            reasons.append("standalone_keyword_noise_penalty")
        if review_signal:
            reasons.append("step4_review_upgrade_signal")
        if ad_cvd_forced:
            reasons.append("ad_cvd_forced")

        if issue == "TRADE_GENERAL":
            tier = "REFERENCE" if score >= MIN_SCORE else "REJECT"
            if tier == "REJECT":
                rejects.append("low_score")
        elif (
            score >= 70
            and (product_samsung or score >= 80)
            and issue in ["TARIFF", "AD_CVD", "EXPORT_CONTROL", "CBAM_CARBON", "ORIGIN_FTA"]
            and title_keyword
            and not unresolved_google
            and not standalone_noise
        ):
            tier = "CORE"
        elif score >= 45:
            tier = "USABLE"
        elif score >= MIN_SCORE:
            tier = "REFERENCE"
        else:
            tier = "REJECT"
            rejects.append("low_score")

        if issue == "CUSTOMS" and not product_samsung and tier == "CORE":
            tier = "USABLE"
        if major_notice_policy and tier == "REFERENCE" and score >= MIN_SCORE:
            tier = "USABLE"
        if review_signal and tier == "REFERENCE" and score >= MIN_SCORE:
            tier = "USABLE"
        if ad_cvd_forced and tier == "REFERENCE" and score >= MIN_SCORE:
            tier = "USABLE"

        tiers.append(tier)
        select_reasons.append("; ".join(reasons))
        reject_reasons.append("; ".join(rejects))

    result["Tier"] = tiers
    result["SelectReason"] = select_reasons
    result["RejectReason"] = reject_reasons
    return result


def enforce_tier_buckets(df: pd.DataFrame) -> pd.DataFrame:
    """Force final STEP3 distribution: CORE 50 / USABLE 100 / REFERENCE 150 by default.

    Why:
      - When all 300 rows are CORE, Step4 loses prioritization.
      - STEP3 should provide a ranked candidate pool, not say every item is executive-critical.

    Rule:
      1) Sort by existing risk/tier/score signals.
      2) Top N -> CORE.
      3) Next N -> USABLE.
      4) Remaining up to MAX_OUTPUT -> REFERENCE.
      5) REJECT rows are not revived.

    Environment override:
      - GTI_STEP3_CORE_LIMIT=50
      - GTI_STEP3_USABLE_LIMIT=100
      - GTI_STEP3_REFERENCE_LIMIT=150
    """
    if df.empty:
        return df

    result = df.copy()
    risk_order = {"상": 3, "중": 2, "하": 1}
    preliminary_tier_order = {"CORE": 3, "USABLE": 2, "REFERENCE": 1}

    result["risk_order"] = result["Risk"].map(risk_order).fillna(0)
    result["tier_order"] = result["Tier"].map(preliminary_tier_order).fillna(0)
    result["title_keyword_order"] = result.apply(lambda r: 1 if has_title_keyword(r) else 0, axis=1)

    sort_cols = ["title_keyword_order", "tier_order", "risk_order", "FinalScore", "TopicScore", "SamsungImpactScore", "FilterDate"]
    result = result.sort_values(sort_cols, ascending=[False, False, False, False, False, False, False]).reset_index(drop=True)

    core_end = min(TIER_CORE_LIMIT, len(result))
    usable_end = min(core_end + TIER_USABLE_LIMIT, len(result))
    reference_end = min(usable_end + TIER_REFERENCE_LIMIT, len(result), MAX_OUTPUT)

    result = result.iloc[:reference_end].copy()
    result["TierOriginal"] = result["Tier"]

    result.loc[:, "Tier"] = "REFERENCE"
    core_candidates = [
        idx for idx, row in result.iterrows()
        if clean(row.get("TierOriginal", "")) == "CORE"
        and has_title_keyword(row)
        and clean(row.get("URLRestoreStatus", "")) not in {"GOOGLE_UNRESOLVED", "NEEDS_SELENIUM_RESOLVE", "GOOGLE_ARTICLE_REDIRECT"}
        and not has_standalone_noise_signal(row)
    ]
    core_selected = core_candidates[:core_end]
    if core_selected:
        result.loc[core_selected, "Tier"] = "CORE"

    non_core = [idx for idx in result.index if idx not in set(core_selected)]
    usable_slots = min(TIER_USABLE_LIMIT, len(non_core))
    if usable_slots > 0:
        result.loc[non_core[:usable_slots], "Tier"] = "USABLE"

    bucket_note = []

    # SelectReason에 강제 배분 흔적을 남겨 감사 가능하게 함
    bucket_note = []
    for i, row in result.iterrows():
        note = f"tier_bucket={row['Tier']}; tier_original={row.get('TierOriginal', '')}; rank_after_cluster={i+1}"
        existing = clean(row.get("SelectReason", ""))
        bucket_note.append(f"{existing}; {note}" if existing else note)
    result["SelectReason"] = bucket_note

    log(
        "Tier bucket applied: "
        f"CORE={int(result['Tier'].eq('CORE').sum())}, "
        f"USABLE={int(result['Tier'].eq('USABLE').sum())}, "
        f"REFERENCE={int(result['Tier'].eq('REFERENCE').sum())}"
    )
    return result


def finalize(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        log("최종 뉴스 후보 없음: 빈 파일 생성")
        return pd.DataFrame(columns=FINAL_COLS)

    result = assign_tier_and_reasons(df)
    result = result[result["Tier"].ne("REJECT")].copy()
    result = limit_trade_general(result)

    if result.empty:
        log("최종 뉴스 후보 없음: score 기준 미달")
        return pd.DataFrame(columns=FINAL_COLS)

    # 모든 후보가 CORE로 분류되는 문제를 방지하기 위해
    # 최종 출력 직전에 CORE/USABLE/REFERENCE를 강제 배분합니다.
    result = enforce_tier_buckets(result)

    final = pd.DataFrame()
    final["Date"] = pd.to_datetime(result["Date"], errors="coerce").dt.strftime("%Y-%m-%d %H:%M:%S")
    final["CollectedAt"] = pd.to_datetime(result["CollectedAt"], errors="coerce").dt.strftime("%Y-%m-%d %H:%M:%S")
    final["Headline"] = result["Headline"]
    final["URL"] = result["URL"]
    final["GoogleURL"] = result["GoogleURL"] if "GoogleURL" in result.columns else ""
    final["OriginalURLCandidate"] = result["OriginalURLCandidate"] if "OriginalURLCandidate" in result.columns else ""
    final["BestLinkURL"] = result["BestLinkURL"] if "BestLinkURL" in result.columns else result["URL"]
    final["Country"] = result["Country"]
    final["Agency"] = result["Agency"]
    final["Risk"] = result["Risk"]
    final["TopicScore"] = result["TopicScore"].astype(int)
    final["SamsungImpactScore"] = result["SamsungImpactScore"].astype(int)
    final["RiskScore"] = result["RiskScore"].astype(int)
    final["FinalScore"] = result["FinalScore"].astype(int)
    final["Score"] = result["Score"].astype(int)
    final["Priority"] = range(1, len(final) + 1)
    final["Tier"] = result["Tier"]
    final["NewsType"] = result["NewsType"]
    final["IssueKey"] = result["IssueKey"]
    final["SamsungSignal"] = result["SamsungSignal"]
    final["SamsungImpact"] = result["SamsungImpact"]
    final["RegulationRelated"] = result["RegulationRelated"]
    final["RegulationTransferType"] = result["RegulationTransferType"]
    final["KeywordMatches"] = result["KeywordMatches"]
    final["TitleKeywordFlag"] = result.get("TitleKeywordFlag", "")
    final["TitleKeywordMatches"] = result.get("TitleKeywordMatches", "")
    final["OriginalURLVerified"] = result.get("OriginalURLVerified", "")
    final["URLQuality"] = result.get("URLQuality", "")
    final["IssueClusterKey"] = result.get("IssueClusterKey", "")
    final["ClusterSize"] = result.get("ClusterSize", 1)
    final["DuplicateCount"] = result.get("DuplicateCount", 0)
    final["ClusterSources"] = result.get("ClusterSources", "")
    final["ClusterHeadlines"] = result.get("ClusterHeadlines", "")
    final["RepresentativeReason"] = result.get("RepresentativeReason", "")
    final["SelectReason"] = result["SelectReason"]
    final["RejectReason"] = result["RejectReason"]
    final["Source"] = result["Source"]
    final["SourceFile"] = result["SourceFile"]
    final["Publisher"] = result["Publisher"]
    final["Importance"] = result["Importance"]
    final["Category"] = result["Category"]
    final["URLRestoreStatus"] = result.get("URLRestoreStatus", "")
    # FINAL CORE v6: Step2 sources use mixed naming (url_decode_status / URLDecodeStatus).
    # Always create URLDecodeStatus before slicing FINAL_COLS to prevent KeyError.
    if "URLDecodeStatus" in result.columns:
        final["URLDecodeStatus"] = result["URLDecodeStatus"]
    elif "url_decode_status" in result.columns:
        final["URLDecodeStatus"] = result["url_decode_status"]
    elif "URLRestoreStatus" in final.columns:
        final["URLDecodeStatus"] = final["URLRestoreStatus"]
    else:
        final["URLDecodeStatus"] = ""
    final["Step4Hint"] = result.get("Step4Hint", "")
    final["SourceScoreReason"] = result.get("SourceScoreReason", "")

    # Safety net: if future schema columns are added, create missing output columns as blank
    # instead of failing the pipeline after a long collection run.
    for _col in FINAL_COLS:
        if _col not in final.columns:
            final[_col] = ""

    return final[FINAL_COLS].reset_index(drop=True)


def standardize_cumulative_columns(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    rename_map = {}

    for col in result.columns:
        lower = str(col).strip().lower()
        if lower == "date":
            rename_map[col] = "Date"
        elif lower in ["collectedat", "collected_at", "last_checked"]:
            rename_map[col] = "CollectedAt"
        elif lower in ["headline", "title"]:
            rename_map[col] = "Headline"
        elif lower in ["url", "link"]:
            rename_map[col] = "URL"
        elif lower == "country":
            rename_map[col] = "Country"
        elif lower == "agency":
            rename_map[col] = "Agency"
        elif lower == "risk":
            rename_map[col] = "Risk"
        elif lower in ["topicscore", "topic_score"]:
            rename_map[col] = "TopicScore"
        elif lower in ["samsungimpactscore", "samsung_impact_score"]:
            rename_map[col] = "SamsungImpactScore"
        elif lower in ["riskscore", "risk_score"]:
            rename_map[col] = "RiskScore"
        elif lower in ["finalscore", "final_score"]:
            rename_map[col] = "FinalScore"
        elif lower == "score":
            rename_map[col] = "Score"
        elif lower == "priority":
            rename_map[col] = "Priority"
        elif lower == "tier":
            rename_map[col] = "Tier"
        elif lower in ["newstype", "news_type"]:
            rename_map[col] = "NewsType"
        elif lower in ["issuekey", "issue_key"]:
            rename_map[col] = "IssueKey"
        elif lower in ["samsungsignal", "samsung_signal"]:
            rename_map[col] = "SamsungSignal"
        elif lower in ["samsungimpact", "samsung_impact"]:
            rename_map[col] = "SamsungImpact"
        elif lower in ["issueclusterkey", "issue_cluster_key"]:
            rename_map[col] = "IssueClusterKey"
        elif lower in ["clustersize", "cluster_size"]:
            rename_map[col] = "ClusterSize"
        elif lower in ["duplicatecount", "duplicate_count"]:
            rename_map[col] = "DuplicateCount"
        elif lower in ["clustersources", "cluster_sources"]:
            rename_map[col] = "ClusterSources"
        elif lower in ["clusterheadlines", "cluster_headlines"]:
            rename_map[col] = "ClusterHeadlines"
        elif lower in ["representativereason", "representative_reason"]:
            rename_map[col] = "RepresentativeReason"
        elif lower in ["keywordmatches", "keyword_matches", "keyword"]:
            rename_map[col] = "KeywordMatches"
        elif lower in ["titlekeywordflag", "title_keyword_flag"]:
            rename_map[col] = "TitleKeywordFlag"
        elif lower in ["titlekeywordmatches", "title_keyword_matches"]:
            rename_map[col] = "TitleKeywordMatches"
        elif lower in ["originalurlverified", "original_url_verified"]:
            rename_map[col] = "OriginalURLVerified"
        elif lower in ["urlquality", "url_quality"]:
            rename_map[col] = "URLQuality"
        elif lower in ["selectreason", "select_reason"]:
            rename_map[col] = "SelectReason"
        elif lower in ["rejectreason", "reject_reason"]:
            rename_map[col] = "RejectReason"
        elif lower == "source":
            rename_map[col] = "Source"
        elif lower in ["sourcefile", "source_file"]:
            rename_map[col] = "SourceFile"
        elif lower == "publisher":
            rename_map[col] = "Publisher"
        elif lower == "importance":
            rename_map[col] = "Importance"
        elif lower == "category":
            rename_map[col] = "Category"
        elif lower in ["googleurl", "google_url"]:
            rename_map[col] = "GoogleURL"
        elif lower in ["originalurlcandidate", "original_url_candidate", "original_url", "article_url"]:
            rename_map[col] = "OriginalURLCandidate"
        elif lower in ["bestlinkurl", "best_link_url", "article_link", "report_url"]:
            rename_map[col] = "BestLinkURL"
        elif lower in ["urlrestorestatus", "url_restore_status"]:
            rename_map[col] = "URLRestoreStatus"
        elif lower in ["step4hint", "step4_hint"]:
            rename_map[col] = "Step4Hint"
        elif lower in ["sourcescorereason", "source_score_reason", "score_reason"]:
            rename_map[col] = "SourceScoreReason"

    result = result.rename(columns=rename_map)
    result = result.loc[:, ~result.columns.duplicated()]

    for col in FINAL_COLS:
        if col not in result.columns:
            result[col] = ""

    return result[FINAL_COLS]


def make_merge_key(row: pd.Series) -> str:
    """Cumulative key: URL based, but prefer stable original/click link columns."""
    for col in ["OriginalURLCandidate", "BestLinkURL", "URL", "GoogleURL"]:
        key = normalize_url(row.get(col, ""))
        if key:
            return key
    return ""


def make_title_source_key(row: pd.Series) -> str:
    headline = clean(row.get("Headline", "")).lower()
    source = clean(row.get("Source", "") or row.get("Publisher", "") or row.get("Agency", "")).lower()
    if not headline:
        return ""
    normalized_headline = re.sub(
        r"\s+",
        " ",
        re.sub(r"[^0-9a-z가-힣一-龥ぁ-ゔァ-ヴー\s]", " ", headline),
    ).strip()
    normalized_source = re.sub(
        r"\s+",
        " ",
        re.sub(r"[^0-9a-z가-힣一-龥ぁ-ゔァ-ヴー\s]", " ", source),
    ).strip()
    return f"{normalized_source}|{normalized_headline}" if normalized_headline else ""


def has_headline_and_link(row: pd.Series) -> bool:
    if not clean(row.get("Headline", "")):
        return False
    for col in ["OriginalURLCandidate", "BestLinkURL", "URL", "GoogleURL"]:
        if clean(row.get(col, "")).startswith("http"):
            return True
    return False


def update_cumulative(final_df: pd.DataFrame, cumulative_file: Path) -> pd.DataFrame:
    new_df = standardize_cumulative_columns(final_df)
    before_required = len(new_df)
    if before_required:
        new_df = new_df[new_df.apply(has_headline_and_link, axis=1)].copy()
        removed_required = before_required - len(new_df)
        if removed_required:
            log(f"cumulative required fields removed: {removed_required} rows without headline/url")

    if cumulative_file.exists():
        try:
            old = standardize_cumulative_columns(pd.read_excel(cumulative_file, engine="openpyxl"))
        except Exception as exc:
            old = pd.DataFrame(columns=FINAL_COLS)
            backup = cumulative_file.with_name(
                f"{cumulative_file.stem}_read_error_{datetime.now().strftime('%Y%m%d_%H%M%S')}{cumulative_file.suffix}"
            )
            try:
                shutil.copy2(cumulative_file, backup)
                log(f"WARN cumulative read failed; backup saved: {backup.name}")
            except Exception as backup_exc:
                log(f"WARN cumulative backup failed: {type(backup_exc).__name__}: {backup_exc}")
            log(f"WARN cumulative read failed -> new create: {cumulative_file} / {type(exc).__name__}: {exc}")
        old_count = len(old)
        log(f"cumulative 기존파일 로드: {old_count} rows")
    else:
        old = pd.DataFrame(columns=FINAL_COLS)
        old_count = 0
        log("cumulative 파일 없음 → 신규 생성")

    if new_df.empty:
        log("cumulative 신규 추가 skip: 0 rows")
        return old

    # URL 기준 누적 비교만 수행한다. 빈 URL은 cumulative 중복 비교 대상에서 제외한다.
    old_keys = set()
    old_title_source_keys = set()
    if not old.empty:
        old_keys = {key for key in old.apply(make_merge_key, axis=1).astype(str) if key}
        old_title_source_keys = {key for key in old.apply(make_title_source_key, axis=1).astype(str) if key}

    new_df = new_df.copy()
    new_df["_merge_key"] = new_df.apply(make_merge_key, axis=1)
    new_df["_title_source_key"] = new_df.apply(make_title_source_key, axis=1)

    additions = new_df[
        (new_df["_merge_key"].astype(str).eq(""))
        | (~new_df["_merge_key"].astype(str).isin(old_keys))
    ].copy()
    additions = additions[
        (additions["_title_source_key"].astype(str).eq(""))
        | (~additions["_title_source_key"].astype(str).isin(old_title_source_keys))
    ].drop(columns=["_merge_key", "_title_source_key"])
    combined = pd.concat([old, additions], ignore_index=True, sort=False)

    if len(combined) < old_count:
        raise RuntimeError(f"cumulative row count decreased: old={old_count}, new={len(combined)}")

    for _col in FINAL_COLS:
        if _col not in combined.columns:
            combined[_col] = ""
    log(f"cumulative update complete: {old_count} + {len(additions)} → {len(combined)} rows")
    return combined[FINAL_COLS]


def write_excel(path: Path, df: pd.DataFrame, sheet_name: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    format_workbook(path)


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    tier_fills = {
        "CORE": "D9EAD3",
        "USABLE": "DDEBF7",
        "REFERENCE": "FFF2CC",
        "REJECT": "F4CCCC",
    }

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        headers = [cell.value for cell in ws[1]]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        tier_idx = headers.index("Tier") if "Tier" in headers else None
        for row in ws.iter_rows(min_row=2):
            if tier_idx is not None:
                tier_value = row[tier_idx].value
                if tier_value in tier_fills:
                    row[tier_idx].fill = PatternFill("solid", fgColor=tier_fills[tier_value])
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for idx in range(1, ws.max_column + 1):
            header = str(ws.cell(1, idx).value or "")
            width = 16
            if header in ["Headline", "URL", "Source", "KeywordMatches", "SelectReason", "RejectReason"]:
                width = 48
            elif header in ["Country", "Agency", "IssueKey", "SamsungSignal"]:
                width = 24
            elif header in ["TopicScore", "SamsungImpactScore", "RiskScore", "FinalScore", "Score", "Priority", "Risk", "Tier", "ClusterSize", "DuplicateCount"]:
                width = 14
            elif header in ["SamsungImpact"]:
                width = 16
            ws.column_dimensions[get_column_letter(idx)].width = width

    wb.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="STEP3-2 news merge")
    parser.add_argument("--output", type=Path, default=OUTPUT_FILE)
    parser.add_argument("--cumulative", type=Path, default=CUMULATIVE_FILE)
    parser.add_argument("--hours", type=int, default=RECENT_HOURS)
    parser.add_argument("--min-score", type=int, default=MIN_SCORE)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    global MIN_SCORE
    MIN_SCORE = args.min_score

    log("STEP3-2 NEWS MERGE START")

    keywords = load_keywords()
    df = load_data(INPUT_FILES)
    df = remove_regulation_rows(df)
    df = remove_old_data(df, args.hours)
    df = remove_noise(df)
    df = add_signals(df, keywords)
    df = filter_relevant_news(df)
    df = add_analysis_fields(df)
    df = dedup_news(df)
    df = cluster_policy_events(df)

    final_df = finalize(df)
    if not final_df.empty:
        title_kw_count = int(final_df["TitleKeywordFlag"].astype(str).str.upper().eq("Y").sum()) if "TitleKeywordFlag" in final_df.columns else 0
        original_verified_count = int(final_df["OriginalURLVerified"].astype(str).str.upper().eq("Y").sum()) if "OriginalURLVerified" in final_df.columns else 0
        needs_selenium_count = int(final_df["URLQuality"].astype(str).eq("NEEDS_SELENIUM_RESOLVE").sum()) if "URLQuality" in final_df.columns else 0
        log(
            "STEP3 quality: "
            f"title_keyword={title_kw_count}/{len(final_df)}, "
            f"original_verified={original_verified_count}/{len(final_df)}, "
            f"needs_selenium={needs_selenium_count}"
        )
    cumulative_df = update_cumulative(final_df, args.cumulative)

    write_excel(args.output, final_df, "news_summary")
    write_excel(args.cumulative, cumulative_df, "news_cumulative")

    log(f"STEP3-2 COMPLETE: summary={len(final_df)} rows / cumulative={len(cumulative_df)} rows")
    log(f"SAVE: {args.output}")
    log(f"SAVE: {args.cumulative}")


if __name__ == "__main__":
    main()


