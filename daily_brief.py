# -*- coding: utf-8 -*-
"""
Samsung Electronics | Customs & Trade Daily Brief
vCurrent STABLE (E2E): Tight Loader + Dedup + Table + Dual Mail + out/ artifacts

목표(안정 운영):
- GitHub Actions에서 매일 자동 실행 (PC 불필요)
- 수집 윈도우: 전일 07:00 ~ 당일 07:00 (KST)
- 발송 시각: 워크플로우(cron)로 통제 (권장: KST 08:00 = UTC 23:00)
- 입력:
  - custom_queries.TXT : 제시어(키워드) 목록(줄 단위)
  - sites.xlsx         : 정부/공인 사이트 목록(시트 자동 탐지, name/url 필수)
- 출력:
  - out/policy_events_YYYY-MM-DD.(csv/xlsx/html) 생성
- 메일:
  - 실무자: 표 중심(요청사항 반영: '출처' 칼럼 삭제, 헤드라인 링크 + 주요내용 한 칸, '비고' 칼럼 제외)
  - 임원: TOP3 + 당사연관성(룰 기반) + Action
  - 임원 TOP3 내용은 실무자 메일에도 동일하게 "Executive Insight" 박스로 포함
"""

import os
import re
import html
import smtplib
import urllib.parse
import datetime as dt
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

import pandas as pd
import feedparser
from dateutil import parser as dtparser
from dateutil.tz import tzoffset

# ===============================
# ENV
# ===============================
SMTP_SERVER = os.getenv("SMTP_SERVER")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_EMAIL = os.getenv("SMTP_EMAIL")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")

RECIPIENTS = [x.strip() for x in os.getenv("RECIPIENTS", "").split(",") if x.strip()]
RECIPIENTS_EXEC = [x.strip() for x in os.getenv("RECIPIENTS_EXEC", "").split(",") if x.strip()]

BASE_DIR = os.getenv("BASE_DIR", os.path.join(os.path.dirname(__file__), "out"))
os.makedirs(BASE_DIR, exist_ok=True)

QUERIES_FILE = os.getenv("QUERIES_FILE", os.path.join(os.path.dirname(__file__), "custom_queries.TXT"))
SITES_FILE = os.getenv("SITES_FILE", os.path.join(os.path.dirname(__file__), "sites.xlsx"))

NEWS_LANGS = [x.strip() for x in os.getenv("NEWS_LANGS", "ko,en,fr").split(",") if x.strip()]
NEWS_REGION = os.getenv("NEWS_REGION", "KR")
NEWS_QUERY_EXPAND = os.getenv("NEWS_QUERY_EXPAND", "1") == "1"

NEWS_MAX_PER_QUERY = int(os.getenv("NEWS_MAX_PER_QUERY", "30"))  # per query per lang
TOTAL_MAX_ITEMS = int(os.getenv("TOTAL_MAX_ITEMS", "250"))       # cap after merge/dedup

# ===============================
# TIME / WINDOW
# ===============================
KST = tzoffset("KST", 9 * 3600)

def now_kst() -> dt.datetime:
    return dt.datetime.now(tz=KST)

def collection_window_kst(ref: dt.datetime):
    """
    수집 윈도우: 전일 07:00 ~ 당일 07:00 (KST)
    - 예: 08:00에 실행되면 end=오늘 07:00, start=어제 07:00
    """
    end = ref.replace(hour=7, minute=0, second=0, microsecond=0)
    if ref.hour < 7:
        end = end - dt.timedelta(days=1)
    start = end - dt.timedelta(days=1)
    return start, end

def parse_pub_kst(published: str):
    if not published:
        return None
    try:
        d = dtparser.parse(published)
        if d.tzinfo is None:
            d = d.replace(tzinfo=dt.timezone.utc)
        return d.astimezone(KST)
    except Exception:
        return None

# ===============================
# TIGHT LOADER
# ===============================
def load_custom_queries(path: str) -> list:
    """
    custom_queries.TXT:
      - 줄 단위 제시어
      - 빈 줄/주석(# 또는 //) 제거
      - 중복 제거(대소문자 무시), 순서 유지
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"custom_queries.TXT not found: {path}")

    with open(path, "r", encoding="utf-8") as f:
        lines = f.read().splitlines()

    out, seen = [], set()
    for line in lines:
        s = line.strip()
        if not s:
            continue
        if s.startswith("#") or s.startswith("//"):
            continue
        s = re.sub(r"\s+", " ", s).strip()
        key = s.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(s)

    if not out:
        raise ValueError("custom_queries.TXT is empty after filtering.")
    return out

def _normalize_url(u) -> str:
    # NaN/None/float 등 어떤 타입이 와도 안전하게 처리
    try:
        import pandas as pd
        if u is None or (isinstance(u, float) and (u != u)):  # NaN 체크
            return ""
        # pandas NaN도 커버
        try:
            if pd.isna(u):
                return ""
        except Exception:
            pass
    except Exception:
        pass

    u = str(u).strip()
    if not u or u.lower() in ("nan", "none"):
        return ""

    # 엑셀 하이퍼링크/텍스트가 URL 아닌 경우도 방어
    # (필요하면 더 확장 가능)
    if not (u.startswith("http://") or u.startswith("https://")):
        return ""

    # 마지막 슬래시 통일 등(선택)
    return u


def _get_domain(u: str) -> str:
    try:
        if not u:
            return ""
        parsed = urllib.parse.urlparse(u)
        host = (parsed.netloc or "").lower()
        host = host.split("@")[-1].split(":")[0]
        if host.startswith("www."):
            host = host[4:]
        return host
    except Exception:
        return ""

def _pick_sites_sheet(xls: pd.ExcelFile) -> str:
    """
    사용자가 올린 sites.xlsx는 'SiteList'가 관찰됨.
    그래도 안정성을 위해 name/url 컬럼이 있는 시트를 자동 탐지.
    """
    preferred = ["SiteList", "sites", "site_list", "Sheet1", "sheet1"]
    for p in preferred:
        for sh in xls.sheet_names:
            if sh == p:
                return sh

    for sh in xls.sheet_names:
        df0 = xls.parse(sh)
        cols = [str(c).strip().lower() for c in df0.columns]
        if "name" in cols and "url" in cols:
            return sh
    return xls.sheet_names[0]

def load_sites_xlsx(path):
    import pandas as pd
    df = pd.read_excel(path, sheet_name=0, dtype=str)  # ⭐ dtype=str 로 강제

    # 컬럼명 표준화(대소문자/공백)
    df.columns = [c.strip().lower() for c in df.columns]

    # 기대 컬럼: name, url
    if "name" not in df.columns or "url" not in df.columns:
        raise ValueError("sites.xlsx에 'name', 'url' 컬럼이 필요합니다.")

    df["name"] = df["name"].fillna("").astype(str).str.strip()
    df["url"]  = df["url"].fillna("").astype(str).apply(_normalize_url)

    # url 비어있으면 제거 (⭐ 여기서 float/NaN 문제 완전 차단)
    df = df[(df["name"] != "") & (df["url"] != "")].copy()

    domain_to_name = {}
    allowed_domains = set()
    for _, r in df.iterrows():
        # 도메인 추출(예: https://www.customs.go.kr/... -> customs.go.kr)
        from urllib.parse import urlparse
        host = urlparse(r["url"]).netloc.lower()
        host = host.replace("www.", "")
        domain_to_name[host] = r["name"]
        allowed_domains.add(host)

    return domain_to_name, allowed_domains


# ===============================
# MULTI-LANG QUERY EXPANSION (deterministic)
# ===============================
KW_TRANSLATIONS = {
    "관세": {"en": ["tariff", "customs duty"], "fr": ["droit de douane", "tarif douanier"]},
    "원산지": {"en": ["rules of origin", "origin"], "fr": ["règles d'origine", "origine"]},
    "fta": {"en": ["FTA", "free trade agreement"], "fr": ["accord de libre-échange"]},
    "세관": {"en": ["customs"], "fr": ["douanes"]},
    "수출통제": {"en": ["export control"], "fr": ["contrôle des exportations"]},
    "제재": {"en": ["sanction"], "fr": ["sanction"]},
}

def expand_query(q: str, lang: str) -> list:
    if not NEWS_QUERY_EXPAND:
        return [q]
    key = q.strip().lower()
    for k, v in KW_TRANSLATIONS.items():
        if k.lower() == key:
            base = [q.strip()]
            for t in v.get(lang, []):
                if t and t.lower() not in {x.lower() for x in base}:
                    base.append(t)
            return base
    return [q.strip()]

# ===============================
# URL / RSS helpers
# ===============================
def unwrap_google_news_url(u: str) -> str:
    """
    Google News RSS는 redirect 링크를 주는 경우가 있어 url= 파라미터가 있으면 복원.
    """
    if not u:
        return u
    try:
        parsed = urllib.parse.urlparse(u)
        qs = urllib.parse.parse_qs(parsed.query)
        if "url" in qs and qs["url"]:
            return qs["url"][0]
        return u
    except Exception:
        return u

def google_news_rss(q: str, lang: str) -> str:
    return "https://news.google.com/rss/search?" + urllib.parse.urlencode({
        "q": q,
        "hl": lang,
        "gl": NEWS_REGION,
        "ceid": f"{NEWS_REGION}:{lang}",
    })

# ===============================
# SCORE / IMPORTANCE / FILTER
# ===============================
RISK_RULES = [
    ("section 301", 6),
    ("section 232", 6),
    ("international emergency economic powers act", 6),
    ("ieepa", 6),
    ("export control", 6),
    ("sanction", 6),
    ("entity list", 5),
    ("anti-dumping", 5),
    ("antidumping", 5),
    ("countervailing", 5),
    ("safeguard", 5),
    ("tariff act", 4),
    ("trade expansion act", 4),
    ("tariff", 4),
    ("duty", 4),
    ("customs duty", 4),
    ("관세", 4),
    ("관세율", 4),
    ("추가관세", 4),
    ("hs code", 3),
    ("hs", 3),
    ("원산지", 3),
    ("rules of origin", 3),
    ("fta", 3),
    ("customs", 3),
    ("통관", 3),
    ("규정", 2),
    ("시행", 2),
    ("개정", 2),
    ("고시", 2),
]

MUST_SHOW = [
    "관세", "관세율", "tariff", "customs duty",
    "tariff act", "trade expansion act",
    "international emergency economic powers act", "ieepa",
    "section 232", "section 301",
    "hs", "hs code"
]

BLOCK = ["시위", "protest", "체포", "arrest", "충돌", "violent", "immigration", "ice raid", "연방정부", "주정부"]

def calc_policy_score(title: str, summary: str) -> int:
    t = f"{title} {summary}".lower()
    score = 1
    for kw, w in RISK_RULES:
        if kw in t:
            score += w
    return min(score, 20)

def importance_label(score: int) -> str:
    if score >= 12:
        return "상"
    if score >= 7:
        return "중"
    return "하"

def is_trade_relevant(title: str, summary: str) -> bool:
    blob = f"{title} {summary}".lower()
    if any(b in blob for b in BLOCK):
        return False
    return any(m.lower() in blob for m in MUST_SHOW) or any(k in blob for k, _ in RISK_RULES)

# ===============================
# COUNTRY TAG (light heuristic)
# ===============================
COUNTRY_KEYWORDS = {
    "USA": ["u.s.", "united states", "america", "section 301", "section 232", "u.s "],
    "India": ["india"],
    "Türkiye": ["turkey", "türkiye"],
    "Vietnam": ["vietnam"],
    "Netherlands": ["netherlands", "dutch"],
    "EU": ["european union", "eu commission", "european commission"],
    "China": ["china"],
    "Mexico": ["mexico"],
    "Brazil": ["brazil"],
    "Korea": ["korea", "korean", "republic of korea"],
    "Japan": ["japan"],
}

def detect_country(text: str) -> str:
    t = (text or "").lower()
    for country, keys in COUNTRY_KEYWORDS.items():
        if any(k in t for k in keys):
            return country
    return ""

# ===============================
# SENSOR
# ===============================
def run_sensor(queries: list, allowed_domains: set, domain_to_name: dict, window_start: dt.datetime, window_end: dt.datetime) -> pd.DataFrame:
    rows = []

    for base_q in queries:
        for lang in NEWS_LANGS:
            for q in expand_query(base_q, lang):
                rss = google_news_rss(q, lang)
                feed = feedparser.parse(rss)

                for e in feed.entries[:NEWS_MAX_PER_QUERY]:
                    title = getattr(e, "title", "").strip()
                    link = unwrap_google_news_url(getattr(e, "link", "").strip())
                    published = getattr(e, "published", "") or getattr(e, "updated", "") or ""
                    summary = getattr(e, "summary", "") or getattr(e, "description", "") or ""
                    summary = re.sub(r"<[^>]+>", "", summary).strip()

                    pub_kst = parse_pub_kst(published)
                    if pub_kst is not None:
                        if not (window_start <= pub_kst < window_end):
                            continue

                    if not is_trade_relevant(title, summary):
                        continue

                    dom = _get_domain(link)
                    agency = domain_to_name.get(dom, "")

                    blob = f"{title} {summary}".lower()
                    must_show = any(m.lower() in blob for m in MUST_SHOW)

                    # Site allow-list: official domains OR must-show keywords
                    if allowed_domains and (dom not in allowed_domains) and (not must_show):
                        continue

                    country = detect_country(f"{title} {summary}")
                    score = calc_policy_score(title, summary)
                    imp = importance_label(score)

                    rows.append({
                        "제시어": base_q,
                        "언어": lang,
                        "헤드라인": title,
                        "주요내용": summary[:700],
                        "발표일": pub_kst.strftime("%Y-%m-%d %H:%M") if pub_kst else "",
                        "대상 국가": country,
                        "관련 기관": agency,
                        "중요도": imp,
                        "점수": score,
                        "URL": link,
                    })

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    if len(df) > TOTAL_MAX_ITEMS:
        df = df.sort_values(["점수"], ascending=False).head(TOTAL_MAX_ITEMS)

    return df

# ===============================
# DEDUP (robust)
# ===============================
def _norm_title(t: str) -> str:
    t = (t or "").lower()
    t = re.sub(r"\s+", " ", t)
    t = re.sub(r"[^\w\s가-힣]", "", t)
    return t.strip()

def dedup_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    df = df.copy()
    df["__url"] = df["URL"].fillna("").astype(str).str.strip()
    df["__title_norm"] = df["헤드라인"].apply(_norm_title)
    df["__domain"] = df["__url"].apply(_get_domain)

    # Key priority: URL if exists else domain+normalized-title
    df["__key"] = df["__url"]
    df.loc[df["__key"] == "", "__key"] = df["__domain"].fillna("") + "|" + df["__title_norm"].fillna("")

    df = df.sort_values(["점수", "발표일"], ascending=[False, False])
    df = df.drop_duplicates(subset=["__key"], keep="first")

    return df.drop(columns=["__url", "__title_norm", "__domain", "__key"], errors="ignore")

# ===============================
# Executive Insight (rule-based)
# ===============================
def exec_insight_top3(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    top3 = df.sort_values("점수", ascending=False).head(3).copy()

    def reason(r):
        blob = f"{r.get('헤드라인','')} {r.get('주요내용','')}".lower()
        hits = [k for k in ["section 301", "section 232", "ieepa", "tariff act", "trade expansion act", "관세율", "추가관세", "hs code", "hs", "원산지", "export control", "sanction"] if k in blob]
        if hits:
            return "핵심 키워드: " + ", ".join(hits[:4])
        return "관세/통상 정책성 이슈 가능"

    def action(r):
        c = r.get("대상 국가", "")
        return f"1) {c or '대상국'}·HS/품목 확인  2) 관세율/적용시점 확인  3) 생산·판매법인 원가/마진 영향 1차 산정"

    top3["당사 연관성(요약)"] = top3.apply(reason, axis=1)
    top3["권고 Action"] = top3.apply(action, axis=1)
    return top3

# ===============================
# HTML BUILD
# - 요청사항:
#   - '출처' 칼럼 삭제
#   - 헤드라인에 링크
#   - '헤드라인/주요내용'을 한 칸에 표기
#   - 실무자 표에서는 '비고' 칼럼 제외
#   - 임원 TOP3 내용은 실무자 메일에도 포함
# ===============================
STYLE = """
<style>
body{font-family:Malgun Gothic,Arial; background:#f6f6f6;}
.page{max-width:1120px;margin:auto;background:white;padding:14px;}
h2{margin:0 0 6px 0;}
.box{border:1px solid #ddd;border-radius:10px;padding:12px;margin:12px 0;}
table{border-collapse:collapse;width:100%;}
th,td{border:1px solid #ccc;padding:8px;font-size:12px;vertical-align:top;}
th{background:#f0f0f0;}
.small{font-size:11px;color:#555;}
</style>
"""

def _esc(x) -> str:
    return html.escape("" if x is None else str(x))

def build_html_practitioner(df: pd.DataFrame) -> str:
    date = now_kst().strftime("%Y-%m-%d")
    top3 = exec_insight_top3(df)

    # Executive box
    exec_items = ""
    for _, r in top3.iterrows():
        exec_items += f"""
        <li>
          <b>[{_esc(r.get('대상 국가',''))} | {_esc(r.get('중요도',''))} | 점수 {_esc(r.get('점수',''))}]</b><br/>
          <a href="{_esc(r.get('URL',''))}" target="_blank">{_esc(r.get('헤드라인',''))}</a><br/>
          <div class="small">{_esc(r.get('당사 연관성(요약)',''))}</div>
          <div class="small"><b>Action:</b> {_esc(r.get('권고 Action',''))}</div>
        </li>
        """

    # Table rows
    rows = ""
    for _, r in df.iterrows():
        headline = _esc(r.get("헤드라인", ""))
        url = _esc(r.get("URL", ""))
        summary = _esc(r.get("주요내용", ""))

        cell = f'<a href="{url}" target="_blank">{headline}</a><br/><div class="small">{summary}</div>'

        rows += f"""
        <tr>
          <td>{_esc(r.get('제시어',''))}</td>
          <td>{cell}</td>
          <td>{_esc(r.get('발표일',''))}</td>
          <td>{_esc(r.get('대상 국가',''))}</td>
          <td>{_esc(r.get('관련 기관',''))}</td>
          <td>{_esc(r.get('중요도',''))}</td>
        </tr>
        """

    return f"""
    <html><head>{STYLE}</head><body>
    <div class="page">
      <h2>관세·통상 정책 센서 (실무) ({date})</h2>

      <div class="box">
        <h3 style="margin:0 0 8px 0;">Executive Insight TOP3 (동일 내용 실무 공유)</h3>
        <ul style="margin:0; padding-left:18px;">{exec_items or "<li>TOP3 후보 없음</li>"}</ul>
      </div>

      <div class="box">
        <h3 style="margin:0 0 8px 0;">② 정책 이벤트 표</h3>
        <table>
          <tr>
            <th>제시어</th>
            <th>헤드라인 / 주요내용</th>
            <th>발표일</th>
            <th>대상 국가</th>
            <th>관련 기관</th>
            <th>중요도</th>
          </tr>
          {rows}
        </table>
      </div>
    </div>
    </body></html>
    """

def build_html_exec(df: pd.DataFrame) -> str:
    date = now_kst().strftime("%Y-%m-%d")
    top3 = exec_insight_top3(df)

    items = ""
    for _, r in top3.iterrows():
        items += f"""
        <li>
          <b>[{_esc(r.get('대상 국가',''))} | {_esc(r.get('중요도',''))} | 점수 {_esc(r.get('점수',''))}]</b><br/>
          <a href="{_esc(r.get('URL',''))}" target="_blank">{_esc(r.get('헤드라인',''))}</a><br/>
          <div class="small">{_esc(r.get('당사 연관성(요약)',''))}</div>
          <div class="small"><b>Action:</b> {_esc(r.get('권고 Action',''))}</div>
        </li>
        """

    return f"""
    <html><head>{STYLE}</head><body>
    <div class="page">
      <h2>[Executive] 관세·통상 핵심 TOP3 ({date})</h2>
      <div class="box">
        <ul style="margin:0; padding-left:18px;">{items or "<li>TOP3 후보 없음</li>"}</ul>
      </div>
    </div>
    </body></html>
    """

# ===============================
# OUTPUTS
# ===============================
def write_outputs(df: pd.DataFrame, html_body: str):
    today = now_kst().strftime("%Y-%m-%d")
    csv_path = os.path.join(BASE_DIR, f"policy_events_{today}.csv")
    xlsx_path = os.path.join(BASE_DIR, f"policy_events_{today}.xlsx")
    html_path = os.path.join(BASE_DIR, f"policy_events_{today}.html")

    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    df.to_excel(xlsx_path, index=False)
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_body)

    return csv_path, xlsx_path, html_path

# ===============================
# MAIL
# ===============================
def send_mail_to(recipients, subject, html_body):
    if not recipients:
        print(f"[WARN] No recipients: {subject}")
        return

    if not SMTP_SERVER or not SMTP_EMAIL or not SMTP_PASSWORD:
        raise ValueError("SMTP env missing (SMTP_SERVER/SMTP_EMAIL/SMTP_PASSWORD).")

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = SMTP_EMAIL
    msg["To"] = ", ".join(recipients)
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as s:
        s.starttls()
        s.login(SMTP_EMAIL, SMTP_PASSWORD)
        s.sendmail(SMTP_EMAIL, recipients, msg.as_string())

# ===============================
# MAIN
# ===============================
def main():
    queries = load_custom_queries(QUERIES_FILE)
    domain_to_name, allowed_domains, sites_df = load_sites_xlsx_strict(SITES_FILE)

    print("[DEBUG] sites loaded:", len(sites_df))
    # 필요 시 sites_df를 out/에 저장해서 검증 가능
    # sites_df.to_csv(os.path.join(BASE_DIR, "sites_cleaned.csv"), index=False, encoding="utf-8-sig")


    ref = now_kst()
    w_start, w_end = collection_window_kst(ref)
    print(f"[INFO] Window(KST): {w_start} ~ {w_end}")
    print(f"[INFO] Queries: {len(queries)} items")
    print(f"[INFO] Allowed domains: {len(allowed_domains)}")

    df = run_sensor(queries, allowed_domains, domain_to_name, w_start, w_end)
    if df is None or df.empty:
        print("[INFO] No items collected.")
        return

    df = dedup_df(df)
    df = df.sort_values(["점수", "발표일"], ascending=[False, False]).reset_index(drop=True)

    html_prac = build_html_practitioner(df)
    write_outputs(df, html_prac)

    today = now_kst().strftime("%Y-%m-%d")
    send_mail_to(RECIPIENTS, f"관세·통상 정책 센서 (실무) ({today})", html_prac)

    html_exec = build_html_exec(df)
    send_mail_to(RECIPIENTS_EXEC, f"[Executive] 관세·통상 핵심 TOP3 ({today})", html_exec)

    print("[OK] vCurrent STABLE completed.")
    print("BASE_DIR =", BASE_DIR)
    print("OUT_FILES =", os.listdir(BASE_DIR))

if __name__ == "__main__":
    main()
from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Dict, Set, Tuple, Optional
from urllib.parse import urlparse

import pandas as pd
from openpyxl import load_workbook


# ===============================
# Helpers
# ===============================
def _safe_str(x) -> str:
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass
    return str(x).strip()


def _normalize_url(u) -> str:
    u = _safe_str(u)
    if not u or u.lower() in ("nan", "none"):
        return ""
    if not (u.startswith("http://") or u.startswith("https://")):
        return ""
    return u


def _extract_host(url: str) -> str:
    try:
        host = urlparse(url).netloc.lower()
    except Exception:
        return ""
    host = host.replace("www.", "").strip()
    return host


def _looks_like_header(v: str) -> str:
    # 컬럼명 유사치 표준화
    v = _safe_str(v).lower()
    v = re.sub(r"\s+", "", v)
    return v


# ===============================
# Strict Loader
# ===============================
@dataclass
class SitesXlsxConfig:
    # 시트명 우선순위 (없으면 첫 시트 사용)
    preferred_sheets: Tuple[str, ...] = ("SiteList", "Sitelist", "sites", "Sites", "Sheet1")
    # 헤더 후보 (name/url 컬럼 인식)
    name_headers: Tuple[str, ...] = ("name", "기관명", "기관", "사이트명", "source", "source_name")
    url_headers: Tuple[str, ...] = ("url", "link", "주소", "사이트", "homepage", "source_url")


def load_sites_xlsx_strict(
    xlsx_path: str,
    cfg: SitesXlsxConfig = SitesXlsxConfig(),
) -> Tuple[Dict[str, str], Set[str], pd.DataFrame]:
    """
    sites.xlsx에서:
      - name / url 컬럼을 '헤더 기준'으로 타이트하게 찾고
      - URL은 "셀 하이퍼링크"를 최우선으로 사용 (표시 텍스트가 URL이 아니어도 OK)
      - http/https 아닌 값/빈칸/NaN은 제거
      - 결과:
          domain_to_name: {domain: name}
          allowed_domains: set(domains)
          cleaned_df: 정제된 DataFrame(name, url, domain)
    """

    wb = load_workbook(xlsx_path, data_only=True)

    # 1) 시트 선택
    sheet = None
    for s in cfg.preferred_sheets:
        if s in wb.sheetnames:
            sheet = wb[s]
            break
    if sheet is None:
        sheet = wb[wb.sheetnames[0]]

    # 2) 헤더 행 찾기: 1~30행에서 name/url 둘 다 발견되는 첫 행 사용
    name_idx = None
    url_idx = None
    header_row = None

    for r in range(1, min(sheet.max_row, 30) + 1):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, min(sheet.max_column, 50) + 1)]
        norm = [_looks_like_header(v) for v in row_vals]

        # name/url 후보 인덱스 찾기
        tmp_name = None
        tmp_url = None

        for i, h in enumerate(norm):
            if any(h == _looks_like_header(x) for x in cfg.name_headers):
                tmp_name = i + 1
            if any(h == _looks_like_header(x) for x in cfg.url_headers):
                tmp_url = i + 1

        if tmp_name and tmp_url:
            name_idx, url_idx = tmp_name, tmp_url
            header_row = r
            break

    if not (name_idx and url_idx and header_row):
        raise ValueError(
            f"sites.xlsx에서 헤더를 찾지 못했습니다. "
            f"시트={sheet.title} / 'name'과 'url' 컬럼(헤더)이 필요합니다."
        )

    # 3) 데이터 읽기 (하이퍼링크 우선)
    items = []
    for r in range(header_row + 1, sheet.max_row + 1):
        name_cell = sheet.cell(row=r, column=name_idx)
        url_cell = sheet.cell(row=r, column=url_idx)

        name = _safe_str(name_cell.value)

        # URL은 1) 하이퍼링크 target 2) 셀 value(텍스트)
        url = ""
        if url_cell.hyperlink and url_cell.hyperlink.target:
            url = _safe_str(url_cell.hyperlink.target)
        else:
            url = _safe_str(url_cell.value)

        url = _normalize_url(url)

        if not name or not url:
            continue

        domain = _extract_host(url)
        if not domain:
            continue

        items.append({"name": name, "url": url, "domain": domain, "sheet": sheet.title, "row": r})

    cleaned_df = pd.DataFrame(items)

    # 4) 도메인 맵 구성
    domain_to_name: Dict[str, str] = {}
    allowed_domains: Set[str] = set()

    for _, r in cleaned_df.iterrows():
        d = r["domain"]
        n = r["name"]
        # 동일 도메인 중복이면 최초 1개 유지(원하면 name merge로 바꿀 수 있음)
        if d not in domain_to_name:
            domain_to_name[d] = n
        allowed_domains.add(d)

    return domain_to_name, allowed_domains, cleaned_df
